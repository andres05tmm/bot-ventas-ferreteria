"""
Google Sheets: pizarra del dia en tiempo real.
Columnas: #, Fecha, Hora, Producto, Cantidad, Precio Unitario, Total, Vendedor, Metodo Pago
"""

import time
from datetime import datetime

import config
from utils import decimal_a_fraccion_legible

# ── Cache del objeto worksheet ────────────────────────────────────────────────
# Evita abrir la conexión + buscar la pestaña en CADA llamada (3 reads evitados
# por operación). TTL de 5 minutos; se invalida en errores 429 o excepciones.
_ws_cache: object = None          # objeto gspread.Worksheet
_ws_cache_ts: float = 0.0         # timestamp del último open exitoso
_WS_CACHE_TTL: float = 300.0      # segundos (5 min, igual que el cache de Claude)


def _invalidar_ws_cache() -> None:
    global _ws_cache, _ws_cache_ts
    _ws_cache = None
    _ws_cache_ts = 0.0


def _col_a_letra(n: int) -> str:
    """
    Convierte número de columna a letra(s): 1→A, 26→Z, 27→AA, 28→AB, etc.
    FIX: Soporta más de 26 columnas (antes fallaba con chr(ord('A') + n - 1))
    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _formato_encabezado(ws) -> None:
    """Aplica formato azul al encabezado. Solo se llama al crear la hoja."""
    num_cols  = len(config.SHEETS_HEADERS)
    col_letra = _col_a_letra(num_cols)
    ws.format(f"A1:{col_letra}1", {
        "backgroundColor": {"red": 0.102, "green": 0.337, "blue": 0.855},
        "textFormat": {
            "bold": True,
            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
        },
        "horizontalAlignment": "CENTER",
    })


def _obtener_hoja_sheets():
    """
    Retorna la worksheet 'Ventas del Dia'.
    - Cachea el objeto ws durante 5 min para evitar múltiples reads por operación.
    - El chequeo de encabezados solo ocurre al crear la hoja por primera vez,
      no en cada llamada (evitaba 1 read extra por operación → 429 en pico).
    - Auto-migra: si falta la columna UNIDAD DE MEDIDA, la inserta en su posición.
    - Retorna None si no hay conexión.
    """
    global _ws_cache, _ws_cache_ts
    if not config.SHEETS_ID:
        return None

    # Devolver caché si sigue vigente
    if _ws_cache is not None and (time.time() - _ws_cache_ts) < _WS_CACHE_TTL:
        return _ws_cache

    try:
        import gspread
        gc          = config.get_sheets_client()
        spreadsheet = gc.open_by_key(config.SHEETS_ID)
        creada_nueva = False
        try:
            ws = spreadsheet.worksheet("Ventas del Dia")
        except gspread.WorksheetNotFound:
            # Primera vez: crear hoja con encabezados y formato
            ws = spreadsheet.add_worksheet(
                "Ventas del Dia", rows=500, cols=len(config.SHEETS_HEADERS)
            )
            ws.append_row(config.SHEETS_HEADERS)
            _formato_encabezado(ws)
            creada_nueva = True

        # ── Migración: insertar UNIDAD DE MEDIDA antes de CANTIDAD si falta ──
        if not creada_nueva:
            try:
                headers_actuales = ws.row_values(1)
                headers_upper = [h.upper().strip() for h in headers_actuales]
                if "UNIDAD DE MEDIDA" not in headers_upper:
                    # Buscar posición de CANTIDAD para insertar antes
                    try:
                        idx_cant = headers_upper.index("CANTIDAD")
                        col_pos = idx_cant + 1  # gspread es 1-indexed
                    except ValueError:
                        col_pos = len(headers_actuales) + 1  # fallback: al final
                    ws.insert_cols(col_pos, number=1)
                    ws.update_cell(1, col_pos, "UNIDAD DE MEDIDA")
                    _formato_encabezado(ws)
                    _invalidar_ws_cache()
                    print(f"✅ Migración Sheets: UNIDAD DE MEDIDA insertada en col {col_pos} (antes de CANTIDAD)")
            except Exception as e_mig:
                print(f"⚠️ Migración Sheets (no crítico): {e_mig}")

        # Guardar en caché
        _ws_cache    = ws
        _ws_cache_ts = time.time()
        config._set_sheets_disponible(True)
        return ws

    except Exception as e:
        print(f"⚠️ Error accediendo a Sheets: {e}")
        _invalidar_ws_cache()
        config._set_sheets_disponible(False)
        config.reset_google_clients()
        return None


def sheets_agregar_venta(num, producto, cantidad, precio_unitario, total, vendedor, metodo,
                          id_cliente="CF", nombre_cliente="Consumidor Final", codigo_producto="",
                          unidad_medida="") -> bool:
    """Agrega una fila de venta al Google Sheets en tiempo real."""
    if not config.SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False

        fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        hora  = datetime.now(config.COLOMBIA_TZ).strftime("%H:%M")

        # Solo mostrar fracciones para galones y unidades genéricas; el resto en decimal
        _um = (unidad_medida or "").lower().replace("ó", "o")
        _usar_fraccion = _um in ("galon", "galón", "unidad", "")
        if _usar_fraccion and not isinstance(cantidad, str):
            cantidad_legible = decimal_a_fraccion_legible(float(cantidad))
        else:
            cantidad_legible = str(cantidad)
        fila = [
            num,                    # CONSECUTIVO DE VENTA
            fecha,                  # FECHA
            hora,                   # HORA
            str(id_cliente),        # ID CLIENTE
            str(nombre_cliente),    # CLIENTE
            str(codigo_producto),   # CODIGO DEL PRODUCTO
            str(producto),          # PRODUCTO
            str(unidad_medida or "Unidad"),  # UNIDAD DE MEDIDA
            cantidad_legible,       # CANTIDAD
            float(precio_unitario), # VALOR UNITARIO
            float(total),           # TOTAL
            str(vendedor),          # VENDEDOR
            str(metodo),            # METODO DE PAGO
        ]
        ws.append_row(fila, value_input_option="USER_ENTERED")

        # Alternar color de fila usando row_count (sin read extra)
        # append_row agrega al final → la nueva fila es la última
        try:
            num_filas = ws.row_count
            num_cols  = len(config.SHEETS_HEADERS)
            col_letra = _col_a_letra(num_cols)

            # Si append_row creció más allá del grid, redimensionar
            try:
                if num_filas % 2 == 0:
                    ws.format(f"A{num_filas}:{col_letra}{num_filas}", {
                        "backgroundColor": {"red": 0.937, "green": 0.961, "blue": 1.0},
                        "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}}
                    })
                else:
                    ws.format(f"A{num_filas}:{col_letra}{num_filas}", {
                        "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                        "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}}
                    })
            except Exception as e_fmt:
                # Grid limit exceeded — resize and retry
                if "exceeds grid limits" in str(e_fmt):
                    try:
                        ws.resize(rows=num_filas + 100, cols=num_cols)
                        ws.format(f"A{num_filas}:{col_letra}{num_filas}", {
                            "backgroundColor": {"red": 0.937, "green": 0.961, "blue": 1.0} if num_filas % 2 == 0 else {"red": 1.0, "green": 1.0, "blue": 1.0},
                            "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}}
                        })
                    except Exception:
                        pass  # Formato no es crítico
                else:
                    print(f"⚠️ Error formato fila: {e_fmt}")
        except Exception:
            pass  # Formato no es crítico — la venta ya se guardó

        config._set_sheets_disponible(True)
        return True
    except Exception as e:
        print(f"⚠️ Error agregando al Sheets: {e}")
        _invalidar_ws_cache()
        config._set_sheets_disponible(False)
        config.reset_google_clients()
        return False


def sheets_borrar_fila(numero_venta) -> bool:
    """Borra del Sheets TODAS las filas cuyo consecutivo sea numero_venta."""
    if not config.SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False
        celdas = ws.get_all_values()

        # Recoger índices de todas las filas que coinciden (en orden inverso para no desplazar)
        filas_a_borrar = []
        for idx, fila in enumerate(celdas):
            if idx == 0:
                continue
            try:
                if int(fila[0]) == int(numero_venta):
                    filas_a_borrar.append(idx + 1)  # +1 porque Sheets es 1-indexed
            except (ValueError, IndexError):
                pass

        if not filas_a_borrar:
            return False

        # Borrar de abajo hacia arriba para no desplazar índices
        for fila_idx in reversed(filas_a_borrar):
            ws.delete_rows(fila_idx)

        _invalidar_ws_cache()
        return True

    except Exception as e:
        print(f"⚠️ Error borrando filas del Sheets: {e}")
        _invalidar_ws_cache()
        return False


def sheets_leer_ventas_del_dia() -> list:
    """
    Lee todas las filas de ventas del Sheets (excluyendo encabezado).
    Retorna lista de dicts. Respeta ediciones manuales.
    """
    if not config.SHEETS_ID:
        return []
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return []
        todas = ws.get_all_records()
        resultado = []
        for fila in todas:
            if not any(fila.values()):
                continue
            resultado.append({
                "num":             fila.get("CONSECUTIVO DE VENTA", fila.get("#", "")),
                "fecha":           fila.get("FECHA", fila.get("Fecha", "")),
                "hora":            fila.get("HORA", fila.get("Hora", "")),
                "id_cliente":      fila.get("ID CLIENTE", "CF"),
                "cliente":         fila.get("CLIENTE", "Consumidor Final"),
                "codigo_producto": fila.get("CODIGO DEL PRODUCTO", fila.get("Código del Producto", "")),
                "producto":        fila.get("PRODUCTO", fila.get("Producto", "")),
                "cantidad":        fila.get("CANTIDAD", fila.get("Cantidad", "")),
                "unidad_medida":   fila.get("UNIDAD DE MEDIDA", fila.get("Unidad de Medida", "")) or "Unidad",
                "precio_unitario": fila.get("VALOR UNITARIO", fila.get("Precio Unitario", 0)),
                "total":           fila.get("TOTAL", fila.get("Total", 0)),
                "alias":           fila.get("ALIAS", ""),
                "vendedor":        fila.get("VENDEDOR", fila.get("Vendedor", "")),
                "metodo":          fila.get("METODO DE PAGO", fila.get("Método Pago", "")),
            })
        return resultado
    except Exception as e:
        print(f"⚠️ Error leyendo Sheets: {e}")
        _invalidar_ws_cache()
        return []


def sheets_detectar_ediciones_vs_excel() -> list[str]:
    """
    Compara el Sheets con el Excel local.
    Retorna lista de strings describiendo diferencias (ediciones manuales).
    """
    from excel import inicializar_excel, obtener_nombre_hoja, detectar_columnas
    import openpyxl
    from config import EXCEL_FILE, EXCEL_FILA_DATOS

    ventas_sheets = sheets_leer_ventas_del_dia()
    if not ventas_sheets:
        return []

    hoy = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
    inicializar_excel()

    import openpyxl
    wb           = openpyxl.load_workbook(EXCEL_FILE, read_only=True)  # solo lectura
    nombre_hoja  = obtener_nombre_hoja()
    ventas_excel = {}

    if nombre_hoja in wb.sheetnames:
        ws   = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        col_num   = next((v for k, v in cols.items() if k == "#"), None)
        col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)
        col_prod  = next((v for k, v in cols.items() if "producto" in k), None)
        col_total = next((v for k, v in cols.items() if "total" in k), None)
        for fila in ws.iter_rows(min_row=EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue
            if col_fecha and str(fila[col_fecha - 1])[:10] == hoy:
                n = fila[col_num - 1] if col_num else None
                if n:
                    ventas_excel[int(n)] = {
                        "producto": fila[col_prod  - 1] if col_prod  else "",
                        "total":    fila[col_total - 1] if col_total else 0,
                    }

    diferencias = []
    for v in ventas_sheets:
        try:
            num = int(v["num"])
        except (ValueError, TypeError):
            continue
        if num not in ventas_excel:
            diferencias.append(
                f"  • Venta #{num} ({v['producto']}) esta en el Sheet pero no en el Excel local"
            )
        else:
            prod_xl = str(ventas_excel[num]["producto"]).lower()
            prod_sh = str(v["producto"]).lower()
            if prod_xl != prod_sh:
                diferencias.append(
                    f"  • #{num}: producto cambiado de '{ventas_excel[num]['producto']}' a '{v['producto']}'"
                )
            try:
                if abs(float(ventas_excel[num]["total"]) - float(v["total"])) > 1:
                    diferencias.append(
                        f"  • #{num} ({v['producto']}): total cambiado de "
                        f"${float(ventas_excel[num]['total']):,.0f} a ${float(v['total']):,.0f}"
                    )
            except (ValueError, TypeError):
                pass
    return diferencias


def sheets_limpiar() -> bool:
    """Limpia todas las ventas del Sheets (deja solo el encabezado)."""
    if not config.SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False
        num_filas = len(ws.get_all_values())
        if num_filas > 1:
            ws.delete_rows(2, num_filas)
        print("🧹 Google Sheets limpiado para el nuevo dia.")
        return True
    except Exception as e:
        print(f"⚠️ Error limpiando Sheets: {e}")
        _invalidar_ws_cache()
        return False


def sheets_obtener_ventas_por_consecutivo(numero_venta) -> list:
    """
    Retorna todas las filas con ese consecutivo desde Google Sheets.
    Usado por /borrar para encontrar ventas a eliminar.
    """
    if not config.SHEETS_ID:
        return []
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return []
        todas = ws.get_all_records()
        filas = []
        for fila in todas:
            try:
                num = fila.get("CONSECUTIVO DE VENTA", fila.get("#", ""))
                if num and int(float(str(num))) == int(numero_venta):
                    filas.append({
                        "producto": fila.get("PRODUCTO", fila.get("Producto", "?")),
                        "total": fila.get("TOTAL", fila.get("Total", 0)),
                        "fecha": fila.get("FECHA", fila.get("Fecha", "?")),
                        "vendedor": fila.get("VENDEDOR", fila.get("Vendedor", "?")),
                        "cantidad": fila.get("CANTIDAD", fila.get("Cantidad", 1)),
                    })
            except (ValueError, TypeError):
                pass
        return filas
    except Exception as e:
        print(f"⚠️ Error buscando consecutivo en Sheets: {e}")
        return []


def sheets_borrar_consecutivo(numero_venta) -> tuple[int, list]:
    """
    Borra TODAS las filas con ese consecutivo del Sheets.
    Retorna (cantidad_borradas, lista_productos_borrados).
    """
    if not config.SHEETS_ID:
        return 0, []
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return 0, []
        
        celdas = ws.get_all_values()
        filas_a_borrar = []  # índices de filas a borrar (1-indexed)
        productos_borrados = []
        
        for idx, fila in enumerate(celdas):
            if idx == 0:  # saltar encabezado
                continue
            try:
                if int(float(str(fila[0]))) == int(numero_venta):
                    filas_a_borrar.append(idx + 1)  # +1 porque Sheets es 1-indexed
                    productos_borrados.append(fila[6] if len(fila) > 6 else "?")  # columna PRODUCTO
            except (ValueError, IndexError, TypeError):
                pass
        
        # Borrar de abajo hacia arriba para no afectar los índices
        for fila_idx in reversed(filas_a_borrar):
            ws.delete_rows(fila_idx)
        
        return len(filas_a_borrar), productos_borrados
    except Exception as e:
        print(f"⚠️ Error borrando consecutivo del Sheets: {e}")
        _invalidar_ws_cache()
        return 0, []


def sheets_editar_consecutivo(numero_venta: int, cambios: dict, producto_original: str = None) -> int:
    """
    Actualiza en Sheets los campos indicados en `cambios` para las filas
    cuyo CONSECUTIVO DE VENTA == numero_venta.
    Si producto_original viene, solo actualiza la fila con ese producto (multi-producto).
    Retorna el número de filas actualizadas.
    """
    if not config.SHEETS_ID or not cambios:
        return 0
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return 0

        todas = ws.get_all_values()
        if not todas:
            return 0

        # Fila 0 = encabezados
        headers = [h.upper().strip() for h in todas[0]]

        # Mapeo campo → nombre de columna en Sheets
        CAMPO_HEADER = {
            "producto":        ["PRODUCTO"],
            "cantidad":        ["CANTIDAD"],
            "precio_unitario": ["VALOR UNITARIO", "PRECIO UNITARIO"],
            "total":           ["TOTAL"],
            "metodo_pago":     ["METODO DE PAGO", "MÉTODO PAGO"],
            "cliente":         ["CLIENTE"],
            "vendedor":        ["VENDEDOR"],
        }

        # Columna del consecutivo y del producto
        col_consec = None
        col_prod_sh = None
        for i, h in enumerate(headers):
            if "CONSECUTIVO" in h or h == "#":
                col_consec = i
            if h == "PRODUCTO":
                col_prod_sh = i
        if col_consec is None:
            return 0

        filtro_prod = producto_original.strip().lower() if producto_original else None

        # Resolver índices de columna para cada campo a cambiar
        col_map = {}
        for campo, valor in cambios.items():
            posibles = CAMPO_HEADER.get(campo, [campo.upper().replace("_", " ")])
            for nombre in posibles:
                for i, h in enumerate(headers):
                    if h == nombre:
                        col_map[campo] = i
                        break
                if campo in col_map:
                    break

        actualizadas = 0
        for idx, fila in enumerate(todas[1:], start=2):   # start=2 → fila real en Sheets
            if not fila:
                continue
            try:
                consec_fila = int(float(str(fila[col_consec]).strip()))
            except (ValueError, IndexError):
                continue
            if consec_fila != int(numero_venta):
                continue

            # Filtrar por producto si se especificó
            if filtro_prod and col_prod_sh is not None:
                prod_fila = str(fila[col_prod_sh]).strip().lower()
                if prod_fila != filtro_prod:
                    continue

            for campo, col_idx in col_map.items():
                valor = cambios[campo]
                # Sheets usa notación fila/col 1-indexed
                ws.update_cell(idx, col_idx + 1, valor)
                actualizadas += 1

        _invalidar_ws_cache()
        return actualizadas
    except Exception as e:
        print(f"⚠️ Error editando Sheets consecutivo #{numero_venta}: {e}")
        _invalidar_ws_cache()
        return 0


def sheets_sincronizar_clientes() -> tuple[bool, str]:
    """
    Copia la hoja 'clientes' del Excel al Sheets, sobreescribiendo siempre
    la misma pestaña 'Clientes'. Retorna (ok, mensaje).
    """
    if not config.SHEETS_ID:
        return False, "Sheets no configurado."
    try:
        import openpyxl, gspread
        from excel import inicializar_excel
        inicializar_excel()
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
        if "clientes" not in [s.lower() for s in wb.sheetnames]:
            wb.close()
            return False, "No encontré la hoja 'clientes' en el Excel."
        # Buscar hoja con nombre exacto (puede ser 'Clientes' o 'clientes')
        nombre_hoja = next(s for s in wb.sheetnames if s.lower() == "clientes")
        ws_excel = wb[nombre_hoja]
        filas = [list(row) for row in ws_excel.iter_rows(values_only=True)]
        wb.close()
        # Limpiar filas vacías al final
        while filas and all(c is None or str(c).strip() == "" for c in filas[-1]):
            filas.pop()
        if not filas:
            return False, "La hoja 'clientes' está vacía."

        gc          = config.get_sheets_client()
        spreadsheet = gc.open_by_key(config.SHEETS_ID)

        # Obtener o crear la pestaña
        try:
            ws_sheets = spreadsheet.worksheet("Clientes")
            ws_sheets.clear()
        except gspread.WorksheetNotFound:
            ws_sheets = spreadsheet.add_worksheet("Clientes", rows=max(500, len(filas)+10), cols=20)

        # Convertir todo a string para Sheets
        datos = [[str(c) if c is not None else "" for c in fila] for fila in filas]
        ws_sheets.update(datos, "A1")

        # Formato encabezado (primera fila)
        if datos:
            num_cols  = len(datos[0])
            col_letra = _col_a_letra(num_cols)
            ws_sheets.format(f"A1:{col_letra}1", {
                "backgroundColor": {"red": 0.102, "green": 0.337, "blue": 0.855},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "horizontalAlignment": "CENTER",
            })

        config._set_sheets_disponible(True)
        url = f"https://docs.google.com/spreadsheets/d/{config.SHEETS_ID}/edit#gid={ws_sheets.id}"
        return True, url

    except Exception as e:
        print(f"Error sincronizando clientes: {e}")
        config._set_sheets_disponible(False)
        return False, f"Error al sincronizar: {e}"
