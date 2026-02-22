"""
Bot Inteligente de Ventas para Telegram con Claude AI
- Conversacion natural con fracciones
- Catalogo completo de 570 productos con precios por fraccion
- Google Sheets en tiempo real (pizarra del dia)
- Excel al cierre del dia (respeta ediciones manuales del Sheet)
- Memoria de precios y catalogo
- Excel por hojas mensuales acumulado
- Guardado permanente en Google Drive
- Generacion de Excel personalizado
- Borrar ventas (con confirmacion)
- Graficas de ventas
- Busqueda de ventas
- Modo offline/fallback
- Webhook (Railway)

CORRECCIONES v6.1:
- [FIX] Importaciones duplicadas de datetime unificadas
- [FIX] gspread.authorize() deprecado -> gspread.service_account_from_dict()
- [FIX] Deteccion de columnas fragil -> matching exacto con _col_para()
- [FIX] Cache en RAM para memoria.json (evita I/O en cada operacion)
- [FIX] Estado global (ventas_pendientes, borrados_pendientes, historiales)
        protegido con threading.Lock para evitar condiciones de carrera
- [FIX] Total incorrecto en botones de pago para fracciones
- [FIX] Archivos PNG de graficas huerfanos -> try/finally garantiza limpieza
- [SEC] exec() con codigo de IA: sandbox restringido sin 'os' ni builtins peligrosos
"""

import os
import io
import json
import re
import tempfile
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from datetime import datetime, date, timezone, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openai
from googleapiclient.discovery import build

COLOMBIA_TZ = timezone(timedelta(hours=-5))

from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from google.oauth2.service_account import Credentials
import gspread

# ============================================================
# CONFIGURACION
# ============================================================
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON")
GOOGLE_FOLDER_ID = os.getenv("GOOGLE_FOLDER_ID")
# ID del Google Sheets del dia (crear uno y pegar el ID aqui como variable de entorno)
# El ID esta en la URL: docs.google.com/spreadsheets/d/ESTE_ES_EL_ID/edit
SHEETS_ID = os.getenv("SHEETS_ID", "")
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "")
WEBHOOK_PORT = int(os.getenv("PORT", "8443"))

claves_requeridas = {
    "TELEGRAM_TOKEN": TELEGRAM_TOKEN,
    "ANTHROPIC_API_KEY": ANTHROPIC_API_KEY,
    "OPENAI_API_KEY": OPENAI_API_KEY,
    "GOOGLE_CREDENTIALS_JSON": GOOGLE_CREDENTIALS_JSON,
    "GOOGLE_FOLDER_ID": GOOGLE_FOLDER_ID,
}
claves_faltantes = [k for k, v in claves_requeridas.items() if not v]
if claves_faltantes:
    print("\n❌ Faltan claves en las variables de entorno:")
    for c in claves_faltantes:
        print(f"   • {c}")
    exit(1)

EXCEL_FILE = "ventas.xlsx"
VERSION = "v6.1-sheets"
MEMORIA_FILE = "memoria.json"

claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
openai_client = openai.OpenAI(api_key=OPENAI_API_KEY)

# Flags globales de disponibilidad
DRIVE_DISPONIBLE = True
SHEETS_DISPONIBLE = bool(SHEETS_ID)


# ============================================================
# GOOGLE DRIVE (con modo offline/fallback)
# ============================================================

def obtener_servicio_drive():
    """Conecta con Google Drive usando las credenciales."""
    credenciales_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    creds = Credentials.from_service_account_info(
        credenciales_dict,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)

def buscar_archivo_en_drive(service, nombre_archivo):
    """Busca un archivo por nombre en la carpeta de Drive."""
    query = f"name='{nombre_archivo}' and '{GOOGLE_FOLDER_ID}' in parents and trashed=false"
    resultado = service.files().list(q=query, fields="files(id, name)").execute()
    archivos = resultado.get("files", [])
    return archivos[0]["id"] if archivos else None

def descargar_de_drive(nombre_archivo):
    """Descarga un archivo de Drive al servidor. Retorna False si falla."""
    global DRIVE_DISPONIBLE
    try:
        service = obtener_servicio_drive()
        file_id = buscar_archivo_en_drive(service, nombre_archivo)
        if not file_id:
            return False
        buffer = io.BytesIO()
        request = service.files().get_media(fileId=file_id)
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buffer.seek(0)
        with open(nombre_archivo, "wb") as f:
            f.write(buffer.read())
        DRIVE_DISPONIBLE = True
        return True
    except Exception as e:
        print(f"⚠️ Error descargando de Drive (modo offline activado): {e}")
        DRIVE_DISPONIBLE = False
        return False

def subir_a_drive(nombre_archivo):
    """
    Sube o actualiza un archivo en Drive.
    Si falla, guarda en cola local para reintentar despues.
    """
    global DRIVE_DISPONIBLE
    try:
        service = obtener_servicio_drive()
        file_id = buscar_archivo_en_drive(service, nombre_archivo)

        if nombre_archivo.endswith(".xlsx"):
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif nombre_archivo.endswith(".json"):
            mime_type = "application/json"
        else:
            mime_type = "application/octet-stream"

        with open(nombre_archivo, "rb") as f:
            contenido = f.read()

        media = MediaIoBaseUpload(io.BytesIO(contenido), mimetype=mime_type)

        if file_id:
            service.files().update(fileId=file_id, media_body=media).execute()
        else:
            metadata = {"name": nombre_archivo, "parents": [GOOGLE_FOLDER_ID]}
            service.files().create(body=metadata, media_body=media).execute()

        DRIVE_DISPONIBLE = True

        # Si habia archivos pendientes por subir, intentar subirlos ahora
        _reintentar_pendientes()
        return True
    except Exception as e:
        print(f"⚠️ Error subiendo a Drive: {e}. Guardando en cola local.")
        DRIVE_DISPONIBLE = False
        _encolar_para_subir(nombre_archivo)
        return False

def _encolar_para_subir(nombre_archivo):
    """Guarda en un JSON local los archivos que no se pudieron subir."""
    cola_file = "cola_drive.json"
    cola = []
    if os.path.exists(cola_file):
        try:
            with open(cola_file, "r") as f:
                cola = json.load(f)
        except Exception:
            cola = []
    if nombre_archivo not in cola:
        cola.append(nombre_archivo)
    with open(cola_file, "w") as f:
        json.dump(cola, f)

def _reintentar_pendientes():
    """Cuando Drive vuelve, sube los archivos que quedaron pendientes."""
    cola_file = "cola_drive.json"
    if not os.path.exists(cola_file):
        return
    try:
        with open(cola_file, "r") as f:
            cola = json.load(f)
        subidos = []
        for nombre in cola:
            if os.path.exists(nombre):
                try:
                    service = obtener_servicio_drive()
                    file_id = buscar_archivo_en_drive(service, nombre)
                    if nombre.endswith(".xlsx"):
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif nombre.endswith(".json"):
                        mime_type = "application/json"
                    else:
                        mime_type = "application/octet-stream"
                    with open(nombre, "rb") as f:
                        contenido = f.read()
                    media = MediaIoBaseUpload(io.BytesIO(contenido), mimetype=mime_type)
                    if file_id:
                        service.files().update(fileId=file_id, media_body=media).execute()
                    else:
                        metadata = {"name": nombre, "parents": [GOOGLE_FOLDER_ID]}
                        service.files().create(body=metadata, media_body=media).execute()
                    subidos.append(nombre)
                    print(f"✅ Archivo pendiente subido a Drive: {nombre}")
                except Exception as e:
                    print(f"⚠️ No se pudo subir archivo pendiente {nombre}: {e}")
        cola_restante = [n for n in cola if n not in subidos]
        with open(cola_file, "w") as f:
            json.dump(cola_restante, f)
    except Exception as e:
        print(f"Error procesando cola Drive: {e}")

def sincronizar_archivos():
    """Al iniciar, descarga los archivos de Drive si existen."""
    print("🔄 Sincronizando con Google Drive...")
    ok_excel = descargar_de_drive(EXCEL_FILE)
    ok_mem = descargar_de_drive(MEMORIA_FILE)
    if ok_mem:
        invalidar_cache_memoria()  # Forzar recarga del JSON recien descargado
    if ok_excel or ok_mem:
        print("✅ Sincronizacion completa.")
    else:
        print("⚠️ No se pudo sincronizar con Drive. Trabajando en modo local.")


# ============================================================
# GOOGLE SHEETS — PIZARRA DEL DIA
# ============================================================
# El Sheet funciona como tabla en tiempo real durante el dia.
# Columnas: #, Fecha, Hora, Producto, Cantidad, Precio Unitario, Total, Vendedor, Metodo Pago
# Al cerrar el dia: se leen los datos (respetando ediciones manuales),
# se pasan al Excel acumulado, y el Sheet se limpia para el dia siguiente.

SHEETS_HEADERS = ["#", "Fecha", "Hora", "Producto", "Cantidad", "Precio Unitario", "Total", "Vendedor", "Método Pago"]
SHEETS_COL_COLORS = {
    "header_bg": (26, 86, 219),    # Azul
    "header_fg": (255, 255, 255),  # Blanco
    "row_par":   (239, 246, 255),  # Azul muy claro
}

def _obtener_cliente_sheets():
    """Retorna cliente gspread autenticado con la service account."""
    credenciales_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    # gspread.authorize() fue deprecado en v5+; usamos service_account_from_dict
    return gspread.service_account_from_dict(
        credenciales_dict,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
    )

def _obtener_hoja_sheets():
    """
    Retorna la worksheet 'Ventas del Dia' del Sheets configurado.
    Si no existe la pestana la crea con encabezados.
    """
    global SHEETS_DISPONIBLE
    if not SHEETS_ID:
        return None
    try:
        gc = _obtener_cliente_sheets()
        spreadsheet = gc.open_by_key(SHEETS_ID)
        try:
            ws = spreadsheet.worksheet("Ventas del Dia")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet("Ventas del Dia", rows=500, cols=len(SHEETS_HEADERS))
            # Escribir encabezados
            ws.append_row(SHEETS_HEADERS)
            # Formatear encabezado (negrita + color azul)
            ws.format("A1:I1", {
                "backgroundColor": {"red": 0.102, "green": 0.337, "blue": 0.855},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "horizontalAlignment": "CENTER"
            })
        SHEETS_DISPONIBLE = True
        return ws
    except Exception as e:
        print(f"⚠️ Error accediendo a Sheets: {e}")
        SHEETS_DISPONIBLE = False
        return None

def sheets_agregar_venta(num, producto, cantidad, precio_unitario, total, vendedor, metodo):
    """
    Agrega una fila de venta al Google Sheets en tiempo real.
    Retorna True si se agrego correctamente.
    """
    global SHEETS_DISPONIBLE
    if not SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False
        fecha = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
        hora  = datetime.now(COLOMBIA_TZ).strftime("%H:%M")
        cantidad_legible = decimal_a_fraccion_legible(float(cantidad)) if not isinstance(cantidad, str) else str(cantidad)
        fila = [
            num,
            fecha,
            hora,
            str(producto),
            cantidad_legible,
            float(precio_unitario),
            float(total),
            str(vendedor),
            str(metodo),
        ]
        ws.append_row(fila, value_input_option="USER_ENTERED")

        # Alternar color de fila
        num_filas = len(ws.get_all_values())
        if num_filas % 2 == 0:
            rango = f"A{num_filas}:I{num_filas}"
            ws.format(rango, {
                "backgroundColor": {"red": 0.937, "green": 0.961, "blue": 1.0}
            })
        SHEETS_DISPONIBLE = True
        return True
    except Exception as e:
        print(f"⚠️ Error agregando al Sheets: {e}")
        SHEETS_DISPONIBLE = False
        return False

def sheets_borrar_fila(numero_venta):
    """
    Borra del Sheets la fila cuyo primer campo sea numero_venta.
    Retorna True si encontro y borro la fila.
    """
    if not SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False
        celdas = ws.get_all_values()
        for idx, fila in enumerate(celdas):
            if idx == 0:
                continue  # saltar encabezado
            try:
                if int(fila[0]) == int(numero_venta):
                    ws.delete_rows(idx + 1)  # gspread es 1-indexed
                    return True
            except (ValueError, IndexError):
                pass
        return False
    except Exception as e:
        print(f"⚠️ Error borrando fila del Sheets: {e}")
        return False

def sheets_leer_ventas_del_dia():
    """
    Lee todas las filas de ventas del Sheets actual (excluyendo el encabezado).
    Retorna lista de dicts. Respeta cualquier edicion manual que se haya hecho.
    """
    if not SHEETS_ID:
        return []
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return []
        todas = ws.get_all_records()  # usa la fila 1 como keys automaticamente
        # Normalizar keys a minusculas sin tildes para manejo uniforme
        resultado = []
        for fila in todas:
            if not any(fila.values()):
                continue
            resultado.append({
                "num":             fila.get("#", fila.get("num", "")),
                "fecha":           fila.get("Fecha", fila.get("fecha", "")),
                "hora":            fila.get("Hora", fila.get("hora", "")),
                "producto":        fila.get("Producto", fila.get("producto", "")),
                "cantidad":        fila.get("Cantidad", fila.get("cantidad", "")),
                "precio_unitario": fila.get("Precio Unitario", fila.get("precio_unitario", 0)),
                "total":           fila.get("Total", fila.get("total", 0)),
                "vendedor":        fila.get("Vendedor", fila.get("vendedor", "")),
                "metodo":          fila.get("Método Pago", fila.get("metodo", "")),
            })
        return resultado
    except Exception as e:
        print(f"⚠️ Error leyendo Sheets: {e}")
        return []

def sheets_detectar_ediciones_vs_excel():
    """
    Compara el Sheets con el Excel local y retorna las diferencias.
    Util para que el bot informe si hay correcciones manuales.
    Retorna lista de strings describiendo las diferencias.
    """
    ventas_sheets = sheets_leer_ventas_del_dia()
    if not ventas_sheets:
        return []

    hoy = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    ventas_excel = {}
    if nombre_hoja in wb.sheetnames:
        ws_xl = wb[nombre_hoja]
        cols = detectar_columnas(ws_xl)
        col_num   = next((v for k, v in cols.items() if k == "#"), None)
        col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)
        col_prod  = next((v for k, v in cols.items() if "producto" in k), None)
        col_total = next((v for k, v in cols.items() if "total" in k), None)
        for fila in ws_xl.iter_rows(min_row=2, values_only=True):
            if not any(fila):
                continue
            if col_fecha and str(fila[col_fecha-1])[:10] == hoy:
                n = fila[col_num-1] if col_num else None
                if n:
                    ventas_excel[int(n)] = {
                        "producto": fila[col_prod-1] if col_prod else "",
                        "total":    fila[col_total-1] if col_total else 0,
                    }

    diferencias = []
    for v in ventas_sheets:
        try:
            num = int(v["num"])
        except (ValueError, TypeError):
            continue
        if num not in ventas_excel:
            diferencias.append(f"  • Venta #{num} ({v['producto']}) esta en el Sheet pero no en el Excel local")
        else:
            prod_xl = str(ventas_excel[num]["producto"]).lower()
            prod_sh = str(v["producto"]).lower()
            if prod_xl != prod_sh:
                diferencias.append(f"  • #{num}: producto cambiado de '{ventas_excel[num]['producto']}' a '{v['producto']}'")
            try:
                total_xl = float(ventas_excel[num]["total"])
                total_sh = float(v["total"])
                if abs(total_xl - total_sh) > 1:
                    diferencias.append(f"  • #{num} ({v['producto']}): total cambiado de ${total_xl:,.0f} a ${total_sh:,.0f}")
            except (ValueError, TypeError):
                pass
    return diferencias

def sheets_limpiar():
    """
    Limpia todas las ventas del Sheets (deja solo el encabezado).
    Se llama al cerrar el dia, despues de haber pasado los datos al Excel.
    """
    if not SHEETS_ID:
        return False
    try:
        ws = _obtener_hoja_sheets()
        if not ws:
            return False
        num_filas = len(ws.get_all_values())
        if num_filas > 1:
            ws.delete_rows(2, num_filas)  # Borra desde fila 2 hasta el final
        print("🧹 Google Sheets limpiado para el nuevo dia.")
        return True
    except Exception as e:
        print(f"⚠️ Error limpiando Sheets: {e}")
        return False


# ============================================================
# MEMORIA  (con cache en RAM para evitar lecturas repetidas)
# ============================================================

_memoria_cache = None  # Cache en RAM del archivo memoria.json

def cargar_memoria():
    global _memoria_cache
    if _memoria_cache is not None:
        return _memoria_cache
    if os.path.exists(MEMORIA_FILE):
        with open(MEMORIA_FILE, "r", encoding="utf-8") as f:
            _memoria_cache = json.load(f)
    else:
        _memoria_cache = {
            "precios": {}, "catalogo": {}, "negocio": {},
            "notas": [], "inventario": {}, "gastos": {},
            "caja_actual": {"abierta": False}
        }
    return _memoria_cache

def guardar_memoria(memoria):
    global _memoria_cache
    _memoria_cache = memoria  # Actualizar cache
    with open(MEMORIA_FILE, "w", encoding="utf-8") as f:
        json.dump(memoria, f, ensure_ascii=False, indent=2)
    subir_a_drive(MEMORIA_FILE)

def invalidar_cache_memoria():
    """Fuerza recarga desde disco en la proxima llamada a cargar_memoria()."""
    global _memoria_cache
    _memoria_cache = None

def buscar_producto_en_catalogo(nombre_buscado):
    """
    Busca un producto en el catalogo por nombre (busqueda flexible).
    Retorna el dict del producto o None.
    Ejemplo: 'vinilo t1 amarillo vivo' -> encuentra 'Vinilo Davinci T1 Amarillo Vivo'
    """
    memoria = cargar_memoria()
    catalogo = memoria.get("catalogo", {})
    if not catalogo:
        return None

    nombre_lower = nombre_buscado.strip().lower()

    # 1. Busqueda exacta en nombre_lower
    for cod, prod in catalogo.items():
        if prod.get("nombre_lower") == nombre_lower:
            return prod

    # 2. Busqueda por contencion — el termino esta contenido en el nombre
    candidatos = []
    palabras = nombre_lower.split()
    for cod, prod in catalogo.items():
        nl = prod.get("nombre_lower", "")
        # Contar cuantas palabras del termino aparecen en el nombre
        coincidencias = sum(1 for p in palabras if p in nl)
        if coincidencias == len(palabras):
            candidatos.append((coincidencias, len(nl), prod))
        elif coincidencias >= max(1, len(palabras) - 1):
            candidatos.append((coincidencias, len(nl), prod))

    if candidatos:
        # Preferir el que tiene mas coincidencias y nombre mas corto (mas especifico)
        candidatos.sort(key=lambda x: (-x[0], x[1]))
        return candidatos[0][2]

    return None

def obtener_precio_para_cantidad(nombre_producto, cantidad_decimal):
    """
    Dado un producto y una cantidad decimal (ej: 0.25 para 1/4),
    retorna el precio correcto segun el catalogo de fracciones.
    Si no hay precio especifico para esa fraccion, retorna precio_unidad * cantidad.
    """
    prod = buscar_producto_en_catalogo(nombre_producto)
    if not prod:
        # Fallback: buscar en precios simples
        memoria = cargar_memoria()
        precios = memoria.get("precios", {})
        nl = nombre_producto.strip().lower()
        precio_u = precios.get(nl, 0)
        return round(precio_u * cantidad_decimal), precio_u

    precio_u = prod.get("precio_unidad", 0)
    fracciones = prod.get("precios_fraccion", {})

    # Buscar si hay un precio especifico para esta fraccion
    for frac_texto, frac_data in fracciones.items():
        if abs(frac_data.get("decimal", 0) - cantidad_decimal) < 0.01:
            return frac_data.get("precio", round(precio_u * cantidad_decimal)), precio_u

    # No hay precio especifico: calcular proporcional
    return round(precio_u * cantidad_decimal), precio_u

def obtener_precios_como_texto():
    """Resumen compacto de precios para el system prompt (solo muestra precio unidad)."""
    memoria = cargar_memoria()
    catalogo = memoria.get("catalogo", {})
    precios = memoria.get("precios", {})

    if catalogo:
        # Usar catalogo completo — solo mostrar nombre y precio unidad para no saturar el prompt
        lineas = []
        for cod, prod in catalogo.items():
            tiene_frac = bool(prod.get("precios_fraccion"))
            sufijo = " [fraccionable]" if tiene_frac else ""
            lineas.append(f"- {prod['nombre']}: ${prod['precio_unidad']:,}{sufijo}")
        return "\n".join(lineas)
    elif precios:
        return "\n".join([f"- {p}: ${v:,}" for p, v in precios.items()])
    return "No hay precios guardados aun."

def obtener_info_fraccion_producto(nombre_producto):
    """
    Retorna texto con los precios por fraccion de un producto, para incluir en el prompt.
    Ej: 'Vinilo T1: unidad=$50000 | 3/4=$53333 | 1/2=$52000 | 1/4=$60000 | 1/8=$72000'
    """
    prod = buscar_producto_en_catalogo(nombre_producto)
    if not prod:
        return None
    fracs = prod.get("precios_fraccion", {})
    if not fracs:
        return f"{prod['nombre']}: unidad=${prod['precio_unidad']:,} (no fraccionable)"
    partes = [f"unidad=${prod['precio_unidad']:,}"]
    for frac_texto, fd in fracs.items():
        partes.append(f"{frac_texto}=${fd['precio']:,}")
    return f"{prod['nombre']}: " + " | ".join(partes)


# ============================================================
# EXCEL PRINCIPAL (VENTAS)
# ============================================================

def obtener_nombre_hoja():
    meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
             7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    ahora = datetime.now(COLOMBIA_TZ)
    return f"{meses[ahora.month]} {ahora.year}"

def inicializar_hoja(ws):
    if ws.max_row > 1:
        return
    encabezados = ["#","Fecha","Hora","Producto","Cantidad","Precio Unitario","Total","Vendedor","Observaciones"]
    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor="1A56DB")
        celda.alignment = Alignment(horizontal="center")
    anchos = [6,12,10,25,12,18,14,20,30]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

def obtener_o_crear_hoja(wb, nombre_hoja):
    if nombre_hoja in wb.sheetnames:
        return wb[nombre_hoja]
    ws = wb.create_sheet(title=nombre_hoja)
    inicializar_hoja(ws)
    return ws

def inicializar_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = obtener_nombre_hoja()
        inicializar_hoja(ws)
        wb.save(EXCEL_FILE)
        subir_a_drive(EXCEL_FILE)
        print("✅ Archivo Excel creado y subido a Drive.")

def detectar_columnas(ws):
    """
    Retorna un dict {nombre_encabezado_lower: numero_columna}.
    Matching exacto para evitar asignaciones incorrectas
    (p.ej. '#' no debe coincidir con 'precio #').
    """
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor:
            encabezados[str(valor).lower().strip()] = col
    return encabezados

def _col_para(cols, *claves_posibles):
    """
    Busca la columna cuyo encabezado coincida exactamente con alguna de las
    claves dadas. Si no hay exacta, hace containment solo para claves de mas
    de 1 caracter (evita falsos positivos con '#').
    Retorna el numero de columna o None.
    """
    # 1. Coincidencia exacta
    for clave in claves_posibles:
        if clave in cols:
            return cols[clave]
    # 2. Containment — solo encabezados/claves de mas de 1 caracter
    for clave in claves_posibles:
        for enc, num in cols.items():
            if len(enc) > 1 and len(clave) > 1 and (clave in enc or enc in clave):
                return num
    return None

def guardar_venta_excel(producto, cantidad, precio_unitario, total, vendedor, observaciones=""):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    ws = obtener_o_crear_hoja(wb, nombre_hoja)
    cols = detectar_columnas(ws)
    fila = ws.max_row + 1
    fecha_hoy = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
    hora_ahora = datetime.now(COLOMBIA_TZ).strftime("%H:%M")
    num_venta = fila - 1

    # Mapa seguro: clave canonica -> columna (coincidencia exacta primero)
    mapa = {
        "#":               _col_para(cols, "#", "num", "numero"),
        "fecha":           _col_para(cols, "fecha"),
        "hora":            _col_para(cols, "hora"),
        "producto":        _col_para(cols, "producto"),
        "cantidad":        _col_para(cols, "cantidad"),
        "precio unitario": _col_para(cols, "precio unitario", "precio_unitario", "precio"),
        "total":           _col_para(cols, "total"),
        "vendedor":        _col_para(cols, "vendedor"),
        "observaciones":   _col_para(cols, "observaciones", "metodo", "metodo pago"),
    }
    datos = {
        "#": num_venta, "fecha": fecha_hoy, "hora": hora_ahora,
        "producto": producto, "cantidad": cantidad,
        "precio unitario": precio_unitario,
        "total": total, "vendedor": vendedor, "observaciones": observaciones,
    }

    if any(mapa.values()):
        for clave, num_col in mapa.items():
            if num_col and clave in datos:
                ws.cell(row=fila, column=num_col, value=datos[clave])
    else:
        valores = [num_venta, fecha_hoy, hora_ahora, producto, cantidad, precio_unitario, total, vendedor, observaciones]
        for col, valor in enumerate(valores, 1):
            ws.cell(row=fila, column=col, value=valor)

    if fila % 2 == 0:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

    wb.save(EXCEL_FILE)
    subir_a_drive(EXCEL_FILE)

    # Escribir tambien en Google Sheets en tiempo real
    sheets_agregar_venta(num_venta, producto, cantidad, precio_unitario, total, vendedor, observaciones)

    return num_venta

def borrar_venta_excel(numero_venta):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return False, "No hay ventas este mes."
    ws = wb[nombre_hoja]
    fila_borrar = None
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == numero_venta:
            fila_borrar = fila
            break
    if not fila_borrar:
        return False, f"No encontre la venta #{numero_venta}."
    ws.delete_rows(fila_borrar)
    wb.save(EXCEL_FILE)
    subir_a_drive(EXCEL_FILE)

    # Borrar tambien del Sheets si la venta es de hoy
    sheets_borrar_fila(numero_venta)

    return True, f"✅ Venta #{numero_venta} borrada del Excel y del Sheets."

def obtener_venta_por_numero(numero_venta):
    """Retorna los datos de una venta especifica para mostrar antes de confirmar borrado."""
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None
    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == numero_venta:
            fila_dict = {}
            for nombre_col, num_col in cols.items():
                fila_dict[nombre_col] = ws.cell(row=fila, column=num_col).value
            return fila_dict
    return None

def obtener_ventas_recientes(limite=10):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return []
    ws = wb[nombre_hoja]
    ventas = [fila for fila in ws.iter_rows(min_row=2, values_only=True) if any(fila)]
    return ventas[-limite:]

def buscar_ventas(termino):
    """
    Busca ventas por producto, vendedor o fecha en todas las hojas.
    Retorna lista de diccionarios con los resultados.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    resultados = []
    termino_lower = termino.lower().strip()
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if not any(fila):
                continue
            fila_texto = " ".join(str(v).lower() for v in fila if v is not None)
            if termino_lower in fila_texto:
                fila_dict = {"hoja": nombre_hoja}
                for nombre_col, num_col in cols.items():
                    fila_dict[nombre_col] = fila[num_col - 1]
                resultados.append(fila_dict)
    return resultados

def obtener_todos_los_datos():
    """Obtiene todos los datos de todas las hojas para analisis."""
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    todos = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if any(fila):
                fila_dict = {"hoja": nombre_hoja}
                for nombre_col, num_col in cols.items():
                    fila_dict[nombre_col] = fila[num_col - 1]
                todos.append(fila_dict)
    return todos

def obtener_resumen_ventas():
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None
    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    col_total = next((v for k, v in cols.items() if "total" in k), None)
    total_general = 0
    num_ventas = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue
        num_ventas += 1
        if col_total and fila[col_total - 1]:
            try:
                total_general += float(fila[col_total - 1])
            except Exception:
                pass
    return {"hoja": nombre_hoja, "total": total_general, "num_ventas": num_ventas}


# ============================================================
# GENERADOR DE EXCEL PERSONALIZADO
# ============================================================

def generar_excel_personalizado(titulo, encabezados, filas, nombre_archivo):
    """Genera un Excel personalizado con los datos que Claude analizo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(len(encabezados))}1")
    celda_titulo = ws.cell(row=1, column=1, value=titulo)
    celda_titulo.font = Font(bold=True, color="FFFFFF", size=14)
    celda_titulo.fill = PatternFill("solid", fgColor="1A56DB")
    celda_titulo.alignment = Alignment(horizontal="center")

    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=2, column=col, value=enc)
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor="374151")
        celda.alignment = Alignment(horizontal="center")

    for i, fila in enumerate(filas, 3):
        for col, valor in enumerate(fila, 1):
            celda = ws.cell(row=i, column=col, value=valor)
            if i % 2 == 0:
                celda.fill = PatternFill("solid", fgColor="EFF6FF")

    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    wb.save(nombre_archivo)
    return nombre_archivo


# ============================================================
# GRAFICAS DE VENTAS
# ============================================================

def generar_grafica_ventas_por_dia():
    """
    Genera una grafica de barras con las ventas por dia del mes actual.
    Retorna la ruta del archivo PNG generado.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None

    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)
    col_total = next((v for k, v in cols.items() if "total" in k), None)

    if not col_fecha or not col_total:
        return None

    ventas_por_dia = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue
        fecha = fila[col_fecha - 1]
        total = fila[col_total - 1]
        if fecha and total:
            try:
                fecha_str = str(fecha)[:10]
                ventas_por_dia[fecha_str] = ventas_por_dia.get(fecha_str, 0) + float(total)
            except Exception:
                pass

    if not ventas_por_dia:
        return None

    fechas = sorted(ventas_por_dia.keys())
    totales = [ventas_por_dia[f] for f in fechas]
    etiquetas = [f[-5:] for f in fechas]  # Solo MM-DD

    fig, ax = plt.subplots(figsize=(10, 5))
    colores = ["#1A56DB"] * len(totales)
    bars = ax.bar(etiquetas, totales, color=colores, edgecolor="white", linewidth=0.5)

    ax.set_title(f"Ventas por día — {nombre_hoja}", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Fecha", fontsize=11)
    ax.set_ylabel("Total ($)", fontsize=11)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    ax.set_facecolor("#F9FAFB")
    fig.patch.set_facecolor("#FFFFFF")

    for bar, valor in zip(bars, totales):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(totales) * 0.01,
                f"${valor:,.0f}", ha="center", va="bottom", fontsize=8, color="#374151")

    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    ruta = f"grafica_dias_{datetime.now(COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    return ruta

def generar_grafica_productos():
    """
    Genera una grafica de torta/pie con los productos mas vendidos por total en pesos.
    Retorna la ruta del archivo PNG generado.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None

    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    col_producto = next((v for k, v in cols.items() if "producto" in k), None)
    col_total = next((v for k, v in cols.items() if "total" in k), None)

    if not col_producto or not col_total:
        return None

    ventas_por_producto = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue
        producto = fila[col_producto - 1]
        total = fila[col_total - 1]
        if producto and total:
            try:
                p = str(producto).strip()
                ventas_por_producto[p] = ventas_por_producto.get(p, 0) + float(total)
            except Exception:
                pass

    if not ventas_por_producto:
        return None

    # Top 7 productos y el resto como "Otros"
    sorted_items = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)
    top = sorted_items[:7]
    otros_total = sum(v for _, v in sorted_items[7:])
    if otros_total > 0:
        top.append(("Otros", otros_total))

    etiquetas = [item[0] for item in top]
    valores = [item[1] for item in top]

    colores = ["#1A56DB","#3B82F6","#60A5FA","#93C5FD","#BFDBFE","#DBEAFE","#EFF6FF","#CBD5E1"]

    fig, ax = plt.subplots(figsize=(8, 6))
    wedges, texts, autotexts = ax.pie(
        valores, labels=None, autopct="%1.1f%%",
        colors=colores[:len(valores)], startangle=90,
        wedgeprops={"edgecolor": "white", "linewidth": 1.5}
    )
    for autotext in autotexts:
        autotext.set_fontsize(9)
        autotext.set_color("white")
        autotext.set_fontweight("bold")

    ax.legend(wedges, [f"{e} (${v:,.0f})" for e, v in zip(etiquetas, valores)],
              loc="lower center", bbox_to_anchor=(0.5, -0.15),
              ncol=2, fontsize=8, frameon=False)

    ax.set_title(f"Productos más vendidos — {nombre_hoja}", fontsize=13, fontweight="bold", pad=15)
    plt.tight_layout()

    ruta = f"grafica_productos_{datetime.now(COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    return ruta

def generar_grafica_metodos_pago():
    """
    Genera una grafica de dona con la distribucion por metodo de pago.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None

    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    col_obs = next((v for k, v in cols.items() if "observa" in k), None)
    col_total = next((v for k, v in cols.items() if "total" in k), None)

    if not col_obs or not col_total:
        return None

    metodos = {"efectivo": 0, "transferencia": 0, "datafono": 0, "otro": 0}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue
        obs = str(fila[col_obs - 1] or "").lower()
        total = fila[col_total - 1]
        if not total:
            continue
        try:
            monto = float(total)
        except Exception:
            continue
        if "efectivo" in obs:
            metodos["efectivo"] += monto
        elif "transfer" in obs:
            metodos["transferencia"] += monto
        elif "datafono" in obs or "datáfono" in obs:
            metodos["datafono"] += monto
        else:
            metodos["otro"] += monto

    metodos = {k: v for k, v in metodos.items() if v > 0}
    if not metodos:
        return None

    etiquetas = list(metodos.keys())
    valores = list(metodos.values())
    colores = {"efectivo": "#22C55E", "transferencia": "#3B82F6", "datafono": "#F59E0B", "otro": "#94A3B8"}
    cols_grafica = [colores.get(e, "#94A3B8") for e in etiquetas]

    fig, ax = plt.subplots(figsize=(7, 5))
    wedges, texts, autotexts = ax.pie(
        valores, labels=None, autopct="%1.1f%%",
        colors=cols_grafica, startangle=90,
        wedgeprops={"edgecolor": "white", "linewidth": 2, "width": 0.6}
    )
    for autotext in autotexts:
        autotext.set_fontsize(10)
        autotext.set_fontweight("bold")

    ax.legend(wedges, [f"{e.capitalize()}: ${v:,.0f}" for e, v in zip(etiquetas, valores)],
              loc="lower center", bbox_to_anchor=(0.5, -0.12), ncol=2, fontsize=9, frameon=False)

    ax.set_title(f"Métodos de pago — {nombre_hoja}", fontsize=13, fontweight="bold", pad=15)
    plt.tight_layout()

    ruta = f"grafica_pagos_{datetime.now(COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    return ruta


# ============================================================
# INVENTARIO, CAJA Y GASTOS
# ============================================================

def cargar_inventario():
    memoria = cargar_memoria()
    return memoria.get("inventario", {})

def guardar_inventario(inventario):
    memoria = cargar_memoria()
    memoria["inventario"] = inventario
    guardar_memoria(memoria)

def cargar_caja():
    memoria = cargar_memoria()
    return memoria.get("caja_actual", {
        "abierta": False, "fecha": None, "monto_apertura": 0,
        "efectivo": 0, "transferencias": 0, "datafono": 0
    })

def guardar_caja(caja):
    memoria = cargar_memoria()
    memoria["caja_actual"] = caja
    guardar_memoria(memoria)

def cargar_gastos_hoy():
    memoria = cargar_memoria()
    hoy = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
    gastos = memoria.get("gastos", {})
    return gastos.get(hoy, [])

def guardar_gasto(gasto):
    memoria = cargar_memoria()
    hoy = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
    if "gastos" not in memoria:
        memoria["gastos"] = {}
    if hoy not in memoria["gastos"]:
        memoria["gastos"][hoy] = []
    memoria["gastos"][hoy].append(gasto)
    guardar_memoria(memoria)

def obtener_resumen_caja():
    caja = cargar_caja()
    if not caja.get("abierta"):
        return "La caja no esta abierta hoy."
    resumen = obtener_resumen_ventas()
    total_ventas = resumen["total"] if resumen else 0
    gastos_hoy = cargar_gastos_hoy()
    total_gastos_caja = sum(g["monto"] for g in gastos_hoy if g.get("origen") == "caja")
    efectivo_esperado = caja["monto_apertura"] + caja["efectivo"] - total_gastos_caja
    texto = (
        f"RESUMEN DE CAJA\n"
        f"Apertura: ${caja['monto_apertura']:,.0f}\n"
        f"Ventas efectivo: ${caja['efectivo']:,.0f}\n"
        f"Transferencias: ${caja['transferencias']:,.0f}\n"
        f"Datafono: ${caja['datafono']:,.0f}\n"
        f"Total ventas: ${total_ventas:,.0f}\n"
        f"Gastos de caja: ${total_gastos_caja:,.0f}\n"
        f"Efectivo esperado en caja: ${efectivo_esperado:,.0f}"
    )
    return texto

def verificar_alertas_inventario():
    inventario = cargar_inventario()
    alertas = []
    for producto, datos in inventario.items():
        if isinstance(datos, dict):
            cantidad = datos.get("cantidad", 0)
            minimo = datos.get("minimo", 3)
            if cantidad <= minimo:
                alertas.append(f"⚠️ STOCK BAJO: {producto} — quedan {cantidad} unidades")
    return alertas

def convertir_fraccion_a_decimal(valor):
    """Convierte fracciones como 1/4, 1/2, 3/4 a decimal."""
    if isinstance(valor, (int, float)):
        return float(valor)
    valor = str(valor).strip()
    fracciones = {"1/8": 0.125, "1/4": 0.25, "3/8": 0.375, "1/2": 0.5,
                  "5/8": 0.625, "3/4": 0.75, "7/8": 0.875}
    if valor in fracciones:
        return fracciones[valor]
    if "/" in valor:
        try:
            partes = valor.split("/")
            return float(partes[0]) / float(partes[1])
        except Exception:
            pass
    if " " in valor:
        try:
            partes = valor.split(" ")
            entero = float(partes[0])
            fraccion = convertir_fraccion_a_decimal(partes[1])
            return entero + fraccion
        except Exception:
            pass
    try:
        return float(valor)
    except Exception:
        return 0.0

def decimal_a_fraccion_legible(valor):
    """Convierte 5.75 a '5 y 3/4'."""
    entero = int(valor)
    decimal = valor - entero
    fracciones = {0.125: "1/8", 0.25: "1/4", 0.375: "3/8", 0.5: "1/2",
                  0.625: "5/8", 0.75: "3/4", 0.875: "7/8"}
    fraccion_texto = ""
    for dec, texto in fracciones.items():
        if abs(decimal - dec) < 0.05:
            fraccion_texto = texto
            break
    if entero == 0 and fraccion_texto:
        return fraccion_texto
    elif fraccion_texto:
        return f"{entero} y {fraccion_texto}"
    elif decimal < 0.05:
        return str(entero)
    else:
        return f"{valor:.2f}"


import threading

# ============================================================
# VENTAS PENDIENTES Y METODO DE PAGO
# ============================================================

# {chat_id: [lista de ventas pendientes de confirmar metodo]}
ventas_pendientes = {}
# {chat_id: numero_venta} — para confirmar borrado
borrados_pendientes = {}
# Lock para proteger acceso concurrente a los dicts globales
_estado_lock = threading.Lock()


def _registrar_ventas_con_metodo(ventas, metodo, vendedor, chat_id):
    """
    Registra una lista de ventas con el metodo de pago dado.
    Respeta precios de fraccion del catalogo: el precio_unitario que viene de Claude
    ya es el precio de esa fraccion especifica (no el de unidad completa).
    El 'total' para el Excel es ese precio (ya que es lo que pago el cliente).
    El inventario se descuenta en la cantidad decimal correspondiente.
    """
    confirmaciones = []
    for venta in ventas:
        producto = venta.get("producto", "Sin nombre")
        cantidad = convertir_fraccion_a_decimal(venta.get("cantidad", 1))
        precio_cobrado = float(venta.get("precio_unitario", 0))

        # Si Claude no puso precio, buscarlo en catalogo con la logica de fracciones
        if precio_cobrado == 0:
            precio_cobrado, _ = obtener_precio_para_cantidad(producto, cantidad)

        # El total es el precio_cobrado (que ya es el precio de la cantidad vendida)
        # Para cantidades >= 1, multiplicamos normalmente
        # Para fracciones, precio_cobrado YA es el precio de esa fraccion
        if cantidad >= 1:
            total = precio_cobrado * cantidad
        else:
            # Es una fraccion: precio_cobrado ya es el total de esa fraccion
            total = precio_cobrado

        cantidad_legible = decimal_a_fraccion_legible(cantidad)
        num = guardar_venta_excel(producto, cantidad, precio_cobrado, total, vendedor, metodo)
        confirmaciones.append(f"• {producto} x{cantidad_legible} = ${total:,.0f}")

        # Actualizar caja
        caja = cargar_caja()
        if caja.get("abierta"):
            campo_map = {"efectivo": "efectivo", "transferencia": "transferencias", "datafono": "datafono"}
            campo = campo_map.get(metodo, "efectivo")
            caja[campo] = caja.get(campo, 0) + total
            guardar_caja(caja)

        # Descontar inventario (en unidades decimales)
        inventario = cargar_inventario()
        prod_lower = producto.lower()
        prod_key = next((k for k in inventario if k in prod_lower or prod_lower in k), None)
        if prod_key and isinstance(inventario[prod_key], dict):
            inv = inventario[prod_key]
            inv["cantidad"] = max(0, round(inv.get("cantidad", 0) - cantidad, 4))
            guardar_inventario(inventario)
            restante = decimal_a_fraccion_legible(inv["cantidad"])
            unidad = inv.get("unidad", "")
            if inv["cantidad"] <= inv.get("minimo", 0.5):
                confirmaciones.append(f"⚠️ Stock bajo: {prod_key} — quedan {restante} {unidad}")

    return confirmaciones


async def manejar_metodo_pago(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja botones de metodo de pago Y confirmacion de borrado."""
    query = update.callback_query
    await query.answer()
    data = query.data
    vendedor = query.from_user.first_name or "Desconocido"

    # ---- Confirmacion de borrado ----
    if data.startswith("borrar_si_") or data.startswith("borrar_no_"):
        partes = data.split("_")
        accion = partes[1]  # "si" o "no"
        chat_id = int(partes[2])
        with _estado_lock:
            numero_venta = borrados_pendientes.pop(chat_id, None)

        if accion == "no" or numero_venta is None:
            await query.edit_message_text("❌ Borrado cancelado.")
            return

        exito, mensaje = borrar_venta_excel(numero_venta)
        await query.edit_message_text(mensaje)
        return

    # ---- Metodo de pago ----
    if data.startswith("pago_"):
        partes = data.split("_")
        if len(partes) < 3:
            return
        metodo = partes[1]
        chat_id = int(partes[2])

        with _estado_lock:
            ventas = ventas_pendientes.pop(chat_id, [])
        if not ventas:
            await query.edit_message_text("Ya no hay ventas pendientes.")
            return

        confirmaciones = _registrar_ventas_con_metodo(ventas, metodo, vendedor, chat_id)
        metodo_emoji = {"efectivo": "💵", "transferencia": "📱", "datafono": "💳"}
        emoji = metodo_emoji.get(metodo, "✅")
        texto = f"✅ Venta registrada — {emoji} {metodo.capitalize()}\n\n" + "\n".join(confirmaciones)
        await query.edit_message_text(texto)


# ============================================================
# HISTORIAL DE CONVERSACION
# ============================================================
historiales = {}

def agregar_al_historial(chat_id, role, content):
    with _estado_lock:
        if chat_id not in historiales:
            historiales[chat_id] = []
        historiales[chat_id].append({"role": role, "content": content})
        if len(historiales[chat_id]) > 20:
            historiales[chat_id] = historiales[chat_id][-20:]


# ============================================================
# CLAUDE — CEREBRO DEL BOT
# ============================================================

async def procesar_con_claude(mensaje_usuario, nombre_usuario, historial_chat):
    memoria = cargar_memoria()
    resumen = obtener_resumen_ventas()
    resumen_texto = f"${resumen['total']:,.0f} en {resumen['num_ventas']} ventas este mes" if resumen else "Sin ventas este mes"

    # Solo carga datos historicos si el mensaje parece un analisis
    palabras_analisis = ["cuanto", "vendimos", "reporte", "analiz", "total", "resumen", "estadistica", "top", "mas vendido"]
    necesita_datos = any(p in mensaje_usuario.lower() for p in palabras_analisis)
    if necesita_datos:
        try:
            todos_los_datos = obtener_todos_los_datos()
            datos_texto = json.dumps(todos_los_datos[-100:], ensure_ascii=False, default=str) if todos_los_datos else "Sin datos aun"
        except Exception:
            datos_texto = "Sin datos aun"
    else:
        datos_texto = "(no cargado)"

    # Detectar si el mensaje menciona fracciones para incluir info especifica
    # Buscar si hay un producto mencionado que tenga precios de fraccion
    info_fracciones_extra = ""
    palabras_frac = ["1/4", "1/2", "3/4", "1/8", "1/16", "cuarto", "medio", "mitad", "octavo"]
    if any(p in mensaje_usuario.lower() for p in palabras_frac):
        # Intentar detectar el producto mencionado y mostrar sus precios de fraccion
        palabras_msg = mensaje_usuario.lower().split()
        # Buscar hasta 3 palabras consecutivas como posible nombre de producto
        for largo in [4, 3, 2]:
            encontrado = False
            for i in range(len(palabras_msg) - largo + 1):
                fragmento = " ".join(palabras_msg[i:i+largo])
                prod = buscar_producto_en_catalogo(fragmento)
                if prod and prod.get("precios_fraccion"):
                    info = obtener_info_fraccion_producto(prod["nombre_lower"])
                    if info:
                        info_fracciones_extra = f"\nPRECIOS POR FRACCION DEL PRODUCTO MENCIONADO:\n{info}"
                    encontrado = True
                    break
            if encontrado:
                break

    # Aviso de modo offline
    aviso_drive = ""
    if not DRIVE_DISPONIBLE:
        aviso_drive = "\n⚠️ AVISO: Google Drive no disponible. Los datos se guardan localmente."

    # Catalogo resumido para el prompt (nombre + precio unidad + indicador fraccionable)
    # Solo incluir si el catalogo no es gigante; si lo es, resumir por categoria
    catalogo = memoria.get("catalogo", {})
    if catalogo:
        # Agrupar por categoria para el prompt, mostrar primero los fraccionables
        lineas_cat = []
        categorias = {}
        for cod, prod in catalogo.items():
            cat = prod.get("categoria", "Otros")
            if cat not in categorias:
                categorias[cat] = []
            tiene_frac = bool(prod.get("precios_fraccion"))
            categorias[cat].append(f"  - {prod['nombre']}: ${prod['precio_unidad']:,}" + (" [fraccionable]" if tiene_frac else ""))
        for cat, items in sorted(categorias.items()):
            lineas_cat.append(f"{cat}:")
            lineas_cat.extend(items[:60])  # max 60 por categoria
        precios_texto = "\n".join(lineas_cat)
    else:
        precios_texto = obtener_precios_como_texto()

    system_prompt = f"""Eres FerreBot, asistente inteligente de una ferreteria colombiana.

==================================================
TUS CAPACIDADES - NUNCA LAS OLVIDES
==================================================
- SI PUEDES registrar ventas con [VENTA]...[/VENTA]
- SI PUEDES crear Excel con [EXCEL]...[/EXCEL]
- SI PUEDES guardar precios con [PRECIO]...[/PRECIO]
- SI PUEDES controlar inventario con [INVENTARIO]...[/INVENTARIO]
- SI PUEDES manejar caja con [CAJA]...[/CAJA]
- SI PUEDES registrar gastos con [GASTO]...[/GASTO]
- TIENES memoria permanente de precios y productos
==================================================

REGLAS CRITICAS DE FRACCIONES Y PRECIOS:
- Muchos productos se venden en fracciones: 1/4, 1/2, 3/4, 1/8 de galon/unidad.
- Los marcados como [fraccionable] tienen precios DIFERENTES segun la cantidad vendida.
- NUNCA calcules el precio de una fraccion multiplicando el precio de unidad.
  Ejemplo INCORRECTO: 1/4 de vinilo T1 = 50000 * 0.25 = 12500 (MAL)
  Ejemplo CORRECTO: 1/4 de vinilo T1 = 60000 (precio especifico de fraccion)
- Cuando alguien pida una fraccion de un producto fraccionable, usa el precio
  especifico que aparece en PRECIOS POR FRACCION DEL PRODUCTO MENCIONADO si esta disponible.
- En el campo "cantidad" del [VENTA] pon el decimal: 1/4=0.25, 1/2=0.5, 3/4=0.75, 1/8=0.125
- En el campo "precio_unitario" pon el precio TOTAL de esa fraccion (no el de unidad completa)
  Ejemplo: venta de 1/4 de vinilo a $60000 -> cantidad: 0.25, precio_unitario: 60000
{info_fracciones_extra}

INFORMACION DEL NEGOCIO:
{json.dumps(memoria.get('negocio', dict()), ensure_ascii=False)}

CATALOGO DE PRODUCTOS (precio de unidad completa):
{precios_texto}

RESUMEN VENTAS DEL MES:
{resumen_texto}

DATOS HISTORICOS (analisis):
{datos_texto}

INVENTARIO ACTUAL:
{json.dumps(cargar_inventario(), ensure_ascii=False)}

ESTADO CAJA:
{obtener_resumen_caja()}

GASTOS DE HOY:
{json.dumps(cargar_gastos_hoy(), ensure_ascii=False, default=str)}
{aviso_drive}

INSTRUCCIONES DE FORMATO:
1. Responde en español, natural y amigable. Sin markdown con ** ni #.
2. Venta detectada — incluye al FINAL uno por producto, SIN repetir:
   [VENTA]{{"producto": "nombre completo", "cantidad": 1, "precio_unitario": 40000}}[/VENTA]
   CRITICO: NUNCA pongas metodo_pago en el JSON. NUNCA repitas [VENTA] para el mismo producto.
   El sistema pregunta el metodo de pago con botones automaticamente.
3. Precio nuevo: [PRECIO]{{"producto": "nombre", "precio": 50000}}[/PRECIO]
4. Info del negocio: [NEGOCIO]{{"clave": "valor"}}[/NEGOCIO]
5. Excel: [EXCEL]{{"titulo": "Titulo", "encabezados": ["Col1"], "filas": [["dato"]]}}[/EXCEL]
6. Apertura caja: [CAJA]{{"accion": "apertura", "monto": 50000}}[/CAJA]
7. Cierre caja: [CAJA]{{"accion": "cierre"}}[/CAJA]
8. Gasto: [GASTO]{{"concepto": "nombre", "monto": 50000, "categoria": "varios", "origen": "caja"}}[/GASTO]
9. Inventario: [INVENTARIO]{{"producto": "nombre", "cantidad": 10, "minimo": 2, "unidad": "galones", "accion": "actualizar"}}[/INVENTARIO]
10. Para borrar: /borrar numero
11. Usuario actual: {nombre_usuario}"""

    messages = []
    for msg in historial_chat[-10:]:
        if isinstance(msg, dict) and "role" in msg and "content" in msg:
            messages.append({"role": str(msg["role"]), "content": str(msg["content"])})
    messages.append({"role": "user", "content": str(mensaje_usuario)})

    respuesta = claude_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        system=system_prompt,
        messages=messages
    )
    return respuesta.content[0].text


def procesar_acciones(texto_respuesta, vendedor, chat_id):
    """
    Extrae y ejecuta todas las acciones del mensaje de Claude.
    Ahora recibe chat_id como parametro (correccion del bug original).
    """
    acciones = []
    archivos_excel = []
    texto_limpio = texto_respuesta

    # Ventas
    ventas_detectadas = []
    for venta_json in re.findall(r'\[VENTA\](.*?)\[/VENTA\]', texto_respuesta, re.DOTALL):
        try:
            venta = json.loads(venta_json.strip())
            ventas_detectadas.append(venta)
        except Exception as e:
            print(f"Error parseando venta: {e}")
        texto_limpio = texto_limpio.replace(f'[VENTA]{venta_json}[/VENTA]', '')

    if ventas_detectadas:
        # SIEMPRE pedir metodo de pago con botones — nunca asumir
        with _estado_lock:
            ventas_pendientes[chat_id] = ventas_detectadas
        acciones.append("PEDIR_METODO_PAGO")

    # Precios
    for precio_json in re.findall(r'\[PRECIO\](.*?)\[/PRECIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(precio_json.strip())
            memoria = cargar_memoria()
            memoria["precios"][datos["producto"].lower()] = float(datos["precio"])
            guardar_memoria(memoria)
            acciones.append(f"🧠 Precio guardado: {datos['producto']} = ${float(datos['precio']):,.0f}")
        except Exception as e:
            print(f"Error precio: {e}")
        texto_limpio = texto_limpio.replace(f'[PRECIO]{precio_json}[/PRECIO]', '')

    # Negocio
    for neg_json in re.findall(r'\[NEGOCIO\](.*?)\[/NEGOCIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(neg_json.strip())
            memoria = cargar_memoria()
            memoria["negocio"].update(datos)
            guardar_memoria(memoria)
        except Exception as e:
            print(f"Error negocio: {e}")
        texto_limpio = texto_limpio.replace(f'[NEGOCIO]{neg_json}[/NEGOCIO]', '')

    # Caja
    for caja_json in re.findall(r'\[CAJA\](.*?)\[/CAJA\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(caja_json.strip())
            caja = cargar_caja()
            if datos.get("accion") == "apertura":
                caja["abierta"] = True
                caja["fecha"] = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d")
                caja["monto_apertura"] = float(datos.get("monto", 0))
                caja["efectivo"] = 0
                caja["transferencias"] = 0
                caja["datafono"] = 0
                guardar_caja(caja)
                acciones.append(f"✅ Caja abierta con ${float(datos.get('monto',0)):,.0f}")
            elif datos.get("accion") == "cierre":
                resumen = obtener_resumen_caja()
                acciones.append(f"🔒 Caja cerrada.\n{resumen}")
                caja["abierta"] = False
                guardar_caja(caja)
        except Exception as e:
            print(f"Error caja: {e}")
        texto_limpio = texto_limpio.replace(f'[CAJA]{caja_json}[/CAJA]', '')

    # Gastos
    for gasto_json in re.findall(r'\[GASTO\](.*?)\[/GASTO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(gasto_json.strip())
            gasto = {
                "concepto": datos.get("concepto", ""),
                "monto": float(datos.get("monto", 0)),
                "categoria": datos.get("categoria", "varios"),
                "origen": datos.get("origen", "externo"),
                "hora": datetime.now(COLOMBIA_TZ).strftime("%H:%M")
            }
            guardar_gasto(gasto)
            acciones.append(f"💸 Gasto registrado: {gasto['concepto']} — ${gasto['monto']:,.0f} ({gasto['origen']})")
        except Exception as e:
            print(f"Error gasto: {e}")
        texto_limpio = texto_limpio.replace(f'[GASTO]{gasto_json}[/GASTO]', '')

    # Inventario
    for inv_json in re.findall(r'\[INVENTARIO\](.*?)\[/INVENTARIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(inv_json.strip())
            inventario = cargar_inventario()
            producto = datos.get("producto", "").lower()
            accion = datos.get("accion", "actualizar")
            if accion == "actualizar":
                cantidad = convertir_fraccion_a_decimal(datos.get("cantidad", 0))
                minimo = convertir_fraccion_a_decimal(datos.get("minimo", 0.5))
                unidad = datos.get("unidad", "unidades")
                inventario[producto] = {
                    "cantidad": cantidad,
                    "minimo": minimo,
                    "unidad": unidad,
                    "nombre_original": datos.get("producto", producto)
                }
                guardar_inventario(inventario)
                cantidad_texto = decimal_a_fraccion_legible(cantidad)
                acciones.append(f"📦 Inventario: {datos['producto']} — {cantidad_texto} {unidad}")
            elif accion == "descontar":
                if producto in inventario:
                    descuento = convertir_fraccion_a_decimal(datos.get("cantidad", 0))
                    inventario[producto]["cantidad"] = max(0, inventario[producto]["cantidad"] - descuento)
                    guardar_inventario(inventario)
            alertas = verificar_alertas_inventario()
            for alerta in alertas:
                acciones.append(alerta)
        except Exception as e:
            print(f"Error inventario: {e}")
        texto_limpio = texto_limpio.replace(f'[INVENTARIO]{inv_json}[/INVENTARIO]', '')

    # Excel personalizado
    for excel_json in re.findall(r'\[EXCEL\](.*?)\[/EXCEL\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(excel_json.strip())
            titulo = datos.get("titulo", "Reporte")
            encabezados = datos.get("encabezados", [])
            filas = datos.get("filas", [])
            nombre_archivo = f"reporte_{datetime.now(COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
            generar_excel_personalizado(titulo, encabezados, filas, nombre_archivo)
            archivos_excel.append(nombre_archivo)
        except Exception as e:
            print(f"Error generando Excel: {e}")
        texto_limpio = texto_limpio.replace(f'[EXCEL]{excel_json}[/EXCEL]', '')

    return texto_limpio.strip(), acciones, archivos_excel


async def _enviar_botones_pago(update_or_query, chat_id, ventas):
    """Funcion auxiliar reutilizable para mostrar botones de metodo de pago."""
    lineas = []
    for v in ventas:
        cantidad_dec = convertir_fraccion_a_decimal(v.get("cantidad", 1))
        precio = float(v.get("precio_unitario", 0))
        # Respetar la misma logica de _registrar_ventas_con_metodo:
        # si cantidad >= 1 el total es precio * cantidad;
        # si es fraccion, precio_unitario YA es el total de esa fraccion.
        if cantidad_dec >= 1:
            total_mostrar = precio * cantidad_dec
        else:
            total_mostrar = precio
        cantidad_legible = decimal_a_fraccion_legible(cantidad_dec) if isinstance(cantidad_dec, float) else v.get("cantidad", 1)
        lineas.append(f"• {v.get('producto')} x{cantidad_legible} = ${total_mostrar:,.0f}")
    resumen = "\n".join(lineas)
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("💵 Efectivo", callback_data=f"pago_efectivo_{chat_id}"),
        InlineKeyboardButton("📱 Transferencia", callback_data=f"pago_transferencia_{chat_id}"),
        InlineKeyboardButton("💳 Datafono", callback_data=f"pago_datafono_{chat_id}"),
    ]])
    await update_or_query.reply_text(f"¿Cómo fue el pago?\n\n{resumen}", reply_markup=keyboard)


# ============================================================
# MANEJADORES DE COMANDOS
# ============================================================

async def comando_inicio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    estado_drive  = "✅ Drive conectado" if DRIVE_DISPONIBLE else "⚠️ Drive offline"
    estado_sheets = "✅ Sheets conectado" if SHEETS_DISPONIBLE else ("⚠️ Sheets no configurado" if not SHEETS_ID else "⚠️ Sheets sin conexion")
    await update.message.reply_text(
        "👋 Hola! Soy tu asistente de la ferreteria.\n\n"
        "Puedo ayudarte con cualquier cosa:\n\n"
        "🛍️ Registrar ventas — 'Vendi 1/4 de vinilo t1 azul'\n"
        "🧠 Recordar precios — ya tengo el catalogo completo\n"
        "📊 Reportes — 'Cuanto vendimos esta semana?'\n"
        "📎 Excel personalizado — 'Hazme un Excel con los productos mas vendidos'\n"
        "📈 Graficas — /grafica\n"
        "🔍 Buscar ventas — /buscar [termino]\n"
        "💬 Cualquier pregunta — Lo que necesites\n\n"
        "Comandos:\n"
        "/ventas — Ver ultimas ventas\n"
        "/buscar [termino] — Buscar ventas\n"
        "/borrar [numero] — Borrar una venta\n"
        "/precios — Ver precios guardados\n"
        "/excel — Descargar archivo acumulado\n"
        "/sheets — Ver estado del Sheet del dia\n"
        "/cerrar — Cierre del dia (genera Excel + limpia Sheets)\n"
        "/grafica — Ver graficas de ventas\n"
        "/caja — Estado de caja\n"
        "/gastos — Gastos de hoy\n"
        "/inventario — Ver inventario\n\n"
        f"{estado_drive} | {estado_sheets}"
    )

async def comando_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    inicializar_excel()
    await update.message.reply_text("📎 Aqui esta tu archivo de ventas:")
    with open(EXCEL_FILE, "rb") as archivo:
        await update.message.reply_document(document=archivo, filename="ventas.xlsx")

async def comando_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ventas = obtener_ventas_recientes(10)
    if not ventas:
        await update.message.reply_text("No hay ventas registradas este mes.")
        return
    texto = "📋 Ultimas ventas:\n\n"
    for v in ventas:
        num = v[0] if v[0] else "?"
        producto = v[3] if len(v) > 3 else "?"
        total = f"${v[6]:,.0f}" if len(v) > 6 and v[6] else "?"
        vendedor = v[7] if len(v) > 7 else "?"
        texto += f"#{num} — {producto} — {total} — {vendedor}\n"
    texto += "\nUsa /borrar [numero] para eliminar una venta."
    await update.message.reply_text(texto)

async def comando_buscar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Busca ventas por producto, vendedor o fecha."""
    if not context.args:
        await update.message.reply_text(
            "Indica que quieres buscar.\n"
            "Ejemplos:\n"
            "/buscar tornillos\n"
            "/buscar Juan\n"
            "/buscar 2025-06"
        )
        return

    termino = " ".join(context.args)
    await update.message.reply_text(f"🔍 Buscando '{termino}'...")

    resultados = buscar_ventas(termino)
    if not resultados:
        await update.message.reply_text(f"No encontre ventas que coincidan con '{termino}'.")
        return

    texto = f"🔍 {len(resultados)} resultado(s) para '{termino}':\n\n"
    for r in resultados[:15]:  # Maximo 15 para no saturar
        num = r.get("#", "?")
        fecha = r.get("fecha", "?")
        producto = r.get("producto", "?")
        total = r.get("total", "?")
        vendedor = r.get("vendedor", "?")
        hoja = r.get("hoja", "")
        try:
            total_fmt = f"${float(total):,.0f}" if total else "?"
        except Exception:
            total_fmt = str(total)
        texto += f"#{num} [{hoja}] {fecha} — {producto} — {total_fmt} — {vendedor}\n"

    if len(resultados) > 15:
        texto += f"\n... y {len(resultados) - 15} mas. Usa un termino mas especifico para filtrar mejor."

    await update.message.reply_text(texto)

async def comando_borrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Borrar venta con confirmacion previa.
    Muestra los datos de la venta y pide confirmacion con botones.
    """
    if not context.args:
        await update.message.reply_text("Indica el numero de venta.\nEjemplo: /borrar 5")
        return
    try:
        numero = int(context.args[0])
    except ValueError:
        await update.message.reply_text("El numero debe ser entero.\nEjemplo: /borrar 5")
        return

    chat_id = update.message.chat_id
    venta = obtener_venta_por_numero(numero)

    if not venta:
        await update.message.reply_text(f"No encontre la venta #{numero}.")
        return

    # Guardar en memoria temporal (protegido con lock)
    with _estado_lock:
        borrados_pendientes[chat_id] = numero

    producto = venta.get("producto", "?")
    fecha = venta.get("fecha", "?")
    total = venta.get("total", "?")
    vendedor = venta.get("vendedor", "?")
    try:
        total_fmt = f"${float(total):,.0f}"
    except Exception:
        total_fmt = str(total)

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Sí, borrar", callback_data=f"borrar_si_{chat_id}"),
        InlineKeyboardButton("❌ Cancelar", callback_data=f"borrar_no_{chat_id}"),
    ]])

    await update.message.reply_text(
        f"⚠️ ¿Confirmas que quieres borrar esta venta?\n\n"
        f"#{numero} — {producto}\n"
        f"Fecha: {fecha}\n"
        f"Total: {total_fmt}\n"
        f"Vendedor: {vendedor}",
        reply_markup=keyboard
    )

async def comando_precios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"🧠 Precios que recuerdo:\n\n{obtener_precios_como_texto()}")

async def comando_caja(update: Update, context: ContextTypes.DEFAULT_TYPE):
    resumen = obtener_resumen_caja()
    await update.message.reply_text(f"💰 {resumen}")

async def comando_gastos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    gastos = cargar_gastos_hoy()
    if not gastos:
        await update.message.reply_text("No hay gastos registrados hoy.")
        return
    texto = "💸 Gastos de hoy:\n\n"
    total = 0
    for g in gastos:
        texto += f"• {g['concepto']}: ${g['monto']:,.0f} ({g['categoria']}) — {g['origen']}\n"
        total += g['monto']
    texto += f"\nTotal gastos: ${total:,.0f}"
    await update.message.reply_text(texto)

async def comando_inventario(update: Update, context: ContextTypes.DEFAULT_TYPE):
    inventario = cargar_inventario()
    if not inventario:
        await update.message.reply_text(
            "No hay productos en inventario aun. Dile al bot cuantas unidades tienes de cada producto."
        )
        return
    texto = "📦 Inventario actual:\n\n"
    alertas = []
    for producto, datos in inventario.items():
        if isinstance(datos, dict):
            cantidad = datos.get("cantidad", 0)
            minimo = datos.get("minimo", 3)
            emoji = "⚠️" if cantidad <= minimo else "✅"
            texto += f"{emoji} {producto}: {cantidad} unidades\n"
            if cantidad <= minimo:
                alertas.append(producto)
    if alertas:
        texto += f"\n⚠️ Stock bajo en: {', '.join(alertas)}"
    await update.message.reply_text(texto)

async def comando_grafica(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra botones para elegir que grafica ver."""
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📅 Ventas por día", callback_data="grafica_dias"),
        InlineKeyboardButton("📦 Productos", callback_data="grafica_productos"),
    ], [
        InlineKeyboardButton("💳 Métodos de pago", callback_data="grafica_pagos"),
    ]])
    await update.message.reply_text("¿Qué gráfica quieres ver?", reply_markup=keyboard)

async def manejar_callback_grafica(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Genera y envía la grafica seleccionada."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id

    tipo = query.data
    await query.edit_message_text("📊 Generando gráfica...")

    ruta = None
    try:
        if tipo == "grafica_dias":
            ruta = generar_grafica_ventas_por_dia()
            titulo = "ventas_por_dia.png"
        elif tipo == "grafica_productos":
            ruta = generar_grafica_productos()
            titulo = "productos_mas_vendidos.png"
        elif tipo == "grafica_pagos":
            ruta = generar_grafica_metodos_pago()
            titulo = "metodos_de_pago.png"
        else:
            return

        if not ruta or not os.path.exists(ruta):
            await context.bot.send_message(chat_id=chat_id, text="No hay datos suficientes para generar esta gráfica aun.")
            return

        with open(ruta, "rb") as f:
            await context.bot.send_photo(chat_id=chat_id, photo=f, filename=titulo)
    except Exception as e:
        import traceback
        print(f"Error generando grafica: {traceback.format_exc()}")
        await context.bot.send_message(chat_id=chat_id, text="Tuve un problema generando la gráfica. Intenta de nuevo.")
    finally:
        # Siempre limpiar el archivo temporal, haya error o no
        if ruta and os.path.exists(ruta):
            os.remove(ruta)


# ============================================================
# MANEJADORES DE MENSAJES
# ============================================================

async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje = update.message.text
    chat_id = update.message.chat_id
    vendedor = update.message.from_user.first_name or "Desconocido"
    if mensaje.startswith("/"):
        return
    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    # Si el usuario tiene un Excel cargado, procesarlo
    excel_temp = context.user_data.get("excel_temp")
    excel_nombre = context.user_data.get("excel_nombre")
    if excel_temp and os.path.exists(excel_temp):
        try:
            await update.message.reply_text("⚙️ Procesando tu Excel...")
            codigo = await editar_excel_con_claude(mensaje, excel_temp, excel_nombre, vendedor, chat_id)

            if codigo.strip() == "IMPOSIBLE":
                await update.message.reply_text("No pude hacer eso con el Excel. Intenta con otra instruccion.")
                return

            # Ejecutar el codigo generado por Claude en un sandbox restringido.
            # Solo se expone openpyxl y json; 'os' queda EXCLUIDO intencionalmente
            # para evitar que codigo malicioso acceda al sistema de archivos o entorno.
            namespace_seguro = {
                "__builtins__": {
                    # Builtins minimos necesarios para manipular Excel
                    "range": range, "len": len, "enumerate": enumerate,
                    "int": int, "float": float, "str": str, "bool": bool,
                    "list": list, "dict": dict, "tuple": tuple, "set": set,
                    "min": min, "max": max, "sum": sum, "abs": abs,
                    "round": round, "sorted": sorted, "zip": zip,
                    "isinstance": isinstance, "print": print,
                    "Exception": Exception, "ValueError": ValueError,
                    "TypeError": TypeError, "KeyError": KeyError,
                },
                "openpyxl": openpyxl,
                "json": json,
            }
            exec(compile(codigo, "<string>", "exec"), namespace_seguro)

            await update.message.reply_text("✅ Excel modificado. Aqui esta el resultado:")
            with open(excel_temp, "rb") as f:
                await update.message.reply_document(document=f, filename=f"modificado_{excel_nombre}")

            context.user_data.pop("excel_temp", None)
            context.user_data.pop("excel_nombre", None)
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
            return
        except Exception as e:
            import traceback
            print(f"Error editando Excel: {traceback.format_exc()}")
            await update.message.reply_text("Tuve un problema editando el Excel. Intenta con una instruccion diferente.")
            return

    try:
        with _estado_lock:
            historial = list(historiales.get(chat_id, []))
        agregar_al_historial(chat_id, "user", f"{vendedor}: {mensaje}")
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {mensaje}", vendedor, historial)
        # CORRECCIÓN: se pasa chat_id a procesar_acciones
        texto_respuesta, acciones, archivos_excel = procesar_acciones(respuesta_raw, vendedor, chat_id)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)

        if texto_respuesta:
            await update.message.reply_text(texto_respuesta)

        pedir_metodo = "PEDIR_METODO_PAGO" in acciones
        for accion in acciones:
            if accion != "PEDIR_METODO_PAGO":
                await update.message.reply_text(accion)

        if pedir_metodo:
            with _estado_lock:
                ventas = ventas_pendientes.get(chat_id, [])
            await _enviar_botones_pago(update.message, chat_id, ventas)

        for archivo in archivos_excel:
            if os.path.exists(archivo):
                await update.message.reply_text("📊 Aqui esta tu reporte:")
                with open(archivo, "rb") as f:
                    await update.message.reply_document(document=f, filename=archivo)
                os.remove(archivo)

    except Exception as e:
        import traceback
        import sys
        print(f"Error completo: {traceback.format_exc()}", file=sys.stderr)
        await update.message.reply_text("Tuve un problema. Intenta de nuevo.")


async def manejar_audio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    vendedor = update.message.from_user.first_name or "Desconocido"
    chat_id = update.message.chat_id
    await update.message.reply_text("🎤 Escuchando...")
    try:
        archivo_voz = await update.message.voice.get_file()
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            await archivo_voz.download_to_drive(tmp.name)
            ruta_audio = tmp.name
        with open(ruta_audio, "rb") as audio_file:
            transcripcion = openai_client.audio.transcriptions.create(
                model="whisper-1", file=audio_file, language="es"
            )
        os.unlink(ruta_audio)
        texto = transcripcion.text
        await update.message.reply_text(f"📝 Escuche: {texto}")
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")
        with _estado_lock:
            historial = list(historiales.get(chat_id, []))
        agregar_al_historial(chat_id, "user", f"{vendedor}: {texto}")
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {texto}", vendedor, historial)
        # CORRECCIÓN: se pasa chat_id
        texto_respuesta, acciones, archivos_excel = procesar_acciones(respuesta_raw, vendedor, chat_id)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)
        if texto_respuesta:
            await update.message.reply_text(texto_respuesta)

        pedir_metodo = "PEDIR_METODO_PAGO" in acciones
        for accion in acciones:
            if accion != "PEDIR_METODO_PAGO":
                await update.message.reply_text(accion)

        if pedir_metodo:
            with _estado_lock:
                ventas = ventas_pendientes.get(chat_id, [])
            await _enviar_botones_pago(update.message, chat_id, ventas)

        for archivo in archivos_excel:
            if os.path.exists(archivo):
                await update.message.reply_text("📊 Aqui esta tu reporte:")
                with open(archivo, "rb") as f:
                    await update.message.reply_document(document=f, filename=archivo)
                os.remove(archivo)
    except Exception as e:
        import traceback
        print(f"Error audio: {traceback.format_exc()}")
        await update.message.reply_text("Problema con el audio. Intenta de nuevo.")


async def manejar_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja archivos Excel que le mandan al bot."""
    vendedor = update.message.from_user.first_name or "Desconocido"
    chat_id = update.message.chat_id
    doc = update.message.document

    if not doc:
        return

    nombre = doc.file_name or "archivo"
    extension = nombre.split(".")[-1].lower() if "." in nombre else ""

    if extension not in ["xlsx", "xls"]:
        await update.message.reply_text(f"Recibí '{nombre}'. Solo puedo procesar archivos Excel (.xlsx). ¿Necesitas algo más?")
        return

    await update.message.reply_text(f"📂 Recibí tu archivo '{nombre}'. Leyendo contenido...")

    try:
        archivo = await doc.get_file()
        ruta_temp = f"temp_{chat_id}_{nombre}"
        await archivo.download_to_drive(ruta_temp)

        wb = openpyxl.load_workbook(ruta_temp)
        resumen_hojas = []
        for hoja_nombre in wb.sheetnames:
            ws = wb[hoja_nombre]
            filas = ws.max_row - 1
            cols = ws.max_column
            encabezados = [ws.cell(row=1, column=c).value for c in range(1, cols+1) if ws.cell(row=1, column=c).value]
            resumen_hojas.append(f"Hoja '{hoja_nombre}': {filas} filas, columnas: {', '.join(str(e) for e in encabezados)}")

        resumen = "\n".join(resumen_hojas)
        context.user_data["excel_temp"] = ruta_temp
        context.user_data["excel_nombre"] = nombre

        await update.message.reply_text(
            f"✅ Excel cargado correctamente.\n\n{resumen}\n\n"
            f"Ahora dime que quieres hacer con el. Por ejemplo:\n"
            f"- Agrega una columna de IVA del 19%\n"
            f"- Ordena de mayor a menor por total\n"
            f"- Cambia los encabezados a color rojo\n"
            f"- Calcula el total de todas las ventas"
        )
    except Exception as e:
        import traceback
        print(f"Error leyendo Excel: {traceback.format_exc()}")
        await update.message.reply_text("Tuve un problema leyendo el archivo. Asegurate de que sea un Excel valido.")


async def editar_excel_con_claude(instruccion, ruta_excel, nombre_excel, vendedor, chat_id):
    """Usa Claude para generar codigo Python que edite el Excel segun la instruccion."""
    wb = openpyxl.load_workbook(ruta_excel)
    info_hojas = []
    for hoja_nombre in wb.sheetnames:
        ws = wb[hoja_nombre]
        encabezados = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        filas_ejemplo = []
        for fila in ws.iter_rows(min_row=2, max_row=min(4, ws.max_row), values_only=True):
            filas_ejemplo.append(list(fila))
        info_hojas.append({
            "hoja": hoja_nombre,
            "encabezados": encabezados,
            "ejemplo_filas": filas_ejemplo,
            "total_filas": ws.max_row - 1
        })

    prompt = f"""Eres un experto en Python y openpyxl. El usuario tiene un archivo Excel llamado '{nombre_excel}' con esta estructura:

{json.dumps(info_hojas, ensure_ascii=False, default=str)}

El usuario quiere: {instruccion}

Genera SOLO el codigo Python necesario para modificar el archivo usando openpyxl.
- El archivo ya esta cargado, usa: wb = openpyxl.load_workbook('{ruta_excel}')
- Al final guarda con: wb.save('{ruta_excel}')
- Usa colores en formato hex sin # (ej: 'FF0000' para rojo)
- Solo tienes disponibles: openpyxl y json. NO uses os, sys, subprocess ni ninguna otra libreria.
- Solo el codigo, sin explicaciones ni comentarios ni bloques ```
- Si la instruccion no tiene sentido para un Excel, devuelve solo la palabra: IMPOSIBLE"""

    respuesta = claude_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )

    codigo = respuesta.content[0].text.strip()
    if "```python" in codigo:
        codigo = codigo.split("```python")[1].split("```")[0].strip()
    elif "```" in codigo:
        codigo = codigo.split("```")[1].split("```")[0].strip()

    return codigo


async def comando_sheets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra el estado del Google Sheets y un enlace directo."""
    if not SHEETS_ID:
        await update.message.reply_text(
            "⚠️ Google Sheets no configurado. Agrega la variable SHEETS_ID en Railway."
        )
        return

    ventas = sheets_leer_ventas_del_dia()
    estado = "✅ Conectado" if SHEETS_DISPONIBLE else "⚠️ Sin conexion"
    url = f"https://docs.google.com/spreadsheets/d/{SHEETS_ID}/edit"

    if not ventas:
        texto = (
            f"📊 Google Sheets — {estado}\n\n"
            f"No hay ventas registradas hoy todavia.\n\n"
            f"🔗 {url}"
        )
    else:
        total_dia = sum(float(v.get("total", 0) or 0) for v in ventas)
        texto = (
            f"📊 Google Sheets — {estado}\n\n"
            f"Ventas de hoy: {len(ventas)}\n"
            f"Total del dia: ${total_dia:,.0f}\n\n"
            f"🔗 {url}"
        )
    await update.message.reply_text(texto)


async def comando_cerrar_dia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Cierre del dia:
    1. Lee las ventas del Google Sheets (respetando ediciones manuales)
    2. Detecta y reporta correcciones manuales
    3. Actualiza ventas.xlsx — reemplaza las filas de hoy con los datos del Sheets
       (una sola pestana por mes, historial acumulado, sin archivos extra)
    4. Sube ventas.xlsx a Drive
    5. Manda ventas.xlsx al chat de Telegram
    6. Limpia el Sheets para el dia siguiente
    """
    chat_id = update.message.chat_id
    await update.message.reply_text("🔒 Iniciando cierre del dia...")

    if not SHEETS_ID:
        await update.message.reply_text(
            "⚠️ Google Sheets no configurado.\n"
            "Usa /excel para descargar el archivo acumulado."
        )
        return

    # Paso 1: leer ventas del Sheets
    ventas_sheets = sheets_leer_ventas_del_dia()
    if not ventas_sheets:
        await update.message.reply_text(
            "📭 El Sheets no tiene ventas hoy. Si las hay en el Excel, usa /excel."
        )
        return

    await update.message.reply_text(f"📋 {len(ventas_sheets)} ventas encontradas en el Sheets...")

    # Paso 2: detectar ediciones manuales
    diferencias = sheets_detectar_ediciones_vs_excel()
    if diferencias:
        aviso = "✏️ Correcciones manuales detectadas (se aplicaran al Excel):\n\n"
        aviso += "\n".join(diferencias)
        await update.message.reply_text(aviso)

    # Paso 3: actualizar ventas.xlsx con los datos del Sheets
    hoy = datetime.now(COLOMBIA_TZ)
    fecha_str = hoy.strftime("%Y-%m-%d")
    meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
             7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

    try:
        inicializar_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        nombre_hoja = obtener_nombre_hoja()
        ws = obtener_o_crear_hoja(wb, nombre_hoja)
        cols = detectar_columnas(ws)

        # Encontrar columna de fecha para identificar filas de hoy
        col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)

        # Borrar todas las filas de hoy del Excel (seran reemplazadas por lo del Sheets)
        filas_a_borrar = []
        if col_fecha:
            for fila in range(2, ws.max_row + 1):
                val_fecha = ws.cell(row=fila, column=col_fecha).value
                if val_fecha and str(val_fecha)[:10] == fecha_str:
                    filas_a_borrar.append(fila)
            # Borrar de abajo hacia arriba para no desplazar indices
            for fila in reversed(filas_a_borrar):
                ws.delete_rows(fila)

        # Insertar las ventas del Sheets (ya con ediciones manuales)
        total_general = 0
        for v in ventas_sheets:
            fila_nueva = ws.max_row + 1
            try:
                cantidad_dec = convertir_fraccion_a_decimal(v.get("cantidad", 1))
            except Exception:
                cantidad_dec = v.get("cantidad", 1)

            datos = {
                "#":              v.get("num", fila_nueva - 1),
                "fecha":          v.get("fecha", fecha_str),
                "hora":           v.get("hora", ""),
                "producto":       v.get("producto", ""),
                "cantidad":       v.get("cantidad", ""),
                "precio unitario":v.get("precio_unitario", 0),
                "precio":         v.get("precio_unitario", 0),
                "total":          v.get("total", 0),
                "vendedor":       v.get("vendedor", ""),
                "observaciones":  v.get("metodo", ""),
            }

            if cols:
                for nombre_col, num_col in cols.items():
                    for clave, valor in datos.items():
                        if clave in nombre_col or nombre_col in clave:
                            ws.cell(row=fila_nueva, column=num_col, value=valor)
                            break
            else:
                vals = [datos["#"], datos["fecha"], datos["hora"], datos["producto"],
                        datos["cantidad"], datos["precio unitario"], datos["total"],
                        datos["vendedor"], datos["observaciones"]]
                for col, val in enumerate(vals, 1):
                    ws.cell(row=fila_nueva, column=col, value=val)

            # Alternar color de fila
            if fila_nueva % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=fila_nueva, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

            try:
                total_general += float(v.get("total", 0) or 0)
            except (ValueError, TypeError):
                pass

        wb.save(EXCEL_FILE)

        # Paso 4: subir a Drive
        subir_a_drive(EXCEL_FILE)
        await update.message.reply_text(
            f"✅ ventas.xlsx actualizado — {len(ventas_sheets)} ventas de hoy\n"
            f"Total del dia: ${total_general:,.0f}\n"
            f"Pestana: {nombre_hoja}"
        )

        # Paso 5: mandar ventas.xlsx al chat
        await update.message.reply_text("📎 Aqui esta el archivo actualizado:")
        with open(EXCEL_FILE, "rb") as f:
            await update.message.reply_document(document=f, filename="ventas.xlsx")

    except Exception as e:
        import traceback
        print(f"Error en cierre: {traceback.format_exc()}")
        await update.message.reply_text(
            "❌ Hubo un error actualizando el Excel. Los datos siguen en el Sheets, no se perdio nada."
        )
        return

    # Paso 6: limpiar el Sheets para manana
    await update.message.reply_text("🧹 Limpiando el Sheets para manana...")
    ok = sheets_limpiar()
    if ok:
        await update.message.reply_text(
            "✅ Cierre completado.\n\n"
            "• ventas.xlsx actualizado en Drive\n"
            "• Sheets limpio y listo para manana"
        )
    else:
        await update.message.reply_text(
            "⚠️ El Excel se actualizo correctamente pero no se pudo limpiar el Sheets.\n"
            "Puedes borrarlo a mano, los datos ya quedaron en el Excel."
        )

def main():
    print(f"🚀 Iniciando FerreBot {VERSION}")
    sincronizar_archivos()
    inicializar_excel()

    if SHEETS_ID:
        print(f"📊 Google Sheets configurado: {SHEETS_ID}")
        # Verificar conexion al Sheets al arrancar
        ws_test = _obtener_hoja_sheets()
        if ws_test:
            print("✅ Conexion a Google Sheets OK")
        else:
            print("⚠️ No se pudo conectar al Sheets")
    else:
        print("ℹ️ SHEETS_ID no configurado — funciones de Sheets desactivadas")

    app = Application.builder().token(TELEGRAM_TOKEN).build()

    # Comandos
    app.add_handler(CommandHandler("start",     comando_inicio))
    app.add_handler(CommandHandler("ayuda",     comando_inicio))
    app.add_handler(CommandHandler("excel",     comando_excel))
    app.add_handler(CommandHandler("ventas",    comando_ventas))
    app.add_handler(CommandHandler("buscar",    comando_buscar))
    app.add_handler(CommandHandler("borrar",    comando_borrar))
    app.add_handler(CommandHandler("precios",   comando_precios))
    app.add_handler(CommandHandler("caja",      comando_caja))
    app.add_handler(CommandHandler("gastos",    comando_gastos))
    app.add_handler(CommandHandler("inventario",comando_inventario))
    app.add_handler(CommandHandler("grafica",   comando_grafica))
    app.add_handler(CommandHandler("sheets",    comando_sheets))
    app.add_handler(CommandHandler("cerrar",    comando_cerrar_dia))

    # Mensajes
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_mensaje))
    app.add_handler(MessageHandler(filters.VOICE, manejar_audio))
    app.add_handler(MessageHandler(filters.Document.ALL, manejar_documento))

    # Callbacks (botones inline)
    app.add_handler(CallbackQueryHandler(manejar_metodo_pago,      pattern="^pago_"))
    app.add_handler(CallbackQueryHandler(manejar_metodo_pago,      pattern="^borrar_"))
    app.add_handler(CallbackQueryHandler(manejar_callback_grafica, pattern="^grafica_"))

    if WEBHOOK_URL:
        print(f"🌐 Iniciando en modo WEBHOOK: {WEBHOOK_URL}")
        app.run_webhook(
            listen="0.0.0.0",
            port=WEBHOOK_PORT,
            url_path=TELEGRAM_TOKEN,
            webhook_url=f"{WEBHOOK_URL}/{TELEGRAM_TOKEN}",
        )
    else:
        print("⚙️ WEBHOOK_URL no configurada. Iniciando en modo polling (desarrollo local).")
        app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
