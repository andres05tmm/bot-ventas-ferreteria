"""
FerreBot Dashboard API — FastAPI
Expone datos de ventas (Google Sheets + Excel) y catálogo (memoria.json).
Corre en el mismo entorno que el bot; reutiliza config.py y sheets.py.
"""

from __future__ import annotations

from dotenv import load_dotenv
load_dotenv()

import json
import logging
import os
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, Union

import config
from sheets import sheets_leer_ventas_del_dia

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="FerreBot Dashboard API",
    description="API de ventas y catálogo para Ferretería Punto Rojo",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Helpers de fecha ──────────────────────────────────────────────────────────
def _hoy() -> str:
    return datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")


def _hace_n_dias(n: int) -> datetime:
    return datetime.now(config.COLOMBIA_TZ) - timedelta(days=n)


# ── Helper: leer Excel histórico ──────────────────────────────────────────────
def _leer_excel_rango(dias: int | None = None, mes_actual: bool = False) -> list[dict]:
    if not os.path.exists(config.EXCEL_FILE):
        return []

    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True, data_only=True)
    except Exception:
        return []

    ahora = datetime.now(config.COLOMBIA_TZ)
    hojas_candidatas = [
        f"{config.MESES[ahora.month]} {ahora.year}",
    ]
    if ahora.month > 1:
        hojas_candidatas.append(f"{config.MESES[ahora.month - 1]} {ahora.year}")
    else:
        hojas_candidatas.append(f"{config.MESES[12]} {ahora.year - 1}")

    if dias is not None:
        limite = (_hace_n_dias(dias)).strftime("%Y-%m-%d")
    else:
        limite = None

    resultado = []
    for nombre_hoja in hojas_candidatas:
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]

        # En modo read_only ws.max_column puede ser None y ws.cell() no es confiable.
        # Leer headers iterando la fila exacta.
        cols: dict[str, int] = {}
        try:
            for fila_hdr in ws.iter_rows(
                min_row=config.EXCEL_FILA_HEADERS,
                max_row=config.EXCEL_FILA_HEADERS,
            ):
                for cell in fila_hdr:
                    if cell.value:
                        cols[str(cell.value).lower().strip()] = cell.column
                break
        except Exception:
            continue

        def _col(*claves) -> int | None:
            for k in claves:
                if k in cols:
                    return cols[k]
            return None

        c_fecha    = _col("fecha")
        c_hora     = _col("hora")
        c_producto = _col("producto")
        c_cantidad = _col("cantidad")
        c_precio   = _col("valor unitario", "precio unitario", "precio")
        c_total    = _col("total")
        c_alias    = _col("alias")
        c_vendedor = _col("vendedor")
        c_metodo   = _col("metodo de pago", "metodo pago", "método pago")
        c_num      = _col("#", "consecutivo", "num", "consecutivo de venta")
        c_unidad   = _col("unidad de medida", "unidad_medida", "unidad")

        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue

            fecha_raw = fila[c_fecha - 1] if (c_fecha and c_fecha <= len(fila)) else None
            if fecha_raw is None:
                continue

            if isinstance(fecha_raw, datetime):
                fecha_str = fecha_raw.strftime("%Y-%m-%d")
            else:
                fecha_str = str(fecha_raw)[:10]

            if limite and fecha_str < limite:
                continue
            if mes_actual and not fecha_str.startswith(f"{ahora.year}-{ahora.month:02d}"):
                continue

            def _v(col_idx):
                if col_idx is None or col_idx > len(fila):
                    return ""
                v = fila[col_idx - 1]
                return v if v is not None else ""

            try:
                total = float(str(_v(c_total)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                total = 0.0

            try:
                precio_unit = float(str(_v(c_precio)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                precio_unit = 0.0

            resultado.append({
                "num":             _v(c_num),
                "fecha":           fecha_str,
                "hora":            str(_v(c_hora)),
                "id_cliente":      "CF",
                "cliente":         "Consumidor Final",
                "codigo_producto": "",
                "producto":        str(_v(c_producto)),
                "cantidad":        str(_v(c_cantidad)),
                "unidad_medida":   str(_v(c_unidad)) or "Unidad",
                "precio_unitario": precio_unit,
                "total":           total,
                "alias":           str(_v(c_alias)),
                "vendedor":        str(_v(c_vendedor)),
                "metodo":          str(_v(c_metodo)),
            })

    try:
        wb.close()
    except Exception:
        pass
    return resultado


# ── Wayper: stock unificado (inventario en unidades, venta en kg o unidades) ──
_WAYPER_KG_A_UNIDAD = 12  # 1 kg = 12 unidades
_WAYPER_KG_KEYS = {
    "wayper_blanco":   "wayper_blanco_unidad",
    "wayper_de_color": "wayper_de_color_unidad",
}

def _stock_wayper(key: str, inventario: dict):
    """
    Para waypers por kg: muestra stock en kg (= unidades / 12).
    Para waypers por unidad: muestra stock en unidades directamente.
    """
    # Si es el producto "por kg", leer el stock de unidades y convertir
    if key in _WAYPER_KG_KEYS:
        inv_und = inventario.get(_WAYPER_KG_KEYS[key])
        if inv_und is not None:
            und = inv_und.get("cantidad") if isinstance(inv_und, dict) else inv_und
            if und is not None:
                return round(und / _WAYPER_KG_A_UNIDAD, 2)  # kg
        return None
    # Stock normal
    raw = inventario.get(key)
    if raw is None:
        return None
    return raw.get("cantidad") if isinstance(raw, dict) else raw


def _leer_excel_compras(dias: int | None = None) -> list[dict]:
    """Lee la hoja 'Compras' del Excel de ventas."""
    if not os.path.exists(config.EXCEL_FILE):
        return []
    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True, data_only=True)
    except Exception:
        return []
    if "Compras" not in wb.sheetnames:
        wb.close()
        return []
    ws     = wb["Compras"]
    ahora  = datetime.now(config.COLOMBIA_TZ)
    limite = (ahora - timedelta(days=dias)).strftime("%Y-%m-%d") if dias else None
    resultado = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        fecha = str(row[0])[:10] if row[0] else ""
        if limite and fecha < limite:
            continue
        resultado.append({
            "fecha":          fecha,
            "hora":           str(row[1] or ""),
            "proveedor":      str(row[2] or "—"),
            "producto":       str(row[3] or ""),
            "cantidad":       _to_float(row[4]),
            "costo_unitario": _to_float(row[5]),
            "costo_total":    _to_float(row[6]),
        })
    wb.close()
    return sorted(resultado, key=lambda x: x["fecha"])


def _to_float(val) -> float:
    try:
        return float(str(val).replace(",", ".") or 0)
    except (ValueError, TypeError):
        return 0.0


def _cantidad_a_float(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        pass
    if "/" in s and " " not in s:
        parts = s.split("/")
        try:
            return float(parts[0]) / float(parts[1])
        except (ValueError, ZeroDivisionError):
            return 0.0
    if " y " in s:
        partes = s.split(" y ")
        try:
            entero = float(partes[0])
            frac_parts = partes[1].split("/")
            frac = float(frac_parts[0]) / float(frac_parts[1])
            return entero + frac
        except (ValueError, IndexError, ZeroDivisionError):
            return 0.0
    return 0.0


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/ventas/hoy")
def ventas_hoy():
    try:
        hoy = _hoy()

        # ── Fuente principal: Google Sheets (datos en tiempo real) ────────────
        filtradas = []
        fuente    = "sheets"
        try:
            ventas    = sheets_leer_ventas_del_dia()
            filtradas = [v for v in ventas if str(v.get("fecha", ""))[:10] == hoy]
        except Exception as e_sheets:
            logging.getLogger("ferrebot.api").warning(
                f"Sheets no disponible, usando Excel como fallback: {e_sheets}"
            )
            fuente = "excel_fallback"

        # ── Fallback al Excel si Sheets devuelve vacío o falló ────────────────
        if not filtradas:
            try:
                ventas_xls = _leer_excel_rango(dias=1)
                filtradas  = [v for v in ventas_xls if str(v.get("fecha", ""))[:10] == hoy]
                if filtradas:
                    fuente = "excel_fallback"
            except Exception:
                pass

        # ── Enriquecer con unidad_medida desde el catálogo (solo si falta) ────
        try:
            # Sheets ahora trae unidad_medida nativo; solo rellenar filas antiguas
            necesitan = [v for v in filtradas if not v.get("unidad_medida") or v["unidad_medida"] == "Unidad"]
            if necesitan and os.path.exists(config.MEMORIA_FILE):
                with open(config.MEMORIA_FILE, encoding="utf-8") as _f:
                    _mem = json.load(_f)
                catalogo = _mem.get("catalogo", {})

                def _unidad_para(nombre_prod: str) -> str:
                    if not nombre_prod:
                        return "Unidad"
                    n = nombre_prod.lower().strip()
                    for key, prod in catalogo.items():
                        if prod.get("nombre", "").lower().strip() == n or key == n.replace(" ", "_"):
                            return prod.get("unidad_medida", "Unidad") or "Unidad"
                    return "Unidad"

                for v in necesitan:
                    v["unidad_medida"] = _unidad_para(v.get("producto", ""))
        except Exception:
            pass

        return {"fecha": hoy, "ventas": filtradas, "total": len(filtradas), "fuente": fuente}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ventas/semana")
def ventas_semana():
    try:
        ventas = _leer_excel_rango(dias=7)
        return {"ventas": ventas, "total": len(ventas)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ventas/top")
def ventas_top(periodo: str = Query(default="semana", pattern="^(semana|mes)$")):
    try:
        dias = 7 if periodo == "semana" else None
        mes = periodo == "mes"
        ventas = _leer_excel_rango(dias=dias, mes_actual=mes)

        por_producto: dict[str, dict] = defaultdict(lambda: {"unidades": 0.0, "ingresos": 0.0})
        for v in ventas:
            nombre = str(v.get("producto", "")).strip()
            if not nombre:
                continue
            cantidad = _cantidad_a_float(v.get("cantidad", 0))
            total    = _to_float(v.get("total", 0))
            por_producto[nombre]["unidades"] += cantidad
            por_producto[nombre]["ingresos"] += total

        ranking = sorted(
            [{"producto": k, **v} for k, v in por_producto.items()],
            key=lambda x: x["unidades"],
            reverse=True,
        )[:10]

        for i, item in enumerate(ranking, 1):
            item["posicion"] = i

        return {"periodo": periodo, "top": ranking}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ventas/resumen")
def ventas_resumen():
    try:
        hoy = _hoy()

        # Sheets — tolerante a fallo
        try:
            ventas_hoy_list = sheets_leer_ventas_del_dia()
            ventas_hoy_list = [v for v in ventas_hoy_list if str(v.get("fecha", ""))[:10] == hoy]
        except Exception:
            ventas_hoy_list = []

        total_hoy   = sum(_to_float(v.get("total", 0)) for v in ventas_hoy_list)
        pedidos_hoy = len({str(v.get("num", i)) for i, v in enumerate(ventas_hoy_list)})

        # Excel semana — tolerante a fallo
        try:
            ventas_sem = _leer_excel_rango(dias=7)
        except Exception:
            ventas_sem = []
        total_sem   = sum(_to_float(v.get("total", 0)) for v in ventas_sem)
        pedidos_sem = len({str(v.get("num", i)) for i, v in enumerate(ventas_sem)}) or 1
        ticket_prom = round(total_sem / pedidos_sem, 0) if pedidos_sem else 0

        ventas_por_dia: dict[str, float] = defaultdict(float)
        for v in ventas_sem:
            fecha = str(v.get("fecha", ""))[:10]
            ventas_por_dia[fecha] += _to_float(v.get("total", 0))

        historico = []
        for i in range(6, -1, -1):
            dia = (_hace_n_dias(i)).strftime("%Y-%m-%d")
            historico.append({"fecha": dia, "total": ventas_por_dia.get(dia, 0)})

        # Excel mes — tolerante a fallo
        try:
            ventas_mes = _leer_excel_rango(mes_actual=True)
        except Exception:
            ventas_mes = []
        total_mes = sum(_to_float(v.get("total", 0)) for v in ventas_mes)

        ventas_mes_por_dia: dict[str, float] = defaultdict(float)
        for v in ventas_mes:
            fecha = str(v.get("fecha", ""))[:10]
            ventas_mes_por_dia[fecha] += _to_float(v.get("total", 0))

        ahora_local = datetime.now(config.COLOMBIA_TZ)
        primer_dia  = ahora_local.replace(day=1)
        historico_mes = []
        current = primer_dia
        while current.date() <= ahora_local.date():
            dia_str = current.strftime("%Y-%m-%d")
            historico_mes.append({"fecha": dia_str, "total": ventas_mes_por_dia.get(dia_str, 0)})
            current += timedelta(days=1)

        return {
            "total_hoy":     total_hoy,
            "pedidos_hoy":   pedidos_hoy,
            "total_semana":  total_sem,
            "ticket_prom":   ticket_prom,
            "historico_7d":  historico,
            "total_mes":     total_mes,
            "historico_mes": historico_mes,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/productos")
def productos():
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"productos": []}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})
        lista = []
        # Fracciones estándar para productos de galón (pinturas/impermeabilizantes)
        _FRACS_GALON = [
            ("3/4", 0.75), ("1/2", 0.5), ("1/4", 0.25),
            ("1/8", 0.125), ("1/16", 0.0625), ("1/10", 0.1),
        ]

        for k, v in catalogo.items():
            ppc = v.get("precio_por_cantidad")
            mayorista = None
            if ppc:
                mayorista = {
                    "umbral": ppc.get("umbral", 50),
                    "precio": ppc.get("precio_sobre_umbral", 0),
                }

            fracs = v.get("precios_fraccion", None)
            precio = v.get("precio_unidad", 0)

            # Auto-generar fracciones para pinturas/impermeabilizantes sin precios_fraccion
            if not fracs and precio > 0:
                cat_lower = (v.get("categoria", "") or "").lower()
                es_galon = "pintura" in cat_lower or "disolvente" in cat_lower or "impermeab" in cat_lower
                if es_galon:
                    fracs = {}
                    for label, decimal in _FRACS_GALON:
                        fracs[label] = {"precio": round(precio * decimal), "decimal": decimal}

            lista.append({
                "key":              k,
                "nombre":           v.get("nombre", k),
                "categoria":        v.get("categoria", "Sin categoría"),
                "precio":           precio,
                "codigo":           v.get("codigo", ""),
                "stock":            _stock_wayper(k, inventario),
                "precios_fraccion": fracs,
                "unidad_medida":    v.get("unidad_medida", "Unidad"),
                "mayorista":        mayorista,
            })
        return {"productos": lista, "total": len(lista)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/inventario/bajo")
def inventario_bajo():
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"alertas": []}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})

        alertas = []
        for key, prod in catalogo.items():
            precio = prod.get("precio_unidad", None)
            raw_stock = inventario.get(key, None)
            stock = raw_stock.get("cantidad") if isinstance(raw_stock, dict) else raw_stock

            sin_precio = precio is None or precio == 0
            sin_stock  = stock is not None and (stock == 0 or stock == "0" or stock == 0.0)

            if sin_precio or sin_stock:
                alertas.append({
                    "key":       key,
                    "nombre":    prod.get("nombre", key),
                    "categoria": prod.get("categoria", ""),
                    "precio":    precio or 0,
                    "stock":     stock,
                    "motivo":    "sin_precio" if sin_precio else "stock_cero",
                })

        return {"alertas": alertas, "total": len(alertas)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Caja del día ─────────────────────────────────────────────────────────────
@app.get("/caja")
def caja():
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"caja": {}, "gastos": []}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        caja_data = mem.get("caja_actual", {
            "abierta": False, "fecha": None,
            "monto_apertura": 0, "efectivo": 0,
            "transferencias": 0, "datafono": 0,
        })

        ahora     = datetime.now(config.COLOMBIA_TZ)
        hoy       = ahora.strftime("%Y-%m-%d")
        gastos_hoy = mem.get("gastos", {}).get(hoy, [])

        total_gastos_caja = sum(
            g.get("monto", 0) for g in gastos_hoy
            if g.get("origen") == "caja"
        )
        total_gastos      = sum(g.get("monto", 0) for g in gastos_hoy)

        abierta = caja_data.get("abierta", False)

        # Si la caja está cerrada, mostrar ceros — los valores guardados
        # corresponden al último día activo y no deben mostrarse como "de hoy"
        if not abierta:
            return {
                "abierta":           False,
                "fecha":             caja_data.get("fecha"),
                "monto_apertura":    0,
                "efectivo":          0,
                "transferencias":    0,
                "datafono":          0,
                "total_ventas":      0,
                "total_gastos_caja": 0,
                "total_gastos":      0,
                "efectivo_esperado": 0,
                "gastos":            [],
            }

        efectivo       = _to_float(caja_data.get("efectivo", 0))
        transferencias = _to_float(caja_data.get("transferencias", 0))
        datafono       = _to_float(caja_data.get("datafono", 0))
        apertura       = _to_float(caja_data.get("monto_apertura", 0))
        total_ventas      = efectivo + transferencias + datafono
        efectivo_esperado = apertura + efectivo - total_gastos_caja

        return {
            "abierta":           True,
            "fecha":             caja_data.get("fecha"),
            "monto_apertura":    apertura,
            "efectivo":          efectivo,
            "transferencias":    transferencias,
            "datafono":          datafono,
            "total_ventas":      total_ventas,
            "total_gastos_caja": total_gastos_caja,
            "total_gastos":      total_gastos,
            "efectivo_esperado": efectivo_esperado,
            "gastos":            gastos_hoy,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Caja: abrir / cerrar desde Dashboard ─────────────────────────────────────

class CajaAbrirBody(BaseModel):
    monto_apertura: Union[float, int] = 0

@app.post("/caja/abrir")
def caja_abrir(body: CajaAbrirBody):
    """Abre la caja del día con un monto inicial."""
    try:
        from memoria import cargar_caja, guardar_caja
        caja = cargar_caja()
        if caja.get("abierta"):
            raise HTTPException(status_code=400, detail="La caja ya está abierta")

        hoy = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        caja = {
            "abierta": True,
            "fecha": hoy,
            "monto_apertura": int(body.monto_apertura),
            "efectivo": 0,
            "transferencias": 0,
            "datafono": 0,
        }
        guardar_caja(caja)
        return {"ok": True, "mensaje": f"Caja abierta con ${int(body.monto_apertura):,}", "caja": caja}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/caja/cerrar")
def caja_cerrar():
    """Cierra la caja del día."""
    try:
        from memoria import cargar_caja, guardar_caja, cargar_gastos_hoy, obtener_resumen_caja
        caja = cargar_caja()
        if not caja.get("abierta"):
            raise HTTPException(status_code=400, detail="La caja ya está cerrada")

        resumen = obtener_resumen_caja()
        caja["abierta"] = False
        guardar_caja(caja)
        return {"ok": True, "mensaje": "Caja cerrada", "resumen": resumen}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Gastos: registrar desde Dashboard ────────────────────────────────────────

class NuevoGastoBody(BaseModel):
    concepto:   str
    monto:      Union[float, int]
    categoria:  str = "General"
    origen:     str = "caja"       # "caja" | "externo"

@app.post("/gastos")
def registrar_gasto(body: NuevoGastoBody):
    """Registra un gasto del día."""
    try:
        from memoria import guardar_gasto

        if not body.concepto.strip():
            raise HTTPException(status_code=400, detail="El concepto es obligatorio")
        if body.monto <= 0:
            raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")

        hora = datetime.now(config.COLOMBIA_TZ).strftime("%H:%M")
        gasto = {
            "concepto":  body.concepto.strip(),
            "monto":     int(body.monto),
            "categoria": body.categoria.strip() or "General",
            "origen":    body.origen,
            "hora":      hora,
        }
        guardar_gasto(gasto)
        return {"ok": True, "gasto": gasto, "mensaje": f"Gasto registrado: {body.concepto} ${int(body.monto):,}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Compras: registrar desde Dashboard ───────────────────────────────────────

class NuevaCompraBody(BaseModel):
    producto:       str
    cantidad:       Union[float, int]
    costo_unitario: Union[float, int]
    proveedor:      str = ""

@app.post("/compras")
def crear_compra(body: NuevaCompraBody):
    """Registra una compra de mercancía (actualiza inventario + kárdex)."""
    try:
        from memoria import registrar_compra

        if not body.producto.strip():
            raise HTTPException(status_code=400, detail="El producto es obligatorio")
        if body.cantidad <= 0:
            raise HTTPException(status_code=400, detail="La cantidad debe ser mayor a 0")
        if body.costo_unitario <= 0:
            raise HTTPException(status_code=400, detail="El costo unitario debe ser mayor a 0")

        ok, mensaje, datos_excel = registrar_compra(
            nombre_producto=body.producto.strip(),
            cantidad=float(body.cantidad),
            costo_unitario=float(body.costo_unitario),
            proveedor=body.proveedor.strip() or None,
        )

        if not ok:
            raise HTTPException(status_code=400, detail=mensaje)

        return {"ok": True, "mensaje": mensaje, "datos": datos_excel}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Ventas Rápidas (desde el Dashboard) ──────────────────────────────────────
class VentaRapidaItem(BaseModel):
    nombre:        str
    cantidad:      Union[float, str] = 1
    total:         float
    unidad_medida: str = ""   # si viene vacío se busca en catálogo

class VentaRapidaPayload(BaseModel):
    productos:       list[VentaRapidaItem]
    metodo:          str = "efectivo"
    vendedor:        str = "Dashboard"
    cliente_nombre:  str = ""
    cliente_id:      str = ""

@app.post("/venta-rapida")
def venta_rapida(payload: VentaRapidaPayload):
    try:
        from excel import guardar_venta_excel, recalcular_caja_desde_excel, obtener_siguiente_consecutivo

        # Cargar catálogo una sola vez para resolver unidad_medida
        _catalogo_cache = {}
        try:
            if os.path.exists(config.MEMORIA_FILE):
                with open(config.MEMORIA_FILE, encoding="utf-8") as _f:
                    _mem = json.load(_f)
                _catalogo_cache = _mem.get("catalogo", {})
        except Exception:
            pass

        def _resolver_unidad(item: VentaRapidaItem) -> str:
            """Devuelve la unidad_medida del item: primero la del payload, si no la del catálogo."""
            if item.unidad_medida and item.unidad_medida not in ("", "Unidad"):
                return item.unidad_medida
            # Buscar en catálogo por nombre normalizado
            nombre_norm = item.nombre.lower().strip()
            for prod_key, prod_val in _catalogo_cache.items():
                if prod_val.get("nombre", "").lower().strip() == nombre_norm or prod_key == nombre_norm.replace(" ", "_"):
                    return prod_val.get("unidad_medida", "Unidad")
            return item.unidad_medida or "Unidad"

        # Un solo consecutivo para toda la venta
        consecutivo = obtener_siguiente_consecutivo()

        filas = []
        for item in payload.productos:
            try:
                from utils import convertir_fraccion_a_decimal
                cant_num = convertir_fraccion_a_decimal(item.cantidad)
            except (ValueError, TypeError):
                cant_num = 1.0
            # Bug fix: cant_num <= 0 causaba precio_unitario=total (incorrecto).
            # Si la cantidad es 0 o invalida, forzamos 1 para que precio_unitario = total.
            if not cant_num or cant_num <= 0:
                cant_num = 1.0
            precio_unitario = round(item.total / cant_num, 2)

            unidad = _resolver_unidad(item)

            fila = guardar_venta_excel(
                producto        = item.nombre,
                cantidad        = cant_num,
                precio_unitario = precio_unitario,
                total           = item.total,
                vendedor        = payload.vendedor,
                observaciones   = "venta-rapida",
                metodo_pago     = payload.metodo,
                consecutivo     = consecutivo,
                unidad_medida   = unidad,
                cliente_nombre  = payload.cliente_nombre or None,
                cliente_id      = payload.cliente_id     or None,
            )
            filas.append(fila)

            # Descontar inventario (igual que hace el bot por Telegram)
            try:
                from memoria import descontar_inventario
                descontar_inventario(item.nombre, cant_num)
            except Exception:
                pass

        recalcular_caja_desde_excel()

        return {
            "ok":          True,
            "consecutivo": consecutivo,
            "productos":   len(filas),
            "total":       sum(i.total for i in payload.productos),
            "metodo":      payload.metodo,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/gastos")
def gastos(dias: int = Query(default=7, ge=1, le=90)):
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"gastos": [], "total": 0, "por_categoria": {}}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        todos_gastos = mem.get("gastos", {})
        ahora  = datetime.now(config.COLOMBIA_TZ)
        limite = (ahora - timedelta(days=dias - 1)).strftime("%Y-%m-%d")

        resultado = []
        por_categoria: dict[str, float] = defaultdict(float)
        por_dia: dict[str, float]       = defaultdict(float)

        for fecha, lista in todos_gastos.items():
            if fecha < limite:
                continue
            for g in lista:
                monto = _to_float(g.get("monto", 0))
                cat   = g.get("categoria", "Sin categoría")
                resultado.append({**g, "fecha": fecha, "monto": monto})
                por_categoria[cat]  += monto
                por_dia[fecha]      += monto

        resultado.sort(key=lambda x: (x["fecha"], x.get("hora", "")), reverse=True)

        # Historico por día para gráfica
        historico = []
        for i in range(dias - 1, -1, -1):
            dia = (ahora - timedelta(days=i)).strftime("%Y-%m-%d")
            historico.append({"fecha": dia, "total": por_dia.get(dia, 0)})

        return {
            "gastos":        resultado,
            "total":         sum(g["monto"] for g in resultado),
            "por_categoria": dict(por_categoria),
            "historico":     historico,
            "dias":          dias,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Historial de compras a proveedores ────────────────────────────────────────
@app.get("/compras")
def compras(dias: int = Query(default=30, ge=1, le=365)):
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"compras": [], "total_invertido": 0, "por_proveedor": {}}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        historial = mem.get("historial_compras", [])
        ahora  = datetime.now(config.COLOMBIA_TZ)
        limite = (ahora - timedelta(days=dias)).strftime("%Y-%m-%d")

        filtradas      = [c for c in historial if str(c.get("fecha", ""))[:10] >= limite]
        por_proveedor: dict[str, float] = defaultdict(float)
        por_producto:  dict[str, float] = defaultdict(float)

        for c in filtradas:
            prov = c.get("proveedor") or "Sin proveedor"
            por_proveedor[prov] += _to_float(c.get("costo_total", 0))
            por_producto[c.get("producto", "")] += _to_float(c.get("costo_total", 0))

        filtradas.sort(key=lambda x: x.get("fecha", ""), reverse=True)

        return {
            "compras":        filtradas,
            "total_invertido": sum(_to_float(c.get("costo_total", 0)) for c in filtradas),
            "por_proveedor":  dict(por_proveedor),
            "por_producto":   dict(sorted(por_producto.items(), key=lambda x: -x[1])[:20]),
            "dias":           dias,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Top ventas (v2 — por ingresos, frecuencia o categoría) ───────────────────
@app.get("/ventas/top2")
def ventas_top2(
    periodo:  str = Query(default="semana", pattern="^(semana|mes)$"),
    criterio: str = Query(default="ingresos", pattern="^(ingresos|frecuencia|categoria)$"),
):
    try:
        dias = 7 if periodo == "semana" else None
        mes  = periodo == "mes"
        ventas = _leer_excel_rango(dias=dias, mes_actual=mes)

        # Leer catálogo para saber la categoría de cada producto
        cat_map: dict[str, str] = {}
        if os.path.exists(config.MEMORIA_FILE):
            with open(config.MEMORIA_FILE, encoding="utf-8") as f:
                mem = json.load(f)
            for v in mem.get("catalogo", {}).values():
                nombre_lower = v.get("nombre_lower", "").strip()
                cat_map[nombre_lower] = v.get("categoria", "Sin categoría")

        # Acumular por producto
        acum: dict[str, dict] = defaultdict(lambda: {
            "ingresos": 0.0, "frecuencia": 0, "categoria": "Sin categoría"
        })
        for v in ventas:
            nombre = str(v.get("producto", "")).strip()
            if not nombre:
                continue
            total = _to_float(v.get("total", 0))
            acum[nombre]["ingresos"]   += total
            acum[nombre]["frecuencia"] += 1
            if acum[nombre]["categoria"] == "Sin categoría":
                acum[nombre]["categoria"] = cat_map.get(nombre.lower(), "Sin categoría")

        if criterio == "ingresos":
            ranking = sorted(acum.items(), key=lambda x: -x[1]["ingresos"])[:10]
            items = [{"producto": k, "valor": v["ingresos"], "frecuencia": v["frecuencia"],
                      "categoria": v["categoria"], "posicion": i+1}
                     for i, (k, v) in enumerate(ranking)]

        elif criterio == "frecuencia":
            ranking = sorted(acum.items(), key=lambda x: -x[1]["frecuencia"])[:10]
            items = [{"producto": k, "valor": v["frecuencia"], "ingresos": v["ingresos"],
                      "categoria": v["categoria"], "posicion": i+1}
                     for i, (k, v) in enumerate(ranking)]

        else:  # categoria
            # Top 5 por ingresos dentro de cada categoría
            por_cat: dict[str, list] = defaultdict(list)
            for nombre, datos in acum.items():
                por_cat[datos["categoria"]].append({"producto": nombre, **datos})
            result_cat = {}
            for cat, prods in por_cat.items():
                top = sorted(prods, key=lambda x: -x["ingresos"])[:5]
                result_cat[cat] = [{"producto": p["producto"], "valor": p["ingresos"],
                                    "frecuencia": p["frecuencia"], "posicion": i+1}
                                   for i, p in enumerate(top)]
            return {"periodo": periodo, "criterio": criterio, "por_categoria": result_cat}

        return {"periodo": periodo, "criterio": criterio, "top": items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Catálogo navegable (para dashboard) ──────────────────────────────────────
@app.get("/catalogo/nav")
def catalogo_nav(q: str = Query(default="")):
    try:
        if not os.path.exists(config.MEMORIA_FILE):
            return {"categorias": {}, "total": 0}
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})

        # Filtro de búsqueda
        q_lower = q.strip().lower()

        # Mapa de normalización: agrupa categorías que difieren solo en mayúsculas
        _cat_canonical: dict[str, str] = {}  # lower → primera aparición original

        categorias: dict[str, list] = defaultdict(list)
        for key, prod in catalogo.items():
            nombre   = prod.get("nombre", key)
            cat_raw  = prod.get("categoria", "Sin categoría")

            # Normalizar: usar siempre la primera forma encontrada
            cat_lower = cat_raw.lower()
            if cat_lower not in _cat_canonical:
                _cat_canonical[cat_lower] = cat_raw
            categoria = _cat_canonical[cat_lower]

            if q_lower and q_lower not in nombre.lower() and q_lower not in (prod.get("codigo","")).lower():
                continue

            # Stock info (wayper por kg usa inventario de unidades)
            stock = _stock_wayper(key, inventario)
            if key in _WAYPER_KG_KEYS:
                costo = None
            else:
                inv_data = inventario.get(key)
                costo = inv_data.get("costo_promedio") if isinstance(inv_data, dict) else None

            # Fracciones
            fracs = {}
            for frac_key, frac_val in (prod.get("precios_fraccion") or {}).items():
                if isinstance(frac_val, dict):
                    fracs[frac_key] = frac_val.get("precio", 0)
                else:
                    fracs[frac_key] = frac_val

            # Precio mayorista
            ppc = prod.get("precio_por_cantidad")
            mayorista = None
            if ppc:
                mayorista = {
                    "umbral":  ppc.get("umbral", 50),
                    "precio":  ppc.get("precio_sobre_umbral", 0),
                }

            categorias[categoria].append({
                "key":           key,
                "nombre":        nombre,
                "codigo":        prod.get("codigo", ""),
                "precio":        prod.get("precio_unidad", 0),
                "stock":         stock,
                "costo":         costo,
                "fracciones":    fracs,
                "mayorista":     mayorista,
                "unidad_medida": prod.get("unidad_medida", "Unidad"),
            })

        # Ordenar por prefijo numérico de categoría y productos por nombre
        result = {}
        for cat in sorted(categorias.keys(), key=lambda c: (int(c.split()[0]) if c[0].isdigit() else 999)):
            prods = sorted(categorias[cat], key=lambda p: p["nombre"].lower())
            result[cat] = prods

        total = sum(len(v) for v in result.values())
        return {"categorias": result, "total": total, "query": q}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Edición de precios desde el dashboard ────────────────────────────────────
class PrecioUpdate(BaseModel):
    precio: Union[float, int]

class FraccionesUpdate(BaseModel):
    fracciones: dict   # { "1/4": 8000, "1/2": 13000, ... }

class MayoristaUpdate(BaseModel):
    precio: Union[float, int]
    umbral: Optional[int] = None   # Si None, conserva el umbral existente

class NuevoProducto(BaseModel):
    nombre:          str
    categoria:       str
    precio_unidad:   Union[float, int]
    unidad_medida:   str  = "Unidad"
    codigo:          str  = ""
    stock_inicial:   Union[float, int, None] = None
    codigo_dian:     str  = "94"
    inventariable:   bool = True
    visible_facturas:bool = True
    stock_minimo:    int  = 0

@app.post("/catalogo")
def crear_producto(body: NuevoProducto):
    """
    Crea un producto nuevo en memoria.json y en BASE_DE_DATOS_PRODUCTOS.xlsx.
    """
    try:
        from utils import _normalizar
        from precio_sync import agregar_producto_a_excel, _normalizar_unidad

        if not body.nombre.strip():
            raise HTTPException(status_code=400, detail="El nombre es obligatorio")

        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)

        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})

        # Generar clave única
        key_base = _normalizar(body.nombre.strip()).replace(" ", "_")
        key = key_base
        sufijo = 2
        while key in catalogo:
            key = f"{key_base}_{sufijo}"
            sufijo += 1

        unidad_norm = _normalizar_unidad(body.unidad_medida)

        nuevo = {
            "nombre":        body.nombre.strip(),
            "nombre_lower":  _normalizar(body.nombre.strip()),
            "categoria":     body.categoria.strip(),
            "precio_unidad": int(body.precio_unidad),
            "unidad_medida": unidad_norm,
        }
        if body.codigo.strip():
            nuevo["codigo"] = body.codigo.strip()

        catalogo[key] = nuevo

        # Stock inicial — guardar en formato dict igual al que usa el bot
        # (float plano hace que descontar_inventario() retorne False silenciosamente)
        if body.stock_inicial is not None:
            from datetime import datetime as _dt
            inventario[key] = {
                "nombre_original": body.nombre.strip(),
                "cantidad":        float(body.stock_inicial),
                "minimo":          body.stock_minimo if hasattr(body, "stock_minimo") else 0,
                "unidad":          "und",
                "fecha_conteo":    _dt.now().strftime("%Y-%m-%d %H:%M"),
            }

        mem["catalogo"]   = catalogo
        mem["inventario"] = inventario

        # guardar_memoria sube a Drive automáticamente (urgente=True evita pérdida en reinicios)
        try:
            from memoria import guardar_memoria, invalidar_cache_memoria
            guardar_memoria(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            # Fallback: escritura directa si el módulo no está disponible
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        # Intentar escribir en el Excel de productos (no bloquea si falla)
        excel_resultado = {"ok": False, "error": "no intentado"}
        try:
            excel_resultado = agregar_producto_a_excel({
                "codigo":          body.codigo.strip() or key,
                "nombre":          body.nombre.strip(),
                "categoria":       body.categoria.strip(),
                "precio_unidad":   int(body.precio_unidad),
                "unidad_medida":   unidad_norm,
                "inventariable":   body.inventariable,
                "visible_facturas":body.visible_facturas,
                "stock_minimo":    body.stock_minimo,
                "codigo_dian":     body.codigo_dian,
            })
        except Exception as e_excel:
            excel_resultado = {"ok": False, "error": str(e_excel)}

        return {
            "ok":             True,
            "key":            key,
            "nombre":         nuevo["nombre"],
            "categoria":      nuevo["categoria"],
            "precio_unidad":  nuevo["precio_unidad"],
            "unidad_medida":  unidad_norm,
            "stock_inicial":  body.stock_inicial,
            "excel_guardado": excel_resultado.get("ok", False),
            "excel_detalle":  excel_resultado,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/catalogo/{key:path}/precio")
def actualizar_precio_endpoint(key: str, body: PrecioUpdate):
    """
    Actualiza precio_unidad de un producto.
    1. Guarda en memoria.json (inmediato).
    2. Encola actualización en BASE_DE_DATOS_PRODUCTOS.xlsx via precio_sync.
    """
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo = mem.get("catalogo", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        nombre_prod     = catalogo[key].get("nombre", key)
        precio_anterior = catalogo[key].get("precio_unidad", 0)
        nuevo_precio    = int(body.precio)

        # 1 ── memoria.json + Drive (guardar_memoria sincroniza ambos)
        catalogo[key]["precio_unidad"] = nuevo_precio
        mem["catalogo"] = catalogo
        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        # 2 ── Excel BASE_DE_DATOS_PRODUCTOS (via precio_sync, cola FIFO)
        try:
            from precio_sync import actualizar_precio as _sync_precio
            _sync_precio(nombre_prod, nuevo_precio, None)
        except Exception as e_sync:
            import logging
            logging.getLogger("ferrebot.api").warning(
                f"precio_sync falló para '{nombre_prod}': {e_sync}"
            )

        return {
            "ok":            True,
            "key":           key,
            "nombre":        nombre_prod,
            "precio_anterior": precio_anterior,
            "precio_nuevo":  nuevo_precio,
            "excel_encolado": True,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/catalogo/{key:path}/fracciones")
def actualizar_fracciones(key: str, body: FraccionesUpdate):
    """
    Actualiza precios_fraccion de un producto.
    1. Guarda en memoria.json (inmediato).
    2. Encola cada fracción en BASE_DE_DATOS_PRODUCTOS.xlsx via precio_sync.
    """
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo = mem.get("catalogo", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        nombre_prod = catalogo[key].get("nombre", key)

        # Convertir a { frac: { "precio": X } } si llegan como { frac: precio }
        fracs_formateadas = {}
        for k, v in body.fracciones.items():
            if isinstance(v, dict):
                fracs_formateadas[k] = v
            else:
                fracs_formateadas[k] = {"precio": int(v)}

        # 1 ── memoria.json + Drive
        catalogo[key]["precios_fraccion"] = fracs_formateadas

        # Si fracción "1" (unidad completa) cambió, sincronizar precio_unidad para que la IA cotice bien
        if "1" in fracs_formateadas:
            precio_unidad_nuevo = fracs_formateadas["1"].get("precio") if isinstance(fracs_formateadas["1"], dict) else int(fracs_formateadas["1"])
            if precio_unidad_nuevo:
                catalogo[key]["precio_unidad"] = precio_unidad_nuevo

        mem["catalogo"] = catalogo
        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        # 2 ── Excel: encolar cada fracción via precio_sync
        try:
            from precio_sync import actualizar_precio as _sync_precio
            for frac_key, frac_val in fracs_formateadas.items():
                precio_frac = frac_val["precio"] if isinstance(frac_val, dict) else int(frac_val)
                if precio_frac > 0:
                    _sync_precio(nombre_prod, precio_frac, frac_key)
        except Exception as e_sync:
            import logging
            logging.getLogger("ferrebot.api").warning(
                f"precio_sync fracciones falló para '{nombre_prod}': {e_sync}"
            )

        return {"ok": True, "key": key, "nombre": nombre_prod, "fracciones": fracs_formateadas}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Sync inverso: Excel → memoria.json ───────────────────────────────────────
@app.post("/catalogo/sync-desde-excel")
def sync_catalogo_desde_excel():
    """
    Descarga BASE_DE_DATOS_PRODUCTOS.xlsx desde Drive y reimporta
    todos los precios a memoria.json.
    Útil cuando el Excel se edita directamente (no desde el dashboard).
    """
    import tempfile, os
    try:
        from drive import descargar_de_drive
        from precio_sync import importar_catalogo_desde_excel

        # Descargar Excel fresco de Drive a un archivo temporal
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            ruta_tmp = tmp.name

        try:
            ok = descargar_de_drive("BASE_DE_DATOS_PRODUCTOS.xlsx", ruta_tmp)
            if not ok:
                raise HTTPException(status_code=502, detail="No se pudo descargar el Excel de Drive")

            resultado = importar_catalogo_desde_excel(ruta_tmp)
        finally:
            try:
                os.unlink(ruta_tmp)
            except Exception:
                pass

        if resultado.get("errores"):
            logging.getLogger("ferrebot.api").warning(
                f"sync-desde-excel errores parciales: {resultado['errores']}"
            )

        return {
            "ok":         True,
            "importados": resultado.get("importados", 0),
            "omitidos":   resultado.get("omitidos", 0),
            "errores":    resultado.get("errores", []),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Kárdex por producto ───────────────────────────────────────────────────────
@app.get("/kardex")
def kardex(q: str = Query(default="")):
    """
    Devuelve el kárdex de movimientos de inventario.
    - Entradas: hoja Compras del Excel + historial_compras de memoria.json
    - Salidas: calculadas como diferencia (entradas - stock actual)
    - Si q != "" filtra por nombre de producto.
    """
    try:
        compras_excel = _leer_excel_compras()

        mem = {}
        if os.path.exists(config.MEMORIA_FILE):
            with open(config.MEMORIA_FILE, encoding="utf-8") as _f:
                mem = json.load(_f)
        inventario = mem.get("inventario", {})
        q_lower    = q.strip().lower()

        # Agrupar entradas por producto
        entradas_por_prod: dict[str, list] = defaultdict(list)
        for c in compras_excel:
            nombre = c["producto"].strip()
            if not nombre:
                continue
            if q_lower and q_lower not in nombre.lower():
                continue
            entradas_por_prod[nombre].append(c)

        kardex_items = []
        for nombre, entradas in entradas_por_prod.items():
            # Buscar stock actual en inventario
            stock_actual = 0.0
            costo_actual = 0.0
            for clave, datos in inventario.items():
                if isinstance(datos, dict):
                    n = datos.get("nombre_original", "").lower()
                    if n == nombre.lower() or clave.replace("_", " ") == nombre.lower():
                        stock_actual = _to_float(datos.get("cantidad", 0))
                        costo_actual = _to_float(datos.get("costo_promedio", 0))
                        break

            # Construir movimientos con saldo running
            movimientos = []
            saldo      = 0.0
            costo_prom = 0.0
            total_entradas = 0.0

            for e in sorted(entradas, key=lambda x: (x["fecha"], x["hora"])):
                cant  = e["cantidad"]
                costo = e["costo_unitario"]
                # Recalcular promedio ponderado
                if saldo + cant > 0:
                    costo_prom = round((saldo * costo_prom + cant * costo) / (saldo + cant))
                saldo         += cant
                total_entradas += cant
                movimientos.append({
                    "tipo":           "entrada",
                    "fecha":          e["fecha"],
                    "hora":           e["hora"],
                    "concepto":       f"Compra — {e['proveedor']}",
                    "entrada":        cant,
                    "salida":         0,
                    "saldo":          round(saldo, 3),
                    "costo_unitario": costo,
                    "costo_promedio": costo_prom,
                    "valor_total":    round(cant * costo),
                })

            # Salidas estimadas = total_entradas - stock_actual (si hay inventario registrado)
            salidas_est = round(max(0.0, total_entradas - stock_actual), 3)

            kardex_items.append({
                "producto":       nombre,
                "total_entradas": round(total_entradas, 3),
                "stock_actual":   stock_actual,
                "salidas_est":    salidas_est,
                "costo_promedio": costo_actual or costo_prom,
                "valor_inventario": round(stock_actual * (costo_actual or costo_prom)),
                "movimientos":    movimientos,
            })

        kardex_items.sort(key=lambda x: x["producto"].lower())
        total_valor_inv = sum(k["valor_inventario"] for k in kardex_items)

        return {
            "kardex":          kardex_items,
            "total_productos": len(kardex_items),
            "valor_inventario_total": total_valor_inv,
            "tiene_datos":     len(kardex_items) > 0,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Estado de Resultados ──────────────────────────────────────────────────────
@app.get("/resultados")
def resultados(periodo: str = Query(default="mes", pattern="^(semana|mes)$")):
    """
    Estado de Resultados:
      Ingresos (ventas)
    — CMV (costo de mercancía vendida = compras del período × promedio ponderado)
    = Utilidad Bruta
    — Gastos operativos
    = Utilidad Neta
    """
    try:
        ahora = datetime.now(config.COLOMBIA_TZ)

        # ── 1. INGRESOS ──────────────────────────────────────────────────────
        dias_rango = 7 if periodo == "semana" else None
        es_mes     = periodo == "mes"
        ventas     = _leer_excel_rango(dias=dias_rango, mes_actual=es_mes)

        total_ventas = sum(_to_float(v.get("total", 0)) for v in ventas)

        # Ventas por día para gráfica
        ventas_por_dia: dict[str, float] = defaultdict(float)
        for v in ventas:
            ventas_por_dia[v["fecha"]] += _to_float(v.get("total", 0))

        # ── 2. CMV ───────────────────────────────────────────────────────────
        # Costo de lo vendido = unidades vendidas × costo_promedio del producto
        mem = {}
        if os.path.exists(config.MEMORIA_FILE):
            with open(config.MEMORIA_FILE, encoding="utf-8") as _f:
                mem = json.load(_f)
        inventario = mem.get("inventario", {})

        # Índice de inventario por nombre normalizado
        inv_idx: dict[str, dict] = {}
        for clave, datos in inventario.items():
            if isinstance(datos, dict):
                nombre_n = datos.get("nombre_original", "").lower().strip()
                inv_idx[nombre_n] = datos
                inv_idx[clave.replace("_", " ")] = datos

        # Agrupar ventas por producto
        ventas_prod: dict[str, dict] = defaultdict(lambda: {"cantidad": 0.0, "ingresos": 0.0})
        for v in ventas:
            nombre = str(v.get("producto", "")).strip()
            if not nombre:
                continue
            cant = _to_float(v.get("cantidad", 1))
            ventas_prod[nombre]["cantidad"] += cant
            ventas_prod[nombre]["ingresos"] += _to_float(v.get("total", 0))

        cmv           = 0.0
        cmv_detalle   = []
        sin_costo     = []

        for nombre, datos_v in ventas_prod.items():
            datos_inv = inv_idx.get(nombre.lower().strip())
            costo_u   = _to_float(datos_inv.get("costo_promedio", 0)) if datos_inv else 0

            if costo_u > 0:
                costo_total_prod = costo_u * datos_v["cantidad"]
                cmv             += costo_total_prod
                margen = round(((datos_v["ingresos"] - costo_total_prod) / datos_v["ingresos"]) * 100, 1) if datos_v["ingresos"] else 0
                cmv_detalle.append({
                    "producto":    nombre,
                    "cantidad":    round(datos_v["cantidad"], 3),
                    "ingresos":    round(datos_v["ingresos"]),
                    "costo_unit":  costo_u,
                    "cmv":         round(costo_total_prod),
                    "margen_pct":  margen,
                })
            else:
                sin_costo.append(nombre)

        cmv_detalle.sort(key=lambda x: -x["cmv"])

        # ── 3. GASTOS ────────────────────────────────────────────────────────
        gastos_mem    = mem.get("gastos", {})
        if periodo == "semana":
            limite_g  = (ahora - timedelta(days=7)).strftime("%Y-%m-%d")
        else:
            limite_g  = f"{ahora.year}-{ahora.month:02d}-01"

        total_gastos      = 0.0
        gastos_por_cat: dict[str, float] = defaultdict(float)
        for fecha_g, lista_g in gastos_mem.items():
            if fecha_g < limite_g:
                continue
            for g in lista_g:
                monto = _to_float(g.get("monto", 0))
                total_gastos += monto
                gastos_por_cat[g.get("categoria", "Sin categoría")] += monto

        # ── 4. RESULTADOS ────────────────────────────────────────────────────
        utilidad_bruta = total_ventas - cmv
        utilidad_neta  = utilidad_bruta - total_gastos
        margen_bruto   = round((utilidad_bruta / total_ventas) * 100, 1) if total_ventas else 0
        margen_neto    = round((utilidad_neta  / total_ventas) * 100, 1) if total_ventas else 0

        # Histórico diario para gráfica
        dias_n = 7 if periodo == "semana" else ahora.day
        historico = []
        for i in range(dias_n - 1, -1, -1):
            dia = (ahora - timedelta(days=i)).strftime("%Y-%m-%d")
            gastos_dia = sum(_to_float(g.get("monto", 0))
                             for g in gastos_mem.get(dia, []))
            historico.append({
                "fecha":   dia,
                "ventas":  ventas_por_dia.get(dia, 0),
                "gastos":  gastos_dia,
            })

        return {
            "periodo":          periodo,
            "total_ventas":     round(total_ventas),
            "cmv":              round(cmv),
            "utilidad_bruta":   round(utilidad_bruta),
            "total_gastos":     round(total_gastos),
            "utilidad_neta":    round(utilidad_neta),
            "margen_bruto_pct": margen_bruto,
            "margen_neto_pct":  margen_neto,
            "cmv_detalle":      cmv_detalle,
            "sin_costo":        sin_costo[:20],
            "gastos_por_cat":   dict(gastos_por_cat),
            "historico":        historico,
            "tiene_cmv":        cmv > 0,
            "cobertura_cmv_pct": round((len(cmv_detalle) / max(len(ventas_prod), 1)) * 100, 1),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Proyección de Caja ────────────────────────────────────────────────────────
@app.get("/proyeccion")
def proyeccion():
    """
    Proyecta el cierre del mes basándose en promedios de los últimos 14 días.
    Proyección = efectivo_actual + (ingreso_diario_prom - gasto_diario_prom) × días_restantes
    """
    try:
        ahora     = datetime.now(config.COLOMBIA_TZ)
        mem = {}
        if os.path.exists(config.MEMORIA_FILE):
            with open(config.MEMORIA_FILE, encoding="utf-8") as _f:
                mem = json.load(_f)
        caja_data = mem.get("caja_actual", {})

        # ── Base de caja actual ───────────────────────────────────────────
        efectivo_actual = (
            _to_float(caja_data.get("efectivo", 0)) +
            _to_float(caja_data.get("transferencias", 0)) +
            _to_float(caja_data.get("datafono", 0)) +
            _to_float(caja_data.get("monto_apertura", 0))
        )

        # ── Ventas últimos 14 días ────────────────────────────────────────
        ventas_14 = _leer_excel_rango(dias=14)
        ventas_por_dia14: dict[str, float] = defaultdict(float)
        for v in ventas_14:
            ventas_por_dia14[v["fecha"]] += _to_float(v.get("total", 0))

        dias_con_ventas = [d for d, t in ventas_por_dia14.items() if t > 0]
        prom_ventas_dia = (
            sum(ventas_por_dia14.values()) / len(dias_con_ventas)
            if dias_con_ventas else 0
        )

        # ── Gastos últimos 14 días ────────────────────────────────────────
        gastos_mem = mem.get("gastos", {})
        limite_14  = (ahora - timedelta(days=14)).strftime("%Y-%m-%d")
        total_gastos_14 = 0.0
        gastos_por_dia14: dict[str, float] = defaultdict(float)
        for fecha_g, lista_g in gastos_mem.items():
            if fecha_g < limite_14:
                continue
            for g in lista_g:
                m = _to_float(g.get("monto", 0))
                total_gastos_14         += m
                gastos_por_dia14[fecha_g] += m

        dias_con_gastos = max(len([d for d, t in gastos_por_dia14.items() if t > 0]), 1)
        prom_gastos_dia = total_gastos_14 / dias_con_gastos if total_gastos_14 else 0

        # ── Días restantes del mes ────────────────────────────────────────
        import calendar
        ultimo_dia   = calendar.monthrange(ahora.year, ahora.month)[1]
        dias_rest    = ultimo_dia - ahora.day
        dias_pasados = ahora.day

        # ── Proyecciones ──────────────────────────────────────────────────
        ventas_proy_rest  = prom_ventas_dia * dias_rest
        gastos_proy_rest  = prom_gastos_dia * dias_rest
        ventas_mes_total  = sum(v for d, v in ventas_por_dia14.items()
                                if d.startswith(f"{ahora.year}-{ahora.month:02d}"))
        gastos_mes_total  = sum(t for d, t in gastos_por_dia14.items()
                                if d >= f"{ahora.year}-{ahora.month:02d}-01")

        proy_ventas_mes   = ventas_mes_total  + ventas_proy_rest
        proy_gastos_mes   = gastos_mes_total  + gastos_proy_rest
        proy_caja_fin_mes = efectivo_actual   + ventas_proy_rest - gastos_proy_rest

        # Serie diaria para gráfica (real + proyectado)
        serie = []
        acum  = _to_float(caja_data.get("monto_apertura", 0))
        for i in range(1, ultimo_dia + 1):
            dia_str = f"{ahora.year}-{ahora.month:02d}-{i:02d}"
            if i < ahora.day:
                # Días pasados — datos reales
                v = ventas_por_dia14.get(dia_str, 0)
                g = sum(_to_float(x.get("monto", 0)) for x in gastos_mem.get(dia_str, []))
                acum += v - g
                serie.append({"dia": i, "valor": round(acum), "real": True})
            elif i == ahora.day:
                serie.append({"dia": i, "valor": round(efectivo_actual), "real": True, "hoy": True})
            else:
                acum_proy = efectivo_actual + (prom_ventas_dia - prom_gastos_dia) * (i - ahora.day)
                serie.append({"dia": i, "valor": round(acum_proy), "real": False})

        return {
            "hoy":                  ahora.strftime("%Y-%m-%d"),
            "dia_del_mes":          ahora.day,
            "dias_restantes":       dias_rest,
            "dias_pasados":         dias_pasados,
            "efectivo_actual":      round(efectivo_actual),
            "prom_ventas_dia":      round(prom_ventas_dia),
            "prom_gastos_dia":      round(prom_gastos_dia),
            "prom_neto_dia":        round(prom_ventas_dia - prom_gastos_dia),
            "ventas_mes_actual":    round(ventas_mes_total),
            "gastos_mes_actual":    round(gastos_mes_total),
            "proy_ventas_mes":      round(proy_ventas_mes),
            "proy_gastos_mes":      round(proy_gastos_mes),
            "proy_caja_fin_mes":    round(proy_caja_fin_mes),
            "serie_diaria":         serie,
            "tiene_datos":          prom_ventas_dia > 0,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class StockUpdate(BaseModel):
    stock: Union[float, int, None]

@app.patch("/catalogo/{key:path}/mayorista")
def actualizar_mayorista(key: str, body: MayoristaUpdate):
    """
    Actualiza el precio mayorista (precio_por_cantidad) de un producto.
    Guarda precio_sobre_umbral en memoria.json y sincroniza al Excel.
    """
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo = mem.get("catalogo", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        prod = catalogo[key]
        nombre_prod = prod.get("nombre", key)
        ppc_actual  = prod.get("precio_por_cantidad")

        # Preservar umbral existente si no se manda uno nuevo
        umbral = body.umbral if body.umbral else (ppc_actual.get("umbral", 50) if ppc_actual else 50)

        prod["precio_por_cantidad"] = {
            "umbral":              umbral,
            "precio_bajo_umbral":  ppc_actual.get("precio_bajo_umbral", prod.get("precio_unidad", 0)) if ppc_actual else prod.get("precio_unidad", 0),
            "precio_sobre_umbral": int(body.precio),
        }
        catalogo[key] = prod
        mem["catalogo"] = catalogo

        # Usar guardar_memoria() para que también suba a Drive (antes solo hacía open+write)
        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            # Fallback: al menos guardar en disco si memoria.py falla
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        return {
            "ok": True, "key": key, "nombre": nombre_prod,
            "precio_mayorista": int(body.precio), "umbral": umbral,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/inventario/{key:path}/stock")
def actualizar_stock(key: str, body: StockUpdate):
    """
    Actualiza cantidad en inventario de un producto (memoria.json).
    Guarda en el mismo formato que usa el bot: {"cantidad": X, "nombre_original": ..., "minimo": N}
    para mantener sincronía completa bot <-> dashboard.
    """
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        nombre_prod    = catalogo[key].get("nombre", key)
        entrada_actual = inventario.get(key)

        # Extraer stock_anterior independientemente del formato (dict o número)
        if isinstance(entrada_actual, dict):
            stock_anterior = entrada_actual.get("cantidad")
        elif entrada_actual is not None:
            stock_anterior = float(entrada_actual)
        else:
            stock_anterior = None

        if body.stock is None:
            inventario.pop(key, None)
        else:
            # Preservar minimo si ya existe, sino usar 0
            minimo_actual = 0
            if isinstance(entrada_actual, dict):
                minimo_actual = entrada_actual.get("minimo", 0)

            from datetime import datetime
            inventario[key] = {
                "nombre_original": nombre_prod,
                "cantidad":        float(body.stock),
                "minimo":          minimo_actual,
                "unidad":          "und",
                "fecha_conteo":    datetime.now().strftime("%Y-%m-%d %H:%M"),
            }

        mem["inventario"] = inventario
        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        return {
            "ok": True, "key": key,
            "nombre":         nombre_prod,
            "stock_anterior": stock_anterior,
            "stock_nuevo":    body.stock,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Health check ──────────────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    return {"estado": "activo", "version": "1.0.0"}


# ── Clientes ──────────────────────────────────────────────────────────────────
@app.get("/clientes/buscar")
def buscar_clientes_endpoint(q: str = Query(default="")):
    """
    Busca clientes en la hoja 'Clientes' del Excel por nombre o identificación.
    Devuelve lista de coincidencias para el autocompletado del dashboard.
    """
    try:
        from excel import cargar_clientes
        from utils import _normalizar
        clientes = cargar_clientes()
        if not q.strip():
            return {"clientes": clientes[:10], "total": len(clientes)}

        q_norm = _normalizar(q.strip())
        resultado = []
        for c in clientes:
            nombre_norm = _normalizar(c.get("Nombre tercero", "") or "")
            id_norm     = _normalizar(str(c.get("Identificacion", "") or ""))
            # Buscar la query completa como substring (para "rene acosta" → encuentra "rene acosta medina")
            # O buscar cada palabra individualmente
            if q_norm in nombre_norm or q_norm in id_norm:
                resultado.append(c)
            else:
                palabras = [p for p in q_norm.split() if p]
                if palabras and all(p in nombre_norm or p in id_norm for p in palabras):
                    resultado.append(c)
        resultado.sort(key=lambda x: len(str(x.get("Nombre tercero", ""))))
        return {"clientes": resultado[:10], "total": len(resultado)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class NuevoCliente(BaseModel):
    nombre:         str
    tipo_id:        str  = "CC"   # CC | NIT | CE | PAS
    identificacion: str  = ""
    tipo_persona:   str  = "Natural"
    correo:         str  = ""
    telefono:       str  = ""
    direccion:      str  = ""

@app.post("/clientes")
def crear_cliente_endpoint(body: NuevoCliente):
    """Crea un cliente nuevo en la hoja Clientes del Excel."""
    try:
        from excel import guardar_cliente_nuevo, buscar_cliente, buscar_clientes_multiples
        from utils import _normalizar

        if not body.nombre.strip():
            raise HTTPException(status_code=400, detail="El nombre es obligatorio")

        # ── Verificar duplicados ──────────────────────────────────────────────
        # 1. Por identificación (exacto) si viene informada
        if body.identificacion.strip():
            existente = buscar_cliente(body.identificacion.strip())
            if existente:
                return {
                    "ok":      True,
                    "existia": True,
                    "cliente": existente,
                    "mensaje": "El cliente ya estaba registrado (misma identificación)",
                }

        # 2. Por nombre (flexible) — evita duplicados cuando no hay cédula
        candidatos_nombre = buscar_clientes_multiples(body.nombre.strip(), limite=3)
        for candidato in candidatos_nombre:
            nombre_existente = _normalizar(candidato.get("Nombre tercero", "") or "")
            nombre_nuevo     = _normalizar(body.nombre.strip())
            # Coincidencia de ≥80 % de palabras → considerar duplicado
            palabras_ex  = set(nombre_existente.split())
            palabras_nu  = set(nombre_nuevo.split())
            if palabras_ex and palabras_nu:
                interseccion = palabras_ex & palabras_nu
                similitud    = len(interseccion) / max(len(palabras_ex), len(palabras_nu))
                if similitud >= 0.6:
                    return {
                        "ok":      True,
                        "existia": True,
                        "cliente": candidato,
                        "mensaje": f"Ya existe un cliente con nombre similar: '{candidato.get('Nombre tercero')}'",
                    }

        ok = guardar_cliente_nuevo(
            nombre         = body.nombre.strip(),
            tipo_id        = body.tipo_id,
            identificacion = body.identificacion.strip(),
            tipo_persona   = body.tipo_persona,
            correo         = body.correo.strip(),
            telefono       = body.telefono.strip(),
            direccion      = body.direccion.strip(),
        )
        if not ok:
            raise HTTPException(status_code=500, detail="Error guardando cliente en Excel")
        return {"ok": True, "existia": False, "mensaje": f"Cliente '{body.nombre.strip().upper()}' creado"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Editar / Eliminar Ventas ──────────────────────────────────────────────────
@app.delete("/ventas/{numero}")
def eliminar_venta(numero: int):
    """
    Elimina todas las filas de un consecutivo de venta del Excel y Sheets.
    También descuenta el total de la caja si era de hoy.
    """
    try:
        from excel import borrar_venta_excel, recalcular_caja_desde_excel
        ok, msg = borrar_venta_excel(numero)
        if ok:
            recalcular_caja_desde_excel()
        # Si no se encontró, devolver 404 para que el frontend lo muestre
        if not ok:
            raise HTTPException(status_code=404, detail=msg)
        return {"ok": ok, "mensaje": msg}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/ventas/{numero}/linea")
def eliminar_linea_venta(numero: int, producto: str = Query(...)):
    """
    Elimina UNA sola línea (producto) de un consecutivo multi-producto.
    Busca por consecutivo + nombre de producto exacto en Excel y Sheets.
    """
    try:
        import openpyxl
        from excel import inicializar_excel, obtener_nombre_hoja, detectar_columnas, recalcular_caja_desde_excel
        from drive import subir_a_drive

        inicializar_excel()
        wb    = openpyxl.load_workbook(config.EXCEL_FILE)
        hojas = [obtener_nombre_hoja(), "Registro de Ventas-Acumulado"]
        total_borradas = 0

        for nombre_sh in hojas:
            if nombre_sh not in wb.sheetnames:
                continue
            ws   = wb[nombre_sh]
            cols = detectar_columnas(ws)
            col_id   = cols.get("consecutivo de venta") or cols.get("consecutivo") or cols.get("alias")
            col_prod = cols.get("producto")
            if not col_id or not col_prod:
                continue

            filas_borrar = []
            for fila in range(config.EXCEL_FILA_DATOS, ws.max_row + 1):
                val_id = ws.cell(row=fila, column=col_id).value
                val_prod = str(ws.cell(row=fila, column=col_prod).value or "").strip()
                try:
                    if val_id is not None and int(float(str(val_id))) == numero:
                        if val_prod.lower() == producto.lower():
                            filas_borrar.append(fila)
                except (ValueError, TypeError):
                    pass

            for fila in reversed(filas_borrar):
                ws.delete_rows(fila)
            total_borradas += len(filas_borrar)

        if total_borradas:
            wb.save(config.EXCEL_FILE)
            try:
                subir_a_drive(config.EXCEL_FILE)
            except Exception:
                pass
            recalcular_caja_desde_excel()

            # Borrar de Sheets también
            try:
                from sheets import _obtener_hoja_sheets, _invalidar_ws_cache
                ws_sh = _obtener_hoja_sheets()
                if ws_sh:
                    todas = ws_sh.get_all_values()
                    headers = [h.upper().strip() for h in todas[0]] if todas else []
                    col_consec = None
                    col_prod_sh = None
                    for i, h in enumerate(headers):
                        if "CONSECUTIVO" in h or h == "#":
                            col_consec = i
                        if h == "PRODUCTO":
                            col_prod_sh = i
                    if col_consec is not None and col_prod_sh is not None:
                        filas_sh = []
                        for idx, fila in enumerate(todas[1:], start=2):
                            try:
                                if int(float(str(fila[col_consec]).strip())) == numero:
                                    if fila[col_prod_sh].strip().lower() == producto.lower():
                                        filas_sh.append(idx)
                            except (ValueError, IndexError):
                                pass
                        for fila_idx in reversed(filas_sh):
                            ws_sh.delete_rows(fila_idx)
                        if filas_sh:
                            _invalidar_ws_cache()
            except Exception:
                pass

            return {"ok": True, "borradas": total_borradas, "mensaje": f"'{producto}' eliminado del consecutivo #{numero}"}

        raise HTTPException(status_code=404, detail=f"No se encontró '{producto}' en consecutivo #{numero}")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class EditarVentaBody(BaseModel):
    producto:          Union[str, None]   = None
    cantidad:          Union[float, None] = None
    precio_unitario:   Union[float, None] = None
    total:             Union[float, None] = None
    metodo_pago:       Union[str, None]   = None
    cliente:           Union[str, None]   = None
    id_cliente:        Union[str, None]   = None
    vendedor:          Union[str, None]   = None
    producto_original: Union[str, None]   = None  # para identificar fila en multi-producto

@app.patch("/ventas/{numero}")
def editar_venta(numero: int, body: EditarVentaBody):
    """
    Edita los campos de un consecutivo en el Excel (hoja mensual + Acumulado)
    y sincroniza los cambios a Google Sheets.
    Si producto_original viene, solo actualiza la fila con ese producto (multi-producto).
    """
    try:
        import openpyxl
        from excel import inicializar_excel, obtener_nombre_hoja, detectar_columnas, recalcular_caja_desde_excel
        from drive import subir_a_drive
        from sheets import sheets_editar_consecutivo

        inicializar_excel()
        wb          = openpyxl.load_workbook(config.EXCEL_FILE)
        hojas       = [obtener_nombre_hoja(), "Registro de Ventas-Acumulado"]
        actualizadas = 0

        cambios = {k: v for k, v in body.dict().items() if v is not None and k != "producto_original"}
        if not cambios:
            raise HTTPException(status_code=400, detail="No hay campos para actualizar")

        # Filtro por producto para multi-producto
        filtro_producto = body.producto_original.strip().lower() if body.producto_original else None

        CAMPO_COL = {
            "producto":        ["producto"],
            "cantidad":        ["cantidad"],
            "precio_unitario": ["valor unitario", "precio unitario"],
            "total":           ["total"],
            "metodo_pago":     ["metodo de pago", "metodo pago"],
            "cliente":         ["cliente"],
            "id_cliente":      ["id cliente"],
            "vendedor":        ["vendedor"],
        }

        for nombre_sh in hojas:
            if nombre_sh not in wb.sheetnames:
                continue
            ws     = wb[nombre_sh]
            cols   = detectar_columnas(ws)
            col_id = cols.get("consecutivo de venta") or cols.get("alias")
            col_prod = cols.get("producto")
            if not col_id:
                continue

            for fila in range(config.EXCEL_FILA_DATOS, ws.max_row + 1):
                val = ws.cell(row=fila, column=col_id).value
                try:
                    if val is None or int(float(str(val))) != numero:
                        continue
                except (ValueError, TypeError):
                    continue

                # Si hay filtro de producto, solo actualizar la fila que coincida
                if filtro_producto and col_prod:
                    prod_fila = str(ws.cell(row=fila, column=col_prod).value or "").strip().lower()
                    if prod_fila != filtro_producto:
                        continue

                for campo, valor in cambios.items():
                    claves = CAMPO_COL.get(campo, [campo.replace("_", " ")])
                    col_destino = None
                    for clave in claves:
                        col_destino = cols.get(clave)
                        if col_destino:
                            break
                    if col_destino:
                        ws.cell(row=fila, column=col_destino).value = valor
                        actualizadas += 1

        if actualizadas:
            wb.save(config.EXCEL_FILE)
            try:
                subir_a_drive(config.EXCEL_FILE)
            except Exception:
                pass
            recalcular_caja_desde_excel()
            # ── Sincronizar a Google Sheets ───────────────────────────────
            try:
                sheets_editar_consecutivo(numero, cambios, producto_original=body.producto_original)
            except Exception:
                pass   # No fallar la respuesta si Sheets falla
            return {"ok": True, "actualizadas": actualizadas, "mensaje": f"Venta #{numero} actualizada"}

        return {"ok": False, "mensaje": f"No se encontró el consecutivo #{numero}"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Editar / Eliminar Productos ───────────────────────────────────────────────
class EditarProductoBody(BaseModel):
    nombre:        Union[str, None]   = None
    categoria:     Union[str, None]   = None
    precio_unidad: Union[float, None] = None
    unidad_medida: Union[str, None]   = None
    codigo:        Union[str, None]   = None

@app.patch("/catalogo/{key:path}")
def editar_producto(key: str, body: EditarProductoBody):
    """Edita nombre, categoría, precio, unidad_medida o código de un producto."""
    try:
        from utils import _normalizar
        from precio_sync import actualizar_precio as _sync_precio, _normalizar_unidad
        from memoria import invalidar_cache_memoria

        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo = mem.get("catalogo", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        prod    = catalogo[key]
        cambios = {k: v for k, v in body.dict().items() if v is not None}
        if not cambios:
            raise HTTPException(status_code=400, detail="Sin campos para actualizar")

        nueva_clave = key
        if "nombre" in cambios:
            prod["nombre"]       = cambios["nombre"].strip()
            prod["nombre_lower"] = _normalizar(cambios["nombre"].strip())
            nueva_clave          = prod["nombre_lower"].replace(" ", "_")

        if "categoria"     in cambios: prod["categoria"]     = cambios["categoria"].strip()
        if "precio_unidad" in cambios: prod["precio_unidad"] = int(cambios["precio_unidad"])
        if "codigo"        in cambios: prod["codigo"]        = cambios["codigo"].strip()
        if "unidad_medida" in cambios: prod["unidad_medida"] = _normalizar_unidad(cambios["unidad_medida"])

        # Si cambió el nombre → mover a nueva clave
        if nueva_clave != key:
            inv = mem.get("inventario", {})
            catalogo[nueva_clave] = prod
            del catalogo[key]
            if key in inv:
                inv[nueva_clave] = inv.pop(key)
            mem["inventario"] = inv
        else:
            catalogo[key] = prod

        mem["catalogo"] = catalogo
        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        # Sync al Excel BASE_DE_DATOS_PRODUCTOS
        # — precio si cambió
        if "precio_unidad" in cambios:
            try:
                _sync_precio(prod["nombre"], int(cambios["precio_unidad"]), None)
            except Exception:
                pass

        # — nombre, categoría o unidad_medida: actualizar fila completa en el Excel
        if any(c in cambios for c in ("nombre", "categoria", "unidad_medida", "codigo")):
            try:
                from precio_sync import _actualizar_metadatos_en_excel
                _actualizar_metadatos_en_excel(
                    nombre_original = catalogo.get(key, prod).get("nombre", prod["nombre"]) if nueva_clave == key else key.replace("_", " "),
                    datos_nuevos    = {
                        "nombre":        prod.get("nombre", ""),
                        "categoria":     prod.get("categoria", ""),
                        "unidad_medida": prod.get("unidad_medida", "Unidad"),
                        "codigo":        prod.get("codigo", ""),
                    },
                )
            except Exception as e_meta:
                logging.getLogger("ferrebot.api").warning(
                    f"_actualizar_metadatos_en_excel falló para '{prod.get('nombre')}': {e_meta}"
                )

        return {"ok": True, "key_nueva": nueva_clave, "producto": prod}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/catalogo/{key:path}")
def eliminar_producto(key: str):
    """Elimina un producto del catálogo e inventario en memoria.json."""
    try:
        from memoria import invalidar_cache_memoria
        with open(config.MEMORIA_FILE, encoding="utf-8") as f:
            mem = json.load(f)
        catalogo   = mem.get("catalogo", {})
        inventario = mem.get("inventario", {})
        if key not in catalogo:
            raise HTTPException(status_code=404, detail=f"Producto '{key}' no encontrado")

        nombre = catalogo[key].get("nombre", key)
        del catalogo[key]
        inventario.pop(key, None)
        mem["catalogo"]   = catalogo
        mem["inventario"] = inventario

        try:
            from memoria import guardar_memoria as _gm, invalidar_cache_memoria
            _gm(mem, urgente=True)
            invalidar_cache_memoria()
        except Exception:
            with open(config.MEMORIA_FILE, "w", encoding="utf-8") as f:
                json.dump(mem, f, ensure_ascii=False, indent=2)

        # Eliminar también de BASE_DE_DATOS_PRODUCTOS.xlsx para evitar que
        # un sync-desde-excel resucite el producto eliminado
        excel_resultado = {"ok": False, "error": "no intentado"}
        try:
            from precio_sync import eliminar_producto_de_excel as _del_xls
            excel_resultado = _del_xls(nombre)
        except Exception as e_xls:
            logging.getLogger("ferrebot.api").warning(
                f"eliminar_producto_de_excel falló para '{nombre}': {e_xls}"
            )

        return {
            "ok":             True,
            "nombre":         nombre,
            "mensaje":        f"'{nombre}' eliminado del catálogo",
            "excel_borrado":  excel_resultado.get("ok", False),
            "excel_detalle":  excel_resultado,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Historial manual de ventas diarias ────────────────────────────────────────

HISTORICO_FILE  = "historico_ventas.json"
HISTORICO_EXCEL = "historico_ventas.xlsx"
_NOMBRES_MES    = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                   "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# ── Helpers: total en vivo desde Sheets / Excel ──────────────────────────────

def _total_ventas_hoy_sheets() -> float:
    """
    Calcula el total de ventas del día actual desde Google Sheets (en vivo).
    Si Sheets falla, intenta Excel como fallback.
    """
    hoy = _hoy()
    try:
        ventas = sheets_leer_ventas_del_dia()
        total = 0.0
        for v in ventas:
            if str(v.get("fecha", ""))[:10] == hoy:
                try:
                    total += float(str(v.get("total", 0)).replace(",", ".") or 0)
                except (ValueError, TypeError):
                    pass
        if total > 0:
            return total
    except Exception:
        pass
    # Fallback: Excel
    try:
        ventas_xls = _leer_excel_rango(dias=1)
        return sum(
            float(str(v.get("total", 0)).replace(",", ".") or 0)
            for v in ventas_xls
            if str(v.get("fecha", ""))[:10] == hoy
        )
    except Exception:
        return 0.0


def _totales_por_dia_excel(dias: int = 30) -> dict:
    """
    Lee ventas del Excel y retorna {fecha: total} para los últimos N días.
    Útil para sincronizar históricos de días pasados.
    """
    try:
        ventas = _leer_excel_rango(dias=dias)
    except Exception:
        return {}
    por_dia: dict[str, float] = defaultdict(float)
    for v in ventas:
        fecha = str(v.get("fecha", ""))[:10]
        if fecha:
            try:
                por_dia[fecha] += float(str(v.get("total", 0)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                pass
    return {k: int(v) for k, v in por_dia.items() if v > 0}


def _sync_historico_hoy() -> dict:
    """
    Sincroniza el total de hoy (desde Sheets) al histórico persistente.
    Retorna {"fecha": ..., "monto": ..., "ok": True/False}.
    """
    hoy = _hoy()
    total = _total_ventas_hoy_sheets()
    if total <= 0:
        return {"fecha": hoy, "monto": 0, "ok": False, "razon": "sin ventas hoy"}
    monto = int(total)
    data = _leer_historico()
    data[hoy] = monto
    _guardar_historico(data)
    return {"fecha": hoy, "monto": monto, "ok": True}


# ── Helpers internos ──────────────────────────────────────────────────────────

def _excel_a_dict(ruta: str) -> dict:
    """
    Lee historico_ventas.xlsx y retorna {fecha: monto}.
    Columna A = Fecha (YYYY-MM-DD), Columna E = Monto.
    Ignora filas con fecha o monto inválidos.
    """
    data = {}
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha = str(row[0] or "").strip()
            monto = row[4]  # Col E
            if not fecha or not monto:
                continue
            try:
                # Validar formato YYYY-MM-DD
                parts = fecha.split("-")
                if len(parts) == 3 and len(parts[0]) == 4:
                    m = int(monto)
                    if m > 0:
                        data[fecha] = m
            except Exception:
                pass
    except Exception as e:
        logging.getLogger("ferrebot.api").warning(f"[historico] No se pudo leer Excel: {e}")
    return data

def _dict_a_excel(data: dict, ruta: str) -> None:
    """
    Escribe {fecha: monto} en historico_ventas.xlsx.
    Columnas: Fecha | Año | Mes | Día | Monto
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Encabezados
    headers = ["Fecha", "Año", "Mes", "Día", "Monto"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C0392B")
        c.alignment = Alignment(horizontal="center")

    # Datos ordenados por fecha
    for row_idx, (fecha, monto) in enumerate(sorted(data.items()), 2):
        try:
            parts   = fecha.split("-")
            año_n   = int(parts[0])
            mes_n   = int(parts[1])
            dia_n   = int(parts[2])
            nom_mes = _NOMBRES_MES[mes_n] if 1 <= mes_n <= 12 else str(mes_n)
            ws.cell(row=row_idx, column=1, value=fecha)
            ws.cell(row=row_idx, column=2, value=año_n)
            ws.cell(row=row_idx, column=3, value=nom_mes)
            ws.cell(row=row_idx, column=4, value=dia_n)
            ws.cell(row=row_idx, column=5, value=int(monto))
        except Exception:
            pass

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 16
    wb.save(ruta)

def _leer_historico_de_excel() -> dict:
    """
    Descarga el Excel de Drive → lee datos → retorna dict.
    Es la fuente de verdad cuando existe.
    """
    try:
        from drive import descargar_de_drive
        ruta_tmp = HISTORICO_EXCEL + ".tmp"
        if descargar_de_drive(HISTORICO_EXCEL, ruta_tmp):
            data = _excel_a_dict(ruta_tmp)
            try:
                import os as _os; _os.remove(ruta_tmp)
            except Exception:
                pass
            return data
    except Exception:
        pass
    return {}

def _leer_historico() -> dict:
    """
    Lee historial: primero intenta Excel de Drive (fuente de verdad),
    luego JSON local como fallback.
    """
    # 1. Excel en Drive (fuente de verdad)
    data = _leer_historico_de_excel()
    if data:
        # Actualizar caché JSON local
        try:
            with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return data

    # 2. JSON local como fallback
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    # 3. JSON en Drive como último recurso
    try:
        from drive import descargar_de_drive
        if descargar_de_drive(HISTORICO_FILE, HISTORICO_FILE):
            with open(HISTORICO_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _guardar_historico(data: dict) -> None:
    """
    Guarda en JSON local + Excel local, y sube ambos a Drive.
    Excel es la fuente de verdad — siempre se regenera completo.
    """
    # 1. JSON local (caché rápida para el API)
    with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 2. Excel local (fuente de verdad editable)
    _dict_a_excel(data, HISTORICO_EXCEL)

    # 3. Subir ambos a Drive
    try:
        from drive import subir_a_drive_urgente
        subir_a_drive_urgente(HISTORICO_FILE)
        subir_a_drive_urgente(HISTORICO_EXCEL)
    except Exception:
        pass

# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/historico/ventas")
def historico_ventas_get(año: int = 0, mes: int = 0):
    """
    Retorna montos del mes/año.
    SIEMPRE inyecta el total en vivo de hoy desde Sheets, para que el
    dashboard muestre el dato real aunque no se haya hecho sync manual.
    """
    data = _leer_historico()

    # ── Inyectar total en vivo del día actual ────────────────────────────
    hoy = _hoy()
    total_hoy = _total_ventas_hoy_sheets()
    if total_hoy > 0:
        data[hoy] = int(total_hoy)

    if not año and not mes:
        return data
    prefijo = f"{año}-{mes:02d}" if mes else str(año)
    return {k: v for k, v in data.items() if k.startswith(prefijo)}

class HistoricoBody(BaseModel):
    año:   int
    mes:   int
    datos: dict   # { "2026-03-01": 850000, ... }

@app.post("/historico/ventas")
def historico_ventas_post(body: HistoricoBody):
    """
    Guarda los montos de un mes.
    1. Lee estado actual del Excel (para no pisar otros meses)
    2. Fusiona con los nuevos datos
    3. Guarda JSON + Excel + sube Drive
    """
    try:
        # Leer estado actual (Excel como fuente de verdad)
        data   = _leer_historico()
        prefijo = f"{body.año}-{body.mes:02d}"

        # Eliminar entradas previas del mes
        data = {k: v for k, v in data.items() if not k.startswith(prefijo)}

        # Agregar los nuevos (solo > 0)
        for fecha, monto in body.datos.items():
            if fecha.startswith(prefijo) and monto and int(monto) > 0:
                data[fecha] = int(monto)

        _guardar_historico(data)
        registros_mes = len([v for k, v in data.items() if k.startswith(prefijo)])
        return {"ok": True, "registros": registros_mes, "total": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/historico/sincronizar-excel")
def historico_sincronizar_excel():
    """
    Lee el Excel de Drive (editado manualmente) y actualiza el JSON.
    Llamar desde dashboard cuando quieras traer cambios hechos en Excel.
    """
    try:
        data = _leer_historico_de_excel()
        if not data:
            return {"ok": False, "error": "No se encontró historico_ventas.xlsx en Drive o está vacío"}
        # Actualizar JSON local con lo que tiene el Excel
        with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            from drive import subir_a_drive_urgente
            subir_a_drive_urgente(HISTORICO_FILE)
        except Exception:
            pass
        return {"ok": True, "registros": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/historico/resumen")
def historico_resumen():
    """Totales mensuales para gráficas — incluye hoy en vivo."""
    data = _leer_historico()

    # Inyectar hoy en vivo
    hoy = _hoy()
    total_hoy = _total_ventas_hoy_sheets()
    if total_hoy > 0:
        data[hoy] = int(total_hoy)

    por_mes: dict = defaultdict(int)
    for fecha, monto in data.items():
        if monto and monto > 0:
            por_mes[fecha[:7]] += monto
    return dict(sorted(por_mes.items()))


@app.post("/historico/auto-sync")
def historico_auto_sync():
    """
    Sincroniza el total de hoy desde Sheets → histórico persistente.
    Llamar desde el dashboard o automáticamente al cierre del día.
    """
    try:
        result = _sync_historico_hoy()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/historico/sync-rango")
def historico_sync_rango(dias: int = Query(default=30, ge=1, le=365)):
    """
    Sincroniza los totales de los últimos N días desde Excel → histórico.
    Útil para llenar histórico con datos de días pasados que no se guardaron.
    No sobreescribe datos existentes (solo agrega los que faltan).
    """
    try:
        data = _leer_historico()
        totales_excel = _totales_por_dia_excel(dias=dias)

        nuevos = 0
        actualizados = 0
        for fecha, monto in totales_excel.items():
            if fecha not in data:
                data[fecha] = monto
                nuevos += 1
            elif data[fecha] != monto:
                # Solo actualizar si el Excel tiene un monto mayor
                # (probablemente el dato manual estaba incompleto)
                if monto > data[fecha]:
                    data[fecha] = monto
                    actualizados += 1

        # Inyectar hoy en vivo desde Sheets (más preciso que Excel para el día actual)
        hoy = _hoy()
        total_hoy = _total_ventas_hoy_sheets()
        if total_hoy > 0:
            if hoy not in data or int(total_hoy) > data.get(hoy, 0):
                data[hoy] = int(total_hoy)
                if hoy not in totales_excel:
                    nuevos += 1

        _guardar_historico(data)
        return {
            "ok": True,
            "dias_escaneados": dias,
            "nuevos": nuevos,
            "actualizados": actualizados,
            "total_registros": len(data),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Chat IA desde el Dashboard ────────────────────────────────────────────────

# Convierte session_id (string) en un chat_id negativo único.
# Los chat_id de Telegram son positivos, los del dashboard son negativos → sin colisión.
def _session_chat_id(session_id: str) -> int:
    return -(abs(hash(session_id)) % (10 ** 9))


def _construir_contexto_dashboard(mensaje: str, tab_activo: str = "") -> str:
    """
    Construye el bloque de contexto enriquecido para el asistente del dashboard.
    Incluye: datos reales del negocio, memoria persistente, estado actual.
    Se renueva en cada llamada para tener datos frescos.
    """
    from memoria import cargar_memoria, cargar_caja, cargar_gastos_hoy
    from datetime import datetime
    import config

    mem   = cargar_memoria()
    ahora = datetime.now(config.COLOMBIA_TZ)
    fecha_hoy = ahora.strftime("%A %d de %B de %Y, %H:%M")

    # ── Memoria persistente del negocio ───────────────────────────────────────
    notas_raw = mem.get("notas", {})
    if isinstance(notas_raw, list):
        # Migración: formato antiguo era lista
        notas_raw = {"observaciones": notas_raw} if notas_raw else {}
    contexto_negocio  = notas_raw.get("contexto_negocio", "")
    decisiones        = notas_raw.get("decisiones", [])
    observaciones     = notas_raw.get("observaciones", [])

    memoria_texto = ""
    if contexto_negocio:
        memoria_texto += f"CONTEXTO DEL NEGOCIO:
{contexto_negocio}

"
    if decisiones:
        memoria_texto += "DECISIONES GUARDADAS:
" + "
".join(f"- {d}" for d in decisiones[-10:]) + "

"
    if observaciones:
        memoria_texto += "OBSERVACIONES:
" + "
".join(f"- {o}" for o in observaciones[-10:]) + "

"

    # ── Caja actual ──────────────────────────────────────────────────────────
    try:
        caja = cargar_caja()
        if caja.get("abierta"):
            ef  = caja.get("efectivo", 0)
            tr  = caja.get("transferencias", 0)
            dat = caja.get("datafono", 0)
            ap  = caja.get("monto_apertura", 0)
            total_caja = ef + tr + dat
            caja_texto = (
                f"CAJA (abierta desde {caja.get('fecha','?')}):
"
                f"  Apertura: ${ap:,.0f} | Efectivo: ${ef:,.0f} | "
                f"Transferencias: ${tr:,.0f} | Datafono: ${dat:,.0f}
"
                f"  Total en caja: ${total_caja:,.0f}"
            )
        else:
            caja_texto = f"CAJA: Cerrada (última fecha: {caja.get('fecha','sin datos')})"
    except Exception:
        caja_texto = "CAJA: Sin datos"

    # ── Gastos del día ───────────────────────────────────────────────────────
    try:
        gastos_hoy = cargar_gastos_hoy()
        if gastos_hoy:
            total_gastos = sum(float(g.get("monto", 0)) for g in gastos_hoy)
            items_gasto  = "
".join(
                f"  {g.get('hora','?')} {g.get('concepto','?')} ${float(g.get('monto',0)):,.0f}"
                for g in gastos_hoy
            )
            gastos_texto = f"GASTOS HOY (total ${total_gastos:,.0f}):
{items_gasto}"
        else:
            gastos_texto = "GASTOS HOY: ninguno registrado"
    except Exception:
        gastos_texto = "GASTOS: Sin datos"

    # ── Fiados activos ───────────────────────────────────────────────────────
    try:
        fiados = mem.get("fiados", {})
        fiados_activos = {
            nombre: datos for nombre, datos in fiados.items()
            if float(datos.get("saldo", 0)) > 0
        }
        if fiados_activos:
            total_fiado = sum(float(d.get("saldo", 0)) for d in fiados_activos.values())
            items_fiado = "
".join(
                f"  {n}: ${float(d.get('saldo',0)):,.0f}"
                for n, d in list(fiados_activos.items())[:15]
            )
            fiados_texto = f"FIADOS ACTIVOS ({len(fiados_activos)} clientes, total ${total_fiado:,.0f}):
{items_fiado}"
        else:
            fiados_texto = "FIADOS: Sin saldos pendientes"
    except Exception:
        fiados_texto = "FIADOS: Sin datos"

    # ── Inventario (si tiene datos) ──────────────────────────────────────────
    try:
        inventario = mem.get("inventario", {})
        if inventario:
            criticos = [
                f"  {k}: {v.get('cantidad',0)} {v.get('unidad','u')} (mín: {v.get('minimo',0)})"
                for k, v in inventario.items()
                if float(v.get("cantidad", 0)) <= float(v.get("minimo", 0)) * 1.2
            ]
            inv_texto = (
                f"INVENTARIO CRÍTICO ({len(criticos)} productos bajo mínimo):
" +
                "
".join(criticos[:10])
            ) if criticos else f"INVENTARIO: {len(inventario)} productos registrados, todos sobre mínimo"
        else:
            inv_texto = "INVENTARIO: Pendiente de configurar (aún no hay stock registrado)"
    except Exception:
        inv_texto = "INVENTARIO: Sin datos"

    # ── Márgenes (si hay precio_compra en el catálogo) ───────────────────────
    try:
        catalogo = mem.get("catalogo", {})
        prods_con_costo = [
            p for p in catalogo.values() if p.get("precio_compra") or p.get("costo")
        ]
        if prods_con_costo:
            margenes_lineas = []
            for p in prods_con_costo[:10]:
                costo  = float(p.get("precio_compra") or p.get("costo", 0))
                venta  = float(p.get("precio_unidad", 0))
                if costo > 0 and venta > 0:
                    margen = ((venta - costo) / venta) * 100
                    margenes_lineas.append(
                        f"  {p['nombre']}: costo ${costo:,.0f} → venta ${venta:,.0f} (margen {margen:.0f}%)"
                    )
            margenes_texto = (
                "MÁRGENES (muestra):
" + "
".join(margenes_lineas)
            ) if margenes_lineas else "MÁRGENES: Pendiente (agrega precio_compra al catálogo)"
        else:
            margenes_texto = "MÁRGENES: Pendiente de configurar (agrega el precio de compra de cada producto)"
    except Exception:
        margenes_texto = "MÁRGENES: Sin datos"

    # ── Tab activo ───────────────────────────────────────────────────────────
    tab_ctx = f"\nTAB ACTIVO EN DASHBOARD: El usuario está mirando '{tab_activo}'. Ten esto en cuenta para dar contexto relevante." if tab_activo else ""

    return f"""CANAL: Dashboard web — modo gerente/asistente avanzado.
FECHA Y HORA ACTUAL: {fecha_hoy}

## PERSONALIDAD Y MODO DE OPERACIÓN
Eres el asistente inteligente de Ferretería Punto Rojo. En este canal tienes un rol dual:
1. REGISTRAR con precisión (ventas, gastos, compras, fiados) — igual que en Telegram
2. SER GERENTE: analizar, opinar, recomendar, advertir, recordar decisiones pasadas

TONO: Directo, claro, con criterio. No eres un bot genérico — conoces este negocio.
Si ves algo raro en los datos, lo dices. Si hay una oportunidad, la señalas.
Si te preguntan tu opinión, la das con base en los datos reales.

FORMATO: Responde con la extensión que el tema requiera. Para análisis, sé detallado.
Para registros (ventas/gastos), usa el mismo formato compacto de siempre con [VENTA]/[GASTO].
No uses markdown (asteriscos, #). Usa texto plano limpio.

## ESTADO ACTUAL DEL NEGOCIO
{caja_texto}

{gastos_texto}

{fiados_texto}

{inv_texto}

{margenes_texto}

{memoria_texto}

## MEMORIA PERSISTENTE
Puedes guardar información importante del negocio usando la acción:
[MEMORIA]{{"tipo":"decision"|"observacion"|"contexto","contenido":"texto"}}[/MEMORIA]
Úsala cuando el usuario mencione algo que debe recordarse: cambios de estrategia,
observaciones sobre clientes, decisiones de precio, metas, etc.

## CAPACIDADES COMPLETAS EN ESTE CANAL
• Registrar ventas, gastos, compras, fiados, abonos
• Analizar ventas por día, semana, mes, producto, vendedor
• Consultar y actualizar precios del catálogo
• Ver márgenes y rentabilidad (cuando esté configurado)
• Gestionar inventario y alertas de stock
• Recordar y recuperar decisiones pasadas del negocio
• Dar opinión y recomendaciones basadas en datos reales{tab_ctx}"""


# ── Endpoint para guardar memoria del negocio ─────────────────────────────────
class MemoriaRequest(BaseModel):
    tipo: str        # "decision" | "observacion" | "contexto_negocio"
    contenido: str


@app.post("/chat/memoria")
def guardar_memoria_negocio(req: MemoriaRequest):
    """Guarda una nota/decisión/observación persistente del negocio."""
    from memoria import cargar_memoria, guardar_memoria as _guardar

    if not req.contenido.strip():
        raise HTTPException(status_code=400, detail="Contenido vacío")

    mem   = cargar_memoria()
    notas = mem.get("notas", {})
    if isinstance(notas, list):
        notas = {"observaciones": notas} if notas else {}

    if req.tipo == "contexto_negocio":
        notas["contexto_negocio"] = req.contenido.strip()
    elif req.tipo == "decision":
        from datetime import datetime
        import config
        fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        notas.setdefault("decisiones", []).append(f"[{fecha}] {req.contenido.strip()}")
        notas["decisiones"] = notas["decisiones"][-30:]  # máx 30 decisiones
    else:
        from datetime import datetime
        import config
        fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        notas.setdefault("observaciones", []).append(f"[{fecha}] {req.contenido.strip()}")
        notas["observaciones"] = notas["observaciones"][-30:]

    mem["notas"] = notas
    _guardar(mem, urgente=True)
    return {"ok": True, "tipo": req.tipo}


class ChatRequest(BaseModel):
    mensaje: str
    nombre: str = "Dashboard"
    historial: list = []
    # Si viene con confirmar_pago, se salta Claude y registra directamente
    confirmar_pago: Optional[str] = None   # "efectivo" | "transferencia" | "datafono"
    # ID de sesión único por pestaña del navegador (evita race condition multi-usuario)
    session_id: str = "default"
    # Tab activo en el dashboard (contexto extra para el asistente)
    tab_activo: str = ""


@app.post("/chat")
async def chat_ia(req: ChatRequest):
    """
    Endpoint de chat IA para el dashboard.

    Flujo normal:
      1. procesar_con_claude()      → respuesta con tags [VENTA]
      2. procesar_acciones_async()  → guarda ventas en ventas_pendientes[0]
      3. Si hay ventas pendientes   → devuelve pendiente=True + botones de pago
      4. El frontend muestra botones; el usuario hace clic
      5. Segunda llamada con confirmar_pago="efectivo"|"transferencia"|"datafono"
      6. registrar_ventas_con_metodo_async() → Excel + Sheets + confirmación

    Flujo con método explícito en el mensaje (ej: "venta 3 tornillos efectivo"):
      - Si Claude detecta el método, devuelve PEDIR_CONFIRMACION:efectivo
      - Se registra directamente sin pedir botones
    """
    from ai import procesar_con_claude, procesar_acciones_async
    from ventas_state import ventas_pendientes, registrar_ventas_con_metodo_async, _estado_lock

    log = logging.getLogger("ferrebot.api")

    # ── RAMA B: Confirmación de pago desde botón ─────────────────────────────
    if req.confirmar_pago:
        metodo = req.confirmar_pago.strip().lower()
        if metodo not in ("efectivo", "transferencia", "datafono"):
            raise HTTPException(status_code=400, detail=f"Método de pago inválido: {metodo}")

        _chat_id = _session_chat_id(req.session_id)
        with _estado_lock:
            ventas_pend = list(ventas_pendientes.get(_chat_id, []))

        if not ventas_pend:
            return {
                "ok": True,
                "respuesta": "⚠️ No hay ventas pendientes de confirmar.",
                "acciones": {"ventas": 0, "gastos": 0},
                "pendiente": False,
            }

        confirmacion = await registrar_ventas_con_metodo_async(
            ventas_pend, metodo, req.nombre, _chat_id
        )
        log.info(f"[/chat] ✅ {len(ventas_pend)} venta(s) confirmadas | método: {metodo}")

        return {
            "ok": True,
            "respuesta": "✅ Venta registrada\n" + "\n".join(confirmacion),
            "acciones": {"ventas": len(ventas_pend), "gastos": 0},
            "pendiente": False,
        }

    # ── RAMA A: Mensaje normal ────────────────────────────────────────────────
    if not req.mensaje or not req.mensaje.strip():
        raise HTTPException(status_code=400, detail="Mensaje vacío")

    try:
        mensaje_formateado = f"{req.nombre}: {req.mensaje.strip()}"

        _chat_id = _session_chat_id(req.session_id)

        # ── Contexto dinámico del dashboard (datos reales en cada llamada) ──
        contexto_dash = _construir_contexto_dashboard(req.mensaje, tab_activo=req.tab_activo)

        # Inyectar flag ##DASHBOARD## para que ai.py active modo dashboard
        mensaje_con_flag = f"##DASHBOARD## {mensaje_formateado}"

        # 1. Claude con contexto enriquecido del dashboard
        respuesta_raw = await procesar_con_claude(
            mensaje_usuario=mensaje_con_flag,
            nombre_usuario=req.nombre,
            historial_chat=req.historial,
            contexto_extra=contexto_dash,
        )

        # 2. Parsear acciones
        texto_limpio, acciones, _ = await procesar_acciones_async(
            respuesta_raw, req.nombre, _chat_id
        )

        pedir_pago   = "PEDIR_METODO_PAGO" in acciones
        confirmacion_accion = next(
            (a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None
        )
        gastos_registrados = sum(
            1 for a in acciones if a.startswith("Gasto registrado:")
        )

        # 3a. Método de pago ya viene en la acción (Claude lo detectó) → registrar directo
        if confirmacion_accion:
            metodo = confirmacion_accion.split(":", 1)[1].strip()
            if metodo not in ("efectivo", "transferencia", "datafono"):
                metodo = "efectivo"

            with _estado_lock:
                ventas_pend = list(ventas_pendientes.get(_chat_id, []))

            if ventas_pend:
                conf = await registrar_ventas_con_metodo_async(
                    ventas_pend, metodo, req.nombre, _chat_id
                )
                return {
                    "ok": True,
                    "respuesta": "✅ Venta registrada\n" + "\n".join(conf),
                    "acciones": {"ventas": len(ventas_pend), "gastos": gastos_registrados},
                    "pendiente": False,
                }

        # 3b. Hay ventas esperando método → devolver botones al frontend
        if pedir_pago:
            with _estado_lock:
                ventas_pend = list(ventas_pendientes.get(_chat_id, []))

            if ventas_pend:
                resumen = "\n".join(
                    f"• {v.get('cantidad',1)} {v.get('producto','?')}  ${float(v.get('total',0)):,.0f}"
                    for v in ventas_pend
                )
                texto_previo = texto_limpio.strip() if texto_limpio and texto_limpio.strip() else ""
                texto_botones = (f"{texto_previo}\n\n" if texto_previo else "") + \
                                f"🧾 {resumen}\n\n¿Cómo pagó?"
                return {
                    "ok": True,
                    "respuesta": texto_botones,
                    "acciones": {"ventas": 0, "gastos": gastos_registrados},
                    "pendiente": True,
                    "opciones_pago": [
                        {"label": "💵 Efectivo",      "valor": "efectivo"},
                        {"label": "📲 Transferencia", "valor": "transferencia"},
                        {"label": "💳 Datafono",      "valor": "datafono"},
                    ],
                }

        # 4. Sin ventas pendientes → respuesta normal
        if texto_limpio and texto_limpio.strip():
            texto_final = texto_limpio.strip()
        elif gastos_registrados:
            gasto_msgs = [a for a in acciones if a.startswith("Gasto registrado:")]
            texto_final = "✅ " + "\n".join(gasto_msgs)
        else:
            otras = [a for a in acciones if a not in (
                "PEDIR_METODO_PAGO", "PAGO_PENDIENTE_AVISO", "INICIAR_FLUJO_CLIENTE",
            ) and not a.startswith("PEDIR_CONFIRMACION:")
              and not a.startswith("CLIENTE_DESCONOCIDO:")]
            texto_final = "\n".join(otras) if otras else "(Sin respuesta)"

        return {
            "ok": True,
            "respuesta": texto_final,
            "acciones": {"ventas": 0, "gastos": gastos_registrados},
            "pendiente": False,
        }

    except Exception as e:
        log.error(f"[/chat] Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/chat/stream")
async def chat_stream(req: ChatRequest):
    """
    Endpoint SSE para el dashboard con streaming token-a-token.
    Emite eventos:
      data: {"type":"chunk","text":"..."}
      data: {"type":"done","respuesta":"...","acciones":{...},"pendiente":bool}
      data: {"type":"error","message":"..."}
    """
    from ai import procesar_con_claude_stream, procesar_acciones_async
    from ventas_state import ventas_pendientes, registrar_ventas_con_metodo_async, _estado_lock

    log = logging.getLogger("ferrebot.api")

    if not req.mensaje or not req.mensaje.strip():
        async def _err():
            yield f"data: {json.dumps({'type':'error','message':'Mensaje vacío'})}\n\n"
        return StreamingResponse(_err(), media_type="text/event-stream")

    _chat_id = _session_chat_id(req.session_id)

    async def generate():
        try:
            mensaje_formateado = f"{req.nombre}: {req.mensaje.strip()}"
            contexto_dash = _construir_contexto_dashboard(req.mensaje, tab_activo=req.tab_activo)
            mensaje_con_flag = f"##DASHBOARD## {mensaje_formateado}"
            full_text = ""

            async for kind, data in procesar_con_claude_stream(
                mensaje_usuario=mensaje_con_flag,
                nombre_usuario=req.nombre,
                historial_chat=req.historial,
                contexto_extra=contexto_dash,
            ):
                if kind == "chunk":
                    full_text += data
                    yield f"data: {json.dumps({'type':'chunk','text':data}, ensure_ascii=False)}\n\n"
                elif kind == "done":
                    full_text = data
                    break
                elif kind == "error":
                    yield f"data: {json.dumps({'type':'error','message':data})}\n\n"
                    return

            texto_limpio, acciones, _ = await procesar_acciones_async(
                full_text, req.nombre, _chat_id
            )

            pedir_pago          = "PEDIR_METODO_PAGO" in acciones
            confirmacion_accion = next((a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None)
            gastos_reg          = sum(1 for a in acciones if a.startswith("Gasto registrado:"))
            ventas_reg          = 0
            pendiente           = False
            opciones_pago       = None

            if confirmacion_accion:
                metodo = confirmacion_accion.split(":", 1)[1].strip()
                if metodo not in ("efectivo", "transferencia", "datafono"):
                    metodo = "efectivo"
                with _estado_lock:
                    vp = list(ventas_pendientes.get(_chat_id, []))
                if vp:
                    conf = await registrar_ventas_con_metodo_async(vp, metodo, req.nombre, _chat_id)
                    texto_limpio = "✅ Venta registrada\n" + "\n".join(conf)
                    ventas_reg = len(vp)

            elif pedir_pago:
                with _estado_lock:
                    vp = list(ventas_pendientes.get(_chat_id, []))
                if vp:
                    resumen = "\n".join(
                        f"• {v.get('cantidad',1)} {v.get('producto','?')}  ${float(v.get('total',0)):,.0f}"
                        for v in vp
                    )
                    tp = texto_limpio.strip() if texto_limpio and texto_limpio.strip() else ""
                    texto_limpio = (f"{tp}\n\n" if tp else "") + f"🧾 {resumen}\n\n¿Cómo pagó?"
                    pendiente = True
                    opciones_pago = [
                        {"label": "💵 Efectivo",      "valor": "efectivo"},
                        {"label": "📲 Transferencia", "valor": "transferencia"},
                        {"label": "💳 Datafono",      "valor": "datafono"},
                    ]

            if not texto_limpio or not texto_limpio.strip():
                if gastos_reg:
                    texto_limpio = "✅ " + "\n".join(a for a in acciones if a.startswith("Gasto registrado:"))
                else:
                    otras = [a for a in acciones if a not in (
                        "PEDIR_METODO_PAGO","PAGO_PENDIENTE_AVISO","INICIAR_FLUJO_CLIENTE"
                    ) and not a.startswith("PEDIR_CONFIRMACION:") and not a.startswith("CLIENTE_DESCONOCIDO:")]
                    texto_limpio = "\n".join(otras) if otras else "(Sin respuesta)"

            payload = {
                "type": "done",
                "respuesta": texto_limpio.strip(),
                "acciones": {"ventas": ventas_reg, "gastos": gastos_reg},
                "pendiente": pendiente,
                "opciones_pago": opciones_pago,
            }
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

        except Exception as exc:
            log.error(f"[/chat/stream] {exc}", exc_info=True)
            yield f"data: {json.dumps({'type':'error','message':str(exc)})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ── Servir dashboard React (build estático) ───────────────────────────────────
# Los archivos del build quedan en dashboard/dist/ después de `npm run build`
_DIST = Path(__file__).parent / "dashboard" / "dist"

if _DIST.exists():
    # Archivos estáticos (JS, CSS, assets)
    app.mount("/assets", StaticFiles(directory=_DIST / "assets"), name="assets")

    # Cualquier ruta que no sea /api/* → devolver index.html (SPA routing)
    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        index = _DIST / "index.html"
        if index.exists():
            return FileResponse(index)
        return {"error": "Dashboard no buildeado. Ejecuta: cd dashboard && npm run build"}
else:
    @app.get("/")
    def root():
        return {
            "servicio": "FerreBot Dashboard API",
            "estado":   "activo",
            "version":  "1.0.0",
            "nota":     "Dashboard no buildeado. Ejecuta: cd dashboard && npm run build",
            "docs":     "/docs",
        }
