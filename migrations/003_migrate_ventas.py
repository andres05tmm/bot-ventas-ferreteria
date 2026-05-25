# Migrado desde migrate_ventas.py (raíz del proyecto)
# Este archivo es una copia exacta. El original se mantiene en raíz hasta que Phase 1 esté verde.
#!/usr/bin/env python3
"""
migrate_ventas.py — Migra ventas historicas de ventas.xlsx a Postgres.
Ejecutar: railway run python migrate_ventas.py
Idempotente: seguro de re-ejecutar (per D-13).

Itera todas las hojas mensuales de ventas.xlsx, agrupa filas por consecutivo,
e inserta en las tablas ventas + ventas_detalle.
"""

# -- stdlib --
import logging
import os
import sys
from datetime import datetime

# Configurar logging basico para ver output
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s -- %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("migrate_ventas")

# Fail-fast si DATABASE_URL no esta configurado
if not os.getenv("DATABASE_URL"):
    print("ERROR: DATABASE_URL no configurado")
    sys.exit(1)

# -- terceros --
import openpyxl

# -- propios --
import config
import db

db.init_db()
if not db.DB_DISPONIBLE:
    print("ERROR: No se pudo conectar a Postgres")
    sys.exit(1)


# ── Hojas a ignorar ────────────────────────────────────────────────────────────
HOJAS_IGNORADAS = {"Compras", "Registro de Ventas-Acumulado", "Productos"}


def _safe_int(val, default=0) -> int:
    """Convierte un valor a entero, con fallback."""
    if val is None or val == "":
        return default
    try:
        return int(float(str(val).replace(",", ".")))
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0) -> float:
    """Convierte un valor a float, con fallback."""
    if val is None or val == "":
        return default
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return default


def _safe_str(val, default="") -> str:
    """Convierte un valor a string limpio."""
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def _parse_fecha(raw) -> str | None:
    """Convierte fecha de Excel a string YYYY-MM-DD."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if len(s) >= 10:
        return s[:10]
    return None


def _parse_hora(raw) -> str | None:
    """Convierte hora de Excel a string HH:MM:SS o None."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%H:%M:%S")
    s = str(raw).strip()
    return s if s else None


def _es_hoja_mensual(nombre: str) -> bool:
    """Heuristica: hoja mensual si su nombre es de la forma '<Mes> <Year>'."""
    if nombre in HOJAS_IGNORADAS:
        return False
    # Los nombres de hojas mensuales son como "Enero 2026", "Marzo 2025", etc.
    partes = nombre.strip().split()
    if len(partes) != 2:
        return False
    mes_nombre, anio = partes
    if not anio.isdigit():
        return False
    if len(anio) != 4:
        return False
    meses_validos = set(config.MESES.values())
    return mes_nombre in meses_validos


def _leer_columnas(ws) -> dict:
    """
    Detecta posicion de columnas relevantes en la fila de headers.
    Retorna dict {nombre_logico: indice_columna_1based}.
    """
    cols: dict[str, int] = {}
    try:
        for fila_hdr in ws.iter_rows(
            min_row=config.EXCEL_FILA_HEADERS,
            max_row=config.EXCEL_FILA_HEADERS,
        ):
            for cell in fila_hdr:
                if cell.value is not None:
                    cols[str(cell.value).lower().strip()] = cell.column
            break
    except Exception:
        pass
    return cols


def _col(cols: dict, *claves) -> int | None:
    """Retorna el primer indice de columna que coincida con las claves dadas."""
    for k in claves:
        if k in cols:
            return cols[k]
    return None


def _v(fila: tuple, col_idx: int | None):
    """Extrae valor de una fila por indice de columna 1-based."""
    if col_idx is None or col_idx > len(fila):
        return None
    return fila[col_idx - 1]


def migrar():
    """Migra todas las ventas de ventas.xlsx a Postgres."""
    excel_file = config.EXCEL_FILE
    if not os.path.exists(excel_file):
        logger.error(f"No se encontro {excel_file}")
        sys.exit(1)

    logger.info(f"Cargando {excel_file} ...")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

    sheets_count  = 0
    ventas_count  = 0
    detalle_count = 0
    skipped_count = 0

    for nombre_hoja in wb.sheetnames:
        if not _es_hoja_mensual(nombre_hoja):
            logger.info(f"Hoja ignorada: {nombre_hoja!r}")
            continue

        logger.info(f"Procesando hoja: {nombre_hoja!r}")
        ws = wb[nombre_hoja]
        sheets_count += 1

        cols = _leer_columnas(ws)
        if not cols:
            logger.warning(f"  No se encontraron columnas en hoja {nombre_hoja!r}, saltando")
            continue

        c_fecha    = _col(cols, "fecha")
        c_hora     = _col(cols, "hora")
        c_producto = _col(cols, "producto")
        c_cantidad = _col(cols, "cantidad")
        c_precio   = _col(cols, "valor unitario", "precio unitario", "precio")
        c_total    = _col(cols, "total")
        c_alias    = _col(cols, "alias")
        c_vendedor = _col(cols, "vendedor")
        c_metodo   = _col(cols, "metodo de pago", "metodo pago", "método pago")
        c_num      = _col(cols, "#", "consecutivo", "num", "consecutivo de venta")
        c_unidad   = _col(cols, "unidad de medida", "unidad_medida", "unidad")
        c_cliente  = _col(cols, "cliente")
        c_id_cli   = _col(cols, "id cliente")

        # Agrupar todas las filas por consecutivo
        grupos: dict[tuple, list] = {}  # (consecutivo_int, fecha_str) -> [filas...]

        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue

            fecha_raw = _v(fila, c_fecha)
            fecha_str = _parse_fecha(fecha_raw)
            if not fecha_str:
                continue

            num_raw = _v(fila, c_num)
            if num_raw is None or str(num_raw).strip() == "":
                continue

            try:
                consecutivo_int = int(float(str(num_raw)))
            except (ValueError, TypeError):
                logger.warning(f"  Consecutivo invalido: {num_raw!r}, fila saltada")
                continue

            clave = (consecutivo_int, fecha_str)
            if clave not in grupos:
                grupos[clave] = []
            grupos[clave].append(fila)

        logger.info(f"  {len(grupos)} grupos de venta encontrados en {nombre_hoja!r}")

        # Procesar cada grupo (un grupo = una venta)
        for (consecutivo_int, fecha_str), filas in grupos.items():
            try:
                # Idempotency check (per D-13)
                existing = db.query_one(
                    "SELECT id FROM ventas WHERE consecutivo = %s AND fecha = %s",
                    (consecutivo_int, fecha_str)
                )
                if existing:
                    skipped_count += 1
                    continue

                # Tomar primera fila para datos de cabecera
                primera = filas[0]

                hora_raw = _v(primera, c_hora)
                hora_str = _parse_hora(hora_raw)

                cliente_nombre = _safe_str(_v(primera, c_cliente), "Consumidor Final")
                vendedor       = _safe_str(_v(primera, c_vendedor), "")
                metodo         = _safe_str(_v(primera, c_metodo), "")

                # cliente_id: None (no resolvemos FK durante migracion)
                cliente_id_pg = None

                # Calcular total del grupo
                group_total = 0
                for f in filas:
                    group_total += _safe_int(_v(f, c_total), 0)

                # INSERT ventas
                row = db.execute_returning(
                    """INSERT INTO ventas (consecutivo, fecha, hora, cliente_id, cliente_nombre, vendedor, metodo_pago, total)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
                    (consecutivo_int, fecha_str, hora_str, cliente_id_pg, cliente_nombre, vendedor, metodo, group_total)
                )
                if not row:
                    logger.warning(f"  INSERT ventas no retorno id para consecutivo={consecutivo_int}, fecha={fecha_str}")
                    continue

                venta_id = row["id"]
                ventas_count += 1

                # INSERT ventas_detalle para cada linea del grupo
                for f in filas:
                    producto   = _safe_str(_v(f, c_producto), "")
                    if not producto:
                        continue

                    cantidad_raw = _v(f, c_cantidad)
                    cantidad     = _safe_float(cantidad_raw, 1.0)

                    unidad       = _safe_str(_v(f, c_unidad), "Unidad")
                    precio_unit  = _safe_int(_v(f, c_precio), 0)
                    total_line   = _safe_int(_v(f, c_total), 0)
                    alias_val    = _v(f, c_alias)
                    alias_usado  = _safe_str(alias_val) if alias_val is not None else None

                    db.execute(
                        """INSERT INTO ventas_detalle (venta_id, producto_nombre, cantidad, unidad_medida, precio_unitario, total, alias_usado)
                           VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                        (venta_id, producto, cantidad, unidad, precio_unit, total_line, alias_usado)
                    )
                    detalle_count += 1

            except Exception as e:
                logger.warning(
                    f"  Error procesando consecutivo={consecutivo_int} fecha={fecha_str} en hoja {nombre_hoja!r}: {e}"
                )
                continue

    try:
        wb.close()
    except Exception:
        pass

    # Resumen final
    print(f"
{'='*50}")
    print(f"Migration complete:")
    print(f"  Sheets processed: {sheets_count}")
    print(f"  Ventas inserted:  {ventas_count}")
    print(f"  Detalle inserted: {detalle_count}")
    print(f"  Skipped (dupes):  {skipped_count}")
    print(f"{'='*50}")


if __name__ == "__main__":
    migrar()
