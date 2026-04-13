"""
Handlers de mensajes: texto, audio (voz) y documentos Excel.

CORRECCIONES v2:
  - manejar_audio: ruta_audio se inicializa a None ANTES del try/finally para
    evitar NameError si download_to_drive falla antes de asignar la variable
  - mensajes_standby: se usa agregar_a_standby() de ventas_state que tiene cap MAX_STANDBY
  - Docstring ANTES del import logging

CORRECCIONES v3:
  - Todos los imports de stdlib (re, json, base64, datetime) movidos al nivel de módulo.
  - Imports de callbacks hoistados (no hay ciclo: callbacks no importa mensajes).
  - Imports que SÍ crean ciclo (handlers.comandos → mensajes) siguen siendo lazy.
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import base64
import collections
import json
import logging
import asyncio
import os
import re
import tempfile
import threading
import time
import traceback
from datetime import datetime

# ── terceros ──────────────────────────────────────────────────────────────────
import openpyxl
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

# ── propios ───────────────────────────────────────────────────────────────────
import config
from ai import procesar_con_claude, procesar_acciones, procesar_acciones_async
from ventas_state import (
    agregar_al_historial, get_historial,
    ventas_pendientes, clientes_en_proceso, _estado_lock,
    get_chat_lock,
    limpiar_pendientes_expirados,
    esperando_correccion,
    agregar_a_standby,
    mensaje_contexto_pendiente,
)
from handlers.comandos import manejar_flujo_agregar_producto
from utils import convertir_fraccion_a_decimal, decimal_a_fraccion_legible, corregir_texto_audio
from memoria import invalidar_cache_memoria, importar_catalogo_desde_excel
# Callbacks: no crean ciclo (callbacks.py no importa mensajes.py en nivel de módulo)
from handlers.callbacks import (
    _enviar_botones_pago as _botones_central,
    _enviar_confirmacion_con_metodo,
)
from handlers.parsing import parsear_actualizacion_masiva as _parsear_actualizacion_masiva
from handlers.cliente_flujo import enviar_pregunta_cliente as _enviar_pregunta_cliente
from handlers.dispatch import (
    manejar_flujo_cliente,
    manejar_flujo_excel,
    manejar_flujo_pago_texto,
    manejar_flujo_correccion,
    manejar_rechazo_cliente,
)

logger = logging.getLogger("ferrebot.mensajes")

# ── Rate limiting ─────────────────────────────────────────────────────────────
_RATE_LIMIT = 18          # mensajes permitidos por ventana
_RATE_WINDOW = 60.0       # segundos de la ventana deslizante
_rate_timestamps: dict[int, collections.deque] = {}  # chat_id → timestamps

def _check_rate_limit(chat_id: int) -> bool:
    """Retorna True si el chat está dentro del límite, False si lo superó."""
    ahora = time.monotonic()
    ventana_inicio = ahora - _RATE_WINDOW
    if chat_id not in _rate_timestamps:
        _rate_timestamps[chat_id] = collections.deque()
    dq = _rate_timestamps[chat_id]
    # Eliminar timestamps fuera de la ventana
    while dq and dq[0] < ventana_inicio:
        dq.popleft()
    if len(dq) >= _RATE_LIMIT:
        return False
    dq.append(ahora)
    return True


async def _enviar_botones_pago(message, chat_id: int, ventas: list):
    """Delega a callbacks._enviar_botones_pago (importado al nivel de módulo)."""
    await _botones_central(message, chat_id, ventas)



async def _manejar_actualizacion_masiva(update, vendedor: str, pares: list):
    """Actualiza todos los precios directamente en PostgreSQL."""
    import db as _db
    from memoria import buscar_producto_en_catalogo, invalidar_cache_memoria

    exitos, errores = [], []
    for item in pares:
        nombre, precio, fraccion, precio_mayorista = item if len(item) == 4 else (*item, None)
        try:
            prod = buscar_producto_en_catalogo(nombre)
            nombre_display = prod["nombre"] if prod else nombre

            if not prod:
                errores.append(f"❌ {nombre}: no encontrado")
                continue

            row = _db.query_one(
                "SELECT id, precio_unidad FROM productos WHERE nombre_lower = %s AND activo = TRUE",
                [prod.get("nombre_lower", nombre.lower())]
            )
            if not row:
                errores.append(f"❌ {nombre}: no encontrado en BD")
                continue

            if precio_mayorista is not None:
                # precio_unidad + precio_por_cantidad (tornillería)
                _db.execute(
                    "UPDATE productos SET precio_unidad = %s, updated_at = NOW() WHERE id = %s",
                    [round(precio), row["id"]],
                )
                _db.execute(
                    """
                    UPDATE productos
                    SET precio_umbral       = 50,
                        precio_bajo_umbral  = %s,
                        precio_sobre_umbral = %s,
                        updated_at          = NOW()
                    WHERE id = %s
                    """,
                    (round(precio), round(precio_mayorista), row["id"]),
                )
                linea = f"✅ {nombre_display}: ${int(precio):,} / ${int(precio_mayorista):,} ×50".replace(",", ".")
            elif fraccion:
                try:
                    # float("1/4") lanza ValueError — parsear manualmente
                    if "/" in str(fraccion):
                        num, den = str(fraccion).split("/", 1)
                        decimal_frac = int(num.strip()) / int(den.strip())
                    else:
                        decimal_frac = float(fraccion)
                    precio_unit = round(precio / decimal_frac, 2) if decimal_frac else precio
                except (ValueError, ZeroDivisionError, TypeError):
                    precio_unit = precio
                _db.execute(
                    """
                    INSERT INTO productos_fracciones
                        (producto_id, fraccion, precio_total, precio_unitario)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT ON CONSTRAINT uq_prod_fraccion DO UPDATE
                    SET precio_total    = EXCLUDED.precio_total,
                        precio_unitario = EXCLUDED.precio_unitario
                    """,
                    (row["id"], fraccion, round(precio), int(precio_unit)),
                )
                linea = f"✅ {nombre_display} {fraccion} → ${int(precio):,}".replace(",", ".")
            else:
                _db.execute(
                    "UPDATE productos SET precio_unidad = %s, updated_at = NOW() WHERE id = %s",
                    [round(precio), row["id"]],
                )
                linea = f"✅ {nombre_display} → ${int(precio):,}".replace(",", ".")

            exitos.append(linea)
        except Exception as e:
            errores.append(f"❌ {nombre}: {e}")

    invalidar_cache_memoria()

    resumen = f"💰 *{len(exitos)} precio(s) actualizado(s):*\n" + "\n".join(exitos)
    if errores:
        resumen += "\n\n" + "\n".join(errores)

    await update.message.reply_text(resumen, parse_mode="Markdown")

async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje  = update.message.text
    chat_id  = update.message.chat_id
    vendedor = update.message.from_user.first_name or "Desconocido"

    if mensaje.startswith("/"):
        return

    # ── "sin foto" en flujo /factura o /abonar ───────────────────────────────
    if mensaje.strip().lower() in ("sin foto", "sinfoto", "sin fotos", "omitir", "skip"):
        if (context.user_data.get("esperando_foto_factura")
                or context.user_data.get("esperando_foto_abono")):
            context.user_data.pop("esperando_foto_factura", None)
            context.user_data.pop("esperando_foto_abono", None)
            await update.message.reply_text("👍 Foto omitida. La factura queda registrada sin foto.")
            return

    if not _check_rate_limit(chat_id):
        await update.message.reply_text("⏳ Demasiados mensajes, espera un momento.")
        return

    async with get_chat_lock(chat_id):
        await _procesar_mensaje(update, context, mensaje, chat_id, vendedor)


async def _procesar_mensaje(update, context, mensaje, chat_id, vendedor):
    # ── Auth gate: verificar que el usuario está registrado ──
    from auth.usuarios import get_usuario
    telegram_id = update.effective_user.id
    usuario = get_usuario(telegram_id)
    if not usuario:
        # Limpiar cualquier estado de wizard pendiente para este chat para no
        # dejarlo atascado — el próximo mensaje válido empezará limpio.
        with _estado_lock:
            clientes_en_proceso.pop(chat_id, None)
        await update.message.reply_text(
            "❌ No estás registrado. Escribe /confirmar TuNombre para activar tu acceso."
        )
        return

    # Store usuario in context for dispatch functions
    context.user_data["usuario"] = usuario

    # Usar siempre el nombre registrado en la BD, no el nombre de Telegram
    vendedor = usuario["nombre"]

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    # ── Flujos con handlers propios (sin cambio) ──
    if await manejar_flujo_agregar_producto(update, context):
        return

    # ── Modo actualización de precios (/actualizar_precio) ──
    from handlers.comandos import manejar_mensaje_precio
    if await manejar_mensaje_precio(update, mensaje):
        return

    # ── Flujos extraídos a dispatch.py ──
    if await manejar_flujo_cliente(update, chat_id, mensaje):
        return
    if await manejar_flujo_excel(update, context, chat_id, mensaje):
        return
    if await manejar_flujo_pago_texto(update, context, chat_id, mensaje, vendedor):
        return
    if await manejar_flujo_correccion(update, context, chat_id, mensaje, vendedor):
        return
    if await manejar_rechazo_cliente(update, chat_id, mensaje):
        return

    # ── Actualización de precios: ahora es SOLO via /actualizar_precio ──
    # (Eliminada la intercepción automática que confundía ventas con precios)

    # ── Crear cliente sin venta: "agregar cliente: nombre" ──
    _msg_lower_cli = mensaje.strip().lower()
    _prefijos_cli = ("agregar cliente:", "nuevo cliente:", "crear cliente:", "add cliente:")
    _match_cli = next((p for p in _prefijos_cli if _msg_lower_cli.startswith(p)), None)
    if _match_cli:
        nombre_nuevo = mensaje.strip()[len(_match_cli):].strip()
        if nombre_nuevo:
            with _estado_lock:
                clientes_en_proceso[chat_id] = {"paso": "tipo_id", "nombre": nombre_nuevo}
            await _enviar_pregunta_flujo_cliente(update.message, chat_id)
        else:
            with _estado_lock:
                clientes_en_proceso[chat_id] = {"paso": "nombre"}
            await _enviar_pregunta_flujo_cliente(update.message, chat_id)
        return

    # ── Flujo normal con Claude ──
    try:
        limpiar_pendientes_expirados()
        # ── Reinyectar contexto pendiente si Claude solo hizo una pregunta antes ──
        _ctx_previo   = None
        _pregunta_bot = None
        with _estado_lock:
            _ctx_data = mensaje_contexto_pendiente.pop(chat_id, None)
            if _ctx_data:
                _ctx_previo   = _ctx_data.get("mensaje", "")
                _pregunta_bot = _ctx_data.get("pregunta", "")

        _mensaje_para_claude = mensaje
        if _ctx_previo and _ctx_previo != mensaje:
            # Armar contexto estructurado: pedido original + pregunta del bot + respuesta del cliente.
            # Así Claude puede resolver la ambigüedad sin perder el pedido completo.
            _mensaje_para_claude = (
                f"[PEDIDO ORIGINAL: {_ctx_previo}]\n"
                f"[PREGUNTA DEL BOT: {_pregunta_bot}]\n"
                f"[RESPUESTA DEL CLIENTE: {mensaje}]\n"
                f"Retoma el pedido original aplicando la respuesta del cliente a la pregunta del bot."
            )
            logging.getLogger("ferrebot.mensajes").info(
                f"[CONTEXTO] Reinyectando — pregunta='{(_pregunta_bot or '')[:60]}' + resp='{mensaje}'"
            )

        historial     = get_historial(chat_id)
        agregar_al_historial(chat_id, "user", f"{vendedor}: {mensaje}")
        _modelo_pref = context.user_data.get("modelo_preferido", None)
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {_mensaje_para_claude}", vendedor, historial, modelo_preferido=_modelo_pref)
        texto_respuesta, acciones, archivos_excel = await procesar_acciones_async(respuesta_raw, vendedor, chat_id)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)

        pedir_metodo        = "PEDIR_METODO_PAGO"    in acciones
        iniciar_cliente     = "INICIAR_FLUJO_CLIENTE" in acciones
        pago_pend_aviso     = "PAGO_PENDIENTE_AVISO"  in acciones
        confirmacion_accion = next((a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None)
        cliente_desconocido = next((a for a in acciones if a.startswith("CLIENTE_DESCONOCIDO:")), None)
        logging.getLogger("ferrebot.mensajes").debug(f"[ACCIONES] acciones={acciones} | pedir_metodo={pedir_metodo}")

        # ── Guardar contexto pendiente si Claude solo hizo una pregunta ──
        # Condición: hay texto (una pregunta), pero NO hay ninguna acción de venta/pago/cliente
        _acciones_venta = (pedir_metodo or confirmacion_accion or cliente_desconocido
                           or pago_pend_aviso or iniciar_cliente
                           or any(a.startswith("[VENTA]") or a.startswith("PRECIO_ACTUALIZADO") for a in acciones))
        if texto_respuesta and not _acciones_venta and "?" in texto_respuesta:
            with _estado_lock:
                mensaje_contexto_pendiente[chat_id] = {
                    "mensaje":  _mensaje_para_claude,
                    "pregunta": texto_respuesta,
                }
            logging.getLogger("ferrebot.mensajes").info(
                f"[CONTEXTO] Guardando — msg='{_mensaje_para_claude[:60]}' | "
                f"pregunta='{texto_respuesta[:60]}'"
            )

        _acciones_internas = ("PEDIR_METODO_PAGO", "INICIAR_FLUJO_CLIENTE", "PAGO_PENDIENTE_AVISO")

        # ── Separar aviso "no encontré en catálogo" del resto del texto ──
        _aviso_no_encontrado = ""
        if texto_respuesta:
            _lineas = texto_respuesta.splitlines()
            def _es_aviso_catalogo(l):
                ls = l.strip()
                return ls.startswith("⚠️") and ("catálogo" in ls.lower() or "catalogo" in ls.lower())

            _aviso_lineas_out = []
            _resto_lineas_out = []
            _en_aviso = False
            for _l in _lineas:
                if _es_aviso_catalogo(_l):
                    _en_aviso = True
                    _aviso_lineas_out.append(_l)
                elif _en_aviso:
                    _ls = _l.strip()
                    _es_nueva_accion = (_ls.startswith("✅") or _ls.startswith("🧠")
                                        or _ls.startswith("[VENTA]") or _ls.startswith("🧾")
                                        or _ls.startswith("💰") or _ls.startswith("📋")
                                        or (_ls.startswith("⚠️") and "catálogo" not in _ls.lower()))
                    if _es_nueva_accion or not _ls:
                        _en_aviso = False
                        _resto_lineas_out.append(_l)
                    else:
                        _aviso_lineas_out.append(_l)
                else:
                    _resto_lineas_out.append(_l)

            _aviso_no_encontrado = "\n".join(_aviso_lineas_out).strip()
            texto_respuesta      = "\n".join(_resto_lineas_out).strip()

        # Enviar aviso de no encontrado PRIMERO, como mensaje separado
        if _aviso_no_encontrado:
            await update.message.reply_text(_aviso_no_encontrado)


        if texto_respuesta and not pago_pend_aviso and not cliente_desconocido and not pedir_metodo and not confirmacion_accion:
            await update.message.reply_text(texto_respuesta)

        # Agrupar precios actualizados en un solo mensaje
        precios_act = [a for a in acciones if a.startswith("🧠 Precio")]
        otras_acciones = [a for a in acciones if not a.startswith("🧠 Precio")
                         and not a.startswith("PEDIR_CONFIRMACION:")
                         and not a.startswith("CLIENTE_DESCONOCIDO:")
                         and a not in _acciones_internas]
        if precios_act:
            await update.message.reply_text("\n".join(precios_act))
        for accion in otras_acciones:
            await update.message.reply_text(accion)

        # ── Cliente desconocido: preguntar si quiere crearlo ──
        if cliente_desconocido:
            nombre_cli = cliente_desconocido.split(":", 1)[1]
            keyboard   = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Sí, crear cliente", callback_data=f"cli_crear_si_{chat_id}"),
                InlineKeyboardButton("➡️ No, registrar así", callback_data=f"cli_crear_no_{chat_id}"),
            ]])
            await update.message.reply_text(
                f"👤 El cliente *{nombre_cli}* no está en la base de datos.\n"
                f"¿Quieres crearlo antes de registrar la venta?",
                reply_markup=keyboard,
                parse_mode="Markdown",
            )

        if pago_pend_aviso:
            agregar_a_standby(chat_id, mensaje)
            with _estado_lock:
                ventas = ventas_pendientes.get(chat_id, [])
            await update.message.reply_text("⚠️ Primero confirma el método de pago de la venta anterior:")
            await _enviar_botones_pago(update.message, chat_id, ventas)
        elif not cliente_desconocido:
            # Solo mostrar botones de pago/confirmar si ya se resolvio el cliente
            if confirmacion_accion:
                metodo_conocido = confirmacion_accion.split(":", 1)[1]
                with _estado_lock:
                    ventas = ventas_pendientes.get(chat_id, [])
                await _enviar_confirmacion_con_metodo(update.message, chat_id, ventas, metodo_conocido)
            elif pedir_metodo:
                with _estado_lock:
                    ventas = ventas_pendientes.get(chat_id, [])
                await _enviar_botones_pago(update.message, chat_id, ventas)

        if iniciar_cliente:
            await _enviar_pregunta_cliente(update.message, chat_id)

        for archivo in archivos_excel:
            if os.path.exists(archivo):
                await update.message.reply_text("📊 Aquí está tu reporte:")
                with open(archivo, "rb") as f:
                    await update.message.reply_document(document=f, filename=archivo)
                os.remove(archivo)

    except Exception:
        _tb = traceback.format_exc()
        logger.error("Error en mensaje: %s", _tb)
        print(f"[ERROR _procesar_mensaje]\n{_tb}")  # visible en Railway
        await update.message.reply_text(
            "⚠️ El asistente IA no está disponible ahora. "
            "Puedes registrar la venta manualmente con /ventas."
        )




async def _manejar_foto_factura_o_abono(update, context) -> bool:
    """
    Intercepta fotos cuando el usuario está en flujo /factura o /abonar.
    Retorna True si la foto fue manejada (y manejar_foto no debe continuar).
    """
    fac_factura = context.user_data.pop("esperando_foto_factura", None)
    fac_abono   = context.user_data.pop("esperando_foto_abono", None)
    fac_id      = fac_factura or fac_abono

    if not fac_id:
        return False

    chat_id = update.message.chat_id
    await context.bot.send_chat_action(chat_id=chat_id, action="upload_photo")

    try:
        foto    = update.message.photo[-1]
        archivo = await foto.get_file()
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            ruta_tmp = tmp.name
        await archivo.download_to_drive(ruta_tmp)

        # Buscar factura en PostgreSQL (fuente de verdad)
        import db as _db
        row = await asyncio.to_thread(
            _db.query_one,
            "SELECT id, proveedor, fecha FROM facturas_proveedores WHERE id = %s",
            (fac_id.upper(),)
        )
        if not row:
            await update.message.reply_text(f"⚠️ Factura {fac_id} no encontrada.")
            return True

        proveedor = row["proveedor"]
        fecha_fac = str(row.get("fecha", ""))[:10]
        hoy       = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")

        if fac_factura:
            nombre_archivo = f"{fecha_fac}_{fac_id}.jpg"
        else:
            # Contar abonos existentes para nombrar el comprobante
            n_abono_row = await asyncio.to_thread(
                _db.query_one,
                "SELECT COUNT(*) AS n FROM facturas_abonos WHERE factura_id = %s",
                (fac_id.upper(),)
            )
            n_abono = int((n_abono_row or {}).get("n", 0))
            nombre_archivo = f"{hoy}_{fac_id}_abono{n_abono}.jpg"

        # Subir a Cloudinary
        from handlers.comandos import upload_foto_cloudinary
        with open(ruta_tmp, "rb") as _f:
            foto_bytes = _f.read()
        try:
            import os as _os; _os.unlink(ruta_tmp)
        except Exception:
            pass

        carpeta   = f"ferreteria/{proveedor.lower().replace(' ', '_')}"
        public_id = nombre_archivo.replace(".jpg", "")
        result    = await upload_foto_cloudinary(foto_bytes, public_id, carpeta)

        if not result["ok"]:
            await update.message.reply_text(
                f"⚠️ No se pudo subir la foto a Cloudinary: {result.get('error','')}"
            )
            return True

        # Guardar URL en Postgres
        if fac_factura:
            await asyncio.to_thread(
                _db.execute,
                "UPDATE facturas_proveedores SET foto_url=%s, foto_nombre=%s WHERE id=%s",
                (result["url"], nombre_archivo, fac_id.upper())
            )
        else:
            await asyncio.to_thread(
                _db.execute,
                "UPDATE facturas_abonos SET foto_url=%s, foto_nombre=%s "
                "WHERE factura_id=%s AND id = ("
                "  SELECT id FROM facturas_abonos WHERE factura_id=%s ORDER BY id DESC LIMIT 1"
                ")",
                (result["url"], nombre_archivo, fac_id.upper(), fac_id.upper())
            )

        tipo_txt = "factura" if fac_factura else "comprobante de abono"
        await update.message.reply_text(
            f"📎 Foto del {tipo_txt} guardada en Cloudinary\n"
            f"🏪 {proveedor} · {fac_id}\n"
            f"[Ver foto]({result['url']})",
            parse_mode="Markdown",
            disable_web_page_preview=True,
        )
        return True

    except Exception as e:
        await update.message.reply_text(f"⚠️ Error guardando foto: {e}")
        return True


async def manejar_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Handler de fotos: transcribe ventas anotadas a mano usando visión de Claude.
    Descarga la foto, la convierte a base64 y la manda a procesar_con_claude
    con el system prompt completo (catálogo incluido).
    """
    from auth.usuarios import get_usuario as _get_usuario
    _u = _get_usuario(update.effective_user.id)
    vendedor = _u["nombre"] if _u else (update.message.from_user.first_name or "Desconocido")
    chat_id  = update.message.chat_id

    async with get_chat_lock(chat_id):
        await _procesar_foto(update, context, vendedor, chat_id)


async def _procesar_foto(update: Update, context: ContextTypes.DEFAULT_TYPE, vendedor: str, chat_id: int):

    ruta_foto = None

    try:
        # ── Interceptar fotos de facturas/abonos ─────────────────────────
        if await _manejar_foto_factura_o_abono(update, context):
            return

        # Chequeo de pago pendiente: si hay venta esperando, la foto va al standby
        with _estado_lock:
            _ventas_pend = list(ventas_pendientes.get(chat_id, []))

        if _ventas_pend:
            await update.message.reply_text("⚠️ Primero confirma el método de pago de la venta anterior:")
            await _enviar_botones_pago(update.message, chat_id, _ventas_pend)
            return

        await update.message.reply_text("📸 Leyendo la foto...")
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")

        # Descargar la foto en la resolución más alta disponible
        foto = update.message.photo[-1]  # última = máxima resolución
        _archivo = None
        for _intento in range(3):
            try:
                _archivo = await foto.get_file()
                break
            except Exception as _e:
                if _intento < 2:
                    await asyncio.sleep(1.5 * (_intento + 1))
                else:
                    raise _e

        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            ruta_foto = tmp.name

        for _intento in range(3):
            try:
                await _archivo.download_to_drive(ruta_foto)
                break
            except Exception as _e:
                if _intento < 2:
                    await asyncio.sleep(1.5 * (_intento + 1))
                else:
                    raise _e

        # Convertir a base64
        with open(ruta_foto, "rb") as f:
            imagen_b64 = base64.b64encode(f.read()).decode("utf-8")

        # Detectar tipo de imagen (siempre jpeg en Telegram)
        media_type = "image/jpeg"

        # Caption de la foto como contexto adicional (si el vendedor escribió algo)
        caption = update.message.caption or ""
        mensaje_usuario = f"{vendedor}: {caption}" if caption else f"{vendedor}: foto de ventas"

        # Procesar con Claude (visión activa)
        historial = get_historial(chat_id)
        agregar_al_historial(chat_id, "user", mensaje_usuario)

        # Fix 1: Forzar Sonnet — Haiku no tiene buena vision de manuscritos
        respuesta_raw = await procesar_con_claude(
            mensaje_usuario,
            vendedor,
            historial,
            imagen_b64=imagen_b64,
            imagen_media_type=media_type,
            modelo_preferido="sonnet",
        )

        # Fix 4/5: Extraer resumen del texto de Claude y mostrar confirmacion
        # ANTES de ejecutar las ventas. El vendedor aprueba o cancela la lectura.
        import re as _re
        import json as _json

        # Extraer el texto visible (sin los tags [VENTA])
        _texto_preview = re.sub(r'\[VENTA\].*?\[/VENTA\]', '', respuesta_raw, flags=re.DOTALL).strip()

        # Contar ventas detectadas para el boton
        _ventas_raw = re.findall(r'\[VENTA\](.*?)\[/VENTA\]', respuesta_raw, re.DOTALL)
        _n_ventas = len(_ventas_raw)

        if _n_ventas == 0:
            # Nada registrable — solo mostrar lo que Claude dijo
            texto_limpio = _texto_preview or "No pude leer productos en la foto. Intenta con mejor iluminacion."
            await update.message.reply_text(texto_limpio)
            agregar_al_historial(chat_id, "assistant", texto_limpio)
            return

        # Guardar respuesta_raw en estado pendiente de confirmacion
        from ventas_state import fotos_pendientes_confirmacion
        with _estado_lock:
            fotos_pendientes_confirmacion[chat_id] = respuesta_raw

        agregar_al_historial(chat_id, "assistant", _texto_preview)

        # Mostrar resumen + botones de confirmar/cancelar
        from telegram import InlineKeyboardButton, InlineKeyboardMarkup
        _msg_confirm = (
            (_texto_preview + "\n\n" if _texto_preview else "") +
            f"¿Registro estas {_n_ventas} {'venta' if _n_ventas == 1 else 'ventas'}?"
        )
        _kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(f"✅ Sí, registrar", callback_data=f"foto_confirmar_{chat_id}"),
            InlineKeyboardButton("❌ Cancelar",        callback_data=f"foto_cancelar_{chat_id}"),
        ]])
        await update.message.reply_text(_msg_confirm, reply_markup=_kb)

    except Exception as e:
        logger.error(f"[foto] Error procesando imagen: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Error procesando la foto: {e}")
    finally:
        if ruta_foto and os.path.exists(ruta_foto):
            try:
                os.unlink(ruta_foto)
            except Exception:
                pass

# ─────────────────────────────────────────────
# HELPERS DE AUDIO
# ─────────────────────────────────────────────

# ── Caché del prompt de Whisper (TTL 30 min) ──────────────────────────────────
# Evita reconstruir el prompt en cada audio — las ventas no cambian en segundos.
# Patrón idéntico al de ai/price_cache.py.
_whisper_prompt_cache: dict = {"prompt": None, "ts": 0.0}
_whisper_prompt_lock  = threading.Lock()
_WHISPER_PROMPT_TTL   = 1800  # 30 minutos


def _top_productos_vendidos(limite: int = 12) -> list[str]:
    """
    Retorna los nombres de los N productos más vendidos en los últimos 30 días,
    ordenados de mayor a menor volumen. Fuente: ventas_detalle JOIN ventas.
    """
    try:
        import db as _db
        filas = _db.query_all(
            """
            SELECT vd.producto_nombre,
                   SUM(vd.cantidad) AS total
            FROM ventas_detalle vd
            JOIN ventas v ON v.id = vd.venta_id
            WHERE v.fecha >= NOW() - INTERVAL '30 days'
              AND vd.producto_nombre IS NOT NULL
              AND vd.sin_detalle IS NOT TRUE
            GROUP BY vd.producto_nombre
            ORDER BY total DESC
            LIMIT %s
            """,
            [limite],
        )
        return [r["producto_nombre"] for r in filas if r.get("producto_nombre")]
    except Exception:
        return []


def _palabras_mas_corregidas(limite: int = 8) -> list[str]:
    """
    Analiza audio_logs del último mes para encontrar las palabras que Whisper
    transcribe mal con más frecuencia — son las que más necesitan el prior.
    Lógica: palabras presentes en texto_corregido pero no en texto_original
    = las que el pipeline tuvo que reparar.
    """
    try:
        import re as _re
        from collections import Counter as _Counter
        import db as _db

        filas = _db.query_all(
            """
            SELECT texto_original, texto_corregido
            FROM audio_logs
            WHERE fecha >= NOW() - INTERVAL '30 days'
              AND texto_original <> texto_corregido
            LIMIT 200
            """,
            [],
        )
        contador: _Counter = _Counter()
        for fila in filas:
            orig = set(_re.findall(r'\w+', fila["texto_original"].lower()))
            corr = set(_re.findall(r'\w+', fila["texto_corregido"].lower()))
            for palabra in (corr - orig):
                if len(palabra) > 3:  # ignorar artículos y partículas
                    contador[palabra] += 1
        return [p for p, _ in contador.most_common(limite)]
    except Exception:
        return []


def _build_whisper_prompt() -> str:
    """
    Construye el prompt de vocabulario para Whisper con tres capas de prioridad:
      1. Top 12 productos más vendidos del mes  → mayor probabilidad de aparecer
      2. Top 8 palabras más corregidas en audio_logs → errores reales detectados
      3. Términos colombianos/técnicos que Whisper no conoce por entrenamiento

    El resultado se cachea 30 min (TTL) para no consultar la BD en cada audio.
    LÍMITE WHISPER: 224 tokens. Este prompt ocupa ~120-150 tokens.
    """
    # ── Verificar caché ───────────────────────────────────────────────────────
    with _whisper_prompt_lock:
        if _whisper_prompt_cache["prompt"] and (
            time.monotonic() - _whisper_prompt_cache["ts"] < _WHISPER_PROMPT_TTL
        ):
            return _whisper_prompt_cache["prompt"]

    # ── Capa 1: más vendidos — van PRIMERO porque sobreviven el truncado ──────
    top_vendidos   = _top_productos_vendidos(12)

    # ── Capa 2: más corregidos — errores reales detectados en producción ──────
    mas_corregidos = _palabras_mas_corregidas(8)

    # ── Capa 3: términos base — solo los que Whisper genuinamente no conoce ───
    # Excluye palabras españolas comunes (metro, litro, caja) que Whisper
    # ya maneja bien. Solo marcas, nombres colombianos y medidas fraccionarias.
    _HARD_TERMS = (
        "drywall, waypercol, latecol, pañete, chazos, bisagra, segueta, "
        "puntilla, thinner, broca 3/8, tornillo 1x6, manguera 1/2, "
        "tres cuartos, cinco octavos, tres octavos"
    )

    # ── Ensamblar en orden de prioridad ───────────────────────────────────────
    partes = []
    if top_vendidos:
        partes.append(", ".join(top_vendidos))
    if mas_corregidos:
        partes.append(", ".join(mas_corregidos))
    partes.append(_HARD_TERMS)

    prompt = f"Ferretería colombiana. {', '.join(partes)}."

    # ── Guardar en caché ──────────────────────────────────────────────────────
    with _whisper_prompt_lock:
        _whisper_prompt_cache["prompt"] = prompt
        _whisper_prompt_cache["ts"]     = time.monotonic()

    logger.info(
        "[audio] Whisper prompt reconstruido — vendidos:%d corregidos:%d",
        len(top_vendidos), len(mas_corregidos),
    )
    return prompt



async def _normalizar_con_haiku(texto: str) -> str:
    """
    Normaliza la transcripción al vocabulario del catálogo usando Haiku.
    Corrige errores residuales que corregir_texto_audio() no alcanza.
    Solo se ejecuta si el texto tiene ≥5 palabras para evitar llamadas innecesarias.
    """
    if not texto or len(texto.split()) < 5:
        return texto
    try:
        from memoria import cargar_memoria as _cm
        catalogo = _cm().get("catalogo", {})
        nombres = sorted(
            {v["nombre"] for v in catalogo.values() if v.get("nombre")},
            key=len
        )[:60]
        vocab = ", ".join(nombres)

        system = (
            "Eres corrector de transcripciones de audio para una ferretería colombiana. "
            "Tu ÚNICA tarea: corregir errores de pronunciación y transcripción de productos. "
            "Devuelve SOLO el texto corregido, sin explicaciones ni comentarios. "
            "Si no hay errores claros, devuelve el texto exactamente igual. "
            "No agregues ni quites productos. No reformules. Solo corrige.\n"
            f"Vocabulario del catálogo: {vocab}"
        )
        respuesta = await asyncio.to_thread(
            lambda: config.claude_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                system=system,
                messages=[{"role": "user", "content": texto}],
            )
        )
        resultado = respuesta.content[0].text.strip()
        # Sanity check: Haiku no debería devolver algo mucho más largo
        if resultado and len(resultado) < len(texto) * 2:
            return resultado
    except Exception as _e:
        logger.warning(f"[audio] Haiku normalization falló: {_e}")
    return texto


def _log_audio(chat_id: int, vendedor: str, texto_original: str, texto_corregido: str) -> None:
    """Guarda la transcripción en audio_logs para análisis posterior. No bloquea."""
    try:
        import db as _db
        _db.execute(
            """
            INSERT INTO audio_logs (chat_id, vendedor, texto_original, texto_corregido)
            VALUES (%s, %s, %s, %s)
            """,
            [chat_id, vendedor, texto_original, texto_corregido],
        )
    except Exception as _e:
        logger.warning(f"[audio] No se pudo guardar en audio_logs: {_e}")


async def manejar_audio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from auth.usuarios import get_usuario as _get_usuario
    _u = _get_usuario(update.effective_user.id)
    vendedor = _u["nombre"] if _u else (update.message.from_user.first_name or "Desconocido")
    chat_id  = update.message.chat_id

    async with get_chat_lock(chat_id):
        await _procesar_audio(update, context, vendedor, chat_id)


async def _procesar_audio(update: Update, context: ContextTypes.DEFAULT_TYPE, vendedor: str, chat_id: int):

    # CORRECCIÓN: inicializar ruta_audio ANTES del try para que el finally
    # no tenga NameError si la descarga falla antes de asignar la variable
    ruta_audio = None

    try:
        # Reintentos ante timeout de red de Telegram (común en Railway con servidor frío)
        _MAX_REINTENTOS = 3
        _archivo_voz    = None
        for _intento in range(_MAX_REINTENTOS):
            try:
                _archivo_voz = await update.message.voice.get_file()
                break
            except Exception as _e:
                if _intento < _MAX_REINTENTOS - 1:
                    await asyncio.sleep(1.5 * (_intento + 1))
                else:
                    raise _e

        archivo_voz = _archivo_voz

        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            ruta_audio = tmp.name  # asignamos aquí, ANTES de la descarga

        # Reintentos también en la descarga del archivo
        for _intento in range(_MAX_REINTENTOS):
            try:
                await archivo_voz.download_to_drive(ruta_audio)
                break
            except Exception as _e:
                if _intento < _MAX_REINTENTOS - 1:
                    await asyncio.sleep(1.5 * (_intento + 1))
                else:
                    raise _e

        # ── [1] Prompt de vocabulario del catálogo para Whisper ───────────────
        _whisper_prompt = await asyncio.to_thread(_build_whisper_prompt)

        # Si hay una pregunta aclaratoria pendiente, añadirla AL INICIO del prompt
        # de Whisper. Esto mejora drásticamente la transcripción de respuestas cortas
        # (ej: "segmentado de 4" se escucha igual que "metro de 4" sin contexto).
        # Usamos get() sin pop() para no consumir el estado antes de que Claude lo use.
        with _estado_lock:
            _ctx_audio = mensaje_contexto_pendiente.get(chat_id)
        if _ctx_audio and _ctx_audio.get("pregunta"):
            _hint = _ctx_audio["pregunta"][:200].replace("\n", " ")
            _whisper_prompt = f"{_hint} {_whisper_prompt}"
            logger.info(f"[audio] Whisper enriquecido con pregunta pendiente: '{_hint[:60]}'")

        def _transcribir():
            with open(ruta_audio, "rb") as audio_file:
                return config.openai_client.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file,
                    language="es",
                    prompt=_whisper_prompt,
                )

        transcripcion = None
        for _w_intento in range(3):
            try:
                transcripcion = await asyncio.to_thread(_transcribir)
                break
            except Exception as _w_e:
                if _w_intento < 2:
                    logger.warning(
                        f"[audio] Whisper reintento {_w_intento + 1}/2 en 2s: {_w_e}"
                    )
                    await asyncio.sleep(2)
                else:
                    logger.error(f"[audio] Whisper sin respuesta tras 3 intentos: {_w_e}")
                    await update.message.reply_text(
                        "⚠️ El asistente IA no está disponible ahora. "
                        "Puedes registrar la venta manualmente con /ventas."
                    )
                    return

        # ── [2] Correcciones regex — rápidas, sin red ────────────────────────
        texto_raw = transcripcion.text
        texto     = corregir_texto_audio(texto_raw)

        # ── [3] Mostrar transcripción al usuario inmediatamente ──────────────
        # Haiku corre DESPUÉS de este punto — el vendedor ve el 📝 sin esperar a Haiku
        await update.message.reply_text(f"📝 {texto}")
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")

        # ── [4] Normalización con Haiku (corre mientras Claude piensa) ───────
        # Claude recibe el texto ya mejorado; si Haiku falla, usa el texto regex
        texto = await _normalizar_con_haiku(texto)

        # ── [5] Log a audio_logs (no bloquea el flujo principal) ─────────────
        _log_audio(chat_id, vendedor, texto_raw, texto)

        # ── Verificar modo corrección/modificación (igual que en texto) ──
        with _estado_lock:
            en_correccion = esperando_correccion.pop(chat_id, False)

        if en_correccion == "modificar":
            # Usar la misma lógica que el handler de texto para modificaciones
            with _estado_lock:
                ventas_actuales = list(ventas_pendientes.get(chat_id, []))

            metodo_original = None
            if ventas_actuales and ventas_actuales[0].get("metodo_pago"):
                metodo_original = ventas_actuales[0]["metodo_pago"]

            resumen_venta = json.dumps(ventas_actuales, ensure_ascii=False)

            prompt_modificacion = (
                "El vendedor tiene esta venta pendiente de confirmar:\n"
                + resumen_venta
                + "\n\nEl vendedor quiere modificarla con esta instrucción: "
                + texto
                + "\n\nAplica EXACTAMENTE los cambios pedidos a la venta (modifica cantidad, precio, "
                "quita o agrega productos según corresponda). "
                "Luego emite los [VENTA] actualizados con los datos correctos y confirma los cambios en texto. "
                "IMPORTANTE: emite [VENTA] para TODOS los productos que quedan en la venta (no solo el modificado). "
                + (f"IMPORTANTE: mantén metodo_pago={metodo_original} en todos los [VENTA]." if metodo_original else "")
            )
            historial = get_historial(chat_id)
            agregar_al_historial(chat_id, "user", prompt_modificacion)

            with _estado_lock:
                ventas_pendientes.pop(chat_id, None)

            respuesta_raw = await procesar_con_claude(prompt_modificacion, vendedor, historial)
            texto_respuesta, acciones, archivos_excel = await procesar_acciones_async(respuesta_raw, vendedor, chat_id)
            agregar_al_historial(chat_id, "assistant", texto_respuesta)

            confirmacion_accion = next((a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None)
            with _estado_lock:
                ventas_nuevas = list(ventas_pendientes.get(chat_id, []))

            if ventas_nuevas:
                nota = texto_respuesta.split("\n")[0] if texto_respuesta else ""
                if confirmacion_accion:
                    metodo_conocido = confirmacion_accion.split(":", 1)[1]
                    await _enviar_confirmacion_con_metodo(update.message, chat_id, ventas_nuevas, metodo_conocido, nota=nota)
                elif metodo_original:
                    with _estado_lock:
                        for v in ventas_pendientes.get(chat_id, []):
                            v["metodo_pago"] = metodo_original
                    await _enviar_confirmacion_con_metodo(update.message, chat_id, ventas_nuevas, metodo_original, nota=nota)
                else:
                    if nota:
                        await update.message.reply_text(nota)
                    await _enviar_botones_pago(update.message, chat_id, ventas_nuevas)
            elif texto_respuesta:
                await update.message.reply_text(texto_respuesta)
            return

        # ── Chequeo de pago pendiente: si hay venta esperando, el audio va al standby ──
        with _estado_lock:
            _ventas_pend = list(ventas_pendientes.get(chat_id, []))

        if _ventas_pend:
            agregar_a_standby(chat_id, texto)
            await update.message.reply_text("⚠️ Primero confirma el método de pago de la venta anterior:")
            await _enviar_botones_pago(update.message, chat_id, _ventas_pend)
            return

        # ── Reinyectar contexto pendiente (misma lógica que en mensajes de texto) ──
        _ctx_previo_audio   = None
        _pregunta_bot_audio = None
        with _estado_lock:
            _ctx_data_audio = mensaje_contexto_pendiente.pop(chat_id, None)
            if _ctx_data_audio:
                _ctx_previo_audio   = _ctx_data_audio.get("mensaje", "")
                _pregunta_bot_audio = _ctx_data_audio.get("pregunta", "")

        _texto_para_claude = texto
        if _ctx_previo_audio and _ctx_previo_audio != texto:
            _texto_para_claude = (
                f"[PEDIDO ORIGINAL: {_ctx_previo_audio}]\n"
                f"[PREGUNTA DEL BOT: {_pregunta_bot_audio}]\n"
                f"[RESPUESTA DEL CLIENTE (audio): {texto}]\n"
                f"Retoma el pedido original aplicando la respuesta del cliente a la pregunta del bot."
            )
            logger.info(
                f"[audio] Reinyectando contexto — pregunta='{(_pregunta_bot_audio or '')[:60]}' "
                f"+ resp='{texto}'"
            )

        historial     = get_historial(chat_id)
        agregar_al_historial(chat_id, "user", f"{vendedor}: {texto}")
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {_texto_para_claude}", vendedor, historial)
        texto_respuesta, acciones, archivos_excel = await procesar_acciones_async(respuesta_raw, vendedor, chat_id)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)

        pedir_metodo        = "PEDIR_METODO_PAGO" in acciones
        confirmacion_accion = next((a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None)
        cliente_desconocido = next((a for a in acciones if a.startswith("CLIENTE_DESCONOCIDO:")), None)
        pago_pend_aviso     = "PAGO_PENDIENTE_AVISO" in acciones
        _acciones_internas  = ("PEDIR_METODO_PAGO", "INICIAR_FLUJO_CLIENTE", "PAGO_PENDIENTE_AVISO")

        # Si Claude volvió a hacer una pregunta (sin registrar venta), guardar contexto
        _acciones_venta_audio = (pedir_metodo or confirmacion_accion or cliente_desconocido
                                 or pago_pend_aviso
                                 or any(a.startswith("[VENTA]") or a.startswith("PRECIO_ACTUALIZADO")
                                        for a in acciones))
        if texto_respuesta and not _acciones_venta_audio and "?" in texto_respuesta:
            with _estado_lock:
                mensaje_contexto_pendiente[chat_id] = {
                    "mensaje":  _texto_para_claude,
                    "pregunta": texto_respuesta,
                }
            logger.info(f"[audio] Guardando contexto pendiente para chat {chat_id}")

        # En audio: mostrar texto de Claude solo si NO hay botones de venta (es una pregunta o error)
        hay_botones_venta = confirmacion_accion or pedir_metodo
        if texto_respuesta and (not hay_botones_venta or cliente_desconocido):
            await update.message.reply_text(texto_respuesta)

        precios_act2 = [a for a in acciones if a.startswith("🧠 Precio")]
        otras_acciones2 = [a for a in acciones if not a.startswith("🧠 Precio")
                          and not a.startswith("PEDIR_CONFIRMACION:")
                          and not a.startswith("CLIENTE_DESCONOCIDO:")
                          and a not in _acciones_internas]
        if precios_act2:
            await update.message.reply_text("\n".join(precios_act2))
        for accion in otras_acciones2:
            await update.message.reply_text(accion)

        if cliente_desconocido:
            nombre_cli = cliente_desconocido.split(":", 1)[1]
            from telegram import InlineKeyboardMarkup, InlineKeyboardButton
            keyboard = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Sí, crear cliente", callback_data=f"cli_crear_si_{chat_id}"),
                InlineKeyboardButton("➡️ No, registrar así", callback_data=f"cli_crear_no_{chat_id}"),
            ]])
            await update.message.reply_text(
                f"👤 *{nombre_cli}* no está en la base. ¿Lo agrego como cliente?",
                reply_markup=keyboard, parse_mode="Markdown",
            )

        if not cliente_desconocido:
            if confirmacion_accion:
                metodo_conocido = confirmacion_accion.split(":", 1)[1]
                with _estado_lock:
                    ventas = ventas_pendientes.get(chat_id, [])
                await _enviar_confirmacion_con_metodo(update.message, chat_id, ventas, metodo_conocido)
            elif pedir_metodo:
                with _estado_lock:
                    ventas = ventas_pendientes.get(chat_id, [])
                await _enviar_botones_pago(update.message, chat_id, ventas)

        for archivo in archivos_excel:
            if os.path.exists(archivo):
                await update.message.reply_text("📊 Aquí está tu reporte:")
                with open(archivo, "rb") as f:
                    await update.message.reply_document(document=f, filename=archivo)
                os.remove(archivo)

    except Exception:
        logger.error("Error en audio: %s", traceback.format_exc())
        await update.message.reply_text("Problema con el audio. Intenta de nuevo.")
    finally:
        if ruta_audio and os.path.exists(ruta_audio):
            try:
                os.unlink(ruta_audio)
            except Exception:
                pass


async def manejar_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja archivos Excel que le mandan al bot."""
    doc = update.message.document
    if not doc:
        return

    nombre    = doc.file_name or "archivo"
    extension = nombre.split(".")[-1].lower() if "." in nombre else ""

    if extension not in ["xlsx", "xls"]:
        await update.message.reply_text(
            f"Recibí '{nombre}'. Solo puedo procesar archivos Excel (.xlsx). ¿Necesitas algo más?"
        )
        return

    # ── Si es BASE_DE_DATOS_PRODUCTOS, importar catálogo automáticamente ──
    if "base_de_datos_productos" in nombre.lower() or "base de datos" in nombre.lower():
        await update.message.reply_text("📦 Detecté el catálogo de productos. Importando...")
        try:
            ruta_temp = f"temp_catalogo_{update.message.chat_id}.xlsx"
            archivo   = await doc.get_file()
            await archivo.download_to_drive(ruta_temp)

            resultado  = await asyncio.to_thread(importar_catalogo_desde_excel, ruta_temp)
            importados = resultado["importados"]
            omitidos   = resultado["omitidos"]
            errores    = resultado["errores"]

            texto = (
                f"✅ Catálogo actualizado exitosamente\n\n"
                f"📦 {importados} productos importados\n"
                f"⏭️ {omitidos} filas omitidas (sin nombre o precio)"
            )
            if errores:
                texto += f"\n⚠️ {len(errores)} errores:\n" + "\n".join(f"  • {e}" for e in errores[:5])
            await update.message.reply_text(texto)

            if os.path.exists(ruta_temp):
                os.remove(ruta_temp)
        except Exception as e:
            await update.message.reply_text(f"❌ Error importando catálogo: {e}")
        return

    await update.message.reply_text(f"📂 Recibí tu archivo '{nombre}'. Leyendo contenido...")
    chat_id = update.message.chat_id

    try:
        excel_temp_anterior = context.user_data.get("excel_temp")
        if excel_temp_anterior and os.path.exists(excel_temp_anterior):
            try:
                os.remove(excel_temp_anterior)
            except Exception:
                pass
        context.user_data.pop("excel_temp", None)
        context.user_data.pop("excel_nombre", None)

        archivo   = await doc.get_file()
        ruta_temp = f"temp_{chat_id}_{nombre}"
        await archivo.download_to_drive(ruta_temp)

        def _leer_excel():
            wb = openpyxl.load_workbook(ruta_temp, read_only=True)
            resumen_hojas = []
            for hoja_nombre in wb.sheetnames:
                ws  = wb[hoja_nombre]
                enc = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value]
                resumen_hojas.append(f"Hoja '{hoja_nombre}': {ws.max_row - 1} filas, columnas: {', '.join(str(e) for e in enc)}")
            wb.close()
            return "\n".join(resumen_hojas)

        resumen = await asyncio.to_thread(_leer_excel)
        context.user_data["excel_temp"]   = ruta_temp
        context.user_data["excel_nombre"] = nombre

        await update.message.reply_text(
            f"✅ Excel cargado correctamente.\n\n{resumen}\n\n"
            f"Ahora dime qué quieres hacer con él. Por ejemplo:\n"
            f"- Agrega una columna de IVA del 19%\n"
            f"- Ordena de mayor a menor por total\n"
            f"- Cambia los encabezados a color rojo\n"
            f"- Calcula el total de todas las ventas"
        )
    except Exception:
        logger.error("Error leyendo Excel: %s", traceback.format_exc())
        await update.message.reply_text("Tuve un problema leyendo el archivo. Asegúrate de que sea un Excel válido.")
