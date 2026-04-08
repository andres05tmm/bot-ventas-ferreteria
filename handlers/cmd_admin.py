"""
handlers/cmd_admin.py — Comandos de administración y sistema.

Handlers: comando_consistencia, comando_exportar_precios,
          comando_keepalive, comando_modelo
"""

# -- stdlib --
import logging
import os
from datetime import datetime

# -- terceros --
from telegram import Update
from telegram.ext import ContextTypes

# -- propios --
import config
import db as _db
from middleware import protegido

logger = logging.getLogger("ferrebot.handlers.cmd_admin")


# ─────────────────────────────────────────────────────────────────────────────
# /consistencia  — compara memoria vs catálogo (sin Drive)
# ─────────────────────────────────────────────────────────────────────────────

@protegido
async def comando_consistencia(update, context):
    """
    Verifica consistencia del catálogo en PostgreSQL.
    Muestra productos sin precio, sin fracciones configuradas y totales por categoría.
    """
    await update.message.reply_text("🔍 Verificando consistencia del catálogo en PG…")
    try:
        if not _db.DB_DISPONIBLE:
            await update.message.reply_text("⚠️ Base de datos no disponible.")
            return

        total = _db.query_one("SELECT COUNT(*) AS n FROM productos WHERE activo = TRUE")["n"]
        sin_precio = _db.query_all(
            "SELECT nombre FROM productos WHERE activo = TRUE AND precio_unidad = 0 ORDER BY nombre LIMIT 20"
        )
        sin_fracciones = _db.query_one(
            """SELECT COUNT(*) AS n FROM productos p
               WHERE p.activo = TRUE
                 AND NOT EXISTS (SELECT 1 FROM productos_fracciones pf WHERE pf.producto_id = p.id)
                 AND p.precio_umbral IS NULL
                 AND p.precio_unidad = 0"""
        )["n"]
        por_cat = _db.query_all(
            """SELECT categoria, COUNT(*) AS n
               FROM productos WHERE activo = TRUE
               GROUP BY categoria ORDER BY n DESC LIMIT 10"""
        )

        lineas = ["📊 CONSISTENCIA DEL CATÁLOGO (PG)", "─" * 32,
                  f"📦 Total productos activos: {total}",
                  f"❌ Sin precio unitario:     {len(sin_precio)}",
                  f"⚠️  Sin precios completos:   {sin_fracciones}",
                  "", "── Por categoría ──"]
        for r in por_cat:
            lineas.append(f"  {r['categoria'] or '(sin cat)':<30} {r['n']:>4}")

        if sin_precio:
            lineas += ["", "── Sin precio (primeros 20) ──"]
            lineas += [f"  • {r['nombre']}" for r in sin_precio]

        if not sin_precio and sin_fracciones == 0:
            lineas += ["", "🎉 ¡Todo consistente!"]

        await update.message.reply_text("\n".join(lineas))
    except Exception as e:
        await update.message.reply_text(f"❌ Error en verificación: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# /exportar_precios  — genera reporte local (sin Drive)
# ─────────────────────────────────────────────────────────────────────────────

@protegido
async def comando_exportar_precios(update, context):
    """
    Genera un Excel con todos los precios directamente desde PostgreSQL y lo envía en el chat.
    """
    await update.message.reply_text("📤 Exportando precios desde PG…")
    try:
        if not _db.DB_DISPONIBLE:
            await update.message.reply_text("⚠️ Base de datos no disponible.")
            return

        productos   = _db.query_all(
            "SELECT id, nombre, categoria, precio_unidad FROM productos WHERE activo = TRUE ORDER BY nombre"
        )
        fracciones  = _db.query_all(
            "SELECT producto_id, fraccion, precio_total FROM productos_fracciones ORDER BY producto_id, fraccion"
        )
        por_cant    = _db.query_all(
            "SELECT id, precio_umbral AS umbral, precio_bajo_umbral, precio_sobre_umbral FROM productos WHERE precio_umbral IS NOT NULL"
        )

        # Índices para lookup rápido
        frac_idx  = {}
        for f in fracciones:
            frac_idx.setdefault(f["producto_id"], []).append(f)
        cant_idx  = {r["id"]: r for r in por_cant}

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import tempfile

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Precios"

        HEADERS = ["Nombre", "Categoría", "Precio Unidad",
                   "3/4", "1/2", "1/4", "1/8", "1/16",
                   "Mayorista (umbral)", "Precio mayorista"]
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1A56DB")
            c.alignment = Alignment(horizontal="center")

        FRACS = ["3/4", "1/2", "1/4", "1/8", "1/16"]
        for i, prod in enumerate(productos, 2):
            pid  = prod["id"]
            fs   = {f["fraccion"]: f["precio_total"] for f in frac_idx.get(pid, [])}
            pc   = cant_idx.get(pid)
            row  = [
                prod["nombre"],
                prod["categoria"] or "",
                prod["precio_unidad"],
                *[fs.get(frac, "") for frac in FRACS],
                pc["umbral"]               if pc else "",
                pc["precio_sobre_umbral"]  if pc else "",
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col, value=val)

        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            ruta = tmp.name
        wb.save(ruta)

        await update.message.reply_text(
            f"📤 EXPORTACIÓN COMPLETADA\n{'─'*30}\n"
            f"✅ Productos exportados: {len(productos)}\n"
            f"📊 Con fracciones:       {len(frac_idx)}\n"
            f"💰 Con precio mayorista: {len(cant_idx)}"
        )
        with open(ruta, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="catalogo_precios_pg.xlsx",
                caption="📎 Catálogo de precios desde PostgreSQL"
            )
        os.remove(ruta)
    except Exception as e:
        await update.message.reply_text(f"❌ Error en exportación: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# /keepalive
# ─────────────────────────────────────────────────────────────────────────────

@protegido
async def comando_keepalive(update, context):
    from keepalive import keepalive_activo, set_keepalive

    args = context.args
    if args:
        arg = args[0].lower().strip()
        if arg == "on":
            set_keepalive(True)
            await update.message.reply_text("✅ Keep-alive ACTIVADO")
        elif arg == "off":
            set_keepalive(False)
            await update.message.reply_text("⏸ Keep-alive DESACTIVADO")
        else:
            await update.message.reply_text("Uso: /keepalive on | /keepalive off")
        return

    from datetime import time
    activo  = keepalive_activo()
    ahora   = datetime.now(config.COLOMBIA_TZ).time()
    horario = time(8, 0) <= ahora <= time(11, 0)

    if activo and horario:
        estado_emoji, estado_texto = "🟢", "ACTIVO y en horario (ping cada 4 min)"
    elif activo:
        estado_emoji, estado_texto = "🟡", "ACTIVADO pero fuera de horario 8-11am"
    else:
        estado_emoji, estado_texto = "🔴", "DESACTIVADO manualmente"

    await update.message.reply_text(
        f"{estado_emoji} Keep-alive: {estado_texto}\n\n"
        "/keepalive on  → activar\n"
        "/keepalive off → desactivar"
    )


# ─────────────────────────────────────────────────────────────────────────────
# /modelo
# ─────────────────────────────────────────────────────────────────────────────

@protegido
async def comando_modelo(update, context):
    args    = context.args
    opciones_validas = ("auto", "haiku", "sonnet")

    if not args:
        actual = context.user_data.get("modelo_preferido", "auto")
        await update.message.reply_text(
            f"🤖 *Modelo actual:* `{actual}`\n\n"
            f"  `/modelo auto` — selección automática\n"
            f"  `/modelo haiku` — ⚡ rápido y eficiente\n"
            f"  `/modelo sonnet` — 🧠 más inteligente",
            parse_mode="Markdown"
        )
        return

    seleccion = args[0].lower()
    if seleccion not in opciones_validas:
        await update.message.reply_text("❌ Opción inválida. Usa: `auto`, `haiku` o `sonnet`", parse_mode="Markdown")
        return

    context.user_data["modelo_preferido"] = seleccion
    emojis       = {"auto":"⚙️","haiku":"⚡","sonnet":"🧠"}
    descripciones = {
        "auto":   "selección automática según el mensaje",
        "haiku":  "rápido y eficiente para ventas simples",
        "sonnet": "más inteligente para consultas complejas",
    }
    await update.message.reply_text(
        f"{emojis[seleccion]} *Modelo cambiado a `{seleccion}`*\n_{descripciones[seleccion]}_",
        parse_mode="Markdown"
    )
