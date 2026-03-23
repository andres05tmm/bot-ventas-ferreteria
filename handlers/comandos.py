"""
Handlers de comandos de Telegram.
Las operaciones de I/O bloqueantes (Excel, Drive) se ejecutan via asyncio.to_thread
para no bloquear el event loop.
"""

import asyncio
import json
import os
import traceback
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

import config
from excel import (
    inicializar_excel, obtener_o_crear_hoja,
    detectar_columnas, buscar_ventas, obtener_ventas_recientes,
    buscar_clientes_multiples, cargar_clientes,
    registrar_compra_en_excel, actualizar_hoja_inventario,
    recalcular_caja_desde_excel,
)
from memoria import (
    cargar_memoria, obtener_resumen_caja, cargar_gastos_hoy,
    cargar_inventario, verificar_alertas_inventario,
    resumen_fiados, detalle_fiado_cliente, abonar_fiado,
    importar_catalogo_desde_excel,
    registrar_conteo_inventario, ajustar_inventario,
    buscar_productos_inventario, buscar_clave_inventario,
    registrar_compra, obtener_resumen_margenes,
)
from sheets import (
    sheets_leer_ventas_del_dia, sheets_detectar_ediciones_vs_excel,
    sheets_limpiar,
)
from drive import subir_a_drive
from utils import convertir_fraccion_a_decimal, obtener_nombre_hoja
from ventas_state import borrados_pendientes, _estado_lock


# ─────────────────────────────────────────────
# /start y /ayuda
# ─────────────────────────────────────────────

async def comando_inicio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    estado_drive  = "✅ Drive conectado" if config._get_drive_disponible() else "⚠️ Drive offline"
    estado_sheets = (
        "✅ Sheets conectado" if config._get_sheets_disponible() else
        ("⚠️ Sheets no configurado" if not config.SHEETS_ID else "⚠️ Sheets sin conexion")
    )
    await update.message.reply_text(
        "👋 Hola! Soy tu asistente de la ferreteria.\n\n"
        "📦 VENTAS\n"
        "/ventas — Ver ventas del dia\n"
        "/buscar [texto] — Buscar ventas por producto o cliente\n"
        "/borrar [#] — Borrar consecutivo completo\n\n"
        "💰 CAJA Y GASTOS\n"
        "/caja — Estado actual de caja\n"
        "/caja abrir [monto] — Abrir caja del dia (ej: /caja abrir 50000)\n"
        "/gastos — Gastos registrados hoy\n"
        "/cerrar — Cierre del dia (Excel + limpia Sheets + cierra caja)\n"
        "/resetventas CONFIRMAR — Limpiar ventas del dia actual\n"
        "/resetventas excel CONFIRMAR DD/MM/YYYY — Borrar ventas de una fecha\n\n"
        "📊 REPORTES\n"
        "/grafica — Graficas de ventas\n"
        "/excel — Descargar Excel acumulado\n"
        "/sheets — Estado del Sheet del dia\n\n"
        "🏪 INVENTARIO Y PRECIOS\n"
        "/inventario — Ver inventario actual\n"
        "/inv [cantidad] [producto] — Registrar conteo de inventario\n"
        "/stock [producto] — Detalle de stock\n"
        "/ajuste [+/-cantidad] [producto] — Ajustar stock (ej: /ajuste +10 brocha)\n"
        "/compra [cantidad] [producto] a [costo] — Registrar compra\n"
        "/precios — Ver catalogo de precios\n"
        "/actualizar_precio — Cambiar precios de productos\n"
        "/margenes — Ver margenes de ganancia\n"
        "/actualizar_catalogo — Recargar catalogo desde Excel\n\n"
        "👥 CLIENTES Y FIADOS\n"
        "/clientes — Ver lista de clientes\n"
        "/fiados — Ver todas las cuentas fiadas\n"
        "/fiados [nombre] — Ver detalle de un cliente\n"
        "/abono [nombre] [monto] — Registrar abono\n\n"
        "⚙️ SISTEMA\n"
        "/agregar_producto — Agregar nuevo producto al catalogo\n"
        "/consistencia — Verificar consistencia del catalogo\n"
        "/exportar_precios — Exportar lista de precios a Excel\n"
        "/alias — Gestionar alias de productos (ver /alias para ayuda)\n"
        "/keepalive on/off — Activar/desactivar cache de prompts\n\n"
        f"{estado_drive} | {estado_sheets}"
    )


# ─────────────────────────────────────────────
# /excel
# ─────────────────────────────────────────────

async def comando_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await asyncio.to_thread(inicializar_excel)
    await update.message.reply_text("📎 Aqui esta tu archivo de ventas:")
    with open(config.EXCEL_FILE, "rb") as archivo:
        await update.message.reply_document(document=archivo, filename="ventas.xlsx")


# ─────────────────────────────────────────────
# /ventas
# ─────────────────────────────────────────────

async def comando_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if config.SHEETS_ID and config._get_sheets_disponible():
        ventas_raw = await asyncio.to_thread(sheets_leer_ventas_del_dia)
        if ventas_raw:
            total_dia = 0
            encabezado = f"📋 Ventas de hoy ({len(ventas_raw)}):\n\n"
            lineas = []
            for v in ventas_raw:
                num      = v.get("num", "?")
                producto = v.get("producto", "?")
                vendedor = v.get("vendedor", "?")
                total_raw = v.get("total")
                try:
                    t = float(total_raw) if total_raw else 0
                    total_dia += t
                    total_fmt = f"${t:,.0f}"
                except (ValueError, TypeError):
                    total_fmt = str(total_raw) if total_raw else "?"
                metodo = v.get("metodo", "")
                linea = f"#{num} — {producto} — {total_fmt} — {vendedor}"
                if metodo:
                    linea += f" ({metodo})"
                lineas.append(linea)
            pie = f"\n💰 Total del día: ${total_dia:,.0f}\n\nUsa /borrar [numero] para eliminar una venta."

            # Partir en bloques de max 4000 chars para no superar el limite de Telegram
            bloque = encabezado
            for linea in lineas:
                if len(bloque) + len(linea) + 1 > 4000:
                    await update.message.reply_text(bloque)
                    bloque = ""
                bloque += linea + "\n"
            bloque += pie
            await update.message.reply_text(bloque)
            return

    await update.message.reply_text("No hay ventas registradas hoy.\nUsa el bot para registrar ventas durante el día.")


# ─────────────────────────────────────────────
# /buscar
# ─────────────────────────────────────────────

async def comando_buscar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
            "Indica que quieres buscar.\nEjemplos:\n/buscar tornillos\n/buscar Juan\n/buscar 2025-06"
        )
        return
    termino = " ".join(context.args)
    await update.message.reply_text(f"🔍 Buscando '{termino}'...")
    resultados = await asyncio.to_thread(buscar_ventas, termino)
    if not resultados:
        await update.message.reply_text(f"No encontre ventas que coincidan con '{termino}'.")
        return
    texto = f"🔍 {len(resultados)} resultado(s) para '{termino}':\n\n"
    for r in resultados[:15]:
        num    = r.get("#", "?")
        fecha  = r.get("fecha", "?")
        prod   = r.get("producto", "?")
        total  = r.get("total", "?")
        vend   = r.get("vendedor", "?")
        hoja   = r.get("hoja", "")
        try:
            total_fmt = f"${float(total):,.0f}" if total else "?"
        except Exception:
            total_fmt = str(total)
        texto += f"#{num} [{hoja}] {fecha} — {prod} — {total_fmt} — {vend}\n"
    if len(resultados) > 15:
        texto += f"\n... y {len(resultados) - 15} mas. Usa un termino mas especifico."
    await update.message.reply_text(texto)


# ─────────────────────────────────────────────
# /borrar
# ─────────────────────────────────────────────

async def comando_borrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Indica el consecutivo a borrar.\nEjemplo: /borrar 5")
        return
    arg = context.args[0].lstrip("#")
    try:
        numero = int(arg)
    except ValueError:
        await update.message.reply_text("El numero debe ser entero.\nEjemplo: /borrar 5")
        return

    chat_id = update.message.chat_id

    # Buscar todas las filas del consecutivo - PRIORIZAR SHEETS
    filas = []
    if config.SHEETS_ID and config._get_sheets_disponible():
        from sheets import sheets_obtener_ventas_por_consecutivo
        filas = await asyncio.to_thread(sheets_obtener_ventas_por_consecutivo, numero)
    
    # Fallback a Excel local si Sheets no tiene datos
    if not filas:
        from excel import obtener_ventas_por_consecutivo
        filas = await asyncio.to_thread(obtener_ventas_por_consecutivo, numero)

    if not filas:
        await update.message.reply_text(f"No encontré el consecutivo #{numero}.")
        return

    with _estado_lock:
        borrados_pendientes[chat_id] = numero

    # Mostrar resumen de todas las filas del consecutivo
    lineas = []
    total_sum = 0
    for f in filas:
        prod  = f.get("producto", f.get(config.COL_PRODUCTO, "?"))
        total = f.get("total", f.get(config.COL_TOTAL, 0))
        try:
            total_sum += float(total)
            lineas.append(f"  • {prod} ${float(total):,.0f}")
        except Exception:
            lineas.append(f"  • {prod}")
    fecha    = filas[0].get("fecha", filas[0].get(config.COL_FECHA, "?"))
    vendedor = filas[0].get("vendedor", filas[0].get(config.COL_VENDEDOR, "?"))

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Sí, borrar todo", callback_data=f"borrar_si_{chat_id}"),
        InlineKeyboardButton("❌ Cancelar",        callback_data=f"borrar_no_{chat_id}"),
    ]])
    await update.message.reply_text(
        f"⚠️ ¿Borrar el consecutivo #{numero} completo?\n"
        f"Fecha: {fecha} | Vendedor: {vendedor}\n\n"
        + "\n".join(lineas)
        + f"\n\nTotal: ${total_sum:,.0f}",
        reply_markup=keyboard,
    )
    # ─────────────────────────────────────────────
# /precios
# ─────────────────────────────────────────────

async def comando_precios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    memoria  = cargar_memoria()
    catalogo = memoria.get("catalogo", {})
    precios  = memoria.get("precios", {})

    if not catalogo and not precios:
        await update.message.reply_text("No hay precios guardados aun.")
        return

    if catalogo:
        categorias: dict = {}
        for prod in catalogo.values():
            cat    = prod.get("categoria", "Otros")
            sufijo = " *" if prod.get("precios_fraccion") else ""
            categorias.setdefault(cat, []).append(f"  • {prod['nombre']}: ${prod['precio_unidad']:,}{sufijo}")

        await update.message.reply_text(
            f"🧠 Catalogo de precios ({len(catalogo)} productos)\n"
            f"* = tiene precios por fraccion\n\nTe envio una categoria a la vez:"
        )
        for cat, items in sorted(categorias.items()):
            encabezado = f"📂 {cat} ({len(items)} productos):\n"
            bloque = encabezado
            for item in items:
                linea = item + "\n"
                if len(bloque) + len(linea) > 4000:
                    await update.message.reply_text(bloque)
                    bloque = f"📂 {cat} (continuacion):\n"
                bloque += linea
            if bloque.strip():
                await update.message.reply_text(bloque)
    else:
        items = [f"  • {p}: ${v:,}" for p, v in sorted(precios.items())]
        await update.message.reply_text(f"🧠 Precios guardados ({len(items)} productos):")
        bloque = ""
        for item in items:
            linea = item + "\n"
            if len(bloque) + len(linea) > 4000:
                await update.message.reply_text(bloque)
                bloque = ""
            bloque += linea
        if bloque.strip():
            await update.message.reply_text(bloque)


# ─────────────────────────────────────────────
# /caja, /gastos, /inventario
# ─────────────────────────────────────────────

async def comando_caja(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args  # palabras despues de /caja
    
    if args and args[0].lower() == "abrir":
        # /caja abrir [monto]
        from memoria import cargar_caja, guardar_caja
        monto = 0
        if len(args) > 1:
            try:
                monto = int(args[1].replace(",", "").replace(".", ""))
            except ValueError:
                await update.message.reply_text("Formato: /caja abrir 50000")
                return
        caja = cargar_caja()
        if caja.get("abierta"):
            await update.message.reply_text("⚠️ La caja ya está abierta.")
            return
        import datetime
        caja["abierta"] = True
        caja["fecha"] = datetime.date.today().isoformat()
        caja["monto_apertura"] = monto
        caja["efectivo"] = 0
        caja["transferencias"] = 0
        caja["datafono"] = 0
        guardar_caja(caja)
        await update.message.reply_text(f"💰 Caja abierta con ${monto:,} de base.")

    elif args and args[0].lower() == "reset":
        # /caja reset — limpia los montos guardados de la última caja cerrada
        from memoria import cargar_caja, guardar_caja
        caja = cargar_caja()
        if caja.get("abierta"):
            await update.message.reply_text("⚠️ La caja está abierta. Ciérrala primero con /cerrar antes de resetear.")
            return
        guardar_caja({
            "abierta": False,
            "fecha": None,
            "monto_apertura": 0,
            "efectivo": 0,
            "transferencias": 0,
            "datafono": 0,
        })
        await update.message.reply_text(
            "🗑️ Caja reseteada. Los montos anteriores fueron borrados.\n"
            "El dashboard ya no mostrará valores de la última sesión.\n\n"
            "Usa /caja abrir [monto] para comenzar un nuevo día."
        )

    else:
        # /caja → ver estado
        resumen = obtener_resumen_caja()
        await update.message.reply_text(f"💰 {resumen}")


async def comando_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📊 *Dashboard Ferretería Punto Rojo*\n\n"
        "🔗 https://bot-ventas-ferreteria-production.up.railway.app/",
        parse_mode="Markdown"
    )


async def comando_gastos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    gastos = cargar_gastos_hoy()
    if not gastos:
        await update.message.reply_text("No hay gastos registrados hoy.")
        return
    texto = "💸 Gastos de hoy:\n\n"
    total = 0
    for g in gastos:
        texto += f"• {g['concepto']}: ${g['monto']:,.0f} ({g['categoria']}) — {g['origen']}\n"
        total += g["monto"]
    texto += f"\nTotal gastos: ${total:,.0f}"
    await update.message.reply_text(texto)


async def comando_inventario(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra inventario. Alias de /stock."""
    await comando_stock(update, context)


# ─────────────────────────────────────────────
# /inv - Registrar conteo de inventario
# ─────────────────────────────────────────────



def _resolver_grm(nombre_producto: str, cantidad: float, es_cajas: bool = False) -> tuple[str, float, str]:
    """
    Para productos GRM (puntillas): convierte cajas a gramos y devuelve
    (nombre_oficial, cantidad_en_gramos, label_para_mostrar).
    Si no es GRM, devuelve los valores originales.
    """
    from memoria import buscar_producto_en_catalogo as _bpc
    _PESO_CAJA_GR = 500
    prod = _bpc(nombre_producto)
    if prod and prod.get("unidad_medida", "").upper() == "GRM":
        if es_cajas:
            gr = cantidad * _PESO_CAJA_GR
            label = f"{int(cantidad)} caja{'s' if cantidad > 1 else ''} ({int(gr)} gr)"
        else:
            gr = cantidad
            cajas = gr / _PESO_CAJA_GR
            label = f"{int(gr)} gr ({cajas:.1f} caja{'s' if cajas != 1 else ''})" if gr >= _PESO_CAJA_GR else f"{int(gr)} gr"
        return prod["nombre"], gr, label
    return nombre_producto, cantidad, str(int(cantidad) if cantidad == int(cantidad) else cantidad)

async def comando_inv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Registra conteo de inventario.
    Uso: /inv [cantidad] [producto]
    Ejemplos:
        /inv 25 brocha 2"
        /inv 18 rodillo de 4 pulgadas
        /inv 50 tornillo drywall 6x1
    """
    args = context.args
    if not args or len(args) < 2:
        await update.message.reply_text(
            "📦 *Registrar inventario*\n\n"
            "Uso: `/inv [cantidad] [producto]`\n\n"
            "Ejemplos:\n"
            "• `/inv 25 brocha 2\"`\n"
            "• `/inv 18 rodillo de 4 pulgadas`\n"
            "• `/inv 50 tornillo drywall 6x1`\n"
            "• `/inv 3.5 galones vinilo t1`",
            parse_mode="Markdown"
        )
        return
    
    # Detectar si viene "N cajas puntilla X" o "N gramos puntilla X"
    import re as _re_inv
    texto_inv = " ".join(args)
    _m_cajas = _re_inv.match(r'^(\d+(?:[.,]\d+)?)\s+cajas?\s+(.+)$', texto_inv, _re_inv.IGNORECASE)
    _m_gramos = _re_inv.match(r'^(\d+(?:[.,]\d+)?)\s+gr(?:amos?)?\s+(.+)$', texto_inv, _re_inv.IGNORECASE)

    if _m_cajas:
        try:
            _n_cajas = float(_m_cajas.group(1).replace(",", "."))
        except ValueError:
            await update.message.reply_text("❌ Cantidad inválida.")
            return
        nombre_producto = _m_cajas.group(2).strip()
        nombre_producto, cantidad, _lbl = _resolver_grm(nombre_producto, _n_cajas, es_cajas=True)
    elif _m_gramos:
        try:
            _gr = float(_m_gramos.group(1).replace(",", "."))
        except ValueError:
            await update.message.reply_text("❌ Cantidad inválida.")
            return
        nombre_producto = _m_gramos.group(2).strip()
        nombre_producto, cantidad, _lbl = _resolver_grm(nombre_producto, _gr, es_cajas=False)
    else:
        # Formato normal: N producto
        try:
            cantidad = float(args[0].replace(",", "."))
        except ValueError:
            await update.message.reply_text(
                f"❌ '{args[0]}' no es una cantidad válida.\n"
                "Usa números: `/inv 25 brocha 2\"`",
                parse_mode="Markdown"
            )
            return
        nombre_producto = " ".join(args[1:])
        _lbl = None

    if len(nombre_producto) < 3:
        await update.message.reply_text("❌ Nombre del producto muy corto.")
        return

    exito, mensaje = await asyncio.to_thread(
        registrar_conteo_inventario, nombre_producto, cantidad
    )
    # Si fue en cajas, añadir aclaración
    if exito and _lbl:
        mensaje = mensaje.replace(str(cantidad), _lbl)
    await update.message.reply_text(mensaje)


# ─────────────────────────────────────────────
# /stock - Ver inventario
# ─────────────────────────────────────────────

async def comando_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Muestra inventario actual.
    Uso: /stock [producto opcional]
    """
    args = context.args
    termino = " ".join(args) if args else None
    
    productos = await asyncio.to_thread(buscar_productos_inventario, termino)
    
    if not productos:
        if termino:
            await update.message.reply_text(f"❌ No encontré '{termino}' en inventario.")
        else:
            await update.message.reply_text(
                "📦 *Inventario vacío*\n\n"
                "Usa `/inv [cantidad] [producto]` para agregar.\n"
                "Ejemplo: `/inv 25 brocha 2\"`",
                parse_mode="Markdown"
            )
        return
    
    # Construir mensaje
    if termino:
        texto = f"📦 *Inventario — '{termino}':*\n\n"
    else:
        texto = f"📦 *Inventario ({len(productos)} productos):*\n\n"
    
    alertas = 0
    for p in productos:
        cantidad = p["cantidad"]
        minimo = p["minimo"]
        unidad = p["unidad"]
        
        if cantidad <= 0:
            emoji = "🔴"
            alertas += 1
        elif cantidad <= minimo:
            emoji = "⚠️"
            alertas += 1
        else:
            emoji = "✅"
        
        # Para puntillas GRM: mostrar en cajas en lugar de gramos crudos
        from memoria import buscar_producto_en_catalogo as _bpc_s
        _prod_s = _bpc_s(p['nombre'])
        if _prod_s and _prod_s.get("unidad_medida", "").upper() == "GRM" and cantidad >= 500:
            _cajas = cantidad / 500
            _cajas_txt = f"{int(_cajas)}" if _cajas == int(_cajas) else f"{_cajas:.1f}"
            _gr_txt = int(cantidad)
            texto += f"{emoji} *{p['nombre']}*: {_cajas_txt} caja(s) ({_gr_txt} gr)\n"
        elif _prod_s and _prod_s.get("unidad_medida", "").upper() == "GRM":
            texto += f"{emoji} *{p['nombre']}*: {int(cantidad)} gr\n"
        else:
            texto += f"{emoji} *{p['nombre']}*: {cantidad} {unidad}\n"
    
    if alertas > 0:
        texto += f"\n⚠️ {alertas} producto(s) con stock bajo"
    
    # Dividir mensaje si es muy largo
    if len(texto) > 4000:
        partes = [texto[i:i+4000] for i in range(0, len(texto), 4000)]
        for parte in partes:
            await update.message.reply_text(parte, parse_mode="Markdown")
    else:
        await update.message.reply_text(texto, parse_mode="Markdown")


# ─────────────────────────────────────────────
# /ajuste - Ajustar inventario
# ─────────────────────────────────────────────

async def comando_ajuste(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Ajusta inventario sumando o restando.
    Uso: /ajuste [+/-cantidad] [producto]
    Ejemplos:
        /ajuste +10 brocha 2"
        /ajuste -5 rodillo 4"
    """
    args = context.args
    if not args or len(args) < 2:
        await update.message.reply_text(
            "🔧 *Ajustar inventario*\n\n"
            "Uso: `/ajuste [+/-cantidad] [producto]`\n\n"
            "Ejemplos:\n"
            "• `/ajuste +10 brocha 2\"` (suma 10)\n"
            "• `/ajuste -5 rodillo 4\"` (resta 5)\n"
            "• `/ajuste +2.5 galones vinilo`",
            parse_mode="Markdown"
        )
        return
    
    # Primer argumento es el ajuste
    ajuste_str = args[0]
    try:
        ajuste = float(ajuste_str.replace(",", "."))
    except ValueError:
        await update.message.reply_text(
            f"❌ '{ajuste_str}' no es válido.\n"
            "Usa +10 o -5 para ajustar.",
            parse_mode="Markdown"
        )
        return
    
    # Resto es el nombre del producto
    nombre_producto = " ".join(args[1:])
    
    # Ajustar inventario
    exito, mensaje = await asyncio.to_thread(
        ajustar_inventario, nombre_producto, ajuste
    )
    
    await update.message.reply_text(mensaje)


# ─────────────────────────────────────────────
# /compra - Registrar compra de mercancía
# ─────────────────────────────────────────────

async def comando_compra(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Registra compra de mercancía con costo.
    Uso: /compra [cantidad] [producto] a [costo] de [proveedor]
    Ejemplos:
        /compra 20 brocha 2" a 2500 de Ferrisariato
        /compra 10 rodillo 4" a 3000
        /compra 50 tornillo drywall 6x1 a 25 de JS Tools
    """
    args = context.args
    if not args or len(args) < 3:
        await update.message.reply_text(
            "📦 *Registrar compra de mercancía*\n\n"
            "Uso: `/compra [cantidad] [producto] a [costo]`\n"
            "Opcional: `de [proveedor]`\n\n"
            "Ejemplos:\n"
            "• `/compra 20 brocha 2\" a 2500`\n"
            "• `/compra 20 brocha 2\" a 2500 de Ferrisariato`\n"
            "• `/compra 50 tornillo 6x1 a 25 de JS Tools`\n\n"
            "Si no ponés proveedor, usa el último registrado para ese producto.",
            parse_mode="Markdown"
        )
        return
    
    import re
    texto_completo = " ".join(args)
    
    # Buscar " a " para separar producto de costo
    if " a " not in texto_completo.lower():
        await update.message.reply_text(
            "❌ Formato incorrecto. Usa: `/compra 20 brocha 2\" a 2500`\n"
            "El 'a' separa el producto del costo.",
            parse_mode="Markdown"
        )
        return
    
    # Separar por " a " (primera ocurrencia)
    idx_a = texto_completo.lower().find(" a ")
    parte_producto = texto_completo[:idx_a].strip()
    resto = texto_completo[idx_a + 3:].strip()  # después de " a "
    
    # Buscar " de " para separar costo de proveedor (si existe)
    proveedor = None
    if " de " in resto.lower():
        # Buscar la última ocurrencia de " de " para el proveedor
        idx_de = resto.lower().rfind(" de ")
        parte_costo = resto[:idx_de].strip()
        proveedor = resto[idx_de + 4:].strip()
    else:
        parte_costo = resto
    
    # Extraer cantidad — soporta "N cajas puntilla X" o "N producto"
    import re as _re_compra
    palabras_producto = parte_producto.split()

    _m_cajas_c = _re_compra.match(r'^(\d+(?:[.,]\d+)?)\s+cajas?\s+(.+)$',
                                    parte_producto, _re_compra.IGNORECASE)
    _m_gramos_c = _re_compra.match(r'^(\d+(?:[.,]\d+)?)\s+gr(?:amos?)?\s+(.+)$',
                                    parte_producto, _re_compra.IGNORECASE)

    if _m_cajas_c:
        try:
            _n_cajas_c = float(_m_cajas_c.group(1).replace(",", "."))
        except ValueError:
            await update.message.reply_text("❌ Cantidad inválida.", parse_mode="Markdown")
            return
        nombre_producto = _m_cajas_c.group(2).strip()
        _prod_c_check = None
        try:
            from memoria import buscar_producto_en_catalogo as _bpc_c
            _prod_c_check = _bpc_c(nombre_producto)
        except Exception:
            pass
        if _prod_c_check and _prod_c_check.get("unidad_medida", "").upper() == "GRM":
            # Convertir cajas → gramos; el costo ya viene "por caja" → dividir a por gramo
            cantidad = float(500 * _n_cajas_c)
            nombre_producto = _prod_c_check["nombre"]
            _label_compra = f"{int(_n_cajas_c)} caja(s) = {int(cantidad)} gr"
            _dividir_costo = True  # costo viene por caja, hay que dividir entre 500
        else:
            # No es GRM, tratar normal
            cantidad = _n_cajas_c
            _label_compra = None
            _dividir_costo = False
    elif _m_gramos_c:
        try:
            cantidad = float(_m_gramos_c.group(1).replace(",", "."))
        except ValueError:
            await update.message.reply_text("❌ Cantidad inválida.", parse_mode="Markdown")
            return
        nombre_producto = _m_gramos_c.group(2).strip()
        _label_compra = None
        _dividir_costo = False
    else:
        try:
            cantidad = float(palabras_producto[0].replace(",", "."))
        except ValueError:
            await update.message.reply_text(
                f"❌ '{palabras_producto[0]}' no es una cantidad válida.",
                parse_mode="Markdown"
            )
            return
        nombre_producto = " ".join(palabras_producto[1:])
        _label_compra = None
        _dividir_costo = False

    if len(nombre_producto) < 2:
        await update.message.reply_text("❌ Nombre del producto muy corto.")
        return
    
    # Extraer costo (limpiar formato)
    try:
        costo_limpio = parte_costo.replace("$", "").replace(",", "").strip()
        # Manejar números con puntos de miles (ej: 2.500)
        if "." in costo_limpio and costo_limpio.replace(".", "").isdigit():
            costo_limpio = costo_limpio.replace(".", "")
        costo = float(costo_limpio)
    except ValueError:
        await update.message.reply_text(
            f"❌ '{parte_costo}' no es un costo válido.",
            parse_mode="Markdown"
        )
        return
    
    # Registrar compra en memoria
    # Si fue "N cajas puntilla a X la caja", costo viene por caja → convertir a por gramo
    if _dividir_costo and costo >= 500:
        costo_gramo = costo / 500
    else:
        costo_gramo = costo

    exito, mensaje, datos_excel = await asyncio.to_thread(
        registrar_compra, nombre_producto, cantidad, costo_gramo, proveedor
    )
    
    # Guardar en Excel
    if exito:
        await asyncio.to_thread(
            registrar_compra_en_excel,
            datos_excel["producto"],
            datos_excel["cantidad"],
            datos_excel["costo_unitario"],
            datos_excel["costo_total"],
            datos_excel["proveedor"],
        )
    
    await update.message.reply_text(mensaje)


# ─────────────────────────────────────────────
# /margenes - Ver márgenes de ganancia
# ─────────────────────────────────────────────

async def comando_margenes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Muestra productos con su margen de ganancia.
    Solo productos que tienen costo registrado.
    Actualiza la hoja Inventario del Excel.
    """
    resultados = await asyncio.to_thread(obtener_resumen_margenes, 30)
    
    if not resultados:
        await update.message.reply_text(
            "📊 *No hay márgenes calculados*\n\n"
            "Para ver márgenes, primero registra compras:\n"
            "`/compra 20 brocha 2\" a 2500`\n\n"
            "El sistema calculará automáticamente el margen\n"
            "comparando el costo con el precio de venta.",
            parse_mode="Markdown"
        )
        return
    
    # Actualizar hoja Inventario en Excel
    await asyncio.to_thread(actualizar_hoja_inventario)
    
    # Separar por rangos de margen
    excelentes = [r for r in resultados if r["margen_porcentaje"] >= 40]
    buenos = [r for r in resultados if 25 <= r["margen_porcentaje"] < 40]
    bajos = [r for r in resultados if r["margen_porcentaje"] < 25]
    
    texto = "📊 *MÁRGENES DE GANANCIA*\n\n"
    
    if excelentes:
        texto += f"🏆 *Excelentes (≥40%):* {len(excelentes)} productos\n"
        for p in excelentes[:5]:
            texto += f"  • {p['nombre']}: {p['margen_porcentaje']}%\n"
        if len(excelentes) > 5:
            texto += f"  _...y {len(excelentes) - 5} más_\n"
        texto += "\n"
    
    if buenos:
        texto += f"✅ *Buenos (25-40%):* {len(buenos)} productos\n"
        for p in buenos[:3]:
            texto += f"  • {p['nombre']}: {p['margen_porcentaje']}%\n"
        if len(buenos) > 3:
            texto += f"  _...y {len(buenos) - 3} más_\n"
        texto += "\n"
    
    if bajos:
        texto += f"⚠️ *Bajos (<25%):* {len(bajos)} productos\n"
        for p in bajos[:5]:
            texto += f"  • {p['nombre']}: {p['margen_porcentaje']}%\n"
    
    texto += "\n📎 Detalle completo en Excel → hoja *Inventario*"
    
    await update.message.reply_text(texto, parse_mode="Markdown")


# ─────────────────────────────────────────────
# /clientes
# ─────────────────────────────────────────────
# /pendientes — productos no encontrados en catálogo
# ─────────────────────────────────────────────

async def comando_pendientes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /pendientes               → lista de hoy
    /pendientes semana        → lista de los últimos 7 días
    /pendientes todo          → lista completa
    /pendientes agregar X     → agrega X manualmente
    /pendientes quitar X      → quita X de la lista
    /pendientes limpiar       → borra toda la lista
    """
    from memoria import cargar_memoria, guardar_memoria
    from datetime import datetime, timedelta
    import re

    args = " ".join(context.args).strip() if context.args else ""
    mem  = cargar_memoria()
    lista = mem.get("productos_pendientes", [])

    # ── AGREGAR ───────────────────────────────────────────────────────────
    if args.lower().startswith("agregar "):
        nombre = args[8:].strip().lower()
        if not nombre:
            await update.message.reply_text("Uso: /pendientes agregar nombre_producto")
            return
        hoy  = datetime.now().strftime("%Y-%m-%d")
        hora = datetime.now().strftime("%H:%M")
        # No duplicar si ya existe hoy
        ya_existe = any(
            p["nombre"].lower() == nombre and p.get("fecha") == hoy
            for p in lista
        )
        if ya_existe:
            await update.message.reply_text(f"ℹ️ *{nombre}* ya está en la lista de hoy.")
            return
        lista.append({"nombre": nombre, "fecha": hoy, "hora": hora})
        mem["productos_pendientes"] = lista
        guardar_memoria(mem, urgente=True)
        await update.message.reply_text(f"✅ *{nombre}* agregado a pendientes.")
        return

    # ── QUITAR ────────────────────────────────────────────────────────────
    if args.lower().startswith("quitar "):
        nombre = args[7:].strip().lower()
        if not nombre:
            await update.message.reply_text("Uso: /pendientes quitar nombre_producto")
            return
        antes = len(lista)
        lista = [p for p in lista if p["nombre"].lower() != nombre]
        if len(lista) == antes:
            await update.message.reply_text(f"ℹ️ No encontré *{nombre}* en la lista.")
            return
        mem["productos_pendientes"] = lista
        guardar_memoria(mem, urgente=True)
        await update.message.reply_text(f"🗑️ *{nombre}* quitado de pendientes.")
        return

    # ── LIMPIAR ───────────────────────────────────────────────────────────
    if args.lower() == "limpiar":
        mem["productos_pendientes"] = []
        guardar_memoria(mem, urgente=True)
        await update.message.reply_text("🧹 Lista de pendientes limpiada.")
        return

    # ── VER LISTA ─────────────────────────────────────────────────────────
    hoy   = datetime.now().strftime("%Y-%m-%d")
    desde = None

    if args.lower() == "semana":
        desde = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        titulo = "📋 *Productos pendientes — últimos 7 días:*"
    elif args.lower() == "todo":
        titulo = "📋 *Todos los productos pendientes:*"
    else:
        desde = hoy
        titulo = "📋 *Productos pendientes de hoy:*"

    if desde:
        filtrados = [p for p in lista if p.get("fecha", "") >= desde]
    else:
        filtrados = lista

    if not filtrados:
        periodo = "hoy" if not args else args
        await update.message.reply_text(
            f"✅ No hay productos pendientes para {periodo}.\n\n"
            f"Comandos disponibles:\n"
            f"• /pendientes semana\n"
            f"• /pendientes agregar nombre\n"
            f"• /pendientes quitar nombre\n"
            f"• /pendientes limpiar"
        )
        return

    # Agrupar por fecha
    por_fecha = {}
    for p in filtrados:
        fecha = p.get("fecha", "?")
        por_fecha.setdefault(fecha, []).append(p)

    lineas = [titulo, ""]
    total_mostrados = 0
    for fecha in sorted(por_fecha.keys(), reverse=True):
        if args.lower() != "" and fecha != hoy:
            lineas.append(f"📅 *{fecha}*")
        items = por_fecha[fecha]
        # Deduplicar: solo mostrar una vez cada nombre exacto por fecha
        vistos = set()
        for p in items:
            nombre = p['nombre'].strip().lower()
            if nombre not in vistos:
                vistos.add(nombre)
                lineas.append(f"  • {p['nombre']}")
                total_mostrados += 1
        lineas.append("")

    lineas.append(f"_Total: {total_mostrados} productos_")
    lineas.append("")
    lineas.append("Para quitar: /pendientes quitar nombre")
    lineas.append("Para limpiar todo: /pendientes limpiar")

    await update.message.reply_text("\n".join(lineas), parse_mode="Markdown")


# ─────────────────────────────────────────────

async def comando_clientes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from sheets import sheets_sincronizar_clientes
    await update.message.reply_text("📋 Sincronizando clientes con Sheets...")
    ok, resultado = await asyncio.to_thread(sheets_sincronizar_clientes)
    if ok:
        clientes = await asyncio.to_thread(cargar_clientes)
        await update.message.reply_text(
            f"👥 {len(clientes)} clientes sincronizados.\n\n"
            f"📊 Ver lista completa:\n{resultado}"
        )
    else:
        await update.message.reply_text(f"⚠️ {resultado}")


# ─────────────────────────────────────────────
# /nuevo_cliente
# ─────────────────────────────────────────────

async def comando_nuevo_cliente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inicia el flujo de creación de cliente sin necesidad de una venta."""
    from ventas_state import clientes_en_proceso, _estado_lock
    from handlers.mensajes import _enviar_pregunta_flujo_cliente

    chat_id = update.effective_chat.id

    # Si viene con nombre como argumento: /nuevo_cliente Juan Pérez
    if context.args:
        nombre = " ".join(context.args).strip()
        with _estado_lock:
            clientes_en_proceso[chat_id] = {"paso": "tipo_id", "nombre": nombre}
        await _enviar_pregunta_flujo_cliente(update.message, chat_id)
    else:
        # Sin nombre → preguntar
        with _estado_lock:
            clientes_en_proceso[chat_id] = {"paso": "nombre"}
        await _enviar_pregunta_flujo_cliente(update.message, chat_id)


# ─────────────────────────────────────────────
# /fiados
# ─────────────────────────────────────────────

async def comando_fiados(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.args:
        cliente = " ".join(context.args)
        texto   = detalle_fiado_cliente(cliente)
    else:
        texto = resumen_fiados()
    await update.message.reply_text(texto, parse_mode="Markdown")


# ─────────────────────────────────────────────
# /abono
# ─────────────────────────────────────────────

async def comando_abono(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /abono [nombre] [monto]
    Ej: /abono Juan Perez 50000
    """
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "⚠️ Uso: /abono [nombre] [monto]\n"
            "Ejemplo: /abono Juan Perez 50000"
        )
        return

    # Último argumento es el monto, el resto es el nombre
    try:
        monto = float(context.args[-1].replace(",", "").replace(".", ""))
    except ValueError:
        await update.message.reply_text("❌ El monto debe ser un número. Ej: /abono Juan Perez 50000")
        return

    nombre = " ".join(context.args[:-1])

    ok, msg = await asyncio.to_thread(abonar_fiado, nombre, monto)
    if ok:
        from memoria import detalle_fiado_cliente
        detalle = detalle_fiado_cliente(nombre)
        await update.message.reply_text(
            f"✅ Abono registrado\n\n"
            f"👤 {nombre.upper()}\n"
            f"💰 Abono: ${monto:,.0f}\n\n"
            f"{detalle}",
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text(f"❌ {msg}")


# ─────────────────────────────────────────────
# /sheets
# ─────────────────────────────────────────────

async def comando_sheets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not config.SHEETS_ID:
        await update.message.reply_text("⚠️ Google Sheets no configurado.")
        return
    ventas = await asyncio.to_thread(sheets_leer_ventas_del_dia)
    estado = "✅ Conectado" if config._get_sheets_disponible() else "⚠️ Sin conexion"
    url    = f"https://docs.google.com/spreadsheets/d/{config.SHEETS_ID}/edit"
    total_dia = sum(float(v.get("total", 0) or 0) for v in ventas)
    texto = (
        f"📊 Google Sheets — {estado}\n\n"
        f"Ventas de hoy: {len(ventas)}\n"
        f"Total del dia: ${total_dia:,.0f}\n\n"
        f"🔗 {url}"
    )
    await update.message.reply_text(texto)


# ─────────────────────────────────────────────
# /grafica
# ─────────────────────────────────────────────

async def comando_grafica(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📅 Ventas por día", callback_data="grafica_dias"),
        InlineKeyboardButton("📦 Productos",      callback_data="grafica_productos"),
    ]])
    await update.message.reply_text("¿Qué gráfica quieres ver?", reply_markup=keyboard)


async def manejar_callback_grafica(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from graficas import (
        generar_grafica_ventas_por_dia_async,
        generar_grafica_productos_async,
    )
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id
    tipo    = query.data
    await query.edit_message_text("📊 Generando gráfica...")

    ruta = None
    try:
        if tipo == "grafica_dias":
            ruta   = await generar_grafica_ventas_por_dia_async()
            titulo = "ventas_por_dia.png"
        elif tipo == "grafica_productos":
            ruta   = await generar_grafica_productos_async()
            titulo = "productos_mas_vendidos.png"
        else:
            return

        if not ruta or not os.path.exists(ruta):
            await context.bot.send_message(chat_id=chat_id, text="No hay datos suficientes para esta gráfica aun.")
            return
        with open(ruta, "rb") as f:
            await context.bot.send_photo(chat_id=chat_id, photo=f, filename=titulo)
    except Exception:
        print(f"Error generando grafica: {traceback.format_exc()}")
        await context.bot.send_message(chat_id=chat_id, text="Tuve un problema generando la gráfica. Intenta de nuevo.")
    finally:
        if ruta and os.path.exists(ruta):
            os.remove(ruta)


# ─────────────────────────────────────────────
# /cerrar
# ─────────────────────────────────────────────

async def comando_cerrar_dia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    await update.message.reply_text("🔒 Iniciando cierre del dia...")

    if not config.SHEETS_ID:
        await update.message.reply_text("⚠️ Google Sheets no configurado.")
        return

    ventas_sheets = await asyncio.to_thread(sheets_leer_ventas_del_dia)
    if not ventas_sheets:
        # Sheets vacío — igual cerrar caja y subir Excel si tiene ventas
        from memoria import cargar_caja, guardar_caja
        caja = cargar_caja()
        resumen_caja = ""
        if caja.get("abierta"):
            resumen_caja = obtener_resumen_caja()
            caja["abierta"] = False
            guardar_caja(caja)
        msg = "📭 Sheets sin ventas hoy — nada que sincronizar."
        if resumen_caja:
            msg += f"\n\n💰 Caja cerrada:\n{resumen_caja}"
        await update.message.reply_text(msg)
        return

    diferencias = await asyncio.to_thread(sheets_detectar_ediciones_vs_excel)

    hoy       = datetime.now(config.COLOMBIA_TZ)
    fecha_str = hoy.strftime("%Y-%m-%d")

    # Detectar las fechas reales de las ventas en Sheets (puede ser ayer si /cerrar
    # se ejecuta después de medianoche)
    fechas_en_sheets = set()
    for v in ventas_sheets:
        f = str(v.get("fecha", ""))[:10]
        if f:
            fechas_en_sheets.add(f)
    if not fechas_en_sheets:
        fechas_en_sheets.add(fecha_str)

    try:
        await asyncio.to_thread(inicializar_excel)
        wb = await asyncio.to_thread(openpyxl.load_workbook, config.EXCEL_FILE)
        hojas_destino = [obtener_nombre_hoja(), "Registro de Ventas-Acumulado"]
        total_general = 0

        for indice_hoja, nombre_hoja in enumerate(hojas_destino):
            ws   = obtener_o_crear_hoja(wb, nombre_hoja)
            cols = detectar_columnas(ws)
            col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)

            # Borrar filas de TODAS las fechas que vienen en Sheets (evita duplicados)
            if col_fecha:
                filas_borrar = [
                    fila for fila in range(config.EXCEL_FILA_DATOS, ws.max_row + 1)
                    if str(ws.cell(row=fila, column=col_fecha).value or "")[:10] in fechas_en_sheets
                ]
                for fila in reversed(filas_borrar):
                    ws.delete_rows(fila)

            for v in ventas_sheets:
                fila_nueva = ws.max_row + 1
                datos = {
                    "fecha":                v.get("fecha", fecha_str),
                    "hora":                 v.get("hora", ""),
                    "id cliente":           v.get("id_cliente", "CF"),
                    "cliente":              v.get("cliente", "Consumidor Final"),
                    "código del producto":  v.get("codigo_producto", ""),
                    "producto":             v.get("producto", ""),
                    "cantidad":             v.get("cantidad", ""),
                    "unidad de medida":     v.get("unidad_medida", "Unidad"),
                    "unidad_medida":        v.get("unidad_medida", "Unidad"),
                    "valor unitario":       v.get("precio_unitario", 0),
                    "total":                v.get("total", 0),
                    "consecutivo de venta": v.get("num", fila_nueva - 1),
                    "vendedor":             v.get("vendedor", ""),
                    "metodo de pago":       v.get("metodo", ""),
                }
                for nombre_col, num_col in cols.items():
                    clave = nombre_col.lower().strip()
                    if clave in datos:
                        ws.cell(row=fila_nueva, column=num_col, value=datos[clave])

                if fila_nueva % 2 == 0:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=fila_nueva, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

                if indice_hoja == 0:
                    total_general += float(v.get("total", 0) or 0)

        await asyncio.to_thread(wb.save, config.EXCEL_FILE)
        await asyncio.to_thread(subir_a_drive, config.EXCEL_FILE)

        await update.message.reply_text(f"✅ Sincronizado: {len(ventas_sheets)} ventas — Total: ${total_general:,.0f}")
        with open(config.EXCEL_FILE, "rb") as f:
            await update.message.reply_document(document=f, filename="ventas.xlsx")

        # ── Guardar total del día en histórico de ventas ─────────────────
        try:
            from api import _leer_historico, _guardar_historico
            historico = _leer_historico()
            # Agrupar totales por fecha real (puede ser ayer si /cerrar es post-medianoche)
            totales_por_fecha = {}
            for v in ventas_sheets:
                f = str(v.get("fecha", fecha_str))[:10]
                totales_por_fecha[f] = totales_por_fecha.get(f, 0) + float(v.get("total", 0) or 0)
            guardados = []
            for f, t_dia in totales_por_fecha.items():
                if t_dia > 0:
                    historico[f] = int(t_dia)
                    guardados.append(f"  {f} → ${t_dia:,.0f}")
            if guardados:
                _guardar_historico(historico)
                await update.message.reply_text(
                    f"📊 Histórico actualizado:\n" + "\n".join(guardados)
                )
        except Exception as e_hist:
            print(f"⚠️ Error guardando histórico: {e_hist}")
            # No bloquear el cierre por un error en el histórico

    except Exception:
        print(traceback.format_exc())
        await update.message.reply_text("❌ Error actualizando el Excel.")
        return

    await update.message.reply_text("🧹 Limpiando Sheets...")
    from sheets import sheets_limpiar
    ok = await asyncio.to_thread(sheets_limpiar)

    # Cerrar caja automaticamente al cerrar el dia
    from memoria import cargar_caja, guardar_caja
    caja = cargar_caja()
    resumen_caja = ""
    if caja.get("abierta"):
        resumen_caja = obtener_resumen_caja()
        caja["abierta"] = False
        guardar_caja(caja)

    if ok:
        msg = "✅ Cierre completado."
        if resumen_caja:
            msg += f"\n\n💰 Caja cerrada:\n{resumen_caja}"
        await update.message.reply_text(msg)
    else:
        await update.message.reply_text("⚠️ Excel actualizado, pero Sheets no se pudo limpiar.")

    # ── Análisis del día con Claude ───────────────────────────────────────────
    try:
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")

        # Construir datos del día para Claude
        total_dia     = sum(float(v.get("total", 0) or 0) for v in ventas_sheets)
        num_ventas    = len(ventas_sheets)
        por_metodo    = {}
        por_vendedor  = {}
        por_producto  = {}
        for v in ventas_sheets:
            m_pago = str(v.get("metodo", "efectivo")).lower()
            por_metodo[m_pago]  = por_metodo.get(m_pago, 0)  + float(v.get("total", 0) or 0)
            vend = str(v.get("vendedor", "?"))
            por_vendedor[vend]  = por_vendedor.get(vend, 0)   + float(v.get("total", 0) or 0)
            prod = str(v.get("producto", "?"))
            por_producto[prod]  = por_producto.get(prod, 0)   + float(v.get("total", 0) or 0)

        # Top 3 productos
        top_prod = sorted(por_producto.items(), key=lambda x: x[1], reverse=True)[:3]
        top_txt  = ", ".join(f"{p} (${t:,.0f})" for p, t in top_prod)

        # Histórico para comparar
        try:
            from api import _leer_historico
            historico = _leer_historico()
            # Últimos 7 días (excluyendo hoy)
            from datetime import timedelta
            ultimos = sorted(historico.keys(), reverse=True)[:8]
            ultimos = [d for d in ultimos if d != fecha_str][:7]
            if ultimos:
                promedio_semana = sum(historico[d] for d in ultimos) / len(ultimos)
                mejor_dia       = max(ultimos, key=lambda d: historico[d])
                hist_txt = (
                    f"Promedio últimos {len(ultimos)} días: ${promedio_semana:,.0f}\n"
                    f"Mejor día reciente: {mejor_dia} con ${historico[mejor_dia]:,.0f}"
                )
            else:
                hist_txt = "Sin histórico previo disponible"
        except Exception:
            hist_txt = "Sin histórico previo disponible"

        prompt_analisis = (
            f"Eres el asistente de Ferretería Punto Rojo. Hoy fue {fecha_str}.\n"
            f"\nDATOS DEL DÍA:\n"
            f"- Total vendido: ${total_dia:,.0f}\n"
            f"- Número de ventas: {num_ventas}\n"
            f"- Por método: {', '.join(f'{k}: ${v:,.0f}' for k,v in por_metodo.items())}\n"
            f"- Por vendedor: {', '.join(f'{k}: ${v:,.0f}' for k,v in por_vendedor.items())}\n"
            f"- Top 3 productos: {top_txt}\n"
            f"\nHISTÓRICO RECIENTE:\n{hist_txt}\n"
            f"\nEscribe un análisis breve del día (máximo 5 líneas). "
            f"Compara con el promedio, destaca lo notable, menciona si fue buen o mal día y por qué. "
            f"Sé directo y concreto. Sin markdown, sin asteriscos. "
            f"Si fue un día excepcional o muy flojo, dilo claramente."
        )

        respuesta_analisis = await asyncio.to_thread(
            lambda: config.claude_client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=400,
                messages=[{"role": "user", "content": prompt_analisis}],
            )
        )
        analisis_txt = respuesta_analisis.content[0].text.strip()
        await update.message.reply_text(f"🧠 Análisis del día:\n\n{analisis_txt}")

    except Exception as e_an:
        # El análisis es opcional — no bloquear el cierre por esto
        print(f"[cerrar] Error en análisis Claude: {e_an}")


# ─────────────────────────────────────────────
# /resetventas
# ─────────────────────────────────────────────

async def comando_reset_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = [a.upper() for a in (context.args or [])]
    if args and args[0] == "EXCEL":
        # Formato: /resetventas excel CONFIRMAR DD/MM/YYYY
        if len(args) < 3 or args[1] != "CONFIRMAR":
            await update.message.reply_text(
                "⚠️ Uso: `/resetventas excel CONFIRMAR DD/MM/YYYY`\nEjemplo: `/resetventas excel CONFIRMAR 24/02/2026`",
                parse_mode="Markdown"
            )
            return
        # Parsear fecha
        from datetime import timedelta
        try:
            fecha_str_raw = context.args[2]  # conservar original con barras
            fecha_obj = datetime.strptime(fecha_str_raw, "%d/%m/%Y")
            fecha_iso = fecha_obj.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            await update.message.reply_text("❌ Fecha inválida. Usa el formato DD/MM/YYYY, ej: 24/02/2026")
            return
        try:
            inicializar_excel()
            wb = await asyncio.to_thread(openpyxl.load_workbook, config.EXCEL_FILE)
            hoja = obtener_nombre_hoja()

            total_borradas = 0
            hojas_limpiar = [hoja, "Registro de Ventas-Acumulado"]
            for nombre_ws in hojas_limpiar:
                if nombre_ws not in wb.sheetnames:
                    continue
                ws_actual = wb[nombre_ws]
                cols_actual = detectar_columnas(ws_actual)
                col_f = next((v for k, v in cols_actual.items() if "fecha" in k), None)
                if not col_f:
                    continue
                filas_borrar = [
                    fila for fila in range(config.EXCEL_FILA_DATOS, ws_actual.max_row + 1)
                    if str(ws_actual.cell(row=fila, column=col_f).value or "")[:10] == fecha_iso
                ]
                for fila in reversed(filas_borrar):
                    ws_actual.delete_rows(fila)
                total_borradas += len(filas_borrar)

            if total_borradas == 0:
                await update.message.reply_text(f"No hay ventas del {fecha_str_raw} en el Excel.")
                return

            await asyncio.to_thread(wb.save, config.EXCEL_FILE)
            await asyncio.to_thread(subir_a_drive, config.EXCEL_FILE)

            # Si la fecha borrada es hoy, recalcular caja y limpiar Sheets
            hoy_iso = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
            sheets_msg = ""
            if fecha_iso == hoy_iso:
                try:
                    await asyncio.to_thread(sheets_limpiar)
                    sheets_msg = " y del Sheets de hoy"
                except Exception:
                    sheets_msg = " (Sheets no pudo limpiarse)"
                # Recalcular caja desde lo que queda en el Excel
                await asyncio.to_thread(recalcular_caja_desde_excel)

            await update.message.reply_text(
                f"✅ Eliminadas {total_borradas} filas del {fecha_str_raw} "
                f"(hoja {hoja} + acumulado{sheets_msg})."
            )
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")
        return

    if not args or args[0] != "CONFIRMAR":
        await update.message.reply_text("⚠️ Escribe `/resetventas CONFIRMAR` para limpiar el dia.", parse_mode="Markdown")
        return

    # 1. Limpiar Google Sheets
    await asyncio.to_thread(sheets_limpiar)

    # 2. Limpiar TODO el estado en memoria (Standbys, ventas a medias, clientes en proceso)
    try:
        from ventas_state import (
            ventas_pendientes, borrados_pendientes, historiales,
            mensajes_standby, clientes_en_proceso, ventas_esperando_cliente,
            _estado_lock
        )
        with _estado_lock:
            ventas_pendientes.clear()
            borrados_pendientes.clear()
            historiales.clear()
            mensajes_standby.clear()
            clientes_en_proceso.clear()
            ventas_esperando_cliente.clear()
    except Exception as e:
        print(f"Error limpiando memoria interna: {e}")

    # 3. Recalcular caja desde lo que queda en el Excel
    await asyncio.to_thread(recalcular_caja_desde_excel)

    await update.message.reply_text("✅ Reset del dia completado. Todos los procesos en standby fueron cancelados.")


async def comando_actualizar_catalogo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /catalogo — Reimporta todos los productos desde BASE_DE_DATOS_PRODUCTOS.xlsx.
    Descarga desde Drive, importa, y reporta problemas detallados.
    """
    import os

    await update.message.reply_text(
        "📦 Actualizando catálogo...\n"
        "Descargando BASE_DE_DATOS_PRODUCTOS.xlsx desde Drive..."
    )

    ruta_local = "BASE_DE_DATOS_PRODUCTOS.xlsx"
    descargado = False

    try:
        from drive import descargar_de_drive
        descargado = await asyncio.to_thread(
            descargar_de_drive, "BASE_DE_DATOS_PRODUCTOS.xlsx", ruta_local
        )
    except Exception as e:
        print(f"[catalogo] Error descargando de Drive: {e}")

    if not descargado:
        await update.message.reply_text(
            "⚠️ No encontré el archivo en Drive.\n\n"
            "Envíame BASE_DE_DATOS_PRODUCTOS.xlsx directamente en este chat."
        )
        return

    # ── Análisis previo a importar ────────────────────────────────────────
    try:
        import openpyxl as _oxl
        import unicodedata, re

        def _norm(s):
            s = s.lower().strip()
            s = unicodedata.normalize("NFD", s)
            s = "".join(c for c in s if unicodedata.category(c) != "Mn")
            s = re.sub(r"[^a-z0-9\s\-_./\"]", "", s)
            return re.sub(r"\s+", " ", s).strip()

        wb_pre = _oxl.load_workbook(ruta_local, data_only=True)
        hoja   = wb_pre["Datos"] if "Datos" in wb_pre.sheetnames else wb_pre.active
        sin_precio, duplicados = [], []
        vistos = {}
        for row in hoja.iter_rows(min_row=2, values_only=True):
            nombre = str(row[1] or "").strip()
            if not nombre or nombre.lower() == "nan":
                continue
            q = row[16] if len(row) > 16 else None
            clave = _norm(nombre).replace(" ", "_")
            # Sin precio
            if q is None or not isinstance(q, (int, float)) or float(q) <= 0:
                sin_precio.append(nombre)
            # Duplicado
            if clave in vistos:
                prev_nombre, prev_precio = vistos[clave]
                prev_q = prev_precio
                curr_q = q if isinstance(q, (int, float)) else 0
                if round(float(prev_q or 0)) != round(float(curr_q or 0)):
                    duplicados.append(
                        f"{nombre}: fila anterior=${int(prev_q or 0):,} vs actual=${int(curr_q or 0):,}"
                    )
            else:
                vistos[clave] = (nombre, q)
    except Exception as e:
        sin_precio, duplicados = [], []
        print(f"[catalogo] análisis previo falló: {e}")

    # ── Importar ──────────────────────────────────────────────────────────
    try:
        resultado = await asyncio.to_thread(importar_catalogo_desde_excel, ruta_local)
    except Exception as e:
        await update.message.reply_text(f"❌ Error importando: {e}")
        if os.path.exists(ruta_local):
            os.remove(ruta_local)
        return

    importados = resultado["importados"]
    omitidos   = resultado["omitidos"]
    errores    = resultado["errores"]

    # ── Resumen principal ─────────────────────────────────────────────────
    texto = (
        f"✅ Catálogo actualizado\n\n"
        f"📦 {importados} productos importados\n"
        f"⏭️ {omitidos} filas sin nombre (ignoradas)\n"
    )
    if duplicados:
        texto += f"🔁 {len(duplicados)} duplicados con precio diferente\n"
    if errores:
        texto += f"❌ {len(errores)} errores de parseo\n"

    await update.message.reply_text(texto)

    # ── Detalle de problemas ──────────────────────────────────────────────
    if duplicados:
        lista = "\n".join(f"  • {d}" for d in duplicados[:20])
        await update.message.reply_text(
            f"🔁 Duplicados con precios distintos (se usó el último):\n{lista}"
        )

    if errores:
        lista = "\n".join(f"  • {e}" for e in errores[:10])
        await update.message.reply_text(f"❌ Errores de parseo:\n{lista}")

    if os.path.exists(ruta_local):
        os.remove(ruta_local)



async def comando_exportar_precios(update, context):
    """
    /exportar_precios — Vuelca todos los precios de memoria.json al Excel en Drive.
    Descarga el Excel una sola vez, actualiza todas las celdas respetando las
    reglas de cada categoría, y lo sube de vuelta. Útil para poner el Excel
    al día cuando la memoria tiene los precios correctos.
    """
    await update.message.reply_text(
        "📤 Exportando precios de memoria → Excel...\n"
        "Esto puede tomar unos segundos."
    )
    try:
        from precio_sync import exportar_catalogo_a_excel
        resultado = await asyncio.to_thread(exportar_catalogo_a_excel)

        actualizados = resultado["actualizados"]
        sin_match    = resultado["sin_match"]
        errores      = resultado["errores"]

        lineas = []
        lineas.append("📤 EXPORTACIÓN COMPLETADA")
        lineas.append("─" * 30)
        lineas.append(f"✅ Productos actualizados: {actualizados}")

        if sin_match:
            lineas.append(f"⚠️  No encontrados en Excel: {len(sin_match)}")
            if len(sin_match) <= 5:
                for nombre in sin_match:
                    lineas.append(f"   • {nombre}")
            else:
                for nombre in sin_match[:3]:
                    lineas.append(f"   • {nombre}")
                lineas.append(f"   …y {len(sin_match)-3} más")
            lineas.append("   (estos productos están en memoria pero no en el Excel)")

        if errores:
            lineas.append(f"❌ Errores: {len(errores)}")
            for e in errores[:3]:
                lineas.append(f"   • {e}")

        if not sin_match and not errores:
            lineas.append("")
            lineas.append("🎉 Excel actualizado correctamente.")
            lineas.append("Usa /consistencia para verificar.")

        await update.message.reply_text("\n".join(lineas))

        # Generar y enviar reporte Excel si hay productos no encontrados
        if sin_match:
            try:
                from precio_sync import generar_reporte_discrepancias
                reporte_data = {"sin_match": sin_match, "diferentes": [], "solo_memoria": [], "solo_excel": []}
                ruta = await asyncio.to_thread(generar_reporte_discrepancias, reporte_data)
                with open(ruta, "rb") as f:
                    await update.message.reply_document(
                        document=f,
                        filename="reporte_exportacion.xlsx",
                        caption="📎 Productos en memoria que no se encontraron en el Excel"
                    )
                import os; os.remove(ruta)
            except Exception as e:
                await update.message.reply_text(f"⚠️ No se pudo generar el reporte: {e}")

    except Exception as e:
        await update.message.reply_text(f"❌ Error en exportación: {e}")



async def comando_consistencia(update, context):
    """
    /consistencia — Compara precios entre memoria.json y BASE_DE_DATOS_PRODUCTOS.xlsx en Drive.
    Útil para detectar desincronizaciones.
    """
    await update.message.reply_text("🔍 Verificando consistencia entre memoria y Excel…")
    try:
        from precio_sync import verificar_consistencia
        resultado = await asyncio.to_thread(verificar_consistencia)

        if "error" in resultado:
            await update.message.reply_text(f"❌ Error: {resultado['error']}")
            return

        iguales   = resultado["iguales"]
        diferentes = resultado["diferentes"]
        solo_mem  = resultado["solo_memoria"]
        solo_xls  = resultado["solo_excel"]

        lineas = []
        lineas.append("📊 CONSISTENCIA DE PRECIOS")
        lineas.append("─" * 30)
        lineas.append(f"✅ Iguales:          {iguales}")
        lineas.append(f"⚠️  Con diferencias: {len(diferentes)}")
        lineas.append(f"🧠 Solo en memoria:  {len(solo_mem)}")
        lineas.append(f"📋 Solo en Excel:    {len(solo_xls)}")

        if diferentes:
            lineas.append("")
            lineas.append("── DIFERENCIAS DE PRECIO ──")
            for d in diferentes[:10]:
                lineas.append(f"\n📦 {d['nombre']}")
                for diff in d["diffs"]:
                    # diff viene como "precio_unidad: mem=7000 xls=8000"
                    # o "fraccion 1/2: mem=None xls=5000"
                    partes = diff.split(": ", 1)
                    if len(partes) == 2:
                        campo, valores = partes
                        # extraer mem y xls
                        mem_val = valores.split(" xls=")[0].replace("mem=", "")
                        xls_val = valores.split(" xls=")[1] if " xls=" in valores else "?"
                        if "fraccion" in campo:
                            frac = campo.replace("fraccion ", "").strip()
                            lineas.append(f"   Fracción {frac}:")
                        else:
                            lineas.append(f"   Precio unidad:")
                        lineas.append(f"     Memoria → ${mem_val}")
                        lineas.append(f"     Excel   → ${xls_val}")
                    else:
                        lineas.append(f"   {diff}")
            if len(diferentes) > 10:
                lineas.append(f"\n   …y {len(diferentes)-10} productos más con diferencias")

        if solo_mem and len(solo_mem) <= 5:
            lineas.append("")
            lineas.append("── SOLO EN MEMORIA (no están en Excel) ──")
            for nombre in solo_mem:
                lineas.append(f"  • {nombre}")
        elif solo_mem:
            lineas.append("")
            lineas.append(f"── {len(solo_mem)} productos en memoria que no están en Excel ──")
            lineas.append("  Usa /actualizar_catalogo para reimportar desde Excel.")

        if solo_xls and len(solo_xls) <= 5:
            lineas.append("")
            lineas.append("── SOLO EN EXCEL (no están en memoria) ──")
            for nombre in solo_xls[:5]:
                lineas.append(f"  • {nombre}")
        elif solo_xls:
            lineas.append("")
            lineas.append(f"── {len(solo_xls)} productos en Excel que no están en memoria ──")
            lineas.append("  Usa /actualizar_catalogo para cargarlos.")

        if not diferentes and not solo_mem and not solo_xls:
            lineas.append("")
            lineas.append("🎉 ¡Todo sincronizado correctamente!")

        await update.message.reply_text("\n".join(lineas))

        # Generar y enviar reporte Excel si hay discrepancias
        hay_discrepancias = diferentes or solo_mem or solo_xls
        if hay_discrepancias:
            try:
                from precio_sync import generar_reporte_discrepancias
                ruta = await asyncio.to_thread(generar_reporte_discrepancias, resultado)
                with open(ruta, "rb") as f_rep:
                    await update.message.reply_document(
                        document=f_rep,
                        filename="reporte_consistencia.xlsx",
                        caption="📎 Detalle completo de discrepancias"
                    )
                import os; os.remove(ruta)
            except Exception as e_rep:
                await update.message.reply_text(f"⚠️ No se pudo generar el reporte: {e_rep}")

    except Exception as e:
        await update.message.reply_text(f"❌ Error en verificación: {e}")


async def comando_keepalive(update, context):
    """
    /keepalive        → muestra estado actual
    /keepalive on     → activa keep-alive (útil en tardes movidas)
    /keepalive off    → desactiva (días de lluvia, festivos, etc.)
    """
    from keepalive import keepalive_activo, set_keepalive

    args = context.args
    if args:
        arg = args[0].lower().strip()
        if arg == "on":
            set_keepalive(True)
            await update.message.reply_text(
                "✅ Keep-alive ACTIVADO\n"
                "El cache se renovará cada 4 min mientras esté activo.\n"
                "Úsalo en tardes movidas o días con muchas ventas seguidas."
            )
        elif arg == "off":
            set_keepalive(False)
            await update.message.reply_text(
                "⏸ Keep-alive DESACTIVADO\n"
                "Se reactiva automáticamente mañana de 8am-11am si lo dejas en ON."
            )
        else:
            await update.message.reply_text("Uso: /keepalive on | /keepalive off")
        return

    # Sin argumentos → mostrar estado
    from datetime import datetime, time
    import config
    activo  = keepalive_activo()
    ahora   = datetime.now(config.COLOMBIA_TZ).time()
    horario = time(8, 0) <= ahora <= time(11, 0)

    if activo and horario:
        estado_emoji = "🟢"
        estado_texto = "ACTIVO y en horario (ping cada 4 min)"
    elif activo and not horario:
        estado_emoji = "🟡"
        estado_texto = "ACTIVADO pero fuera de horario 8-11am"
    else:
        estado_emoji = "🔴"
        estado_texto = "DESACTIVADO manualmente"

    await update.message.reply_text(
        f"{estado_emoji} Keep-alive: {estado_texto}\n\n"
        f"Horario automático: 8:00am - 11:00am\n"
        f"Apagado automático (si ON manual): L-S 5:00pm | D 1:00pm\n"
        f"Intervalo: cada 4 minutos\n\n"
        f"/keepalive on  → activar\n"
        f"/keepalive off → desactivar"
    )


# ─────────────────────────────────────────────
# AGREGAR PRODUCTO AL CATÁLOGO
# ─────────────────────────────────────────────

CATEGORIAS_DISPONIBLES = {
    "1": "1 Artículos de Ferreteria",
    "2": "2 Pinturas y Disolventes",
    "3": "3 Tornilleria",
    "4": "4 Impermeabilizantes y Materiales de construcción",
    "5": "5 Materiales Electricos",
}

# Nombres limpios para mostrar al usuario (sin el prefijo numérico)
CATEGORIAS_DISPLAY = {
    "1": "Artículos de Ferreteria",
    "2": "Pinturas y Disolventes",
    "3": "Tornilleria",
    "4": "Impermeabilizantes y Materiales de construcción",
    "5": "Materiales Eléctricos",
}

# Orden de pasos para poder retroceder
_PASOS_ORDEN = ["nombre", "categoria", "precio"]

CATEGORIAS_CON_FRACCIONES = {"2 pinturas y disolventes"}
CATEGORIAS_TORNILLERIA    = {"3 tornilleria"}


async def comando_agregar_producto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /agregar_producto — Inicia flujo para agregar un producto nuevo al catálogo y al Excel.
    """
    context.user_data["nuevo_producto"] = {}
    context.user_data["paso_producto"]  = "nombre"

    await update.message.reply_text(
        "➕ Agregar producto nuevo\n\n"
        "Escribe el nombre del producto:\n\n"
        "_(Escribe 'cancelar' en cualquier momento para salir)_",
        parse_mode="Markdown"
    )


def _texto_categoria_prompt(nombre_prod: str) -> str:
    cats = "\n".join(f"  {k}. {v}" for k, v in CATEGORIAS_DISPLAY.items())
    return (
        f"Producto: *{nombre_prod}*\n\n"
        f"Elige la categoría:\n{cats}\n\n"
        f"Responde con el número (1-5):\n\n"
        f"_(Escribe 'volver' para cambiar el nombre o 'cancelar' para salir)_"
    )


async def manejar_flujo_agregar_producto(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """
    Maneja los pasos del flujo de agregar producto.
    Retorna True si el mensaje fue consumido por este flujo, False si no.
    """
    paso = context.user_data.get("paso_producto")
    if not paso:
        return False

    texto = update.message.text.strip()

    # Cancelar en cualquier momento
    if texto.lower() in {"cancelar", "/cancelar"}:
        context.user_data.pop("nuevo_producto", None)
        context.user_data.pop("paso_producto", None)
        await update.message.reply_text("❌ Registro cancelado.")
        return True

    prod = context.user_data.get("nuevo_producto", {})

    # ── Volver al paso anterior ──
    if texto.lower() in {"volver", "atras", "atrás"}:
        if paso == "categoria":
            context.user_data["paso_producto"] = "codigo"
            await update.message.reply_text(
                f"↩️ Volvemos al código.\n\n"
                f"¿Cuál es el código del producto?\n_(Escribe 'omitir' si no tiene código)_",
                parse_mode="Markdown"
            )
            return True
        elif paso == "precio":
            context.user_data["paso_producto"] = "categoria"
            await update.message.reply_text(
                _texto_categoria_prompt(prod.get("nombre", "")),
                parse_mode="Markdown"
            )
            return True
        elif paso in {"fracciones_3_4", "mayorista"}:
            context.user_data["paso_producto"] = "precio"
            await update.message.reply_text(
                f"↩️ Volvemos al precio.\n\n"
                f"¿Cuál es el precio de la unidad completa?\n(solo el número, ej: 50000)\n\n"
                f"_(Escribe 'volver' para cambiar la categoría)_",
                parse_mode="Markdown"
            )
            return True
        elif paso.startswith("fracciones_"):
            # Dentro de fracciones, volver a la fracción anterior
            orden_fracs = ["fracciones_3_4", "fracciones_1_2", "fracciones_1_4", "fracciones_1_8", "fracciones_1_16"]
            idx = orden_fracs.index(paso) if paso in orden_fracs else -1
            if idx > 0:
                paso_ant = orden_fracs[idx - 1]
                context.user_data["paso_producto"] = paso_ant
                fracs_labels = {"fracciones_3_4": "3/4", "fracciones_1_2": "1/2",
                                "fracciones_1_4": "1/4", "fracciones_1_8": "1/8", "fracciones_1_16": "1/16"}
                frac_ant = fracs_labels.get(paso_ant, "")
                # Limpiar la fracción anterior para re-ingresarla
                prod.get("fracciones", {}).pop(frac_ant, None)
                context.user_data["nuevo_producto"] = prod
                await update.message.reply_text(
                    f"↩️ Volvemos a la fracción {frac_ant}.\n\n"
                    f"¿Precio unitario para vender {frac_ant}?\n(Escribe 0 si no aplica)"
                )
            else:
                context.user_data["paso_producto"] = "precio"
                await update.message.reply_text(
                    f"↩️ Volvemos al precio.\n\n"
                    f"¿Cuál es el precio de la unidad completa?\n(solo el número, ej: 50000)"
                )
            return True
        elif paso == "confirmar":
            # Volver al último paso de datos
            cat_lower = prod.get("categoria", "").lower()
            if cat_lower in CATEGORIAS_CON_FRACCIONES:
                context.user_data["paso_producto"] = "fracciones_1_16"
                await update.message.reply_text(
                    "↩️ Volvemos a la última fracción.\n\n"
                    "¿Precio unitario para vender 1/16?\n(Escribe 0 si no aplica)"
                )
            elif cat_lower in CATEGORIAS_TORNILLERIA:
                context.user_data["paso_producto"] = "mayorista"
                await update.message.reply_text(
                    "↩️ Volvemos al precio mayorista.\n\n"
                    "¿Cuál es el precio unitario para compras de 50 o más unidades?\n(Escribe 0 si no aplica)"
                )
            else:
                context.user_data["paso_producto"] = "precio"
                await update.message.reply_text(
                    "↩️ Volvemos al precio.\n\n"
                    "¿Cuál es el precio de la unidad completa?\n(solo el número, ej: 50000)"
                )
            return True
        else:
            await update.message.reply_text("Ya estás en el primer paso. Escribe 'cancelar' para salir.")
            return True

    # ── Paso 1: nombre ──
    if paso == "nombre":
        if len(texto) < 2:
            await update.message.reply_text("Nombre muy corto. Escribe el nombre del producto:")
            return True
        prod["nombre"] = texto
        context.user_data["nuevo_producto"] = prod
        context.user_data["paso_producto"]  = "codigo"

        await update.message.reply_text(
            f"Nombre: *{texto}*\n\n"
            f"¿Cuál es el código del producto?\n"
            f"(ej: `1cepilloacero`, `2viniloT1blanco`)\n\n"
            f"_Escribe 'omitir' si no tiene código_",
            parse_mode="Markdown"
        )
        return True

    # ── Paso 1b: código ──
    if paso == "codigo":
        if texto.lower() in {"omitir", "no", "ninguno", "-"}:
            prod["codigo"] = ""
        else:
            prod["codigo"] = texto.strip()
        context.user_data["nuevo_producto"] = prod
        context.user_data["paso_producto"]  = "categoria"

        await update.message.reply_text(
            _texto_categoria_prompt(prod["nombre"]),
            parse_mode="Markdown"
        )
        return True

    # ── Paso 2: categoría ──
    if paso == "categoria":
        if texto not in CATEGORIAS_DISPONIBLES:
            await update.message.reply_text("Responde con un número del 1 al 5:")
            return True
        categoria = CATEGORIAS_DISPONIBLES[texto]
        prod["categoria"] = categoria
        context.user_data["nuevo_producto"] = prod
        context.user_data["paso_producto"]  = "precio"
        nombre_display = CATEGORIAS_DISPLAY[texto]

        await update.message.reply_text(
            f"Categoría: *{nombre_display}*\n\n"
            f"¿Cuál es el precio de la unidad completa?\n"
            f"(solo el número, ej: 50000)\n\n"
            f"_(Escribe 'volver' para cambiar la categoría)_",
            parse_mode="Markdown"
        )
        return True

    # ── Paso 3: precio base ──
    if paso == "precio":
        try:
            precio = float(texto.replace(",", "").replace(".", "").replace("$", ""))
            if precio <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text("Precio inválido. Escribe solo el número (ej: 50000):")
            return True

        prod["precio_unidad"] = precio
        context.user_data["nuevo_producto"] = prod
        cat_lower = prod["categoria"].lower()

        # Según categoría, decidir siguiente paso
        if cat_lower in CATEGORIAS_CON_FRACCIONES:
            context.user_data["paso_producto"] = "fracciones_3_4"
            await update.message.reply_text(
                f"Precio base: ${precio:,.0f}\n\n"
                f"Es de Pinturas/Disolventes — necesito los precios por fracción.\n"
                f"(Escribe 0 si no aplica esa fracción)\n\n"
                f"¿Precio unitario para vender 3/4?\n"
                f"_(Escribe 'volver' para corregir el precio base)_",
                parse_mode="Markdown"
            )
        elif cat_lower in CATEGORIAS_TORNILLERIA:
            context.user_data["paso_producto"] = "mayorista"
            await update.message.reply_text(
                f"Precio base: ${precio:,.0f}\n\n"
                f"Es Tornillería — ¿cuál es el precio unitario para compras de 50 o más unidades?\n"
                f"(Escribe 0 si no aplica precio mayorista)\n\n"
                f"_(Escribe 'volver' para corregir el precio base)_",
                parse_mode="Markdown"
            )
        else:
            context.user_data["paso_producto"] = "confirmar"
            await _mostrar_confirmacion(update, prod)
        return True

    # ── Paso 4a: fracciones (pinturas) ──
    for frac_paso, frac_key, frac_mult, siguiente_paso, siguiente_texto in [
        ("fracciones_3_4", "3/4",  0.75,   "fracciones_1_2", "¿Precio unitario para vender 1/2?\n(× 0.5 = total)"),
        ("fracciones_1_2", "1/2",  0.5,    "fracciones_1_4", "¿Precio unitario para vender 1/4?\n(× 0.25 = total)"),
        ("fracciones_1_4", "1/4",  0.25,   "fracciones_1_8", "¿Precio unitario para vender 1/8?\n(× 0.125 = total)"),
        ("fracciones_1_8", "1/8",  0.125,  "fracciones_1_16","¿Precio unitario para vender 1/16?\n(× 0.0625 = total)"),
        ("fracciones_1_16","1/16", 0.0625, "confirmar",       None),
    ]:
        if paso == frac_paso:
            try:
                val = float(texto.replace(",", "").replace(".", "").replace("$", ""))
            except ValueError:
                await update.message.reply_text("Escribe solo el número (ej: 52000 o 0):")
                return True

            if val > 0:
                prod.setdefault("fracciones", {})[frac_key] = val
            context.user_data["nuevo_producto"] = prod
            context.user_data["paso_producto"]  = siguiente_paso

            if siguiente_paso == "confirmar":
                await _mostrar_confirmacion(update, prod)
            else:
                total_ejemplo = round(val * frac_mult) if val > 0 else 0
                ejemplo = f"(ej: ${val:,.0f} × {frac_mult} = ${total_ejemplo:,.0f})\n\n" if val > 0 else ""
                await update.message.reply_text(f"{ejemplo}{siguiente_texto}")
            return True

    # ── Paso 4b: mayorista (tornillería) ──
    if paso == "mayorista":
        try:
            val = float(texto.replace(",", "").replace(".", "").replace("$", ""))
        except ValueError:
            await update.message.reply_text("Escribe solo el número (ej: 150 o 0):")
            return True

        if val > 0:
            prod["precio_mayorista"] = val
        context.user_data["nuevo_producto"] = prod
        context.user_data["paso_producto"]  = "confirmar"
        await _mostrar_confirmacion(update, prod)
        return True

    # ── Paso final: confirmación ──
    if paso == "confirmar":
        if texto.lower() in {"si", "sí", "s", "yes"}:
            await _guardar_producto(update, context, prod)
        else:
            context.user_data.pop("nuevo_producto", None)
            context.user_data.pop("paso_producto", None)
            await update.message.reply_text("❌ Cancelado. El producto no fue guardado.")
        return True

    return False


async def _mostrar_confirmacion(update, prod: dict):
    """Muestra resumen del producto antes de guardar."""
    lineas = [
        f"📦 Confirmar producto nuevo:\n",
        f"Nombre:    {prod['nombre']}",
        f"Categoría: {prod['categoria']}",
        f"Precio:    ${prod['precio_unidad']:,.0f}",
    ]
    fracs = prod.get("fracciones", {})
    if fracs:
        lineas.append("Fracciones:")
        for frac, p_unit in fracs.items():
            mult = {"3/4":0.75,"1/2":0.5,"1/4":0.25,"1/8":0.125,"1/16":0.0625}[frac]
            total = round(p_unit * mult)
            lineas.append(f"  {frac}: ${total:,.0f}  (unitario: ${p_unit:,.0f})")
    if prod.get("precio_mayorista"):
        lineas.append(f"Mayorista: ${prod['precio_mayorista']:,.0f} (x50+)")

    lineas.append("\n¿Confirmas? (si / no)\n_(Escribe 'volver' para corregir el último dato)_")
    await update.message.reply_text("\n".join(lineas), parse_mode="Markdown")


async def _guardar_producto(update, context, prod: dict):
    """Guarda el producto en memoria.json y en BASE_DE_DATOS_PRODUCTOS.xlsx."""
    from utils import _normalizar
    from memoria import (cargar_memoria, guardar_memoria, invalidar_cache_memoria,
                         _es_producto_con_fracciones, _es_tornillo_drywall)

    mem      = cargar_memoria()
    catalogo = mem.get("catalogo", {})

    nombre       = prod["nombre"]
    categoria    = prod["categoria"]
    precio_unidad = prod["precio_unidad"]
    fracs_input  = prod.get("fracciones", {})
    p_mayorista  = prod.get("precio_mayorista")

    nombre_lower = _normalizar(nombre)
    clave        = nombre_lower.replace(" ", "_")

    # Construir entrada del catálogo
    entrada = {
        "nombre":       nombre,
        "nombre_lower": nombre_lower,
        "categoria":    categoria,
        "precio_unidad": round(precio_unidad),
    }
    codigo_prod = prod.get("codigo", "").strip()
    if codigo_prod:
        entrada["codigo"] = codigo_prod

    if _es_tornillo_drywall(nombre) and p_mayorista:
        entrada["precio_por_cantidad"] = {
            "umbral":              50,
            "precio_bajo_umbral":  round(precio_unidad),
            "precio_sobre_umbral": round(p_mayorista),
        }
    elif fracs_input:
        mult_map = {"3/4":0.75,"1/2":0.5,"1/4":0.25,"1/8":0.125,"1/16":0.0625}
        fracs_cat = {}
        for frac, p_unit in fracs_input.items():
            mult = mult_map[frac]
            fracs_cat[frac] = {"precio": round(p_unit * mult)}
        entrada["precios_fraccion"] = fracs_cat

    catalogo[clave] = entrada
    mem["catalogo"] = catalogo
    guardar_memoria(mem, urgente=True)   # urgente=True: sube a Drive sin debounce
    invalidar_cache_memoria()

    # Agregar al Excel BASE_DE_DATOS_PRODUCTOS.xlsx
    excel_ok  = False
    excel_msg = ""
    try:
        import openpyxl
        from drive import descargar_de_drive, subir_a_drive
        ruta = "BASE_DE_DATOS_PRODUCTOS_temp.xlsx"

        descargado = await asyncio.to_thread(descargar_de_drive, "BASE_DE_DATOS_PRODUCTOS.xlsx", ruta)
        if descargado:
            wb = openpyxl.load_workbook(ruta)
            ws = wb["Datos"]

            # Construir fila con el mismo formato de columnas del importador
            fila = [""] * 22  # columnas A-V
            codigo_prod = prod.get("codigo", "").strip()
            fila[0]  = codigo_prod if codigo_prod else nombre.lower().replace(" ", "")[:20]  # col A — código
            fila[1]  = nombre                 # col B — nombre
            fila[2]  = "P-Producto"           # col C — tipo
            fila[3]  = categoria              # col D — categoría
            fila[4]  = "SI"                   # col E — inventariable
            fila[5]  = "SI"                   # col F — visible facturas
            fila[6]  = 0                      # col G — stock mínimo
            fila[7]  = "94"                   # col H — código DIAN
            fila[13] = "22-IVA 0%"            # col N — código impuesto
            fila[15] = "SI"                   # col P — incluye IVA
            fila[16] = round(precio_unidad)   # col Q — precio unidad

            mult_map = {"3/4":17,"1/2":18,"1/4":19,"1/8":20,"1/16":21}
            for frac, col_idx in mult_map.items():
                if frac in fracs_input:
                    fila[col_idx] = round(fracs_input[frac])

            if p_mayorista:
                fila[17] = round(p_mayorista)  # col R — precio mayorista (3/4 slot)

            # Buscar primera fila vacía en col B (igual que dashboard)
            fila_nueva_prod = ws.max_row + 1
            for r in range(2, ws.max_row + 2):
                if not ws.cell(row=r, column=2).value:
                    fila_nueva_prod = r
                    break
            for col_idx, valor in enumerate(fila, 0):
                if valor != "":
                    ws.cell(row=fila_nueva_prod, column=col_idx + 1, value=valor)
            wb.save(ruta)

            # Renombrar archivo temporal al nombre original antes de subir
            import shutil
            ruta_final = "BASE_DE_DATOS_PRODUCTOS.xlsx"
            shutil.copy(ruta, ruta_final)
            subido = await asyncio.to_thread(subir_a_drive, ruta_final)
            excel_ok  = subido
            excel_msg = "✅ También agregado al Excel en Drive." if subido else "⚠️ No se pudo subir el Excel a Drive."

            import os
            if os.path.exists(ruta):
                os.remove(ruta)
        else:
            excel_msg = "⚠️ No se encontró BASE_DE_DATOS_PRODUCTOS.xlsx en Drive. Guardado solo en catálogo."
    except Exception as e:
        excel_msg = f"⚠️ Error actualizando Excel: {e}"

    context.user_data.pop("nuevo_producto", None)
    context.user_data.pop("paso_producto",  None)

    await update.message.reply_text(
        f"✅ Producto guardado en el catálogo.\n"
        f"{excel_msg}\n\n"
        f"Ya puedes registrar ventas de '{nombre}'."
    )


# ─────────────────────────────────────────────
# /actualizar_precio
# ─────────────────────────────────────────────

async def comando_actualizar_precio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /actualizar_precio — Entra en modo actualización de precios.
    El usuario envía líneas "producto= precio" y se actualizan.
    Escribe "listo" o "salir" para salir del modo.
    """
    from ventas_state import actualizando_precios, _estado_lock

    chat_id = update.effective_chat.id

    # Si viene con argumentos directos: /actualizar_precio tornillo 6x1= 50
    args_text = " ".join(context.args) if context.args else ""
    if args_text and "=" in args_text:
        # Procesar directamente sin entrar en modo
        resultado = await _procesar_linea_precio(args_text, update)
        return

    with _estado_lock:
        actualizando_precios[chat_id] = True

    await update.message.reply_text(
        "💰 Modo actualización de precios\n\n"
        "Envía los precios así:\n"
        "  producto= precio\n"
        "  producto= precio / precio_mayorista\n"
        "  producto 1/4= precio_fraccion\n\n"
        "Ejemplo:\n"
        "  Tornillo drywall 6x1= 42/30\n"
        "  Vinilo T1 Blanco= 50000\n"
        "  Thinner 1/4= 8000\n\n"
        "Escribe 'listo' para salir."
    )


async def _procesar_linea_precio(linea: str, update):
    """Procesa una línea de actualización de precio."""
    import re as _re
    from precio_sync import actualizar_precio as _ap
    from memoria import buscar_producto_en_catalogo, invalidar_cache_memoria, cargar_memoria, guardar_memoria

    linea = linea.strip()
    if not linea:
        return

    _FRACCIONES = {"1/16", "1/8", "1/4", "1/3", "3/8", "1/2", "3/4"}

    def _parse_precio(s):
        return float(s.replace(".", "").replace(",", ""))

    # Patrón con dos precios: nombre= p1 / p2
    PAT_DOS = _re.compile(r"^(.+?)\s*(?:=|:|→|->)\s*\$?\s*([\d][\d.,]*)\s*/\s*\$?\s*([\d][\d.,]*)$")
    # Patrón un precio: nombre= precio
    PAT_UNO = _re.compile(r"^(.+?)\s*(?:=|:|→|->)\s*\$?\s*([\d][\d.,]*)$")

    precio_mayorista = None

    m = PAT_DOS.match(linea)
    if m:
        nombre_raw = m.group(1).strip()
        try:
            precio = _parse_precio(m.group(2))
            precio_mayorista = _parse_precio(m.group(3))
        except ValueError:
            await update.message.reply_text(f"❌ No entendí el precio: {linea}")
            return
    else:
        m = PAT_UNO.match(linea)
        if not m:
            await update.message.reply_text(f"❌ Formato: producto= precio. Ej: Vinilo T1= 50000")
            return
        nombre_raw = m.group(1).strip()
        try:
            precio = _parse_precio(m.group(2))
        except ValueError:
            await update.message.reply_text(f"❌ No entendí el precio: {linea}")
            return

    if precio <= 0:
        await update.message.reply_text("❌ El precio debe ser mayor a 0.")
        return

    # Detectar fracción al final del nombre
    fraccion = None
    nombre_lower = nombre_raw.lower()
    for frac in _FRACCIONES:
        if nombre_lower.endswith(" " + frac):
            fraccion = frac
            nombre_raw = nombre_raw[:-(len(frac)+1)].strip()
            break

    # Buscar producto en catálogo
    prod = buscar_producto_en_catalogo(nombre_raw)
    if not prod:
        await update.message.reply_text(f"⚠️ No encontré '{nombre_raw}' en el catálogo.")
        return

    nombre_display = prod["nombre"]

    # Si hay precio mayorista → actualizar precio_por_cantidad (tornillos)
    if precio_mayorista is not None:
        mem = cargar_memoria()
        cat = mem.get("catalogo", {})
        clave = next((k for k, v in cat.items()
                      if v.get("nombre_lower") == prod.get("nombre_lower", "")), None)
        if clave:
            cat[clave]["precio_unidad"] = round(precio)
            pxc = cat[clave].get("precio_por_cantidad", {})
            pxc["precio_bajo_umbral"] = round(precio)
            pxc["precio_sobre_umbral"] = round(precio_mayorista)
            if "umbral" not in pxc:
                pxc["umbral"] = 50
            cat[clave]["precio_por_cantidad"] = pxc
            mem["catalogo"] = cat
            guardar_memoria(mem)
            invalidar_cache_memoria()

            # Sync al Excel
            try:
                _ap(nombre_display, round(precio), fraccion)
            except Exception:
                pass

            await update.message.reply_text(
                f"✅ {nombre_display}\n"
                f"   Unitario: ${round(precio):,} → Mayorista: ${round(precio_mayorista):,}"
            )
        else:
            await update.message.reply_text(f"⚠️ No encontré '{nombre_raw}' en el catálogo.")
        return

    # Precio simple o fracción
    ok, desc = _ap(nombre_display, round(precio), fraccion)
    if ok:
        frac_txt = f" ({fraccion})" if fraccion else ""
        await update.message.reply_text(f"✅ {nombre_display}{frac_txt} → ${round(precio):,}")
    else:
        await update.message.reply_text(f"⚠️ {desc}")


async def manejar_mensaje_precio(update, mensaje: str) -> bool:
    """
    Si el chat está en modo actualización de precios, procesa el mensaje.
    Retorna True si fue procesado, False si no estaba en modo precios.
    """
    from ventas_state import actualizando_precios, _estado_lock

    chat_id = update.effective_chat.id

    with _estado_lock:
        if not actualizando_precios.get(chat_id):
            return False

    msg = mensaje.strip().lower()

    # Salir del modo
    if msg in ("listo", "salir", "exit", "ok", "ya", "fin"):
        with _estado_lock:
            actualizando_precios.pop(chat_id, None)
        await update.message.reply_text("✅ Modo actualización de precios finalizado.")
        return True

    # Procesar como línea(s) de precio
    lineas = [l.strip() for l in mensaje.strip().split("\n") if l.strip()]
    # También separar por comas si hay múltiples
    if len(lineas) == 1 and "," in lineas[0]:
        lineas = [l.strip() for l in lineas[0].split(",") if l.strip()]

    for linea in lineas:
        await _procesar_linea_precio(linea, update)

    return True
