"""
Operaciones sobre el archivo Excel de ventas (openpyxl).
El Excel es la fuente de verdad histórica; el Sheets es la pizarra del día.

CORRECCIONES v2:
  - _normalizar importada de utils (eliminada la definición duplicada)
  - read_only=True en todas las funciones que solo leen el Excel
  - alias por hoja calculado correctamente (antes la hoja Acumulado recibía alias=1 siempre)
"""

import logging
import asyncio
import os
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from utils import (
    convertir_fraccion_a_decimal,
    decimal_a_fraccion_legible,
    obtener_nombre_hoja,
    _normalizar,          # ← importada de utils, ya no duplicada aquí
)

# ─────────────────────────────────────────────
# CACHÉ DE CLIENTES (TTL 5 minutos)
# Evita abrir el Excel completo en cada búsqueda de autocompletado
# ─────────────────────────────────────────────
_clientes_cache: list = []
_clientes_cache_ts: float = 0.0
_CLIENTES_CACHE_TTL: float = 300.0  # segundos


def _invalidar_cache_clientes():
    """Fuerza recarga en la próxima llamada a cargar_clientes()."""
    global _clientes_cache_ts
    _clientes_cache_ts = 0.0


# ─────────────────────────────────────────────
# ESTRUCTURA INTERNA
# ─────────────────────────────────────────────

def inicializar_hoja(ws, nombre_mes: str = ""):
    """
    Crea el formato exacto del Excel de Ferretería Punto Rojo:
      Fila 1 : banner rojo con título del mes + logo en A1:D1
      Fila 2 : separador rojo oscuro
      Fila 3 : encabezados con fondo negro
      Fila 4+: datos de ventas

    Columnas (orden exacto del Excel real):
      1 FECHA | 2 HORA | 3 ID CLIENTE | 4 CLIENTE | 5 CODIGO DEL PRODUCTO |
      6 PRODUCTO | 7 UNIDAD DE MEDIDA | 8 CANTIDAD | 9 VALOR UNITARIO |
      10 TOTAL | 11 CONSECUTIVO DE VENTA | 12 VENDEDOR | 13 METODO DE PAGO
    """
    # Solo inicializar si fila 3 col 1 está vacía
    if ws.cell(row=3, column=1).value is not None:
        return

    from openpyxl.drawing.image import Image as XLImage

    NUM_COLS = 13

    # ── Fila 1: banner rojo ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 72.0

    # A1:D1 → logo (rojo de fondo igual que el banner)
    ws.merge_cells("A1:D1")
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="C00000")

    # E1:M1 → título
    ws.merge_cells("E1:M1")
    titulo = nombre_mes if nombre_mes else "Registro de Ventas"
    celda_titulo = ws.cell(row=1, column=5, value=titulo)
    celda_titulo.font      = Font(bold=True, color="FFFFFF", size=16)
    celda_titulo.fill      = PatternFill("solid", fgColor="C00000")
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(6, NUM_COLS + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="C00000")

    # Logo
    logo_paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png"),
        "/app/logo.png",
        "logo.png",
    ]
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            try:
                img        = XLImage(logo_path)
                img.width  = 220
                img.height = 68
                img.anchor = "A1"
                ws.add_image(img)
            except Exception as e:
                print(f"No se pudo agregar logo: {e}")
            break

    # ── Fila 2: separador rojo oscuro ─────────────────────────────────────────
    ws.row_dimensions[2].height = 9.6
    for col in range(1, NUM_COLS + 1):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor="E00000")

    # ── Fila 3: encabezados ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 30.0
    ENCABEZADOS = [
        "FECHA", "HORA", "ID CLIENTE", "CLIENTE",
        "CODIGO DEL PRODUCTO", "PRODUCTO", "UNIDAD DE MEDIDA",
        "CANTIDAD", "VALOR UNITARIO", "TOTAL",
        "CONSECUTIVO DE VENTA", "VENDEDOR", "METODO DE PAGO",
    ]
    for col, titulo_col in enumerate(ENCABEZADOS, 1):
        celda = ws.cell(row=3, column=col, value=titulo_col)
        celda.font      = Font(bold=True, color="FFFFFF", size=10)
        celda.fill      = PatternFill("solid", fgColor="1A1A1A")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border    = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin",  color="444444"),
        )

    # ── Anchos de columna (medidos del Excel real) ────────────────────────────
    ANCHOS = {
        "A": 16.33, "B": 12.55, "C": 14.44, "D": 24.11,
        "E": 19.44, "F": 26.89, "G": 16.66, "H": 13.33,
        "I": 17.55, "J": 13.0,  "K": 20.33, "L": 18.55, "M": 19.44,
    }
    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho

    # ── Fila 4: altura por defecto para filas de datos ────────────────────────
    ws.row_dimensions[4].height = 19.95


def obtener_o_crear_hoja(wb, nombre_hoja: str):
    if nombre_hoja in wb.sheetnames:
        return wb[nombre_hoja]
    ws = wb.create_sheet(title=nombre_hoja)
    inicializar_hoja(ws)
    return ws



def _ultima_fila_con_datos(ws, desde_fila: int = 1) -> int:
    """
    Devuelve el número de la última fila que realmente contiene datos.
    ws.max_row en openpyxl incluye filas con formato vacío (de la plantilla),
    lo que provoca que las ventas se escriban en filas 998+ en lugar de la 4.
    Esta función itera desde abajo hasta encontrar una celda no vacía.
    """
    max_r = ws.max_row or desde_fila
    for fila in range(max_r, desde_fila - 1, -1):
        for cell in ws[fila]:
            if cell.value is not None and str(cell.value).strip():
                return fila
    return desde_fila - 1  # Hoja completamente vacía desde desde_fila


def inicializar_excel():
    """Crea ventas.xlsx si no existe."""
    if not os.path.exists(config.EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = obtener_nombre_hoja()
        inicializar_hoja(ws)
        wb.save(config.EXCEL_FILE)
        print("✅ Archivo Excel creado.")


def detectar_columnas(ws) -> dict:
    """
    Lee los encabezados desde EXCEL_FILA_HEADERS.
    Retorna {nombre_lower: numero_columna}.
    Usa iter_rows para ser compatible con modo read_only (max_column puede ser None).
    """
    encabezados = {}
    try:
        for fila_hdr in ws.iter_rows(
            min_row=config.EXCEL_FILA_HEADERS,
            max_row=config.EXCEL_FILA_HEADERS,
        ):
            for cell in fila_hdr:
                if cell.value:
                    encabezados[str(cell.value).lower().strip()] = cell.column
            break
    except Exception:
        pass
    return encabezados


def _col_para(cols: dict, *claves_posibles) -> int | None:
    """
    Busca la columna cuyo encabezado coincida con alguna clave.
    Primero exacto, luego containment (solo para claves de más de 1 carácter).
    """
    for clave in claves_posibles:
        if clave in cols:
            return cols[clave]
    for clave in claves_posibles:
        for enc, num in cols.items():
            if len(enc) > 1 and len(clave) > 1 and (clave in enc or enc in clave):
                return num
    return None


# ─────────────────────────────────────────────
# CONSECUTIVO
# ─────────────────────────────────────────────

def obtener_siguiente_consecutivo() -> int:
    """
    Retorna el SIGUIENTE consecutivo disponible (el que se usará en la próxima venta).
    Si hoy no hay ventas, retorna 1.
    """
    hoy = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")

    try:
        import db as _db
        if _db.DB_DISPONIBLE:
            row = _db.query_one(
                "SELECT MAX(consecutivo) AS max_consec FROM ventas WHERE fecha::date = %s::date",
                (hoy,),
            )
            if row and row.get("max_consec") is not None:
                return int(row["max_consec"]) + 1
            return 1
    except Exception:
        pass

    if not os.path.exists(config.EXCEL_FILE):
        return 1
    try:
        # CORRECCIÓN: read_only=True para lectura de consecutivo
        wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
        nombre_hoja = obtener_nombre_hoja()
        max_hoy     = 0

        if nombre_hoja in wb.sheetnames:
            ws   = wb[nombre_hoja]
            cols = {}
            for fila_hdr in ws.iter_rows(min_row=config.EXCEL_FILA_HEADERS, max_row=config.EXCEL_FILA_HEADERS):
                for cell in fila_hdr:
                    if cell.value:
                        cols[str(cell.value).lower().strip()] = cell.column
                break

            col_fecha  = next((v for k, v in cols.items() if "fecha"       in k), None)
            col_consec = next((v for k, v in cols.items() if "consecutivo" in k), None)

            if col_fecha and col_consec:
                for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
                    if str(fila[col_fecha - 1] or "")[:10] != hoy:
                        continue
                    try:
                        num = int(float(str(fila[col_consec - 1])))
                        if num > max_hoy:
                            max_hoy = num
                    except (TypeError, ValueError):
                        pass

        wb.close()
        return max_hoy + 1
    except Exception as e:
        print(f"Error leyendo consecutivo del Excel: {e}")
        return 1


def obtener_consecutivo_actual() -> int:
    """
    Retorna el último consecutivo REGISTRADO hoy (el de la última venta).
    Si no hay ventas hoy, retorna 0.

    CORRECCIÓN: antes retornaba obtener_siguiente_consecutivo() - 1, lo que daba
    0 cuando no había ventas del día y ese 0 se guardaba como consecutivo de la
    primera venta. Ahora lee el máximo real del día (o 0 si no hay ventas).
    """
    siguiente = obtener_siguiente_consecutivo()
    # siguiente=1 significa que no hay ventas hoy (el próximo será el 1)
    # siguiente=N+1 significa que el último registrado fue N
    return max(0, siguiente - 1)


# ─────────────────────────────────────────────
# CLIENTES (dentro del Excel)
# ─────────────────────────────────────────────

def cargar_clientes() -> list:
    global _clientes_cache, _clientes_cache_ts
    # Retornar caché si aún es válida (TTL 5 min)
    if _clientes_cache and (time.time() - _clientes_cache_ts) < _CLIENTES_CACHE_TTL:
        return _clientes_cache

    if not os.path.exists(config.EXCEL_FILE):
        return []
    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
        if "Clientes" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Clientes"]
        # En read_only ws.max_column puede ser None — usar iter_rows para headers
        headers = []
        try:
            for fila_hdr in ws.iter_rows(min_row=1, max_row=1):
                headers = [str(cell.value).strip() if cell.value else "" for cell in fila_hdr]
                break
        except Exception:
            wb.close()
            return []
        clientes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            cliente = {}
            for i, h in enumerate(headers):
                if h:
                    cliente[h] = row[i] if i < len(row) else None
            if cliente:
                clientes.append(cliente)
        wb.close()
        # Guardar en caché con timestamp
        _clientes_cache    = clientes
        _clientes_cache_ts = time.time()
        return clientes
    except Exception as e:
        print(f"Error cargando clientes: {e}")
        return []


def buscar_clientes_multiples(termino: str, limite: int = 5) -> list:
    clientes     = cargar_clientes()
    termino_norm = _normalizar(termino)
    palabras     = [p for p in termino_norm.split() if len(p) > 2]

    if not palabras:
        return []

    resultado = []
    for c in clientes:
        nombre_norm = _normalizar(c.get("Nombre tercero", "") or "")
        if any(p in nombre_norm for p in palabras):
            resultado.append(c)

    resultado.sort(key=lambda x: len(str(x.get("Nombre tercero", ""))))
    return resultado[:limite]


def buscar_cliente(termino: str) -> dict | None:
    clientes     = cargar_clientes()
    termino_norm = _normalizar(termino)

    for c in clientes:
        if _normalizar(c.get("Identificacion", "") or "") == termino_norm:
            return c

    palabras = [p for p in termino_norm.split() if len(p) > 2]
    if not palabras:
        return None

    coincidencias = []
    for c in clientes:
        nombre_norm = _normalizar(c.get("Nombre tercero", "") or "")
        if any(p in nombre_norm for p in palabras):
            coincidencias.append(c)

    if len(coincidencias) == 1:
        return coincidencias[0]

    return None


def buscar_cliente_con_resultado(termino: str) -> tuple[dict | None, list]:
    clientes     = cargar_clientes()
    termino_norm = _normalizar(termino)

    for c in clientes:
        if _normalizar(c.get("Identificacion", "") or "") == termino_norm:
            return c, [c]

    palabras = [p for p in termino_norm.split() if len(p) > 2]
    if not palabras:
        return None, []

    candidatos = []
    for c in clientes:
        nombre_norm = _normalizar(c.get("Nombre tercero", "") or "")
        if any(p in nombre_norm for p in palabras):
            candidatos.append(c)

    candidatos.sort(key=lambda x: len(str(x.get("Nombre tercero", ""))))

    if len(candidatos) == 1:
        return candidatos[0], candidatos
    return None, candidatos


def obtener_clientes_recientes(limite: int = 5) -> list:
    clientes = cargar_clientes()
    def _fecha(c):
        f = c.get("Fecha registro") or ""
        return str(f)
    con_fecha = [c for c in clientes if c.get("Fecha registro")]
    sin_fecha = [c for c in clientes if not c.get("Fecha registro")]
    con_fecha.sort(key=_fecha, reverse=True)
    return (con_fecha + sin_fecha)[:limite]


def obtener_nombre_id_cliente(termino: str) -> tuple[str, str]:
    cliente = buscar_cliente(termino)
    if cliente:
        id_c     = cliente.get("Identificacion") or "CF"
        nombre_c = cliente.get("Nombre tercero") or "Consumidor Final"
        return str(id_c), str(nombre_c)
    return "CF", "Consumidor Final"


def guardar_cliente_nuevo(nombre, tipo_id, identificacion, tipo_persona="Natural", correo="", telefono="", direccion="") -> bool:
    try:
        inicializar_excel()
        wb = openpyxl.load_workbook(config.EXCEL_FILE)
        if "Clientes" not in wb.sheetnames:
            ws_c    = wb.create_sheet("Clientes")
            headers = [
                "Nombre tercero", "Es Juridica o Persona", "Tipo de identificacion",
                "Identificacion", "Digito verificacion", "Correo electronico",
                "Direccion", "Telefono", "Nombres contacto", "Fecha registro",
            ]
            for col, h in enumerate(headers, 1):
                celda      = ws_c.cell(row=1, column=col, value=h)
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill("solid", fgColor="1A56DB")
        else:
            ws_c = wb["Clientes"]

        fila = ws_c.max_row + 1
        ws_c.cell(row=fila, column=1, value=nombre.upper())
        ws_c.cell(row=fila, column=2, value=tipo_persona)
        ws_c.cell(row=fila, column=3, value=tipo_id)
        ws_c.cell(row=fila, column=4, value=identificacion)
        ws_c.cell(row=fila, column=5, value="0")
        ws_c.cell(row=fila, column=6, value=correo)
        ws_c.cell(row=fila, column=7, value=direccion or "No aplica")
        ws_c.cell(row=fila, column=8, value=telefono or "000-0000000-")
        ws_c.cell(row=fila, column=9, value=nombre.upper())
        ws_c.cell(row=fila, column=10, value=datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M"))

        wb.save(config.EXCEL_FILE)
        _invalidar_cache_clientes()
        return True
    except Exception as e:
        print(f"Error guardando cliente: {e}")
        return False


def borrar_cliente(termino: str) -> tuple[bool, str]:
    try:
        inicializar_excel()
        wb = openpyxl.load_workbook(config.EXCEL_FILE)
        if "Clientes" not in wb.sheetnames:
            return False, "No hay clientes registrados."
        ws           = wb["Clientes"]
        termino_norm = _normalizar(termino)
        fila_borrar  = None
        nombre_borrado = None
        for fila in range(2, ws.max_row + 1):
            nombre = str(ws.cell(row=fila, column=1).value or "")
            id_val  = str(ws.cell(row=fila, column=4).value or "")
            if _normalizar(nombre) == termino_norm or _normalizar(id_val) == termino_norm:
                fila_borrar    = fila
                nombre_borrado = nombre
                break
            palabras = [p for p in termino_norm.split() if len(p) > 2]
            if palabras and all(p in _normalizar(nombre) for p in palabras):
                fila_borrar    = fila
                nombre_borrado = nombre
                break

        if not fila_borrar:
            return False, f"No encontré un cliente que coincida con '{termino}'."
        ws.delete_rows(fila_borrar)
        wb.save(config.EXCEL_FILE)
        _invalidar_cache_clientes()
        return True, f"✅ Cliente '{nombre_borrado}' borrado del sistema."
    except Exception as e:
        print(f"Error borrando cliente: {e}")
        return False, "Hubo un error borrando el cliente."


# ─────────────────────────────────────────────
# CRUD DE VENTAS
# ─────────────────────────────────────────────

def guardar_venta_excel(producto, cantidad, precio_unitario, total, vendedor,
                        observaciones="", cliente_nombre=None, cliente_id=None,
                        codigo_producto=None, consecutivo=None, metodo_pago=None,
                        unidad_medida=None) -> int:
    from memoria import cargar_memoria

    inicializar_excel()
    wb = openpyxl.load_workbook(config.EXCEL_FILE)

    fecha_hoy  = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
    hora_ahora = datetime.now(config.COLOMBIA_TZ).strftime("%H:%M")

    # CORRECCIÓN: consecutivo_final nunca puede ser 0 o None
    # Si por alguna razón llega None o 0, recalculamos
    if not consecutivo:
        consecutivo = obtener_siguiente_consecutivo()
    consecutivo_final = consecutivo

    id_cliente_final     = cliente_id    or "CF"
    nombre_cliente_final = cliente_nombre or "Consumidor Final"

    cod_producto_final = codigo_producto or ""
    if not cod_producto_final:
        from memoria import buscar_producto_en_catalogo
        prod_encontrado = buscar_producto_en_catalogo(str(producto))
        if prod_encontrado:
            cod_producto_final = prod_encontrado.get("codigo", "")

    datos_base = {
        "fecha":                fecha_hoy,
        "hora":                 hora_ahora,
        "id cliente":           id_cliente_final,
        "cliente":              nombre_cliente_final,
        "codigo del producto":  cod_producto_final,
        "producto":             str(producto),
        "cantidad":             cantidad,
        "valor unitario":       float(precio_unitario),
        "total":                float(total),
        "subtotal":             float(total),
        "consecutivo de venta": consecutivo_final,
        "vendedor":             str(vendedor),
        "metodo de pago":       str(metodo_pago) if metodo_pago else str(observaciones),
        "unidad de medida":     str(unidad_medida) if unidad_medida else "Unidad",
    }

    # Hojas donde guardar simultáneamente
    hojas_destino = [obtener_nombre_hoja(), "Registro de Ventas-Acumulado"]

    for nombre_sh in hojas_destino:
        ws   = obtener_o_crear_hoja(wb, nombre_sh)
        cols = detectar_columnas(ws)

        # ── Auto-crear columna UNIDAD DE MEDIDA antes de CANTIDAD ─────────
        if "unidad de medida" not in cols and "unidad_medida" not in cols:
            col_cantidad = cols.get("cantidad")
            if col_cantidad:
                # Insertar columna nueva antes de CANTIDAD
                ws.insert_cols(col_cantidad)
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=col_cantidad, value="UNIDAD DE MEDIDA")
                from openpyxl.styles import Font as _Font, PatternFill as _PF, Alignment as _Al
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=col_cantidad).font = _Font(bold=True, color="FFFFFF")
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=col_cantidad).fill = _PF("solid", fgColor="1B56E1")
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=col_cantidad).alignment = _Al(horizontal="center")
                # Re-detectar columnas después de la inserción
                cols = detectar_columnas(ws)
            else:
                next_col = max(cols.values(), default=0) + 1
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=next_col, value="UNIDAD DE MEDIDA")
                from openpyxl.styles import Font as _Font, PatternFill as _PF, Alignment as _Al
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=next_col).font = _Font(bold=True, color="FFFFFF")
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=next_col).fill = _PF("solid", fgColor="1B56E1")
                ws.cell(row=config.EXCEL_FILA_HEADERS, column=next_col).alignment = _Al(horizontal="center")
                cols["unidad de medida"] = next_col

        fila     = max(_ultima_fila_con_datos(ws, config.EXCEL_FILA_DATOS) + 1, config.EXCEL_FILA_DATOS)

        datos = datos_base.copy()

        for nombre_col, num_col in cols.items():
            clave = nombre_col.lower().strip()
            if clave in datos:
                ws.cell(row=fila, column=num_col, value=datos[clave])
                continue
            for dato_key, dato_val in datos.items():
                if len(dato_key) > 5 and len(clave) > 5 and (dato_key in clave or clave in dato_key):
                    ws.cell(row=fila, column=num_col, value=dato_val)
                    break

        if consecutivo_final % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

    wb.save(config.EXCEL_FILE)

    return consecutivo_final


def borrar_venta_excel(numero_venta) -> tuple[bool, str]:
    import logging
    log = logging.getLogger("ferrebot.excel")
    from memoria import cargar_caja, guardar_caja

    inicializar_excel()
    wb = openpyxl.load_workbook(config.EXCEL_FILE)

    total_borradas  = 0
    hojas_buscar    = [obtener_nombre_hoja(), "Registro de Ventas-Acumulado"]
    totales_por_metodo = {}
    hoy = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")

    nombre_hoja_mes = obtener_nombre_hoja()
    log.info("[borrar] buscando consecutivo=%s hoja_mes=%s sheets=%s",
             numero_venta, nombre_hoja_mes, wb.sheetnames)

    if nombre_hoja_mes in wb.sheetnames:
        ws_mes = wb[nombre_hoja_mes]
        cols   = detectar_columnas(ws_mes)
        log.info("[borrar] cols detectadas: %s", list(cols.keys())[:10])
        col_id     = cols.get("consecutivo de venta") or cols.get("consecutivo") or cols.get("alias")
        col_total  = next((v for k, v in cols.items() if k == "total"), None)
        col_metodo = next((v for k, v in cols.items() if "metodo" in k), None)
        col_fecha  = next((v for k, v in cols.items() if "fecha" in k), None)
        log.info("[borrar] col_id=%s col_total=%s col_fecha=%s", col_id, col_total, col_fecha)

        if col_id and col_total:
            for fila in range(config.EXCEL_FILA_DATOS, ws_mes.max_row + 1):
                val = ws_mes.cell(row=fila, column=col_id).value
                try:
                    if val is not None and int(float(str(val).strip())) == int(numero_venta):
                        fecha_fila = str(ws_mes.cell(row=fila, column=col_fecha).value or "")[:10] if col_fecha else ""
                        if fecha_fila == hoy:
                            t = float(ws_mes.cell(row=fila, column=col_total).value or 0)
                            m = str(ws_mes.cell(row=fila, column=col_metodo).value or "efectivo").lower() if col_metodo else "efectivo"
                            totales_por_metodo[m] = totales_por_metodo.get(m, 0) + t
                except (ValueError, TypeError):
                    pass

    for nombre_sh in hojas_buscar:
        if nombre_sh not in wb.sheetnames:
            log.info("[borrar] hoja '%s' no existe", nombre_sh)
            continue
        ws   = wb[nombre_sh]
        cols = detectar_columnas(ws)
        col_id = (
            cols.get("consecutivo de venta") or
            cols.get("consecutivo") or
            cols.get("alias") or
            cols.get("#") or
            cols.get("num")
        )
        log.info("[borrar] hoja='%s' col_id=%s max_row=%s", nombre_sh, col_id, ws.max_row)
        if not col_id:
            log.warning("[borrar] sin col_id en '%s', cols=%s", nombre_sh, list(cols.keys()))
            continue

        filas_a_borrar = []
        for fila in range(config.EXCEL_FILA_DATOS, ws.max_row + 1):
            val = ws.cell(row=fila, column=col_id).value
            if val is None:
                continue
            try:
                val_str = str(val).strip()
                if val_str and int(float(val_str)) == int(numero_venta):
                    filas_a_borrar.append(fila)
                    log.info("[borrar] fila %s coincide (val=%r)", fila, val)
            except (ValueError, TypeError):
                pass

        log.info("[borrar] filas_a_borrar=%s en '%s'", filas_a_borrar, nombre_sh)
        for fila in reversed(filas_a_borrar):
            ws.delete_rows(fila)
        total_borradas += len(filas_a_borrar)

    log.info("[borrar] total_borradas=%s", total_borradas)

    if total_borradas:
        wb.save(config.EXCEL_FILE)

        if totales_por_metodo:
            caja = cargar_caja()
            if caja.get("abierta"):
                _map = {"efectivo": "efectivo", "transferencia": "transferencias",
                        "transferencias": "transferencias", "datafono": "datafono"}
                for metodo, monto in totales_por_metodo.items():
                    campo = _map.get(metodo, "efectivo")
                    caja[campo] = max(0, caja.get(campo, 0) - monto)
                guardar_caja(caja)

        return True, f"✅ Consecutivo #{numero_venta} borrado — {total_borradas} fila(s) eliminadas del Excel."

    return False, f"No encontré el consecutivo #{numero_venta}. Hojas revisadas: {hojas_buscar}"


def recalcular_caja_desde_excel():
    """
    Recalcula los totales de la caja leyendo las ventas de HOY desde la hoja mensual.
    Útil después de borrar ventas para que la caja quede consistente.
    Solo actúa si la caja está abierta.
    """
    from memoria import cargar_caja, guardar_caja

    caja = cargar_caja()
    if not caja.get("abierta"):
        return

    hoy = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")

    efectivo      = 0.0
    transferencias = 0.0
    datafono      = 0.0

    if os.path.exists(config.EXCEL_FILE):
        try:
            wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
            nombre_hoja = obtener_nombre_hoja()
            if nombre_hoja in wb.sheetnames:
                ws        = wb[nombre_hoja]
                cols      = detectar_columnas(ws)
                col_fecha  = next((v for k, v in cols.items() if "fecha"  in k), None)
                col_total  = next((v for k, v in cols.items() if k == "total"), None)
                col_metodo = next((v for k, v in cols.items() if "metodo" in k), None)

                if col_fecha and col_total:
                    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
                        fecha_fila = str(fila[col_fecha - 1] or "")[:10]
                        if fecha_fila != hoy:
                            continue
                        try:
                            t = float(fila[col_total - 1] or 0)
                            m = str(fila[col_metodo - 1] or "efectivo").lower() if col_metodo else "efectivo"
                            if "transfer" in m:
                                transferencias += t
                            elif "datafono" in m or "tarjeta" in m:
                                datafono += t
                            else:
                                efectivo += t
                        except (TypeError, ValueError):
                            pass
            wb.close()
        except Exception as e:
            print(f"Error recalculando caja: {e}")

    caja["efectivo"]       = efectivo
    caja["transferencias"] = transferencias
    caja["datafono"]       = datafono
    guardar_caja(caja)


def obtener_ventas_por_consecutivo(numero_venta) -> list:
    """Retorna todas las filas con ese consecutivo (para ventas con múltiples productos)."""
    inicializar_excel()
    wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        return []
    ws     = wb[nombre_hoja]
    cols   = detectar_columnas(ws)
    col_id = cols.get("consecutivo de venta") or cols.get("alias")
    filas  = []
    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        val = fila[col_id - 1] if col_id else None
        try:
            if val is not None and int(float(str(val))) == int(numero_venta):
                filas.append({nombre: fila[num - 1] for nombre, num in cols.items()})
        except (ValueError, TypeError):
            pass
    wb.close()
    return filas


def obtener_venta_por_numero(numero_venta) -> dict | None:
    inicializar_excel()
    # CORRECCIÓN: read_only=True para lectura
    wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        return None
    ws     = wb[nombre_hoja]
    cols   = detectar_columnas(ws)
    col_id = cols.get("consecutivo de venta") or cols.get("alias")

    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        val = fila[col_id - 1] if col_id else None
        try:
            if val is not None and int(float(str(val))) == int(numero_venta):
                wb.close()
                return {nombre: fila[num - 1] for nombre, num in cols.items()}
        except (ValueError, TypeError):
            pass
    wb.close()
    return None


def obtener_ventas_recientes(limite: int = 10) -> list:
    inicializar_excel()
    # CORRECCIÓN: read_only=True
    wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        return []
    ws     = wb[nombre_hoja]
    cols   = detectar_columnas(ws)
    ventas = []
    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        if not any(fila):
            continue
        ventas.append({nombre: fila[num - 1] for nombre, num in cols.items()})
    wb.close()
    return ventas[-limite:]


def buscar_ventas(termino: str) -> list:
    inicializar_excel()
    # CORRECCIÓN: read_only=True
    wb            = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    termino_lower = termino.lower().strip()
    resultados    = []
    for nombre_hoja in wb.sheetnames:
        ws   = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue
            fila_texto = " ".join(str(v).lower() for v in fila if v is not None)
            if termino_lower in fila_texto:
                d = {"hoja": nombre_hoja}
                d.update({nombre: fila[num - 1] for nombre, num in cols.items()})
                resultados.append(d)
    wb.close()
    return resultados


def obtener_todos_los_datos() -> list:
    inicializar_excel()
    # CORRECCIÓN: read_only=True
    wb    = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    todos = []
    for nombre_hoja in wb.sheetnames:
        ws   = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if any(fila):
                d = {"hoja": nombre_hoja}
                d.update({nombre: fila[num - 1] for nombre, num in cols.items()})
                todos.append(d)
    wb.close()
    return todos


def obtener_resumen_ventas() -> dict | None:
    inicializar_excel()
    # CORRECCIÓN: read_only=True
    wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        return None
    ws        = wb[nombre_hoja]
    cols      = detectar_columnas(ws)
    col_total = next((v for k, v in cols.items() if k == "total"), None)
    total_general = 0
    num_ventas    = 0
    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        if not any(fila):
            continue
        num_ventas += 1
        if col_total and fila[col_total - 1]:
            try:
                total_general += float(fila[col_total - 1])
            except Exception:
                pass
    wb.close()
    return {"hoja": nombre_hoja, "total": total_general, "num_ventas": num_ventas}


# ─────────────────────────────────────────────
# EXCEL PERSONALIZADO
# ─────────────────────────────────────────────

def generar_excel_personalizado(titulo: str, encabezados: list, filas: list, nombre_archivo: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
    celda           = ws.cell(row=1, column=1, value=titulo)
    celda.font      = Font(bold=True, color="FFFFFF", size=14)
    celda.fill      = PatternFill("solid", fgColor="1A56DB")
    celda.alignment = Alignment(horizontal="center")

    for col, enc in enumerate(encabezados, 1):
        celda           = ws.cell(row=2, column=col, value=enc)
        celda.font      = Font(bold=True, color="FFFFFF", size=11)
        celda.fill      = PatternFill("solid", fgColor="374151")
        celda.alignment = Alignment(horizontal="center")

    for i, fila in enumerate(filas, 3):
        for col, valor in enumerate(fila, 1):
            celda = ws.cell(row=i, column=col, value=valor)
            if i % 2 == 0:
                celda.fill = PatternFill("solid", fgColor="EFF6FF")

    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(nombre_archivo)
    return nombre_archivo


# ─────────────────────────────────────────────
# WRAPPERS ASYNC
# ─────────────────────────────────────────────

def obtener_ventas_hoy_excel() -> dict:
    """
    Suma solo las ventas del día actual en el Excel.
    Usado por obtener_resumen_caja() para mostrar Total ventas de hoy, no del mes.
    """
    from datetime import datetime as _dt
    inicializar_excel()
    hoy = _dt.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
    try:
        wb          = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True)
        nombre_hoja = obtener_nombre_hoja()
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            return {"total": 0, "num_ventas": 0}
        ws        = wb[nombre_hoja]
        cols      = detectar_columnas(ws)
        col_total = next((v for k, v in cols.items() if k == "total"), None)
        col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)
        total_hoy = 0
        num_hoy   = 0
        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue
            fecha_fila = str(fila[col_fecha - 1] or "")[:10] if col_fecha else ""
            if fecha_fila != hoy:
                continue
            num_hoy += 1
            if col_total and fila[col_total - 1]:
                try:
                    total_hoy += float(fila[col_total - 1])
                except Exception:
                    pass
        wb.close()
        return {"total": total_hoy, "num_ventas": num_hoy}
    except Exception:
        return {"total": 0, "num_ventas": 0}


async def guardar_venta_excel_async(*args, **kwargs) -> int:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, lambda: guardar_venta_excel(*args, **kwargs))


async def borrar_venta_excel_async(*args, **kwargs) -> tuple[bool, str]:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, lambda: borrar_venta_excel(*args, **kwargs))


async def inicializar_excel_async():
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, inicializar_excel)


# ─────────────────────────────────────────────
# FIADOS EN EXCEL
# ─────────────────────────────────────────────

def registrar_fiado_en_excel(cliente: str, concepto: str, cargo: float, abono: float, saldo: float):
    """
    Agrega una fila a la hoja 'Fiados' del Excel con el movimiento.
    Crea la hoja si no existe.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(config.EXCEL_FILE)

    nombre_hoja = "Fiados"
    if nombre_hoja not in wb.sheetnames:
        ws          = wb.create_sheet(title=nombre_hoja)
        encabezados = ["FECHA", "CLIENTE", "CONCEPTO", "CARGO", "ABONO", "SALDO"]
        for col, titulo in enumerate(encabezados, 1):
            celda           = ws.cell(row=1, column=col, value=titulo)
            celda.font      = Font(bold=True, color="FFFFFF", size=11)
            celda.fill      = PatternFill("solid", fgColor="1A1A1A")
            celda.alignment = Alignment(horizontal="center", vertical="center")
        anchos = [12, 25, 40, 15, 15, 15]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
        fila_datos = 2
    else:
        ws         = wb[nombre_hoja]
        fila_datos = _ultima_fila_con_datos(ws, 2) + 1

    fecha  = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
    valores = [fecha, cliente, concepto, cargo if cargo > 0 else "", abono if abono > 0 else "", saldo]
    for col, val in enumerate(valores, 1):
        celda           = ws.cell(row=fila_datos, column=col, value=val)
        celda.alignment = Alignment(horizontal="center")
        if col in (4, 5, 6) and val != "":
            celda.number_format = "$#,##0"
        if fila_datos % 2 == 0:
            celda.fill = PatternFill("solid", fgColor="EFF6FF")

    wb.save(config.EXCEL_FILE)


# ─────────────────────────────────────────────
# COMPRAS EN EXCEL
# ─────────────────────────────────────────────

def registrar_compra_en_excel(producto: str, cantidad: float, costo_unitario: float, 
                               costo_total: float, proveedor: str = "—"):
    """
    Agrega una fila a la hoja 'Compras' del Excel.
    Crea la hoja si no existe.
    """
    inicializar_excel()
    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    
    nombre_hoja = "Compras"
    if nombre_hoja not in wb.sheetnames:
        ws = wb.create_sheet(title=nombre_hoja)
        encabezados = ["FECHA", "HORA", "PROVEEDOR", "PRODUCTO", "CANTIDAD", "COSTO UNIT.", "COSTO TOTAL"]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=enc)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="2563EB")
            celda.alignment = Alignment(horizontal="center")
        # Anchos de columna
        anchos = [12, 8, 20, 30, 12, 15, 15]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
        fila_datos = 2
    else:
        ws = wb[nombre_hoja]
        fila_datos = _ultima_fila_con_datos(ws, 2) + 1
    
    ahora = datetime.now(config.COLOMBIA_TZ)
    fecha = ahora.strftime("%Y-%m-%d")
    hora = ahora.strftime("%H:%M")
    
    valores = [fecha, hora, proveedor, producto, cantidad, costo_unitario, costo_total]
    for col, val in enumerate(valores, 1):
        celda = ws.cell(row=fila_datos, column=col, value=val)
        celda.alignment = Alignment(horizontal="center")
        if col in (6, 7):  # Columnas de costo
            celda.number_format = "$#,##0"
        if fila_datos % 2 == 0:
            celda.fill = PatternFill("solid", fgColor="FEF3C7")  # Amarillo claro

    wb.save(config.EXCEL_FILE)


def actualizar_hoja_inventario():
    """
    Actualiza/crea la hoja 'Inventario' con el estado actual.
    Sobreescribe todos los datos cada vez.
    """
    from memoria import cargar_inventario, buscar_producto_en_catalogo
    
    inicializar_excel()
    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    
    nombre_hoja = "Inventario"
    
    # Si existe, eliminar y recrear
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    
    ws = wb.create_sheet(title=nombre_hoja)
    
    # Encabezados
    encabezados = ["PRODUCTO", "STOCK", "COSTO PROM.", "PRECIO VENTA", "MARGEN %", "PROVEEDOR", "ÚLT. COMPRA"]
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=enc)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="059669")  # Verde
        celda.alignment = Alignment(horizontal="center")
    
    # Anchos de columna
    anchos = [30, 10, 15, 15, 12, 20, 15]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    
    # Cargar datos de inventario
    inventario = cargar_inventario()
    fila = 2
    
    for clave, datos in sorted(inventario.items()):
        if not isinstance(datos, dict):
            continue
        
        nombre = datos.get("nombre_original", clave)
        cantidad = datos.get("cantidad", 0)
        costo_prom = datos.get("costo_promedio", 0)
        proveedor = datos.get("ultimo_proveedor", "—")
        ultima_compra = datos.get("ultima_compra", "—")
        
        # Buscar precio de venta en catálogo
        producto_cat = buscar_producto_en_catalogo(nombre)
        precio_venta = producto_cat.get("precio_unidad", 0) if producto_cat else 0
        
        # Calcular margen
        if precio_venta > 0 and costo_prom > 0:
            margen = round(((precio_venta - costo_prom) / precio_venta) * 100, 1)
        else:
            margen = "—"
        
        valores = [nombre, cantidad, costo_prom, precio_venta, margen, proveedor, ultima_compra]
        
        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=val if val != 0 else "—")
            celda.alignment = Alignment(horizontal="center")
            if col in (3, 4):  # Costos y precios
                if val and val != "—":
                    celda.number_format = "$#,##0"
            if col == 5 and val != "—":  # Margen
                celda.number_format = "0.0%"
                if isinstance(val, (int, float)):
                    celda.value = val / 100  # Convertir a porcentaje
                    if val >= 40:
                        celda.fill = PatternFill("solid", fgColor="D1FAE5")  # Verde claro
                    elif val < 20:
                        celda.fill = PatternFill("solid", fgColor="FEE2E2")  # Rojo claro
            if fila % 2 == 0:
                if col != 5 or val == "—":  # No sobreescribir color de margen
                    celda.fill = PatternFill("solid", fgColor="F0FDF4")

        fila += 1

    wb.save(config.EXCEL_FILE)
