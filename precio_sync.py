"""
precio_sync.py — Sincronización robusta y bidireccional de precios.

ARQUITECTURA:
  memoria.json (catálogo en RAM) ←→ BASE_DE_DATOS_PRODUCTOS.xlsx (Drive)

══════════════════════════════════════════════════════════
ESTRUCTURA DEL EXCEL (hoja "Datos")
══════════════════════════════════════════════════════════
  Col A  (idx 0)  : Código del Producto
  Col B  (idx 1)  : Nombre del Producto
  Col D  (idx 3)  : Categoría
  Col Q  (idx 16) : UNIDAD  → precio por unidad completa
  Col R  (idx 17) : 0.75
  Col S  (idx 18) : 0.5
  Col T  (idx 19) : 0.25
  Col U  (idx 20) : 0.13  → real decimal = 1/8 = 0.125
  Col V  (idx 21) : 0.06  → real decimal = 1/16 = 0.0625
  Col W  (idx 22) : 0.1   → reservada (actualmente vacía)

INTERPRETACIÓN POR CATEGORÍA:
  Cat 2 - Pinturas / Cat 4 - Impermeabilizantes:
    Cols R-V = precio UNITARIO para esa fracción de galón.
    precio_total = valor_celda × decimal_real
    Ej: col S = 52000, decimal=0.5  →  total 1/2 galón = 26000

  Cat 3 - Tornillería:
    Col R = precio mayorista por unidad cuando qty >= UMBRAL_TORNILLERIA (50).
    Si col R == col Q → sin descuento (igual en ambos niveles).
    Se guarda como precio_por_cantidad {umbral, precio_bajo, precio_sobre}.

  Resto (Cat 1 - Ferretería, Cat 5 - Eléctricos, etc.):
    Cols R-V se usan solo si valor < precio_unidad.
    En ese caso el valor de la celda ES el precio total de la fracción
    (no se multiplica). Ej: WAYPER Q=10000, S=5000 → 1/2 vale $5.000.

GARANTÍAS:
  1. Campo "decimal" SIEMPRE presente en precios_fraccion.
  2. Cola FIFO serializada → sin condición de carrera al escribir el Excel.
  3. Una sola función pública: actualizar_precio(nombre, precio, fraccion).
  4. importar_catalogo_desde_excel() es la única entrada para bulk import.
"""

import logging
import os
import queue
import re
import shutil
import threading
from typing import Optional

import openpyxl
import config

log = logging.getLogger("ferrebot.precio_sync")

# ─────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────

NOMBRE_EXCEL_PRODUCTOS = "BASE_DE_DATOS_PRODUCTOS.xlsx"
UMBRAL_TORNILLERIA     = 50

_IDX_CODIGO         = 0
_IDX_NOMBRE         = 1
_IDX_CATEGORIA      = 3
_IDX_UNIDAD_MEDIDA  = 8   # Col I — Unidad de Medida Impresión Factura (DIAN)
_IDX_UNIDAD         = 16  # Col Q — Precio por unidad completa

# header_str → (col_idx_base0, decimal_real, label)
# "0.13" y "0.06" son aproximaciones de 1/8 y 1/16 (los headers del Excel
# están redondeados, pero los decimales reales son 0.125 y 0.0625)
_HEADER_MAP: dict[str, tuple[int, float, str]] = {
    "0.75":             (17, 0.75,   "3/4"),
    "0.5":              (18, 0.5,    "1/2"),
    "0.25":             (19, 0.25,   "1/4"),
    "0.13":             (20, 0.125,  "1/8"),
    "0.06":             (21, 0.0625, "1/16"),
    "0.1":              (22, 0.1,    "1/10"),
    # "Precio de venta 8" (col X) eliminado — el precio por unidad suelta
    # ahora se maneja como un producto separado en el catálogo (ej: "Wayper Blanco Unidad")
}

# label → (decimal_real, col_idx_base0) — índice inverso
_LABEL_MAP: dict[str, tuple[float, int]] = {
    label: (dec, idx) for _, (idx, dec, label) in _HEADER_MAP.items()
}

_CATS_GALON = {
    "2 pinturas y disolventes",
    "4 impermeabilizantes y materiales de construccion",
    "4 impermeabilizantes y materiales de construcción",
}

_CATS_TORNILLERIA = {"3 tornilleria", "3 tornillería"}


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────

def _norm_cat(cat: str) -> str:
    return (
        (cat or "").lower()
        .replace("á","a").replace("é","e").replace("í","i")
        .replace("ó","o").replace("ú","u").replace("ñ","n")
        .strip()
    )

def _es_galon(cat: str) -> bool:
    return _norm_cat(cat) in _CATS_GALON

def _es_tornilleria(cat: str) -> bool:
    return _norm_cat(cat) in _CATS_TORNILLERIA

def _num(v) -> Optional[float]:
    """Celda → float positivo, o None."""
    if v is None:
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None

def _limpiar(*rutas):
    for r in rutas:
        try:
            if r and os.path.exists(r):
                os.remove(r)
        except Exception:
            pass


# Mapa de normalización de unidades para factura electrónica (DIAN)
_UNIDAD_MAP: dict[str, str] = {
    # galón y variantes
    "galon":   "Galón", "galón":  "Galón", "gal":    "Galón",
    # kilogramo
    "kg":      "Kg",    "kgs":    "Kg",    "kilo":   "Kg",
    "kilos":   "Kg",    "kilogramo": "Kg", "25 kg":  "Kg",
    # metro
    "mts":     "Mts",   "mt":     "Mts",   "metro":  "Mts",
    "metros":  "Mts",   "m":      "Mts",
    # centímetro
    "cms":     "Cms",   "cm":     "Cms",   "centimetro": "Cms",
    # litro
    "lt":      "Lt",    "lts":    "Lts",   "litro":  "Lt",
    "litros":  "Lts",
    # mililitro — código DIAN: MLT (tabla 13.3.6 anexo técnico v1.8)
    "ml":      "MLT",   "mlt":    "MLT",   "mililitro":  "MLT",
    "mililitros": "MLT", "cc":    "MLT",   "centimetro cubico": "MLT",
    # unidad (por defecto)
    "unidad":  "Unidad","und":    "Unidad","un":     "Unidad",
    "unidades":"Unidad","uni":    "Unidad",
}

def _normalizar_unidad(raw: str) -> str:
    """Normaliza el texto de la col I → etiqueta canónica para DIAN."""
    if not raw:
        return "Unidad"
    clave = raw.strip().lower().replace("á","a").replace("é","e").replace("ó","o")
    return _UNIDAD_MAP.get(clave, raw.strip()) or "Unidad"


# ─────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DE UN PRODUCTO DESDE UNA FILA
# ─────────────────────────────────────────────────────────────────

def construir_producto_desde_fila(row: tuple, col_headers: list) -> Optional[dict]:
    """
    Convierte una fila del Excel en un dict listo para el catálogo.
    Siempre incluye campo "decimal" en precios_fraccion.
    Retorna None si la fila no tiene nombre o precio válido.
    """
    from utils import _normalizar

    nombre    = str(row[_IDX_NOMBRE] or "").strip()
    if not nombre or nombre.lower() == "nan":
        return None

    cat      = str(row[_IDX_CATEGORIA] or "").strip()
    codigo   = str(row[_IDX_CODIGO] or "").strip()
    p_unidad = _num(row[_IDX_UNIDAD]) if _IDX_UNIDAD < len(row) else None

    # ── Col I: Unidad de Medida Impresión Factura (DIAN) ─────────────────────
    unidad_raw = str(row[_IDX_UNIDAD_MEDIDA] or "").strip() if len(row) > _IDX_UNIDAD_MEDIDA else ""
    unidad_medida = _normalizar_unidad(unidad_raw)

    if p_unidad is None:
        return None

    nombre_lower = _normalizar(nombre)

    prod = {
        "nombre":        nombre,
        "nombre_lower":  nombre_lower,
        "categoria":     cat,
        "precio_unidad": round(p_unidad),
        "unidad_medida": unidad_medida,
    }
    if codigo:
        prod["codigo"] = codigo

    # ── Cat 2 / Cat 4: fracciones de galón ──────────────────────────────────
    if _es_galon(cat):
        fracs = {}
        for i, header in enumerate(col_headers):
            if i == _IDX_UNIDAD:
                continue
            info = _HEADER_MAP.get(str(header).strip())
            if not info:
                continue
            idx, decimal_real, label = info
            v = _num(row[i]) if i < len(row) else None
            if v is None:
                continue
            # Valor en celda = precio UNITARIO. Total = v × decimal_real
            fracs[label] = {
                "precio":  round(v * decimal_real),
                "decimal": decimal_real,
            }
        if fracs:
            prod["precios_fraccion"] = fracs

    # ── Cat 3: tornillería → precio mayorista ────────────────────────────────
    elif _es_tornilleria(cat):
        idx_r  = _HEADER_MAP["0.75"][0]   # col R = idx 17
        p_may  = _num(row[idx_r]) if idx_r < len(row) else None
        if p_may is not None and round(p_may) != round(p_unidad):
            prod["precio_por_cantidad"] = {
                "umbral":              UMBRAL_TORNILLERIA,
                "precio_bajo_umbral":  round(p_unidad),
                "precio_sobre_umbral": round(p_may),
            }

    # ── Resto: fracciones directas (celda = total, solo si < precio_unidad) ──
    else:
        fracs = {}
        for i, header in enumerate(col_headers):
            if i == _IDX_UNIDAD:
                continue
            info = _HEADER_MAP.get(str(header).strip().lower())
            if not info:
                info = _HEADER_MAP.get(str(header).strip())
            if not info:
                continue
            idx, decimal_real, label = info
            v = _num(row[i]) if i < len(row) else None
            if v is None:
                continue
            # Solo fracciones con decimal definido y valor menor al precio unidad
            if decimal_real is not None and v < p_unidad:
                fracs[label] = {
                    "precio":  round(v),
                    "decimal": decimal_real,
                }
        if fracs:
            prod["precios_fraccion"] = fracs

    return prod


# ─────────────────────────────────────────────────────────────────
# IMPORTAR CATÁLOGO COMPLETO (Excel → memoria.json)
# ─────────────────────────────────────────────────────────────────

def importar_catalogo_desde_excel(ruta_excel: str) -> dict:
    """
    Lee BASE_DE_DATOS_PRODUCTOS.xlsx e importa todos los productos.
    Reemplaza COMPLETAMENTE el catálogo en memoria (limpia claves legacy).

    Retorna {"importados": N, "omitidos": N, "errores": [...]}
    """
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        ws = wb["Datos"]
    except Exception as e:
        return {"importados": 0, "omitidos": 0, "errores": [str(e)]}

    col_headers = [
        str(ws.cell(1, c).value or "")
        for c in range(1, ws.max_column + 1)
    ]

    from memoria import cargar_memoria, guardar_memoria, invalidar_cache_memoria

    mem      = cargar_memoria()
    catalogo = {}        # siempre limpio — evita duplicados de claves antiguas
    importados = 0
    omitidos   = 0
    errores    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            prod = construir_producto_desde_fila(row, col_headers)
            if prod is None:
                omitidos += 1
                continue
            # Sanitizar clave: reemplazar espacios con _ y eliminar chars que rompen URLs/JSON
            # Ej: 'brocha de 1"' → 'brocha_de_1' (sin comilla doble)
            clave = re.sub(r'["\'\°\#\!\?\(\)\[\]\{\}]', '', prod["nombre_lower"]).replace(" ", "_")
            clave = re.sub(r'_+', '_', clave).strip('_')  # limpiar guiones dobles o extremos
            catalogo[clave] = prod
            importados += 1
        except Exception as e:
            nombre_raw = row[_IDX_NOMBRE] if row and len(row) > _IDX_NOMBRE else "?"
            errores.append(f"{nombre_raw}: {e}")

    mem["catalogo"] = catalogo
    # Limpiar precios simples para que no contradigan al catálogo
    mem["precios"] = {}

    guardar_memoria(mem)
    invalidar_cache_memoria()

    return {"importados": importados, "omitidos": omitidos, "errores": errores[:10]}



# ─────────────────────────────────────────────────────────────────
# CREAR PRODUCTO NUEVO (Dashboard → Excel + memoria.json)
# ─────────────────────────────────────────────────────────────────

def agregar_producto_a_excel(datos: dict) -> dict:
    """
    Agrega una nueva fila al Excel BASE_DE_DATOS_PRODUCTOS.xlsx (hoja 'Datos')
    con los datos del producto recibidos desde el dashboard.

    datos = {
        "codigo":        str,          # opcional
        "nombre":        str,          # obligatorio
        "categoria":     str,          # obligatorio
        "precio_unidad": int/float,    # obligatorio
        "unidad_medida": str,          # default "Unidad"
        "inventariable": bool,         # default True
        "visible_facturas": bool,      # default True
        "stock_minimo":  int,          # default 0
        "codigo_dian":   str,          # default "94"
    }
    Retorna {"ok": True, "fila": N} o {"ok": False, "error": "..."}
    """
    import shutil
    import tempfile
    try:
        from drive import descargar_de_drive, subir_a_drive, subir_archivo_a_drive
    except ImportError:
        descargar_de_drive = subir_a_drive = subir_archivo_a_drive = None

    ruta = config.EXCEL_PRODUCTOS if hasattr(config, "EXCEL_PRODUCTOS") else None
    # Intentar localizar el Excel de productos
    for candidato in [
        getattr(config, "EXCEL_PRODUCTOS", None),
        getattr(config, "BASE_DATOS_FILE", None),
        os.path.join(os.path.dirname(config.EXCEL_FILE), "BASE_DE_DATOS_PRODUCTOS.xlsx"),
        "BASE_DE_DATOS_PRODUCTOS.xlsx",
    ]:
        if candidato and os.path.exists(candidato):
            ruta = candidato
            break

    # Si no existe localmente, intentar descargar de Drive
    ruta_tmp = None
    if not ruta and descargar_de_drive:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                ruta_tmp = tmp.name
            ok = descargar_de_drive("BASE_DE_DATOS_PRODUCTOS.xlsx", ruta_tmp)
            if ok:
                ruta = ruta_tmp
        except Exception:
            pass

    if not ruta or not os.path.exists(ruta):
        return {"ok": False, "error": "No se encontró el Excel de productos"}

    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active

        # Encontrar la primera fila vacía (col B = Nombre)
        fila_nueva = ws.max_row + 1
        for r in range(2, ws.max_row + 2):
            if not ws.cell(row=r, column=2).value:
                fila_nueva = r
                break

        nombre    = str(datos.get("nombre", "")).strip()
        codigo    = str(datos.get("codigo", "")).strip()
        categoria = str(datos.get("categoria", "")).strip()
        precio    = datos.get("precio_unidad", 0)
        unidad    = datos.get("unidad_medida", "Unidad")
        inv       = "SI" if datos.get("inventariable", True) else "NO"
        vis       = "SI" if datos.get("visible_facturas", True) else "NO"
        stock_min = datos.get("stock_minimo", 0)
        cod_dian  = str(datos.get("codigo_dian", "94"))

        # Col A: Código, B: Nombre, C: Tipo, D: Categoría,
        # E: Inventariable, F: Visible, G: Stock mínimo,
        # H: Código DIAN, I: Unidad medida, Q: Precio unidad
        ws.cell(row=fila_nueva, column=1,  value=codigo or nombre.lower().replace(" ", "")[:20])
        ws.cell(row=fila_nueva, column=2,  value=nombre)
        ws.cell(row=fila_nueva, column=3,  value="P-Producto")
        ws.cell(row=fila_nueva, column=4,  value=categoria)
        ws.cell(row=fila_nueva, column=5,  value=inv)
        ws.cell(row=fila_nueva, column=6,  value=vis)
        ws.cell(row=fila_nueva, column=7,  value=stock_min)
        ws.cell(row=fila_nueva, column=8,  value=cod_dian)
        ws.cell(row=fila_nueva, column=9,  value=unidad)
        ws.cell(row=fila_nueva, column=14, value="22-IVA 0%")  # Col N — código impuesto
        ws.cell(row=fila_nueva, column=16, value="SI")          # Col P — incluye IVA
        ws.cell(row=fila_nueva, column=17, value=int(precio) if precio is not None else 0)  # Col Q

        wb.save(ruta)

        # Subir a Drive siempre que esté disponible
        # Si ruta es un archivo temporal, usar subir_archivo_a_drive que maneja la diferencia
        # entre ruta local y nombre en Drive
        if subir_a_drive or subir_archivo_a_drive:
            try:
                if ruta_tmp and ruta == ruta_tmp:
                    # Archivo temporal → subir con nombre correcto en Drive
                    if subir_archivo_a_drive:
                        subir_archivo_a_drive(ruta_tmp, "BASE_DE_DATOS_PRODUCTOS.xlsx")
                else:
                    # Archivo local con nombre correcto → subir directo
                    subir_a_drive(ruta)
            except Exception:
                pass

        return {"ok": True, "fila": fila_nueva}

    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if ruta_tmp:
            try:
                os.unlink(ruta_tmp)
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────
# COLA FIFO — ESCRITURAS AL EXCEL (serializa todos los hilos)
# ─────────────────────────────────────────────────────────────────

class _ExcelWorker:
    """
    Hilo daemon que procesa actualizaciones al Excel de productos en serie.
    Elimina la condición de carrera: aunque ai.py lance N hilos de precio
    simultáneos, todos esperan su turno en esta cola.
    """
    def __init__(self):
        self._q = queue.Queue()
        t = threading.Thread(target=self._loop, daemon=True, name="excel-prod-worker")
        t.start()

    def encolar(self, nombre: str, precio: float, fraccion: Optional[str]):
        self._q.put({"nombre": nombre, "precio": precio, "fraccion": fraccion})
        log.debug("[cola] encolado: %s frac=%s precio=%s", nombre, fraccion, precio)

    def _loop(self):
        while True:
            tarea = self._q.get()
            _MAX_REINTENTOS = 3
            _ESPERA_INICIAL = 2  # segundos
            try:
                ok, msg = False, "no intentado"
                for intento in range(1, _MAX_REINTENTOS + 1):
                    try:
                        ok, msg = _escribir_en_excel(
                            tarea["nombre"], tarea["precio"], tarea["fraccion"]
                        )
                        if ok:
                            break  # éxito → salir del loop de reintentos
                        # fallo no-excepción (ej: archivo no encontrado en Drive)
                        if intento < _MAX_REINTENTOS:
                            import time as _time
                            espera = _ESPERA_INICIAL * intento
                            log.warning("[excel] reintento %d/%d en %ds: %s → %s",
                                        intento, _MAX_REINTENTOS, espera, tarea["nombre"], msg)
                            _time.sleep(espera)
                    except Exception as exc:
                        if intento < _MAX_REINTENTOS:
                            import time as _time
                            espera = _ESPERA_INICIAL * intento
                            log.warning("[excel] excepción reintento %d/%d en %ds: %s → %s",
                                        intento, _MAX_REINTENTOS, espera, tarea["nombre"], exc)
                            _time.sleep(espera)
                        else:
                            log.error("[excel] ❌ excepción tras %d intentos: %s → %s",
                                      _MAX_REINTENTOS, tarea["nombre"], exc)
                nivel = log.info if ok else log.warning
                nivel("[excel] %s %s → %s", "✅" if ok else "⚠️", tarea["nombre"], msg)
            except Exception as e:
                log.error("[excel] ❌ excepción en worker: %s", e)
            finally:
                self._q.task_done()


_worker = _ExcelWorker()


# ─────────────────────────────────────────────────────────────────
# ESCRITURA REAL AL EXCEL (ejecutada por el worker, nunca concurrente)
# ─────────────────────────────────────────────────────────────────

def _col_idx_para(fraccion: Optional[str]) -> int:
    """Retorna el índice base-0 de la columna para la fracción dada."""
    if not fraccion or fraccion == "1":
        return _IDX_UNIDAD
    info = _LABEL_MAP.get(fraccion)
    return info[1] if info else _IDX_UNIDAD


def _valor_para_celda(precio_total: float, fraccion: Optional[str], cat: str) -> float:
    """
    Calcula el valor a escribir en la celda.

    Para pinturas/impermeabilizantes la celda almacena precio UNITARIO:
      valor_celda = precio_total / decimal
    Para el resto (tornillería, ferretería) la celda almacena el valor tal cual.
    """
    if not fraccion or fraccion == "1":
        return round(precio_total)
    info = _LABEL_MAP.get(fraccion)
    if info is None:
        return round(precio_total)
    decimal_real = info[0]
    if decimal_real is None:
        # precio directo
        return round(precio_total)
    if _es_galon(cat) and decimal_real > 0:
        return round(precio_total / decimal_real)
    return round(precio_total)


def _escribir_en_excel(nombre: str, precio: float, fraccion: Optional[str]) -> tuple[bool, str]:
    """Descarga el Excel, actualiza la celda correcta y lo sube."""
    ruta_tmp   = "BASE_DE_DATOS_PRODUCTOS_tmp.xlsx"
    ruta_final = NOMBRE_EXCEL_PRODUCTOS

    try:
        from drive import descargar_de_drive, subir_a_drive_urgente
    except ImportError:
        return False, "módulo drive no disponible"

    # 1 ── Descargar
    try:
        if not descargar_de_drive(NOMBRE_EXCEL_PRODUCTOS, ruta_tmp):
            return False, "archivo no encontrado en Drive"
    except Exception as e:
        return False, f"error descargando: {e}"

    # 2 ── Abrir
    try:
        wb = openpyxl.load_workbook(ruta_tmp)
        ws = wb["Datos"]
    except Exception as e:
        _limpiar(ruta_tmp)
        return False, f"no se pudo abrir: {e}"

    col_headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]

    # 3 ── Buscar fila
    from utils import _normalizar
    nombre_norm     = _normalizar(nombre)
    fila_encontrada = None
    cat_prod        = ""

    for row in ws.iter_rows(min_row=2):
        v = row[_IDX_NOMBRE].value
        if v and _normalizar(str(v)) == nombre_norm:
            fila_encontrada = row
            cat_prod = str(row[_IDX_CATEGORIA].value or "")
            break

    if fila_encontrada is None:
        _limpiar(ruta_tmp)
        return False, f"'{nombre}' no encontrado en el Excel"

    # 4 ── Calcular columna y valor
    col_idx   = _col_idx_para(fraccion)
    val_celda = _valor_para_celda(precio, fraccion, cat_prod)

    if col_idx >= len(fila_encontrada):
        _limpiar(ruta_tmp)
        return False, f"índice de columna {col_idx} fuera de rango"

    fila_encontrada[col_idx].value = val_celda

    # 5 ── Guardar y subir (SÍNCRONO — el worker es hilo único, subida síncrona
    #      garantiza que el próximo job descargue la versión ya actualizada)
    try:
        wb.save(NOMBRE_EXCEL_PRODUCTOS)
        from drive import _ejecutar_subida_real
        _ejecutar_subida_real(NOMBRE_EXCEL_PRODUCTOS)  # síncrono, no lanza hilo
    except Exception as e:
        return False, f"error guardando/subiendo: {e}"
    finally:
        _limpiar(ruta_tmp)   # solo el temporal de descarga

    col_letra = chr(ord("A") + col_idx)
    return True, f"col {col_letra} = {val_celda:,}"


# ─────────────────────────────────────────────────────────────────
# ACTUALIZAR METADATOS EN EXCEL (nombre, categoría, unidad, código)
# ─────────────────────────────────────────────────────────────────

def _actualizar_metadatos_en_excel(nombre_original: str, datos_nuevos: dict) -> dict:
    """
    Actualiza nombre, categoría, unidad_medida y/o código de un producto
    en la hoja 'Datos' de BASE_DE_DATOS_PRODUCTOS.xlsx.
    Busca la fila por nombre_original (flexible).
    datos_nuevos puede tener: nombre, categoria, unidad_medida, codigo.
    Retorna {"ok": True} o {"ok": False, "error": "..."}.
    """
    try:
        from drive import descargar_de_drive, subir_a_drive
    except ImportError:
        descargar_de_drive = subir_a_drive = None

    import tempfile

    # Localizar el Excel de productos
    ruta = None
    for candidato in [
        getattr(config, "EXCEL_PRODUCTOS", None),
        getattr(config, "BASE_DATOS_FILE", None),
        os.path.join(os.path.dirname(config.EXCEL_FILE), "BASE_DE_DATOS_PRODUCTOS.xlsx"),
        "BASE_DE_DATOS_PRODUCTOS.xlsx",
    ]:
        if candidato and os.path.exists(candidato):
            ruta = candidato
            break

    ruta_tmp = None
    if not ruta and descargar_de_drive:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                ruta_tmp = tmp.name
            if descargar_de_drive("BASE_DE_DATOS_PRODUCTOS.xlsx", ruta_tmp):
                ruta = ruta_tmp
        except Exception:
            pass

    if not ruta or not os.path.exists(ruta):
        return {"ok": False, "error": "No se encontró el Excel de productos"}

    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active

        # Detectar índices de columna leyendo la fila de encabezados (fila 1)
        col_map = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                col_map[_norm_cat(str(cell.value))] = cell.column

        # Mapear campos del dict a los nombres de columna del Excel
        _CAMPO_COL = {
            "nombre":        ["nombre del producto", "nombre", "product name"],
            "categoria":     ["categoria", "categoría", "category"],
            "unidad_medida": ["unidad de medida", "unidad medida", "unidad", "unit"],
            "codigo":        ["codigo del producto", "código del producto", "codigo", "código", "code"],
        }

        nombre_norm = _norm_cat(nombre_original)
        fila_prod = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=_IDX_NOMBRE + 1).value  # Col B (idx 1 → col 2)
            if val and _norm_cat(str(val)) == nombre_norm:
                fila_prod = row
                break

        # Búsqueda flexible si no hay coincidencia exacta
        if not fila_prod:
            palabras = [p for p in nombre_norm.split() if len(p) > 2]
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=_IDX_NOMBRE + 1).value
                if val and palabras and all(p in _norm_cat(str(val)) for p in palabras):
                    fila_prod = row
                    break

        if not fila_prod:
            _limpiar(ruta_tmp)
            return {"ok": False, "error": f"Fila de '{nombre_original}' no encontrada en Excel"}

        actualizados = []
        for campo, valor in datos_nuevos.items():
            if not valor:
                continue
            claves_posibles = _CAMPO_COL.get(campo, [campo])
            col_idx = None
            for clave in claves_posibles:
                col_idx = col_map.get(clave)
                if col_idx:
                    break
            if col_idx:
                ws.cell(row=fila_prod, column=col_idx).value = valor.strip() if isinstance(valor, str) else valor
                actualizados.append(campo)

        if actualizados:
            wb.save(ruta)
            if subir_a_drive:
                try:
                    subir_a_drive(ruta)
                except Exception as e:
                    log.warning("_actualizar_metadatos_en_excel: Drive upload falló: %s", e)

        _limpiar(ruta_tmp)
        return {"ok": True, "actualizados": actualizados, "fila": fila_prod}

    except Exception as e:
        _limpiar(ruta_tmp)
        return {"ok": False, "error": str(e)}


# ─────────────────────────────────────────────────────────────────
# ELIMINAR PRODUCTO DEL EXCEL
# ─────────────────────────────────────────────────────────────────

def eliminar_producto_de_excel(nombre_producto: str) -> dict:
    """
    Elimina la fila del producto en BASE_DE_DATOS_PRODUCTOS.xlsx (hoja 'Datos').
    Busca por nombre exacto o flexible (igual que el sistema de búsqueda del bot).
    Retorna {"ok": True, "fila": N} o {"ok": False, "error": "..."}.
    """
    try:
        from drive import descargar_de_drive, subir_a_drive
    except ImportError:
        descargar_de_drive = subir_a_drive = None

    import shutil, tempfile

    # Localizar el Excel de productos
    ruta = None
    for candidato in [
        getattr(config, "EXCEL_PRODUCTOS", None),
        getattr(config, "BASE_DATOS_FILE", None),
        os.path.join(os.path.dirname(config.EXCEL_FILE), "BASE_DE_DATOS_PRODUCTOS.xlsx"),
        "BASE_DE_DATOS_PRODUCTOS.xlsx",
    ]:
        if candidato and os.path.exists(candidato):
            ruta = candidato
            break

    ruta_tmp = None
    if not ruta and descargar_de_drive:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                ruta_tmp = tmp.name
            ok = descargar_de_drive("BASE_DE_DATOS_PRODUCTOS.xlsx", ruta_tmp)
            if ok:
                ruta = ruta_tmp
        except Exception:
            pass

    if not ruta or not os.path.exists(ruta):
        return {"ok": False, "error": "No se encontró el Excel de productos"}

    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active

        nombre_lower = _norm_cat(nombre_producto)
        fila_borrar = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value  # Col B = Nombre
            if val and _norm_cat(str(val)) == nombre_lower:
                fila_borrar = row
                break

        # Búsqueda flexible si no hay coincidencia exacta
        if not fila_borrar:
            palabras = [p for p in nombre_lower.split() if len(p) > 2]
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=2).value
                if val and palabras and all(p in _norm_cat(str(val)) for p in palabras):
                    fila_borrar = row
                    break

        if not fila_borrar:
            _limpiar(ruta_tmp)
            return {"ok": False, "error": f"Producto '{nombre_producto}' no encontrado en el Excel"}

        ws.delete_rows(fila_borrar)
        wb.save(ruta)

        # Subir a Drive
        if subir_a_drive:
            try:
                subir_a_drive(ruta)
            except Exception as e:
                log.warning("eliminar_producto_de_excel: fallo Drive upload: %s", e)

        # Si usamos temporal, copiar de vuelta al destino canónico
        if ruta_tmp and ruta == ruta_tmp:
            destino = os.path.join(os.path.dirname(config.EXCEL_FILE), "BASE_DE_DATOS_PRODUCTOS.xlsx")
            try:
                shutil.copy2(ruta_tmp, destino)
            except Exception:
                pass

        _limpiar(ruta_tmp)
        return {"ok": True, "fila": fila_borrar}

    except Exception as e:
        _limpiar(ruta_tmp)
        return {"ok": False, "error": str(e)}


# ─────────────────────────────────────────────────────────────────
# FUNCIÓN PÚBLICA PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def actualizar_precio(
    nombre_producto: str,
    nuevo_precio: float,
    fraccion: Optional[str] = None,
) -> tuple[bool, str]:
    """
    Actualiza el precio en memoria.json (inmediato) y encola la actualización
    del Excel en Drive (asíncrono, serializado).

    Args:
        nombre_producto : nombre del producto (como en el catálogo).
        nuevo_precio    : precio total de la unidad o de la fracción.
        fraccion        : "3/4" | "1/2" | "1/4" | "1/8" | "1/16" | None/"1" para unidad.

    Returns:
        (True, descripcion) si el producto existe y se actualizó la memoria.
        (False, mensaje_error) si el producto no está en el catálogo.
    """
    from memoria import (
        actualizar_precio_en_catalogo,
        invalidar_cache_memoria,
        buscar_producto_en_catalogo,
    )

    frac = fraccion.strip() if fraccion and fraccion.strip() not in ("", "1") else None

    # 1 ── Actualizar memoria.json
    if not actualizar_precio_en_catalogo(nombre_producto, nuevo_precio, frac):
        return False, f"Producto '{nombre_producto}' no encontrado en catálogo."

    invalidar_cache_memoria()

    # 2 ── Encolar actualización del Excel
    prod = buscar_producto_en_catalogo(nombre_producto)
    nombre_oficial = prod["nombre"] if prod else nombre_producto
    _worker.encolar(nombre_oficial, nuevo_precio, frac)

    desc = nombre_oficial
    if frac:
        desc += f" {frac}"
    desc += f" = ${nuevo_precio:,.0f}"
    return True, desc


# ─────────────────────────────────────────────────────────────────
# VERIFICACIÓN DE CONSISTENCIA (bajo demanda)
# ─────────────────────────────────────────────────────────────────


def exportar_catalogo_a_excel() -> dict:
    """
    Vuelca TODOS los precios de memoria.json al Excel BASE_DE_DATOS_PRODUCTOS.xlsx
    en una sola operación: descarga una vez, actualiza todas las celdas, sube una vez.

    Reglas por categoría (idénticas al importador, sentido inverso):

      Cat 2 Pinturas / Cat 4 Impermeabilizantes:
        Col Q (idx 16) = precio_unidad
        Col R (idx 17) = precio_fraccion["3/4"].precio  / 0.75   (precio unitario)
        Col S (idx 18) = precio_fraccion["1/2"].precio  / 0.5
        Col T (idx 19) = precio_fraccion["1/4"].precio  / 0.25
        Col U (idx 20) = precio_fraccion["1/8"].precio  / 0.125
        Col V (idx 21) = precio_fraccion["1/16"].precio / 0.0625

      Cat 3 Tornillería con precio_por_cantidad:
        Col Q (idx 16) = precio_bajo_umbral   (precio unitario normal)
        Col R (idx 17) = precio_sobre_umbral  (precio mayorista >= umbral)

      Resto (ferretería, eléctricos, etc.) con precios_fraccion:
        Col Q (idx 16) = precio_unidad
        Cols R-V       = precio_fraccion[label].precio  (valor total directo)

    Retorna:
        {
          "actualizados": N,   productos con al menos una celda modificada
          "sin_match":  [...], productos en memoria no encontrados en el Excel
          "errores":    [...],
        }
    """
    ruta_tmp   = "BASE_DE_DATOS_PRODUCTOS_tmp.xlsx"
    ruta_final = NOMBRE_EXCEL_PRODUCTOS

    try:
        from drive import descargar_de_drive, subir_a_drive_urgente
    except ImportError:
        return {"actualizados": 0, "sin_match": [], "errores": ["módulo drive no disponible"]}

    # 1 ── Descargar una sola vez
    try:
        if not descargar_de_drive(NOMBRE_EXCEL_PRODUCTOS, ruta_tmp):
            return {"actualizados": 0, "sin_match": [], "errores": [f"{NOMBRE_EXCEL_PRODUCTOS} no encontrado en Drive"]}
    except Exception as e:
        return {"actualizados": 0, "sin_match": [], "errores": [f"error descargando: {e}"]}

    try:
        wb = openpyxl.load_workbook(ruta_tmp)
        ws = wb["Datos"]
    except Exception as e:
        _limpiar(ruta_tmp)
        return {"actualizados": 0, "sin_match": [], "errores": [f"no se pudo abrir: {e}"]}

    from utils import _normalizar
    from memoria import cargar_memoria

    catalogo = cargar_memoria().get("catalogo", {})

    # 2 ── Construir índice nombre_lower → fila del Excel
    filas_excel: dict[str, list] = {}   # nombre_lower → lista de celdas de la fila
    for row in ws.iter_rows(min_row=2):
        v = row[_IDX_NOMBRE].value
        if v:
            filas_excel[_normalizar(str(v))] = row

    actualizados = 0
    sin_match    = []
    errores      = []

    # 3 ── Iterar catálogo y escribir cada precio en su celda
    # nombre_lower en el JSON puede haber sido generado con la versión vieja de
    # _normalizar (reemplazos manuales, no eliminaba "°" ni otros no-ASCII).
    # Normalizamos aquí con la versión actual para que coincida con el índice.
    for clave, prod in catalogo.items():
        nombre_lower = _normalizar(prod.get("nombre_lower", "") or prod.get("nombre", clave))
        row = filas_excel.get(nombre_lower)
        if row is None:
            sin_match.append(prod.get("nombre", clave))
            continue

        cat = prod.get("categoria", "")
        modificado = False

        try:
            # ── Precio unidad (col Q) ────────────────────────────────────────
            p_unidad = prod.get("precio_unidad")
            if p_unidad:
                row[_IDX_UNIDAD].value = round(p_unidad)
                modificado = True

            # ── Tornillería: precio_por_cantidad (col R = mayorista) ─────────
            pxc = prod.get("precio_por_cantidad")
            if pxc and _es_tornilleria(cat):
                p_bajo  = pxc.get("precio_bajo_umbral")
                p_sobre = pxc.get("precio_sobre_umbral")
                if p_bajo:
                    row[_IDX_UNIDAD].value = round(p_bajo)   # col Q
                if p_sobre:
                    idx_r = _HEADER_MAP["0.75"][0]            # col R = idx 17
                    if idx_r < len(row):
                        row[idx_r].value = round(p_sobre)
                        modificado = True

            # ── Fracciones de galón (pinturas) o directas (ferretería) ───────
            fracs = prod.get("precios_fraccion", {})
            for label, frac_data in fracs.items():
                precio_total = frac_data.get("precio")
                if not precio_total:
                    continue
                info = _LABEL_MAP.get(label)
                if info is None:
                    continue
                decimal_real, col_idx = info
                if col_idx >= len(row):
                    continue
                val_celda = _valor_para_celda(precio_total, label, cat)
                row[col_idx].value = val_celda
                modificado = True

        except Exception as e:
            errores.append(f"{prod.get('nombre', clave)}: {e}")
            continue

        if modificado:
            actualizados += 1

    # 4 ── Guardar y subir una sola vez
    # subir_a_drive_urgente lanza un hilo background que lee el archivo.
    # Por eso guardamos en NOMBRE_EXCEL_PRODUCTOS y NO lo borramos en finally
    # — el hilo lo necesita vivo. Solo borramos el temporal de descarga (ruta_tmp).
    try:
        wb.save(NOMBRE_EXCEL_PRODUCTOS)
        subir_a_drive_urgente(NOMBRE_EXCEL_PRODUCTOS)
    except Exception as e:
        errores.append(f"error guardando/subiendo: {e}")
    finally:
        _limpiar(ruta_tmp)   # solo el temporal de descarga

    return {
        "actualizados": actualizados,
        "sin_match":    sin_match,
        "errores":      errores[:10],
    }

def verificar_consistencia() -> dict:
    """
    Descarga el Excel de Drive y compara precios contra memoria.json.
    Útil para detectar si hubo alguna desincronización.

    Retorna:
        {
          "iguales":      int,
          "diferentes":   [{"nombre": ..., "diffs": [...]}],
          "solo_memoria": [nombre, ...],   # en JSON pero no en Excel
          "solo_excel":   [nombre, ...],   # en Excel pero no en JSON
        }
    """
    ruta_tmp = "BASE_DE_DATOS_PRODUCTOS_check.xlsx"

    try:
        from drive import descargar_de_drive
        if not descargar_de_drive(NOMBRE_EXCEL_PRODUCTOS, ruta_tmp):
            return {"error": "No se pudo descargar el Excel de Drive"}
    except Exception as e:
        return {"error": str(e)}

    resultado = {"iguales": 0, "diferentes": [], "solo_memoria": [], "solo_excel": []}

    try:
        wb = openpyxl.load_workbook(ruta_tmp, data_only=True)
        ws = wb["Datos"]
        col_headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    except Exception as e:
        _limpiar(ruta_tmp)
        return {"error": f"No se pudo leer el Excel: {e}"}

    from utils import _normalizar
    from memoria import cargar_memoria

    catalogo = cargar_memoria().get("catalogo", {})
    excel_prods = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod = construir_producto_desde_fila(row, col_headers)
        if prod:
            excel_prods[prod["nombre_lower"]] = prod

    _limpiar(ruta_tmp)

    for nl, pm in catalogo.items():
        # Normalizar con _normalizar actual por si nombre_lower fue guardado con versión vieja
        nl_norm = _normalizar(pm.get("nombre_lower", "") or pm.get("nombre", nl))
        pe = excel_prods.get(nl_norm)
        if pe is None:
            resultado["solo_memoria"].append(pm["nombre"])
            continue

        diffs = []

        # ── Precio unidad ────────────────────────────────────────────────────
        pu_m = pm.get("precio_unidad")
        pu_x = pe.get("precio_unidad")
        if pu_m != pu_x:
            diffs.append(f"precio_unidad: mem={pu_m} xls={pu_x}")

        # ── Precio por cantidad / mayorista (tornillería) ────────────────────
        pxc_m = pm.get("precio_por_cantidad", {})
        pxc_x = pe.get("precio_por_cantidad", {})
        if pxc_m or pxc_x:
            # Comparar precio normal (bajo umbral)
            pb_m = pxc_m.get("precio_bajo_umbral") if pxc_m else None
            pb_x = pxc_x.get("precio_bajo_umbral") if pxc_x else None
            if pb_m != pb_x:
                diffs.append(f"precio_normal: mem={pb_m} xls={pb_x}")
            # Comparar precio mayorista (sobre umbral)
            ps_m = pxc_m.get("precio_sobre_umbral") if pxc_m else None
            ps_x = pxc_x.get("precio_sobre_umbral") if pxc_x else None
            if ps_m != ps_x:
                umbral = pxc_m.get("umbral") or pxc_x.get("umbral") or UMBRAL_TORNILLERIA
                diffs.append(f"precio_mayorista (x{umbral}+): mem={ps_m} xls={ps_x}")

        # ── Fracciones (pinturas, ferretería) ────────────────────────────────
        fracs_m = pm.get("precios_fraccion", {})
        fracs_x = pe.get("precios_fraccion", {})
        for lbl in set(list(fracs_m) + list(fracs_x)):
            pm_p = fracs_m.get(lbl, {}).get("precio")
            px_p = fracs_x.get(lbl, {}).get("precio")
            if pm_p != px_p:
                diffs.append(f"fraccion {lbl}: mem={pm_p} xls={px_p}")

        if diffs:
            resultado["diferentes"].append({"nombre": pm["nombre"], "diffs": diffs})
        else:
            resultado["iguales"] += 1

    # Construir set de claves normalizadas del catálogo para comparar
    claves_mem_norm = {
        _normalizar(pm.get("nombre_lower", "") or pm.get("nombre", k))
        for k, pm in catalogo.items()
    }
    for nl, pe in excel_prods.items():
        if nl not in claves_mem_norm:
            resultado["solo_excel"].append(pe["nombre"])

    return resultado


def generar_reporte_discrepancias(resultado: dict, ruta: str = "reporte_discrepancias.xlsx") -> str:
    """
    Genera un Excel con una pestaña por tipo de discrepancia.
    Recibe el dict de verificar_consistencia() o cualquier función que devuelva
    {diferentes, solo_memoria, solo_excel, sin_match (opcional)}.
    Retorna la ruta del archivo generado.

    Pestañas generadas (solo si tienen datos):
      - Diferencias de precio   : productos con precio distinto entre memoria y Excel
      - Solo en memoria         : están en el bot pero no en el Excel
      - Solo en Excel           : están en el Excel pero no en el bot
      - No encontrados en Excel : solo para exportar_precios (sin_match)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import config

    def _estilo_header(celda, color="1A1A1A"):
        celda.font      = Font(bold=True, color="FFFFFF", size=11)
        celda.fill      = PatternFill("solid", fgColor=color)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border    = Border(bottom=Side(style="thin", color="FFFFFF"))

    def _crear_hoja(wb, titulo, encabezados, filas, color_header="1A56DB"):
        ws = wb.create_sheet(title=titulo[:31])
        # Título fusionado
        ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
        c = ws.cell(row=1, column=1, value=titulo)
        c.font      = Font(bold=True, color="FFFFFF", size=13)
        c.fill      = PatternFill("solid", fgColor=color_header)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Encabezados
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=2, column=col, value=enc)
            _estilo_header(celda)
        ws.row_dimensions[2].height = 22

        # Datos
        for i, fila in enumerate(filas, 3):
            for col, val in enumerate(fila, 1):
                celda = ws.cell(row=i, column=col, value=val)
                celda.alignment = Alignment(horizontal="left", vertical="center")
                if i % 2 == 0:
                    celda.fill = PatternFill("solid", fgColor="EFF6FF")

        # Anchos automáticos
        for col in range(1, len(encabezados) + 1):
            max_len = max(
                (len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

        return ws

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M")
    hojas_creadas = 0

    # ── Pestaña 1: Diferencias de precio ────────────────────────────────────
    diferentes = resultado.get("diferentes", [])
    if diferentes:
        filas = []
        for d in diferentes:
            nombre = d["nombre"]
            for diff in d["diffs"]:
                # Formato: "campo_label: mem=VALOR xls=VALOR"
                partes  = diff.split(": ", 1)
                campo   = partes[0] if partes else diff
                valores = partes[1] if len(partes) > 1 else ""
                mem_val = valores.split(" xls=")[0].replace("mem=", "") if " xls=" in valores else ""
                xls_val = valores.split(" xls=")[1] if " xls=" in valores else ""
                # Etiqueta legible según campo
                if "fraccion" in campo:
                    tipo = f"Fracción {campo.replace('fraccion ', '').strip()}"
                elif "mayorista" in campo:
                    tipo = campo  # ya viene legible: "precio_mayorista (x50+)"
                elif "precio_normal" in campo:
                    tipo = "Precio normal (unidad)"
                else:
                    tipo = "Precio unidad"
                filas.append([nombre, tipo, mem_val, xls_val])
        _crear_hoja(wb, "Diferencias de precio",
                    ["Producto", "Campo", "Precio Memoria", "Precio Excel"],
                    filas, color_header="B45309")
        hojas_creadas += 1

    # ── Pestaña 2: Solo en memoria ───────────────────────────────────────────
    solo_mem = resultado.get("solo_memoria", [])
    if solo_mem:
        filas = [[nombre] for nombre in sorted(solo_mem)]
        _crear_hoja(wb, "Solo en memoria",
                    ["Producto (en bot, no en Excel)"],
                    filas, color_header="7C3AED")
        hojas_creadas += 1

    # ── Pestaña 3: Solo en Excel ─────────────────────────────────────────────
    solo_xls = resultado.get("solo_excel", [])
    if solo_xls:
        filas = [[nombre] for nombre in sorted(solo_xls)]
        _crear_hoja(wb, "Solo en Excel",
                    ["Producto (en Excel, no en bot)"],
                    filas, color_header="065F46")
        hojas_creadas += 1

    # ── Pestaña 4: No encontrados en Excel (exportar_precios) ───────────────
    sin_match = resultado.get("sin_match", [])
    if sin_match:
        filas = [[nombre] for nombre in sorted(sin_match)]
        _crear_hoja(wb, "No encontrados en Excel",
                    ["Producto (en bot, sin fila en Excel)"],
                    filas, color_header="9F1239")
        hojas_creadas += 1

    if hojas_creadas == 0:
        # Todo OK — crear hoja de confirmación
        ws = wb.create_sheet(title="Todo sincronizado")
        ws.cell(1, 1, f"✅ Sin discrepancias al {fecha}")
        ws.column_dimensions["A"].width = 40

    wb.save(ruta)
    return ruta
