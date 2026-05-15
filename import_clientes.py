"""
Importación masiva de clientes desde Excel → PostgreSQL.
Uso: python3 import_clientes.py [ruta_excel]

Si no se pasa ruta, busca 'clientes.xlsx' en el mismo directorio.
Requiere DATABASE_URL en el entorno (igual que el bot en Railway).
"""

import os
import sys
import openpyxl
import psycopg2
import psycopg2.extras

# ── Configuración ─────────────────────────────────────────────────────────────
ARCHIVO_DEFAULT = os.path.join(os.path.dirname(__file__), "clientes.xlsx")

# Mapeo tipo de identificación → código interno
TIPO_ID_MAP = {
    "cédula de ciudadanía": "CC",
    "cedula de ciudadania":  "CC",
    "cc":                    "CC",
    "nit":                   "NIT",
    "nuip":                  "NIT",   # NUIP es NIT para personas jurídicas
    "cédula de extranjería": "CE",
    "cedula de extranjeria": "CE",
    "ce":                    "CE",
    "pasaporte":             "PAS",
    "pas":                   "PAS",
    "tarjeta de identidad":  "TI",
    "ti":                    "TI",
}

# Teléfonos que son placeholders (no datos reales)
TELEFONOS_FALSOS = {"000-0000000-", "-0000000-", "031-0000000-", "605-0000000-", "=0-0", "--", ""}


def limpiar_telefono(val) -> str:
    if not val:
        return ""
    s = str(val).strip()
    # Quitar caracteres no numéricos para evaluar
    solo_num = "".join(c for c in s if c.isdigit())
    if s in TELEFONOS_FALSOS or len(solo_num) < 7:
        return ""
    return s


def limpiar_direccion(val) -> str:
    if not val:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("no aplica", "n/a", "na", "-") else s


def limpiar_correo(val) -> str:
    if not val:
        return ""
    s = str(val).strip().lower()
    return "" if "@" not in s else s


def mapear_tipo_id(val) -> str:
    if not val:
        return "CC"
    return TIPO_ID_MAP.get(str(val).strip().lower(), "CC")


def mapear_tipo_persona(val) -> str:
    if not val:
        return "Natural"
    s = str(val).strip().lower()
    return "Jurídica" if "jur" in s else "Natural"


def leer_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    clientes = []

    for r in range(2, ws.max_row + 1):
        nombre_raw = ws.cell(r, 1).value
        if not nombre_raw:
            continue

        nombre = str(nombre_raw).strip().upper()
        if not nombre:
            continue

        identificacion_raw = ws.cell(r, 4).value
        identificacion = str(int(identificacion_raw)) if isinstance(identificacion_raw, float) \
            else str(identificacion_raw).strip() if identificacion_raw else ""

        clientes.append({
            "nombre":         nombre,
            "tipo_persona":   mapear_tipo_persona(ws.cell(r, 2).value),
            "tipo_id":        mapear_tipo_id(ws.cell(r, 3).value),
            "identificacion": identificacion,
            "correo":         limpiar_correo(ws.cell(r, 6).value),
            "direccion":      limpiar_direccion(ws.cell(r, 7).value),
            "telefono":       limpiar_telefono(ws.cell(r, 8).value),
            "municipio_dian": 13001,  # Cartagena por defecto
        })

    return clientes


def importar(clientes: list[dict], conn) -> tuple[int, int, int]:
    """Retorna (insertados, duplicados, errores)."""
    insertados = duplicados = errores = 0

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        for c in clientes:
            try:
                # Verificar duplicado por identificación
                if c["identificacion"]:
                    cur.execute(
                        "SELECT id FROM clientes WHERE identificacion = %s LIMIT 1",
                        (c["identificacion"],)
                    )
                    if cur.fetchone():
                        print(f"  ⚠  DUPLICADO  {c['nombre']} ({c['identificacion']})")
                        duplicados += 1
                        continue

                # Insertar
                cur.execute(
                    """
                    INSERT INTO clientes
                        (nombre, tipo_id, identificacion, tipo_persona,
                         correo, telefono, direccion, municipio_dian)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        c["nombre"],
                        c["tipo_id"],
                        c["identificacion"],
                        c["tipo_persona"],
                        c["correo"],
                        c["telefono"],
                        c["direccion"],
                        c["municipio_dian"],
                    ),
                )
                print(f"  ✅ {c['nombre']} ({c['tipo_id']} {c['identificacion']})")
                insertados += 1

            except Exception as e:
                print(f"  ❌ ERROR en {c['nombre']}: {e}")
                errores += 1
                conn.rollback()
                continue

    conn.commit()
    return insertados, duplicados, errores


def main():
    archivo = sys.argv[1] if len(sys.argv) > 1 else ARCHIVO_DEFAULT

    if not os.path.exists(archivo):
        print(f"❌ No se encontró el archivo: {archivo}")
        sys.exit(1)

    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        print("❌ DATABASE_URL no está definida en el entorno.")
        sys.exit(1)

    print(f"📂 Leyendo: {archivo}")
    clientes = leer_excel(archivo)
    print(f"   → {len(clientes)} clientes encontrados en el Excel\n")

    print("🔌 Conectando a la base de datos…")
    conn = psycopg2.connect(db_url)

    print("📥 Importando clientes:\n")
    insertados, duplicados, errores = importar(clientes, conn)
    conn.close()

    print(f"""
{'─'*50}
Resultado final:
  ✅ Insertados : {insertados}
  ⚠  Duplicados : {duplicados}
  ❌ Errores    : {errores}
{'─'*50}
""")


if __name__ == "__main__":
    main()
