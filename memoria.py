"""
Memoria persistente del bot: precios, catálogo, negocio, inventario, caja, gastos.
100% PostgreSQL — sin JSON local ni Google Drive.

CORRECCIONES v4:
  - _cargar_desde_postgres() lee negocio y notas desde config_sistema PG
  - guardar_memoria() solo escribe a PG (sin json.dump a disco)
  - construir_producto_desde_fila() y constantes de columnas extraídas de precio_sync.py
  - Lazy import de db dentro de funciones para evitar importación circular
  - Firmas públicas cargar_memoria() y guardar_memoria() sin cambios (~151 referencias externas)
"""

# ═══════════════════════════════════════════════════════════════════════════════
# ⚠️  ADVERTENCIA DE ARQUITECTURA — LEER ANTES DE MODIFICAR ESTE ARCHIVO
# ═══════════════════════════════════════════════════════════════════════════════
#
# memoria.py es un THIN WRAPPER de re-export sobre los services reales.
# Existen ~151 callers de `from memoria import X` en el proyecto.
# Cambiar una firma aquí sin actualizar el service original rompe todo
# silenciosamente, sin error de importación.
#
# TABLA DE RE-EXPORTS:
#   cargar_memoria, buscar_producto_*, buscar_multiples_*,
#   obtener_precios_como_texto, obtener_info_fraccion_producto,
#   importar_catalogo_desde_excel, actualizar_precio_en_catalogo
#       → services/catalogo_service.py
#
#   cargar_inventario
#       → services/inventario_service.py
#
#   cargar_caja, guardar_gasto, obtener_resumen_caja, cargar_gastos_hoy
#       → services/caja_service.py
#
#   guardar_fiado_movimiento, abonar_fiado
#       → services/fiados_service.py
#
# MIGRACIÓN PROGRESIVA (no hacer todo de una vez):
#   Preferir importar directamente desde el service en código nuevo:
#     ✅  from services.catalogo_service import buscar_producto_en_catalogo
#     ⚠️  from memoria import buscar_producto_en_catalogo  (funciona pero oculta dependencia)
#
# ═══════════════════════════════════════════════════════════════════════════════

import logging
import json
import os
import threading
import time
from datetime import datetime

import config
from utils import _normalizar  # única definición centralizada

logger = logging.getLogger("ferrebot.memoria")

_cache: dict | None = None
_cache_ts: float | None = None          # epoch del último load completo desde PG
_CACHE_TTL: int = 600                   # segundos — recarga automática si el cache es más viejo
_bloquear_subida_drive: bool = False  # True durante la sincronización inicial
_cache_lock = threading.Lock()        # Protege _cache en entornos multi-hilo


def bloquear_subida_drive(bloquear: bool):
    global _bloquear_subida_drive
    _bloquear_subida_drive = bloquear


def _leer_catalogo_postgres(db_module) -> dict:
    """Reconstruye memoria["catalogo"] desde Postgres.
    La estructura dict es IDENTICA a la que tenia el JSON."""
    productos = db_module.query_all("SELECT * FROM productos WHERE activo = TRUE")
    fracciones = db_module.query_all("SELECT * FROM productos_fracciones")
    precios_cant = db_module.query_all("SELECT * FROM productos_precio_cantidad")
    aliases = db_module.query_all("SELECT * FROM productos_alias")

    # Indexar por producto_id para joins eficientes en Python
    frac_by_prod = {}
    for f in fracciones:
        frac_by_prod.setdefault(f["producto_id"], []).append(f)

    pxc_by_prod = {}
    for p in precios_cant:
        pxc_by_prod[p["producto_id"]] = p

    alias_by_prod = {}
    for a in aliases:
        alias_by_prod.setdefault(a["producto_id"], []).append(a["alias"])

    catalogo = {}
    for p in productos:
        prod_dict = {
            "nombre": p["nombre"],
            "nombre_lower": p["nombre_lower"],
            "categoria": p["categoria"] or "",
            "precio_unidad": p["precio_unidad"],
            "unidad_medida": p["unidad_medida"] or "Unidad",
        }
        # Codigo (si existe)
        if p.get("codigo"):
            prod_dict["codigo"] = p["codigo"]
        # Fracciones
        if p["id"] in frac_by_prod:
            prod_dict["precios_fraccion"] = {
                f["fraccion"]: {
                    "precio": f["precio_total"],
                    "precio_unitario": f["precio_unitario"],
                }
                for f in frac_by_prod[p["id"]]
            }
        # Precio por cantidad
        if p["id"] in pxc_by_prod:
            pxc = pxc_by_prod[p["id"]]
            prod_dict["precio_por_cantidad"] = {
                "umbral": pxc["umbral"],
                "precio_bajo_umbral": pxc["precio_bajo_umbral"],
                "precio_sobre_umbral": pxc["precio_sobre_umbral"],
            }
        # Alias
        if p["id"] in alias_by_prod:
            prod_dict["alias"] = alias_by_prod[p["id"]]

        catalogo[p["clave"]] = prod_dict

    return catalogo


def _leer_inventario_postgres(db_module) -> dict:
    """Reconstruye memoria["inventario"] desde Postgres."""
    rows = db_module.query_all("""
        SELECT p.clave,
               i.cantidad, i.minimo, i.unidad,
               i.nombre_original, i.costo_promedio, i.ultimo_costo,
               i.ultimo_proveedor, i.ultima_compra, i.ultima_venta,
               i.ultimo_ajuste, i.fecha_conteo
        FROM inventario i
        JOIN productos p ON p.id = i.producto_id
    """)
    inventario = {}
    for r in rows:
        entrada = {
            "cantidad":         float(r["cantidad"])       if r["cantidad"]       is not None else 0,
            "minimo":           float(r["minimo"])         if r["minimo"]         is not None else 0,
            "unidad":           r["unidad"]                or "Unidad",
            "nombre_original":  r["nombre_original"]       or "",
            "costo_promedio":   float(r["costo_promedio"]) if r["costo_promedio"] is not None else None,
            "ultimo_costo":     float(r["ultimo_costo"])   if r["ultimo_costo"]   is not None else None,
            "ultimo_proveedor": r["ultimo_proveedor"]      or "",
        }
        for ts_field in ("ultima_compra", "ultima_venta", "ultimo_ajuste", "fecha_conteo"):
            val = r[ts_field]
            entrada[ts_field] = val.strftime("%Y-%m-%d %H:%M") if val else ""
        inventario[r["clave"]] = entrada
    return inventario


def _cargar_desde_postgres() -> dict:
    """Construye el dict de memoria 100% desde Postgres — sin JSON local."""
    import db as _db
    return {
        "precios":     {},
        "catalogo":    _leer_catalogo_postgres(_db),
        "negocio":     _leer_negocio_postgres(_db),
        "notas":       _leer_notas_postgres(_db),
        "inventario":  _leer_inventario_postgres(_db),
        "gastos":      {},        # cargar_gastos_hoy() lee PG directamente
        "caja_actual": {},        # cargar_caja() lee PG directamente
    }


def _reload_cache_background() -> None:
    """
    Recarga el cache desde PostgreSQL en un hilo daemon.
    Patrón stale-while-revalidate: el request ya fue servido con el dato viejo;
    este hilo actualiza el cache para el siguiente request.
    """
    global _cache, _cache_ts
    import db as _db
    if not _db.DB_DISPONIBLE:
        return
    try:
        nuevo = _cargar_desde_postgres()
        with _cache_lock:
            _cache = nuevo
            _cache_ts = time.monotonic()
        logger.debug("[CACHE] Recargado en background (stale-while-revalidate)")
    except Exception as e:
        logger.warning(f"[CACHE] Error en recarga background: {e}")


def cargar_memoria() -> dict:
    global _cache, _cache_ts
    with _cache_lock:
        ahora = time.monotonic()

        if _cache is not None and _cache_ts is not None:
            edad = ahora - _cache_ts
            if edad < _CACHE_TTL:
                # Cache vigente → retorno inmediato
                return _cache
            # TTL expirado: servir el dato viejo AHORA y recargar en background
            # El usuario no espera; el próximo request ya tendrá el dato fresco.
            logger.debug("[CACHE] TTL expirado — sirviendo stale, recargando en background")
            import threading as _t
            _t.Thread(target=_reload_cache_background, daemon=True, name="cache-reload").start()
            return _cache  # retorno inmediato con dato stale

        # Primera carga (cache es None) — bloqueante, solo ocurre al arranque
        import db as _db
        if _db.DB_DISPONIBLE:
            _cache = _cargar_desde_postgres()
            _cache_ts = ahora
        else:
            logger.warning("DB no disponible — cargar_memoria() retorna estructura vacía")
            _cache = {
                "precios": {}, "catalogo": {}, "negocio": {},
                "notas": {}, "inventario": {}, "gastos": {},
                "caja_actual": {"abierta": False},
            }
        return _cache


def _leer_negocio_postgres(db_module) -> dict:
    """Lee config_sistema.negocio (JSON) → dict. Retorna {} si aún no existe."""
    row = db_module.query_one(
        "SELECT valor FROM config_sistema WHERE clave = 'negocio'"
    )
    if not row or not row["valor"]:
        return {}
    try:
        return json.loads(row["valor"])
    except Exception:
        return {}


def _guardar_negocio_postgres(negocio: dict, db_module) -> None:
    db_module.execute(
        """INSERT INTO config_sistema (clave, valor, updated_at)
           VALUES ('negocio', %s, NOW())
           ON CONFLICT (clave) DO UPDATE
           SET valor = EXCLUDED.valor, updated_at = NOW()""",
        (json.dumps(negocio, ensure_ascii=False),),
    )


def _leer_notas_postgres(db_module) -> dict:
    """Lee config_sistema.notas (JSON) → dict. Retorna {} si aún no existe."""
    row = db_module.query_one(
        "SELECT valor FROM config_sistema WHERE clave = 'notas'"
    )
    if not row or not row["valor"]:
        return {}
    try:
        val = json.loads(row["valor"])
        # Compatibilidad: el campo era lista en versiones antiguas del JSON
        if isinstance(val, list):
            return {"observaciones": val} if val else {}
        return val
    except Exception:
        return {}


def _guardar_notas_postgres(notas: dict, db_module) -> None:
    db_module.execute(
        """INSERT INTO config_sistema (clave, valor, updated_at)
           VALUES ('notas', %s, NOW())
           ON CONFLICT (clave) DO UPDATE
           SET valor = EXCLUDED.valor, updated_at = NOW()""",
        (json.dumps(notas, ensure_ascii=False),),
    )


def _sincronizar_catalogo_postgres(catalogo: dict, db_module):
    """Sincroniza catalogo dict a Postgres via UPSERT (D-06, D-07)."""
    for clave, prod in catalogo.items():
        row = db_module.execute_returning("""
            INSERT INTO productos (clave, nombre, nombre_lower, categoria, precio_unidad, unidad_medida)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (clave) DO UPDATE SET
                nombre = EXCLUDED.nombre,
                nombre_lower = EXCLUDED.nombre_lower,
                precio_unidad = EXCLUDED.precio_unidad,
                unidad_medida = EXCLUDED.unidad_medida,
                updated_at = NOW()
            RETURNING id
        """, (
            clave,
            prod.get("nombre", ""),
            prod.get("nombre_lower", prod.get("nombre", "").lower()),
            prod.get("categoria", ""),
            prod.get("precio_unidad", 0),
            prod.get("unidad_medida", "Unidad"),
        ))
        if not row:
            continue
        prod_id = row["id"]

        # Fracciones: delete + insert (simpler than individual UPSERTs)
        db_module.execute("DELETE FROM productos_fracciones WHERE producto_id = %s", (prod_id,))
        for frac, datos in prod.get("precios_fraccion", {}).items():
            db_module.execute("""
                INSERT INTO productos_fracciones (producto_id, fraccion, precio_total, precio_unitario)
                VALUES (%s, %s, %s, %s)
            """, (prod_id, frac, datos.get("precio", 0), datos.get("precio_unitario", 0)))

        # Precio por cantidad
        pxc = prod.get("precio_por_cantidad", {})
        if pxc:
            db_module.execute("""
                INSERT INTO productos_precio_cantidad (producto_id, umbral, precio_bajo_umbral, precio_sobre_umbral)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (producto_id) DO UPDATE SET
                    umbral = EXCLUDED.umbral,
                    precio_bajo_umbral = EXCLUDED.precio_bajo_umbral,
                    precio_sobre_umbral = EXCLUDED.precio_sobre_umbral
            """, (
                prod_id,
                pxc.get("umbral", 50),
                pxc.get("precio_bajo_umbral", prod.get("precio_unidad", 0)),
                pxc.get("precio_sobre_umbral", 0),
            ))

        # Alias
        for alias_str in prod.get("alias", []):
            if alias_str and isinstance(alias_str, str) and alias_str.strip():
                db_module.execute("""
                    INSERT INTO productos_alias (producto_id, alias)
                    VALUES (%s, %s)
                    ON CONFLICT (alias) DO NOTHING
                """, (prod_id, alias_str.strip()))


def _upsert_precio_producto_postgres(clave: str, datos_prod: dict, fraccion: str = None):
    """Upsert quirúrgico de precios de un producto.
    Solo toca las filas que cambiaron — no hace sync completo del catálogo.
    Raises si PG falla.
    """
    import db as _db
    prod_row = _db.query_one("SELECT id FROM productos WHERE clave = %s", (clave,))
    if not prod_row:
        raise ValueError(f"Producto con clave '{clave}' no existe en productos")
    prod_id = prod_row["id"]

    # Actualizar precio_unidad en productos si cambió
    _db.execute(
        "UPDATE productos SET precio_unidad = %s, updated_at = NOW() WHERE id = %s",
        (datos_prod.get("precio_unidad", 0), prod_id)
    )

    # Actualizar fracción específica si se proporcionó
    if fraccion:
        datos_frac = datos_prod.get("precios_fraccion", {}).get(fraccion, {})
        if datos_frac:
            _db.execute("""
                INSERT INTO productos_fracciones (producto_id, fraccion, precio_total, precio_unitario)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (producto_id, fraccion) DO UPDATE SET
                    precio_total    = EXCLUDED.precio_total,
                    precio_unitario = EXCLUDED.precio_unitario
            """, (
                prod_id,
                fraccion,
                datos_frac.get("precio", 0),
                datos_frac.get("precio_unitario", 0),
            ))

    # Actualizar precio_por_cantidad si existe en el producto
    pxc = datos_prod.get("precio_por_cantidad", {})
    if pxc:
        _db.execute("""
            INSERT INTO productos_precio_cantidad
                (producto_id, umbral, precio_bajo_umbral, precio_sobre_umbral)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (producto_id) DO UPDATE SET
                precio_bajo_umbral  = EXCLUDED.precio_bajo_umbral,
                precio_sobre_umbral = EXCLUDED.precio_sobre_umbral
        """, (
            prod_id,
            pxc.get("umbral", 50),
            pxc.get("precio_bajo_umbral", datos_prod.get("precio_unidad", 0)),
            pxc.get("precio_sobre_umbral", 0),
        ))


def _sincronizar_inventario_postgres(inventario: dict, db_module):
    """DEPRECADO — usar _upsert_inventario_producto_postgres para escrituras."""
    for clave, datos in inventario.items():
        _upsert_inventario_producto_postgres(clave, datos)


def _upsert_inventario_producto_postgres(clave: str, datos: dict):
    """Upsert quirúrgico de un solo producto en inventario. Raises si PG falla."""
    import db as _db

    def _parse_ts(val):
        """Convierte string 'YYYY-MM-DD HH:MM' a datetime, o None si vacío."""
        if not val:
            return None
        try:
            return datetime.strptime(val, "%Y-%m-%d %H:%M")
        except (ValueError, TypeError):
            return None

    prod_row = _db.query_one("SELECT id FROM productos WHERE clave = %s", (clave,))
    if not prod_row:
        raise ValueError(f"Producto con clave '{clave}' no existe en productos")
    _db.execute("""
        INSERT INTO inventario (
            producto_id, cantidad, minimo, unidad,
            nombre_original, costo_promedio, ultimo_costo, ultimo_proveedor,
            ultima_compra, ultima_venta, ultimo_ajuste, fecha_conteo,
            updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON CONFLICT (producto_id) DO UPDATE SET
            cantidad        = EXCLUDED.cantidad,
            minimo          = EXCLUDED.minimo,
            unidad          = EXCLUDED.unidad,
            nombre_original = EXCLUDED.nombre_original,
            costo_promedio  = EXCLUDED.costo_promedio,
            ultimo_costo    = EXCLUDED.ultimo_costo,
            ultimo_proveedor= EXCLUDED.ultimo_proveedor,
            ultima_compra   = EXCLUDED.ultima_compra,
            ultima_venta    = EXCLUDED.ultima_venta,
            ultimo_ajuste   = EXCLUDED.ultimo_ajuste,
            fecha_conteo    = EXCLUDED.fecha_conteo,
            updated_at      = NOW()
    """, (
        prod_row["id"],
        datos.get("cantidad", 0),
        datos.get("minimo", 0),
        datos.get("unidad", "Unidad"),
        datos.get("nombre_original") or None,
        datos.get("costo_promedio")  or None,
        datos.get("ultimo_costo")    or None,
        datos.get("ultimo_proveedor") or None,
        _parse_ts(datos.get("ultima_compra")),
        _parse_ts(datos.get("ultima_venta")),
        _parse_ts(datos.get("ultimo_ajuste")),
        _parse_ts(datos.get("fecha_conteo")),
    ))


def guardar_memoria(memoria: dict, urgente: bool = False):
    """
    Persiste memoria en Postgres — sin escritura a JSON ni a Drive (100% PG).
    urgente=True: parámetro mantenido por compatibilidad con callers existentes.
    """
    global _cache, _cache_ts
    with _cache_lock:
        _cache = memoria
        _cache_ts = time.monotonic()
    import db as _db
    if _db.DB_DISPONIBLE:
        # Catálogo — D-06: inventario ya no se sincroniza aquí (upsert directo)
        try:
            _sincronizar_catalogo_postgres(memoria.get("catalogo", {}), _db)
        except Exception as e:
            logger.warning("Error sincronizando catalogo a Postgres (no critico): %s", e)
        # Negocio
        try:
            _guardar_negocio_postgres(memoria.get("negocio", {}), _db)
        except Exception as e:
            logger.warning("Error guardando negocio a Postgres (no critico): %s", e)
        # Notas
        try:
            notas = memoria.get("notas", {})
            if isinstance(notas, list):
                notas = {"observaciones": notas} if notas else {}
            _guardar_notas_postgres(notas, _db)
        except Exception as e:
            logger.warning("Error guardando notas a Postgres (no critico): %s", e)


def invalidar_cache_memoria():
    global _cache, _cache_ts
    with _cache_lock:
        _cache = None
        _cache_ts = None
    # Reconstruir índice fuzzy para que productos nuevos sean encontrables
    try:
        from fuzzy_match import construir_indice
        mem = cargar_memoria()
        construir_indice(mem.get("catalogo", {}))
    except Exception:
        pass


# ─────────────────────────────────────────────
# RE-EXPORTS DESDE SERVICES/
# (Tarea H: memoria.py como thin wrapper)
# ─────────────────────────────────────────────

from services.catalogo_service import (
    buscar_producto_en_catalogo,
    buscar_multiples_en_catalogo,
    expandir_con_alias,
    buscar_multiples_con_alias,
    obtener_precio_para_cantidad,
    obtener_precios_como_texto,
    obtener_info_fraccion_producto,
    actualizar_precio_en_catalogo,
)

from services.inventario_service import (
    cargar_inventario,
    guardar_inventario,
    verificar_alertas_inventario,
    buscar_clave_inventario,
    registrar_conteo_inventario,
    ajustar_inventario,
    descontar_inventario,
    buscar_productos_inventario,
)

from services.caja_service import (
    cargar_caja,
    guardar_caja,
    obtener_resumen_caja,
    cargar_gastos_hoy,
    guardar_gasto,
)

from services.fiados_service import (
    cargar_fiados,
    guardar_fiado_movimiento,
    abonar_fiado,
    resumen_fiados,
    detalle_fiado_cliente,
)

def registrar_compra(nombre_producto: str, cantidad: float, costo_unitario: float, proveedor: str = None, incluye_iva: bool = False, tarifa_iva: int = 0) -> tuple[bool, str, dict]:
    """
    Registra una compra de mercancía:
    - Suma cantidad al inventario
    - Actualiza costo promedio ponderado
    - Guarda proveedor (usa el último si no se especifica)
    - Registra en historial de compras
    Retorna (éxito, mensaje, datos_para_excel).
    """
    inventario = cargar_inventario()
    
    # Buscar producto en catálogo
    producto_catalogo = buscar_producto_en_catalogo(nombre_producto)
    if producto_catalogo:
        nombre_oficial = producto_catalogo.get("nombre", nombre_producto)
    else:
        nombre_oficial = nombre_producto.strip()
    
    # Buscar si ya existe en inventario
    clave = buscar_clave_inventario(nombre_producto)
    if not clave:
        clave = _normalizar_clave_inventario(nombre_oficial)
    
    # Obtener datos actuales o crear nuevo
    datos = inventario.get(clave, {})
    if not isinstance(datos, dict):
        datos = {}
    
    cantidad_anterior = datos.get("cantidad", 0)
    costo_anterior = datos.get("costo_promedio", costo_unitario)
    
    # Determinar proveedor: usar el especificado o el último registrado
    if proveedor:
        proveedor_final = proveedor.strip()
    else:
        proveedor_final = datos.get("ultimo_proveedor", "—")
    
    # Calcular costo promedio ponderado
    # (cantidad_anterior * costo_anterior + cantidad_nueva * costo_nuevo) / cantidad_total
    cantidad_total = cantidad_anterior + cantidad
    if cantidad_total > 0:
        costo_promedio = round(
            (cantidad_anterior * costo_anterior + cantidad * costo_unitario) / cantidad_total
        )
    else:
        costo_promedio = costo_unitario
    
    # Actualizar inventario
    datos["nombre_original"] = nombre_oficial
    datos["cantidad"] = round(cantidad_total, 4)
    datos["costo_promedio"] = costo_promedio
    datos["ultima_compra"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    datos["ultimo_costo"] = costo_unitario
    datos["ultimo_proveedor"] = proveedor_final
    if "minimo" not in datos:
        datos["minimo"] = 5
    if "unidad" not in datos:
        datos["unidad"] = "unidades"
    
    guardar_inventario(clave, datos)
    
    # Registrar en historial de compras (para reportes)
    _registrar_historial_compra(nombre_oficial, cantidad, costo_unitario, proveedor_final, incluye_iva=incluye_iva, tarifa_iva=tarifa_iva)
    
    total_compra = round(cantidad * costo_unitario)
    
    proveedor_txt = f"   🏪 Proveedor: {proveedor_final}\n" if proveedor_final != "—" else ""
    
    mensaje = (
        f"✅ Compra registrada:\n"
        f"   📦 +{cantidad} {nombre_oficial}\n"
        f"   📊 Inventario: {cantidad_anterior} → {cantidad_total}\n"
        f"   💰 Costo: ${costo_unitario:,.0f} c/u (${total_compra:,.0f} total)\n"
        f"{proveedor_txt}"
        f"   📈 Costo promedio: ${costo_promedio:,.0f}"
    )
    
    datos_excel = {
        "producto": nombre_oficial,
        "cantidad": cantidad,
        "costo_unitario": costo_unitario,
        "costo_total": total_compra,
        "proveedor": proveedor_final,
    }
    
    return True, mensaje, datos_excel


def _registrar_historial_compra(producto: str, cantidad: float, costo_unitario: float, proveedor: str = "—", incluye_iva: bool = False, tarifa_iva: int = 0):
    """Persiste la compra en PostgreSQL (fuente única de verdad)."""
    import db as _db
    if not _db.DB_DISPONIBLE:
        logger.warning("DB no disponible — historial_compra no registrado para: %s", producto)
        return
    ahora = datetime.now(config.COLOMBIA_TZ)
    # Intentar buscar producto_id y auto-detectar IVA del catálogo si no se especificó
    prod_row = _db.query_one(
        "SELECT id, tiene_iva, porcentaje_iva FROM productos WHERE LOWER(nombre) = LOWER(%s) OR LOWER(nombre_lower) = LOWER(%s) LIMIT 1",
        (producto, producto)
    )
    prod_id       = prod_row["id"]            if prod_row else None
    iva_final     = incluye_iva
    tarifa_final  = tarifa_iva
    # Si no se mandó IVA explícito pero el producto está en catálogo con IVA, auto-detectar
    if not incluye_iva and prod_row and prod_row.get("tiene_iva"):
        iva_final    = True
        tarifa_final = int(prod_row.get("porcentaje_iva") or 0)

    _db.execute(
        """INSERT INTO compras
           (fecha, hora, proveedor, producto_id, producto_nombre,
            cantidad, costo_unitario, costo_total, incluye_iva, tarifa_iva)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (ahora.strftime("%Y-%m-%d"),
         ahora.strftime("%H:%M"),
         proveedor, prod_id, producto, cantidad,
         int(costo_unitario), round(cantidad * costo_unitario),
         iva_final, tarifa_final)
    )


def obtener_costo_producto(nombre_producto: str) -> float | None:
    """Obtiene el costo promedio de un producto del inventario."""
    clave = buscar_clave_inventario(nombre_producto)
    if not clave:
        return None
    
    inventario = cargar_inventario()
    datos = inventario.get(clave, {})
    return datos.get("costo_promedio")


def calcular_margen(nombre_producto: str, precio_venta: float) -> dict | None:
    """
    Calcula el margen de ganancia de un producto.
    Retorna dict con costo, margen_valor, margen_porcentaje o None.
    """
    costo = obtener_costo_producto(nombre_producto)
    if costo is None or costo == 0:
        return None
    
    margen_valor = precio_venta - costo
    margen_porcentaje = round((margen_valor / precio_venta) * 100, 1)
    
    return {
        "costo": costo,
        "precio_venta": precio_venta,
        "margen_valor": margen_valor,
        "margen_porcentaje": margen_porcentaje,
    }


def obtener_resumen_margenes(limite: int = 20) -> list[dict]:
    """
    Obtiene productos con su margen calculado.
    Requiere que el producto tenga costo_promedio y precio en catálogo.
    """
    inventario = cargar_inventario()
    catalogo = cargar_memoria().get("catalogo", {})
    
    resultados = []
    
    for clave, datos in inventario.items():
        if not isinstance(datos, dict):
            continue
        
        costo = datos.get("costo_promedio")
        if not costo:
            continue
        
        nombre = datos.get("nombre_original", clave)
        
        # Buscar precio de venta en catálogo
        producto_cat = buscar_producto_en_catalogo(nombre)
        if not producto_cat:
            continue
        
        precio_venta = producto_cat.get("precio_unidad", 0)
        if not precio_venta or precio_venta <= costo:
            continue
        
        margen_valor = precio_venta - costo
        margen_porcentaje = round((margen_valor / precio_venta) * 100, 1)
        
        resultados.append({
            "nombre": nombre,
            "costo": costo,
            "precio_venta": precio_venta,
            "margen_valor": margen_valor,
            "margen_porcentaje": margen_porcentaje,
            "cantidad": datos.get("cantidad", 0),
        })
    
    # Ordenar por margen porcentaje descendente
    resultados.sort(key=lambda x: -x["margen_porcentaje"])
    return resultados[:limite]



# ─────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DE PRODUCTO DESDE FILA EXCEL
# (extraído de precio_sync.py — sin dependencia de ese módulo)
# ─────────────────────────────────────────────────────────────────

_UMBRAL_TORNILLERIA = 50
_IDX_CODIGO        = 0
_IDX_NOMBRE        = 1
_IDX_CATEGORIA     = 3
_IDX_UNIDAD_MEDIDA = 8   # Col I — Unidad de Medida (DIAN)
_IDX_UNIDAD        = 16  # Col Q — Precio por unidad completa

# header_str → (col_idx_base0, decimal_real, label)
_HEADER_MAP: dict = {
    "0.75": (17, 0.75,   "3/4"),
    "0.5":  (18, 0.5,    "1/2"),
    "0.25": (19, 0.25,   "1/4"),
    "0.13": (20, 0.125,  "1/8"),
    "0.06": (21, 0.0625, "1/16"),
    "0.1":  (22, 0.1,    "1/10"),
}

_CATS_GALON = {
    "2 pinturas y disolventes",
    "4 impermeabilizantes y materiales de construccion",
    "4 impermeabilizantes y materiales de construcción",
}
_CATS_TORNILLERIA = {"3 tornilleria", "3 tornillería"}

_UNIDAD_MAP: dict = {
    "galon": "Galón", "galón": "Galón", "gal": "Galón",
    "kg": "Kg", "kgs": "Kg", "kilo": "Kg", "kilos": "Kg",
    "kilogramo": "Kg", "25 kg": "Kg",
    "mts": "Mts", "mt": "Mts", "metro": "Mts", "metros": "Mts", "m": "Mts",
    "cms": "Cms", "cm": "Cms", "centimetro": "Cms",
    "lt": "Lt", "lts": "Lts", "litro": "Lt", "litros": "Lts",
    "ml": "MLT", "mlt": "MLT", "mililitro": "MLT", "mililitros": "MLT",
    "cc": "MLT", "centimetro cubico": "MLT",
    "unidad": "Unidad", "und": "Unidad", "un": "Unidad",
    "unidades": "Unidad", "uni": "Unidad",
}


def _norm_cat(cat: str) -> str:
    return (
        (cat or "").lower()
        .replace("á","a").replace("é","e").replace("í","i")
        .replace("ó","o").replace("ú","u").replace("ñ","n")
        .strip()
    )

def _es_galon(cat: str) -> bool:
    return _norm_cat(cat) in _CATS_GALON

def _es_tornilleria(cat: str) -> bool:
    return _norm_cat(cat) in _CATS_TORNILLERIA

def _num_excel(v):
    """Celda Excel → float positivo, o None."""
    if v is None:
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None

def _normalizar_unidad_excel(raw: str) -> str:
    if not raw:
        return "Unidad"
    clave = raw.strip().lower().replace("á","a").replace("é","e").replace("ó","o")
    return _UNIDAD_MAP.get(clave, raw.strip()) or "Unidad"


def construir_producto_desde_fila(row: tuple, col_headers: list):
    """
    Convierte una fila del Excel BASE_DE_DATOS_PRODUCTOS en un dict de catálogo.
    Retorna None si la fila no tiene nombre o precio válido.
    Campo \"decimal\" SIEMPRE presente en precios_fraccion.
    """
    nombre = str(row[_IDX_NOMBRE] or "").strip()
    if not nombre or nombre.lower() == "nan":
        return None

    cat      = str(row[_IDX_CATEGORIA] or "").strip()
    codigo   = str(row[_IDX_CODIGO] or "").strip()
    p_unidad = _num_excel(row[_IDX_UNIDAD]) if _IDX_UNIDAD < len(row) else None

    unidad_raw    = str(row[_IDX_UNIDAD_MEDIDA] or "").strip() if len(row) > _IDX_UNIDAD_MEDIDA else ""
    unidad_medida = _normalizar_unidad_excel(unidad_raw)

    if p_unidad is None:
        return None

    nombre_lower = _normalizar(nombre)
    prod = {
        "nombre":        nombre,
        "nombre_lower":  nombre_lower,
        "categoria":     cat,
        "precio_unidad": round(p_unidad),
        "unidad_medida": unidad_medida,
    }
    if codigo:
        prod["codigo"] = codigo

    if _es_galon(cat):
        fracs = {}
        for i, header in enumerate(col_headers):
            if i == _IDX_UNIDAD:
                continue
            info = _HEADER_MAP.get(str(header).strip())
            if not info:
                continue
            _, decimal_real, label = info
            v = _num_excel(row[i]) if i < len(row) else None
            if v is None:
                continue
            fracs[label] = {"precio": round(v * decimal_real), "decimal": decimal_real}
        if fracs:
            prod["precios_fraccion"] = fracs

    elif _es_tornilleria(cat):
        idx_r = _HEADER_MAP["0.75"][0]
        p_may = _num_excel(row[idx_r]) if idx_r < len(row) else None
        if p_may is not None and round(p_may) != round(p_unidad):
            prod["precio_por_cantidad"] = {
                "umbral":              _UMBRAL_TORNILLERIA,
                "precio_bajo_umbral":  round(p_unidad),
                "precio_sobre_umbral": round(p_may),
            }

    else:
        fracs = {}
        for i, header in enumerate(col_headers):
            if i == _IDX_UNIDAD:
                continue
            info = _HEADER_MAP.get(str(header).strip().lower()) or _HEADER_MAP.get(str(header).strip())
            if not info:
                continue
            _, decimal_real, label = info
            v = _num_excel(row[i]) if i < len(row) else None
            if v is None:
                continue
            if decimal_real is not None and v < p_unidad:
                fracs[label] = {"precio": round(v), "decimal": decimal_real}
        if fracs:
            prod["precios_fraccion"] = fracs

    return prod


def importar_catalogo_desde_excel(ruta_excel: str) -> dict:
    """
    Lee BASE_DE_DATOS_PRODUCTOS.xlsx e importa todos los productos directamente a
    PostgreSQL. NO escribe a JSON ni a Drive.

    Reglas de construcción (misma lógica que precio_sync.construir_producto_desde_fila):
      - Campo "decimal" SIEMPRE presente en precios_fraccion.
      - Pinturas/Impermeabilizantes (Cat 2/4): total = col_value × decimal_real.
      - Tornillería (Cat 3): precio_por_cantidad con umbral=50.
      - Resto: precios_fraccion directos (valor celda = total si < precio_unidad).

    Robustez: usa UPSERT — nunca elimina productos añadidos vía dashboard.
    Tras el sync, invalida el cache para que el próximo cargar_memoria() lea
    TODOS los productos activos de PG (Excel + dashboard).

    Retorna {"importados": N, "omitidos": N, "errores": [...]}
    """
    import re as _re
    import db as _db
    try:
        import openpyxl as _openpyxl
    except ImportError:
        return {"importados": 0, "omitidos": 0, "errores": ["openpyxl no disponible"]}

    # construir_producto_desde_fila definida en este mismo módulo (sin precio_sync)
    _construir = construir_producto_desde_fila

    if not _db.DB_DISPONIBLE:
        raise RuntimeError("⚠️ Base de datos no disponible — importar_catalogo_desde_excel requiere PostgreSQL.")

    try:
        wb = _openpyxl.load_workbook(ruta_excel, data_only=True)
        ws = wb["Datos"]
    except Exception as e:
        return {"importados": 0, "omitidos": 0, "errores": [str(e)]}

    col_headers = [
        str(ws.cell(1, c).value or "")
        for c in range(1, ws.max_column + 1)
    ]

    catalogo   = {}   # sólo los productos del Excel (para el UPSERT)
    importados = 0
    omitidos   = 0
    errores    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            prod = _construir(row, col_headers)
            if prod is None:
                omitidos += 1
                continue
            clave = _re.sub('[^\\w\\s/]', '', prod["nombre_lower"]).replace(" ", "_")
            clave = _re.sub(r'_+', '_', clave).strip('_')
            catalogo[clave] = prod
            importados += 1
        except Exception as e:
            nombre_raw = row[1] if row and len(row) > 1 else "?"
            errores.append(f"{nombre_raw}: {e}")

    # ── UPSERT a PG — nunca hace DELETE, productos del dashboard sobreviven ──
    try:
        _sincronizar_catalogo_postgres(catalogo, _db)
    except Exception as e:
        errores.append(f"Error PG sync: {e}")

    # ── Invalidar cache: próximo cargar_memoria() lee PG completo ────────────
    # No pisamos _cache aquí con solo los productos del Excel —
    # invalidar_cache_memoria() limpia y reconstruye desde PG incluyendo
    # productos añadidos por el dashboard que no están en el Excel.
    invalidar_cache_memoria()

    logger.info("[importar_excel] ✅ %d importados, %d omitidos, %d errores",
                importados, omitidos, len(errores))
    return {"importados": importados, "omitidos": omitidos, "errores": errores[:10]}


# ─────────────────────────────────────────────
# SINCRONIZACIÓN DE PRECIO → PostgreSQL (sin Excel/Drive)
# ─────────────────────────────────────────────

def actualizar_precio_en_excel_drive(
    nombre_producto: str,
    nuevo_precio: float,
    fraccion: str = None,
) -> tuple[bool, str]:
    """
    Actualiza el precio en PostgreSQL + cache. Nombre mantenido por
    compatibilidad con callers existentes — ya NO escribe al Excel de Drive.

    Delega a actualizar_precio_en_catalogo() que hace el upsert en PG de
    forma atómica (precio_unidad, fracciones, precio_por_cantidad).

    Returns:
        (True, descripción)  — producto encontrado y precio actualizado.
        (False, msg_error)   — producto no encontrado en catálogo.
    """
    frac = fraccion.strip() if fraccion and fraccion.strip() not in ("", "1") else None

    ok = actualizar_precio_en_catalogo(nombre_producto, nuevo_precio, frac)
    if not ok:
        return False, f"Producto '{nombre_producto}' no encontrado en catálogo."

    invalidar_cache_memoria()

    prod = buscar_producto_en_catalogo(nombre_producto)
    nombre_oficial = prod["nombre"] if prod else nombre_producto
    desc = nombre_oficial + (f" {frac}" if frac else "") + f" = ${nuevo_precio:,.0f}"
    return True, desc


# ─────────────────────────────────────────────────────────────────────────────
# CUENTAS POR PAGAR (facturas de proveedores)
# ─────────────────────────────────────────────────────────────────────────────

def _siguiente_id_factura() -> str:
    """Genera el próximo ID secuencial desde PostgreSQL: FAC-001, FAC-002, ..."""
    import db as _db
    import re as _re
    rows = _db.query_all("SELECT id FROM facturas_proveedores", ())
    nums = []
    for r in rows:
        m = _re.match(r"FAC-(\d+)", r["id"])
        if m:
            nums.append(int(m.group(1)))
    siguiente = max(nums) + 1 if nums else 1
    return f"FAC-{siguiente:03d}"


def registrar_factura_proveedor(
    proveedor: str,
    descripcion: str,
    total: float,
    fecha: str = None,
    foto_url: str = "",
    foto_nombre: str = "",
) -> dict:
    """
    Registra una nueva factura de proveedor en memoria.json.
    Retorna el dict de la factura creada.
    """
    import db as _db
    from datetime import datetime as _dt
    if not _db.DB_DISPONIBLE:
        raise RuntimeError("⚠️ Base de datos no disponible. Intenta de nuevo en un momento.")

    fac_id = _siguiente_id_factura()
    hoy    = fecha or _dt.now(import_config_tz()).strftime("%Y-%m-%d")

    factura = {
        "id":          fac_id,
        "proveedor":   proveedor.strip(),
        "descripcion": descripcion.strip(),
        "total":       float(total),
        "pagado":      0.0,
        "pendiente":   float(total),
        "estado":      "pendiente",   # pendiente | parcial | pagada
        "fecha":       hoy,
        "foto_url":    foto_url,
        "foto_nombre": foto_nombre,
        "abonos":      [],             # historial de abonos
    }
    _db.execute(
        """INSERT INTO facturas_proveedores
           (id, proveedor, descripcion, total, pagado, pendiente, estado, fecha, foto_url, foto_nombre)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
           ON CONFLICT (id) DO NOTHING""",
        (fac_id, proveedor.strip(), descripcion.strip(),
         int(float(total)), 0, int(float(total)), "pendiente", hoy,
         foto_url, foto_nombre)
    )
    return factura


def registrar_abono_factura(
    fac_id: str,
    monto: float,
    fecha: str = None,
    foto_url: str = "",
    foto_nombre: str = "",
) -> dict:
    """
    Registra un abono a una factura existente.
    Actualiza pagado/pendiente/estado.
    Retorna {"ok": True/False, "factura": {...}, "error": "..."}
    """
    import db as _db
    from datetime import datetime as _dt
    if not _db.DB_DISPONIBLE:
        raise RuntimeError("⚠️ Base de datos no disponible. Intenta de nuevo en un momento.")

    # Leer factura desde Postgres
    row = _db.query_one(
        "SELECT * FROM facturas_proveedores WHERE id = %s",
        (fac_id.upper(),)
    )
    if not row:
        return {"ok": False, "error": f"Factura {fac_id} no encontrada"}

    hoy      = fecha or _dt.now(import_config_tz()).strftime("%Y-%m-%d")
    pagado   = int(row["pagado"]) + int(float(monto))
    total    = int(row["total"])
    pendiente = max(total - pagado, 0)
    estado   = "pagada" if pendiente <= 0 else ("parcial" if pagado > 0 else "pendiente")

    _db.execute(
        """INSERT INTO facturas_abonos (factura_id, monto, fecha, foto_url, foto_nombre)
           VALUES (%s, %s, %s, %s, %s)""",
        (fac_id.upper(), int(float(monto)), hoy, foto_url, foto_nombre)
    )
    _db.execute(
        """UPDATE facturas_proveedores
           SET pagado=%s, pendiente=%s, estado=%s
           WHERE id=%s""",
        (pagado, pendiente, estado, fac_id.upper())
    )

    # Registrar el gasto del abono
    guardar_gasto({
        "concepto":  f"Abono {fac_id} - {row['proveedor']}",
        "monto":     float(monto),
        "categoria": "abono_proveedor",
        "origen":    "proveedor",
        "hora":      _dt.now(import_config_tz()).strftime("%H:%M"),
        "fac_id":    fac_id,
    })

    factura = {
        "id":          fac_id.upper(),
        "proveedor":   row["proveedor"],
        "descripcion": row.get("descripcion", ""),
        "total":       total,
        "pagado":      pagado,
        "pendiente":   pendiente,
        "estado":      estado,
        "fecha":       str(row["fecha"]),
    }
    return {"ok": True, "factura": factura}


def import_config_tz():
    """Helper para no importar config a nivel de módulo aquí."""
    try:
        import config as _c
        return _c.COLOMBIA_TZ
    except Exception:
        import pytz
        return pytz.timezone("America/Bogota")


def listar_facturas(solo_pendientes: bool = False) -> list:
    """Retorna la lista de facturas desde PostgreSQL.
    Lanza RuntimeError si la base de datos no está disponible.
    """
    import db as _db
    if not _db.DB_DISPONIBLE:
        raise RuntimeError(
            "Base de datos no disponible — listar_facturas requiere PostgreSQL."
        )
    rows = _db.query_all(
        """SELECT fp.id, fp.proveedor, fp.descripcion, fp.total, fp.pagado,
                  fp.pendiente, fp.estado, fp.fecha::text, fp.foto_url, fp.foto_nombre,
                  COALESCE(
                      json_agg(
                          json_build_object(
                              'fecha', fa.fecha::text,
                              'monto', fa.monto,
                              'foto_url', COALESCE(fa.foto_url, ''),
                              'foto_nombre', COALESCE(fa.foto_nombre, '')
                          ) ORDER BY fa.created_at
                      ) FILTER (WHERE fa.id IS NOT NULL),
                      '[]'
                  ) AS abonos
           FROM facturas_proveedores fp
           LEFT JOIN facturas_abonos fa ON fa.factura_id = fp.id
           GROUP BY fp.id
           ORDER BY fp.fecha DESC""",
        (),
    )
    facturas = [
        {**dict(r), "abonos": r["abonos"] if r["abonos"] else []}
        for r in rows
    ]
    if solo_pendientes:
        return [f for f in facturas if f["estado"] != "pagada"]
    return facturas
