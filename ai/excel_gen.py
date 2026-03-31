"""
ai/excel_gen.py — Generación y edición de Excel con Claude.
Extraído de ai.py (Tarea G).

Imports: openpyxl (inside functions), config, asyncio, json.
No imports de db ni memoria a nivel de módulo (PRM-04).

Modelo de seguridad
-------------------
Claude nunca genera código ejecutable. En su lugar elige UNA operación de
OPERACIONES_DISPONIBLES y devuelve un dict {"operacion": ..., "params": {...}}.
ejecutar_operacion_excel() aplica la operación usando funciones Python puras.
El exec() fue eliminado completamente.
"""

# -- stdlib --
import asyncio
import json
import logging

# -- propios --
import config

logger = logging.getLogger("ferrebot.ai.excel_gen")


# ── Utilidades internas ───────────────────────────────────────────────────────

def _col_por_nombre(ws, nombre: str) -> int | None:
    """Retorna el índice de columna (1-based) para un nombre de encabezado."""
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v and str(v).strip().lower() == nombre.strip().lower():
            return col
    return None


def _comparar(valor_celda, operador: str, valor_ref) -> bool:
    """Evalúa valor_celda <operador> valor_ref."""
    try:
        vc = float(str(valor_celda).replace(",", ".")) if valor_celda is not None else 0.0
        vr = float(str(valor_ref).replace(",", "."))
        if operador == "gt":  return vc > vr
        if operador == "lt":  return vc < vr
        if operador == "gte": return vc >= vr
        if operador == "lte": return vc <= vr
    except (ValueError, TypeError):
        pass
    vc_str = str(valor_celda or "").lower()
    vr_str = str(valor_ref or "").lower()
    if operador in ("eq", "=", "=="):   return vc_str == vr_str
    if operador in ("ne", "!=", "<>"): return vc_str != vr_str
    if operador == "contiene":          return vr_str in vc_str
    return False


# ── Operaciones predefinidas ──────────────────────────────────────────────────
# Cada función recibe (ws, params: dict) y modifica ws en memoria.
# Retorna un string descriptivo del resultado.

def _op_calcular_total(ws, params: dict) -> str:
    """Suma una columna numérica y escribe el total al final."""
    columna = params.get("columna", "")
    col_idx = _col_por_nombre(ws, columna)
    if col_idx is None:
        return f"Columna '{columna}' no encontrada"
    total = 0.0
    for fila in range(2, ws.max_row + 1):
        v = ws.cell(row=fila, column=col_idx).value
        try:
            total += float(str(v).replace(",", ".")) if v is not None else 0.0
        except (ValueError, TypeError):
            pass
    fila_res = ws.max_row + 1
    etiqueta = params.get("etiqueta", "TOTAL")
    etiqueta_col = col_idx - 1 if col_idx > 1 else col_idx
    ws.cell(row=fila_res, column=etiqueta_col, value=etiqueta)
    ws.cell(row=fila_res, column=col_idx, value=total)
    return f"Total de '{columna}': {total:,.2f}"


def _op_filtrar_filas(ws, params: dict) -> str:
    """Conserva solo las filas que cumplen la condición (elimina las que no)."""
    col_idx = _col_por_nombre(ws, params.get("columna_filtro", ""))
    if col_idx is None:
        return f"Columna '{params.get('columna_filtro')}' no encontrada"
    operador  = params.get("operador", "eq")
    valor_ref = params.get("valor_filtro", "")
    filas_borrar = [
        fila for fila in range(2, ws.max_row + 1)
        if not _comparar(ws.cell(row=fila, column=col_idx).value, operador, valor_ref)
    ]
    for fila in reversed(filas_borrar):
        ws.delete_rows(fila)
    return f"Filtrado: {len(filas_borrar)} filas eliminadas, quedan {ws.max_row - 1}"


def _op_eliminar_filas(ws, params: dict) -> str:
    """Elimina las filas que SÍ cumplen la condición."""
    col_idx = _col_por_nombre(ws, params.get("columna_filtro", ""))
    if col_idx is None:
        return f"Columna '{params.get('columna_filtro')}' no encontrada"
    operador  = params.get("operador", "eq")
    valor_ref = params.get("valor_filtro", "")
    filas_borrar = [
        fila for fila in range(2, ws.max_row + 1)
        if _comparar(ws.cell(row=fila, column=col_idx).value, operador, valor_ref)
    ]
    for fila in reversed(filas_borrar):
        ws.delete_rows(fila)
    return f"Eliminadas {len(filas_borrar)} filas"


def _op_ordenar_filas(ws, params: dict) -> str:
    """Ordena las filas de datos por una columna."""
    col_idx = _col_por_nombre(ws, params.get("columna_orden", ""))
    if col_idx is None:
        return f"Columna '{params.get('columna_orden')}' no encontrada"
    descendente = bool(params.get("descendente", False))
    datos = [list(fila) for fila in ws.iter_rows(min_row=2, values_only=True)]

    def _clave(fila):
        v = fila[col_idx - 1]
        try:
            return (0, float(str(v).replace(",", ".")))
        except (ValueError, TypeError):
            return (1, str(v or "").lower())

    datos.sort(key=_clave, reverse=descendente)
    for i, fila_datos in enumerate(datos, start=2):
        for j, valor in enumerate(fila_datos, start=1):
            ws.cell(row=i, column=j, value=valor)
    return f"Filas ordenadas por '{params.get('columna_orden')}'"


def _op_agregar_fila(ws, params: dict) -> str:
    """Agrega una fila al final con los valores indicados."""
    valores = params.get("valores", [])
    ws.append(valores)
    return f"Fila agregada con {len(valores)} valores"


def _op_resaltar_filas(ws, params: dict) -> str:
    """Colorea el fondo de las filas que cumplen la condición."""
    from openpyxl.styles import PatternFill
    col_idx = _col_por_nombre(ws, params.get("columna_filtro", ""))
    if col_idx is None:
        return f"Columna '{params.get('columna_filtro')}' no encontrada"
    operador  = params.get("operador", "eq")
    valor_ref = params.get("valor_filtro", "")
    color     = str(params.get("color", "FFFF00")).lstrip("#")
    fill      = PatternFill("solid", fgColor=color)
    resaltadas = 0
    for fila in range(2, ws.max_row + 1):
        v = ws.cell(row=fila, column=col_idx).value
        if _comparar(v, operador, valor_ref):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=fila, column=col).fill = fill
            resaltadas += 1
    return f"{resaltadas} filas resaltadas"


def _op_formatear_encabezados(ws, params: dict) -> str:
    """Aplica formato de fuente y fondo a la fila de encabezados."""
    from openpyxl.styles import Font, PatternFill, Alignment
    color_fondo = str(params.get("color_fondo", "1A56DB")).lstrip("#")
    color_texto = str(params.get("color_texto", "FFFFFF")).lstrip("#")
    negrita     = bool(params.get("negrita", True))
    fill = PatternFill("solid", fgColor=color_fondo)
    font = Font(bold=negrita, color=color_texto)
    alin = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        celda           = ws.cell(row=1, column=col)
        celda.fill      = fill
        celda.font      = font
        celda.alignment = alin
    return "Encabezados formateados"


def _op_agregar_columna(ws, params: dict) -> str:
    """Agrega una columna calculada como operación entre dos columnas existentes."""
    nombre       = params.get("nombre_columna", "Nueva")
    col_a_nombre = params.get("columna_a", "")
    col_b_nombre = params.get("columna_b", "")
    op           = params.get("operacion", "multiplicar")
    col_a        = _col_por_nombre(ws, col_a_nombre) if col_a_nombre else None
    col_b        = _col_por_nombre(ws, col_b_nombre) if col_b_nombre else None
    nueva_col    = ws.max_column + 1
    ws.cell(row=1, column=nueva_col, value=nombre)
    for fila in range(2, ws.max_row + 1):
        try:
            va = float(str(ws.cell(row=fila, column=col_a).value or 0).replace(",", ".")) if col_a else 0.0
            vb = float(str(ws.cell(row=fila, column=col_b).value or 0).replace(",", ".")) if col_b else 0.0
            if op == "multiplicar": res = va * vb
            elif op == "sumar":     res = va + vb
            elif op == "restar":    res = va - vb
            elif op == "dividir":   res = va / vb if vb != 0 else 0.0
            else:                   res = va * vb
            ws.cell(row=fila, column=nueva_col, value=round(res, 2))
        except (ValueError, TypeError):
            ws.cell(row=fila, column=nueva_col, value=None)
    return f"Columna '{nombre}' agregada"


def _op_renombrar_hoja(ws, params: dict) -> str:
    """Renombra la hoja activa."""
    nuevo_nombre = str(params.get("nuevo_nombre", "Hoja"))[:31]
    ws.title     = nuevo_nombre
    return f"Hoja renombrada a '{nuevo_nombre}'"


def _op_limpiar_duplicados(ws, params: dict) -> str:
    """Elimina filas duplicadas según columnas clave."""
    columnas_clave = params.get("columnas_clave", [])
    if not columnas_clave:
        return "No se especificaron columnas clave"
    indices = [_col_por_nombre(ws, c) for c in columnas_clave]
    indices = [i for i in indices if i is not None]
    if not indices:
        return "Ninguna columna clave encontrada"
    vistos: set = set()
    filas_borrar = []
    for fila in range(2, ws.max_row + 1):
        clave = tuple(ws.cell(row=fila, column=i).value for i in indices)
        if clave in vistos:
            filas_borrar.append(fila)
        else:
            vistos.add(clave)
    for fila in reversed(filas_borrar):
        ws.delete_rows(fila)
    return f"{len(filas_borrar)} duplicados eliminados"


# Mapa de nombre → función. Solo estas operaciones son ejecutables.
_OPERACIONES: dict = {
    "calcular_total":       _op_calcular_total,
    "filtrar_filas":        _op_filtrar_filas,
    "eliminar_filas":       _op_eliminar_filas,
    "ordenar_filas":        _op_ordenar_filas,
    "agregar_fila":         _op_agregar_fila,
    "resaltar_filas":       _op_resaltar_filas,
    "formatear_encabezados": _op_formatear_encabezados,
    "agregar_columna":      _op_agregar_columna,
    "renombrar_hoja":       _op_renombrar_hoja,
    "limpiar_duplicados":   _op_limpiar_duplicados,
}

_DESCRIPCION_OPERACIONES = """\
Operaciones disponibles (elige UNA y devuelve solo JSON válido):

1. calcular_total       — Suma una columna numérica y escribe el total al final
   params: hoja, columna (nombre exacto del encabezado), etiqueta (texto del rótulo, opcional)

2. filtrar_filas        — Conserva SOLO las filas que cumplen la condición (elimina las demás)
   params: hoja, columna_filtro, valor_filtro, operador (eq|ne|gt|lt|gte|lte|contiene)

3. eliminar_filas       — Elimina las filas que cumplen la condición
   params: hoja, columna_filtro, valor_filtro, operador (eq|ne|gt|lt|gte|lte|contiene)

4. ordenar_filas        — Ordena las filas de datos por una columna
   params: hoja, columna_orden, descendente (true|false)

5. agregar_fila         — Agrega una fila al final con los valores indicados
   params: hoja, valores (lista ordenada igual que las columnas)

6. resaltar_filas       — Colorea las filas que cumplen la condición
   params: hoja, columna_filtro, valor_filtro, operador, color (hex sin #, ej: FFFF00)

7. formatear_encabezados — Aplica formato a la fila de encabezados
   params: hoja, color_fondo (hex sin #), color_texto (hex sin #), negrita (true|false)

8. agregar_columna      — Agrega una columna calculada (operación entre otras dos)
   params: hoja, nombre_columna, columna_a, columna_b,
           operacion (multiplicar|sumar|restar|dividir)

9. renombrar_hoja       — Cambia el nombre de una hoja
   params: hoja, nuevo_nombre

10. limpiar_duplicados  — Elimina filas duplicadas según columnas clave
    params: hoja, columnas_clave (lista de nombres de columna)
"""


# ── API pública ───────────────────────────────────────────────────────────────

def ejecutar_operacion_excel(ruta_excel: str, operacion_dict: dict) -> str:
    """
    Aplica la operación descrita en operacion_dict al archivo Excel y lo guarda.
    Retorna un mensaje descriptivo del resultado.

    Raises ValueError si la operación no está en la lista permitida.
    """
    import openpyxl
    operacion = operacion_dict.get("operacion", "")
    params    = operacion_dict.get("params", {})
    hoja      = params.get("hoja")

    fn = _OPERACIONES.get(operacion)
    if fn is None:
        raise ValueError(f"Operación no permitida: '{operacion}'")

    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
    resultado = fn(ws, params)
    wb.save(ruta_excel)
    return resultado


async def editar_excel_con_claude(instruccion: str, ruta_excel: str, nombre_excel: str,
                                   vendedor: str, chat_id: int) -> dict:
    """
    Pide a Claude que elija una operación predefinida para editar el Excel.
    Devuelve {"operacion": ..., "params": {...}} o {"operacion": "IMPOSIBLE"}.

    Claude NUNCA genera código ejecutable — solo elige de la lista y devuelve JSON.
    """
    import openpyxl
    wb = openpyxl.load_workbook(ruta_excel)
    info_hojas = []
    for hoja_nombre in wb.sheetnames:
        ws = wb[hoja_nombre]
        encabezados   = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        filas_ejemplo = []
        for fila in ws.iter_rows(
            min_row=config.EXCEL_FILA_DATOS,
            max_row=min(config.EXCEL_FILA_DATOS + 2, ws.max_row),
            values_only=True,
        ):
            filas_ejemplo.append(list(fila))
        info_hojas.append({
            "hoja": hoja_nombre, "encabezados": encabezados,
            "ejemplo_filas": filas_ejemplo, "total_filas": ws.max_row - 1,
        })

    prompt = f"""Tienes un archivo Excel '{nombre_excel}' con esta estructura:

{json.dumps(info_hojas, ensure_ascii=False, default=str)}

El usuario quiere: {instruccion}

{_DESCRIPCION_OPERACIONES}
Responde SOLO con un JSON válido. Ejemplos:
{{"operacion": "calcular_total", "params": {{"hoja": "Sheet1", "columna": "Monto"}}}}
{{"operacion": "ordenar_filas", "params": {{"hoja": "Ventas", "columna_orden": "Fecha", "descendente": true}}}}

Si la instrucción no puede expresarse con ninguna operación disponible, responde:
{{"operacion": "IMPOSIBLE"}}

Solo el JSON, sin explicaciones ni bloques de código."""

    loop      = asyncio.get_event_loop()
    respuesta = await loop.run_in_executor(
        None,
        lambda: config.claude_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
    )
    texto = respuesta.content[0].text.strip()
    # Limpiar bloque de código si Claude lo incluyó igualmente
    if "```" in texto:
        partes = texto.split("```")
        texto  = partes[1] if len(partes) > 1 else partes[0]
        if texto.startswith("json"):
            texto = texto[4:].strip()

    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        logger.warning(f"Claude devolvió JSON inválido para operación Excel: {texto!r}")
        return {"operacion": "IMPOSIBLE"}


def generar_excel_personalizado(titulo: str, encabezados: list, filas: list, nombre_archivo: str) -> str:
    """Genera un .xlsx con cabecera azul y filas alternas. Reemplaza excel.generar_excel_personalizado."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
    celda           = ws.cell(row=1, column=1, value=titulo)
    celda.font      = Font(bold=True, color="FFFFFF", size=14)
    celda.fill      = PatternFill("solid", fgColor="1A56DB")
    celda.alignment = Alignment(horizontal="center")
    for col, enc in enumerate(encabezados, 1):
        celda           = ws.cell(row=2, column=col, value=enc)
        celda.font      = Font(bold=True, color="FFFFFF", size=11)
        celda.fill      = PatternFill("solid", fgColor="374151")
        celda.alignment = Alignment(horizontal="center")
    for i, fila in enumerate(filas, 3):
        for col, valor in enumerate(fila, 1):
            celda = ws.cell(row=i, column=col, value=valor)
            if i % 2 == 0:
                celda.fill = PatternFill("solid", fgColor="EFF6FF")
    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    wb.save(nombre_archivo)
    return nombre_archivo
