"""
Generacion de graficas de ventas con matplotlib.
"""

import asyncio
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

import config
from excel import inicializar_excel, detectar_columnas
from utils import obtener_nombre_hoja
import openpyxl


def _cargar_ws():
    """Helper: carga y retorna (ws, cols, nombre_hoja) o None si no hay datos."""
    inicializar_excel()
    wb          = openpyxl.load_workbook(config.EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None, None, nombre_hoja
    ws   = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    return ws, cols, nombre_hoja


def _col_total(cols: dict):
    """
    Busca la columna de totales de forma flexible:
    acepta 'total' o 'subtotal' como encabezado.
    """
    for k, v in cols.items():
        if k in ("total", "subtotal"):
            return v
    # Fallback: cualquier columna que contenga 'total'
    for k, v in cols.items():
        if "total" in k:
            return v
    return None


def generar_grafica_ventas_por_dia() -> str | None:
    ws, cols, nombre_hoja = _cargar_ws()
    if ws is None:
        return None

    col_fecha = next((v for k, v in cols.items() if "fecha" in k), None)
    col_total = _col_total(cols)
    if not col_fecha or not col_total:
        return None

    ventas_por_dia: dict[str, float] = {}
    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        if not any(fila):
            continue
        fecha = fila[col_fecha - 1]
        total = fila[col_total - 1]
        if fecha and total:
            try:
                fecha_str = str(fecha)[:10]
                ventas_por_dia[fecha_str] = ventas_por_dia.get(fecha_str, 0) + float(total)
            except Exception:
                pass

    if not ventas_por_dia:
        return None

    fechas    = sorted(ventas_por_dia.keys())
    totales   = [ventas_por_dia[f] for f in fechas]
    etiquetas = [f[-5:] for f in fechas]

    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(etiquetas, totales, color="#1A56DB", edgecolor="white", linewidth=0.5)
    ax.set_title(f"Ventas por día — {nombre_hoja}", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Fecha", fontsize=11)
    ax.set_ylabel("Total ($)", fontsize=11)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    ax.set_facecolor("#F9FAFB")
    fig.patch.set_facecolor("#FFFFFF")
    for bar, valor in zip(bars, totales):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(totales) * 0.01,
            f"${valor:,.0f}", ha="center", va="bottom", fontsize=8, color="#374151",
        )
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    ruta = f"grafica_dias_{datetime.now(config.COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    return ruta


def generar_grafica_productos() -> str | None:
    ws, cols, nombre_hoja = _cargar_ws()
    if ws is None:
        return None

    col_producto = next((v for k, v in cols.items() if "producto" in k), None)
    col_total    = _col_total(cols)
    if not col_producto or not col_total:
        return None

    ventas_por_producto: dict[str, float] = {}
    for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
        if not any(fila):
            continue
        producto = fila[col_producto - 1]
        total    = fila[col_total - 1]
        if producto and total:
            try:
                p = str(producto).strip()
                ventas_por_producto[p] = ventas_por_producto.get(p, 0) + float(total)
            except Exception:
                pass

    if not ventas_por_producto:
        return None

    sorted_items = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)
    top          = list(sorted_items[:7])
    otros_total  = sum(v for _, v in sorted_items[7:])
    if otros_total > 0:
        top.append(("Otros", otros_total))

    etiquetas = [item[0] for item in top]
    valores   = [item[1] for item in top]
    colores   = ["#1A56DB","#3B82F6","#60A5FA","#93C5FD","#BFDBFE","#DBEAFE","#EFF6FF","#CBD5E1"]

    fig, ax = plt.subplots(figsize=(8, 6))
    wedges, _, autotexts = ax.pie(
        valores, labels=None, autopct="%1.1f%%",
        colors=colores[:len(valores)], startangle=90,
        wedgeprops={"edgecolor": "white", "linewidth": 1.5},
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_color("white")
        at.set_fontweight("bold")
    ax.legend(wedges, [f"{e} (${v:,.0f})" for e, v in zip(etiquetas, valores)],
              loc="lower center", bbox_to_anchor=(0.5, -0.15), ncol=2, fontsize=8, frameon=False)
    ax.set_title(f"Productos más vendidos — {nombre_hoja}", fontsize=13, fontweight="bold", pad=15)
    plt.tight_layout()

    ruta = f"grafica_productos_{datetime.now(config.COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    return ruta


# Wrappers async
async def generar_grafica_ventas_por_dia_async():
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, generar_grafica_ventas_por_dia)

async def generar_grafica_productos_async():
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, generar_grafica_productos)
