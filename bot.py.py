"""
Bot Inteligente de Ventas para Telegram con Claude AI
=====================================================
- Conversación natural (no solo ventas)
- Memoria de precios de productos
- Excel organizado por hojas mensuales
- Puede borrar ventas
- Se adapta al formato del Excel
"""

import os
import json
import re
import tempfile
from datetime import datetime
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openai

# ============================================================
# CONFIGURACIÓN
# ============================================================
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

claves_requeridas = {
    "TELEGRAM_TOKEN": TELEGRAM_TOKEN,
    "ANTHROPIC_API_KEY": ANTHROPIC_API_KEY,
    "OPENAI_API_KEY": OPENAI_API_KEY,
}
claves_faltantes = [k for k, v in claves_requeridas.items() if not v]
if claves_faltantes:
    print("\n❌ Faltan claves en las variables de entorno:")
    for c in claves_faltantes:
        print(f"   • {c}")
    exit(1)

EXCEL_FILE = "ventas.xlsx"
MEMORIA_FILE = "memoria.json"

claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
openai_client = openai.OpenAI(api_key=OPENAI_API_KEY)


# ============================================================
# MEMORIA
# ============================================================

def cargar_memoria():
    if os.path.exists(MEMORIA_FILE):
        with open(MEMORIA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"precios": {}, "negocio": {}, "notas": []}

def guardar_memoria(memoria):
    with open(MEMORIA_FILE, "w", encoding="utf-8") as f:
        json.dump(memoria, f, ensure_ascii=False, indent=2)

def obtener_precios_como_texto():
    memoria = cargar_memoria()
    precios = memoria.get("precios", {})
    if not precios:
        return "No hay precios guardados aun."
    return "\n".join([f"- {p}: ${v:,}" for p, v in precios.items()])


# ============================================================
# EXCEL
# ============================================================

def obtener_nombre_hoja():
    meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
             7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    ahora = datetime.now()
    return f"{meses[ahora.month]} {ahora.year}"

def inicializar_hoja(ws):
    if ws.max_row > 1:
        return
    encabezados = ["#","Fecha","Hora","Producto","Cantidad","Precio Unitario","Total","Vendedor","Observaciones"]
    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor="1A56DB")
        celda.alignment = Alignment(horizontal="center")
    anchos = [6,12,10,25,12,18,14,20,30]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

def obtener_o_crear_hoja(wb, nombre_hoja):
    if nombre_hoja in wb.sheetnames:
        return wb[nombre_hoja]
    ws = wb.create_sheet(title=nombre_hoja)
    inicializar_hoja(ws)
    return ws

def inicializar_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = obtener_nombre_hoja()
        inicializar_hoja(ws)
        wb.save(EXCEL_FILE)
        print("✅ Archivo Excel creado correctamente.")

def detectar_columnas(ws):
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor:
            encabezados[str(valor).lower().strip()] = col
    return encabezados

def guardar_venta_excel(producto, cantidad, precio_unitario, total, vendedor, observaciones=""):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    ws = obtener_o_crear_hoja(wb, nombre_hoja)
    cols = detectar_columnas(ws)
    fila = ws.max_row + 1
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    hora_ahora = datetime.now().strftime("%H:%M")
    num_venta = fila - 1

    datos = {
        "#": num_venta, "fecha": fecha_hoy, "hora": hora_ahora,
        "producto": producto, "cantidad": cantidad,
        "precio unitario": precio_unitario, "precio": precio_unitario,
        "total": total, "vendedor": vendedor, "observaciones": observaciones,
    }

    if cols:
        for nombre_col, num_col in cols.items():
            for clave, valor in datos.items():
                if clave in nombre_col or nombre_col in clave:
                    ws.cell(row=fila, column=num_col, value=valor)
                    break
    else:
        valores = [num_venta, fecha_hoy, hora_ahora, producto, cantidad, precio_unitario, total, vendedor, observaciones]
        for col, valor in enumerate(valores, 1):
            ws.cell(row=fila, column=col, value=valor)

    if fila % 2 == 0:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

    wb.save(EXCEL_FILE)
    return num_venta

def borrar_venta_excel(numero_venta):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return False, "No hay ventas este mes."
    ws = wb[nombre_hoja]
    fila_borrar = None
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == numero_venta:
            fila_borrar = fila
            break
    if not fila_borrar:
        return False, f"No encontre la venta #{numero_venta}."
    ws.delete_rows(fila_borrar)
    wb.save(EXCEL_FILE)
    return True, f"✅ Venta #{numero_venta} borrada correctamente."

def obtener_ventas_recientes(limite=10):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return []
    ws = wb[nombre_hoja]
    ventas = [fila for fila in ws.iter_rows(min_row=2, values_only=True) if any(fila)]
    return ventas[-limite:]

def obtener_resumen_ventas():
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    nombre_hoja = obtener_nombre_hoja()
    if nombre_hoja not in wb.sheetnames:
        return None
    ws = wb[nombre_hoja]
    cols = detectar_columnas(ws)
    col_total = next((v for k, v in cols.items() if "total" in k), None)
    total_general = 0
    num_ventas = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue
        num_ventas += 1
        if col_total and fila[col_total - 1]:
            try:
                total_general += float(fila[col_total - 1])
            except:
                pass
    return {"hoja": nombre_hoja, "total": total_general, "num_ventas": num_ventas}


# ============================================================
# HISTORIAL
# ============================================================
historiales = {}

def agregar_al_historial(chat_id, role, content):
    if chat_id not in historiales:
        historiales[chat_id] = []
    historiales[chat_id].append({"role": role, "content": content})
    if len(historiales[chat_id]) > 20:
        historiales[chat_id] = historiales[chat_id][-20:]


# ============================================================
# CLAUDE
# ============================================================

async def procesar_con_claude(mensaje_usuario, nombre_usuario, historial_chat):
    memoria = cargar_memoria()
    precios = obtener_precios_como_texto()
    resumen = obtener_resumen_ventas()
    resumen_texto = f"${resumen['total']:,.0f} en {resumen['num_ventas']} ventas este mes" if resumen else "Sin ventas este mes"

    system_prompt = f"""Eres un asistente inteligente para un negocio. Puedes ayudar con cualquier cosa: responder preguntas, hacer calculos, redactar mensajes, dar consejos, y tambien registrar ventas.

INFORMACION DEL NEGOCIO:
{json.dumps(memoria.get('negocio', {}), ensure_ascii=False)}

PRECIOS GUARDADOS EN MEMORIA:
{precios}

RESUMEN ACTUAL:
{resumen_texto}

INSTRUCCIONES:
1. Responde en español, de forma natural y amigable. Sin formato markdown.
2. Si el mensaje contiene una venta, incluye al FINAL este bloque JSON:
   [VENTA]{{"producto": "nombre", "cantidad": 2, "precio_unitario": 50000}}[/VENTA]
   Puedes incluir varios bloques si hay multiples ventas.
3. Si el producto ya esta en PRECIOS GUARDADOS y no mencionan precio, usa ese precio.
4. Si te dicen el precio de un producto, incluye:
   [PRECIO]{{"producto": "nombre", "precio": 50000}}[/PRECIO]
5. Si te dan info del negocio, incluye:
   [NEGOCIO]{{"clave": "valor"}}[/NEGOCIO]
6. Para borrar ventas, indícale que use el comando /borrar numero.
7. El usuario se llama: {nombre_usuario}"""

    messages = list(historial_chat[-10:])
    messages.append({"role": "user", "content": mensaje_usuario})

    respuesta = claude_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        system=system_prompt,
        messages=messages
    )
    return respuesta.content[0].text


def procesar_acciones(texto_respuesta, vendedor):
    acciones = []
    texto_limpio = texto_respuesta

    # Ventas
    for venta_json in re.findall(r'\[VENTA\](.*?)\[/VENTA\]', texto_respuesta, re.DOTALL):
        try:
            venta = json.loads(venta_json.strip())
            producto = venta.get("producto", "Sin nombre")
            cantidad = float(venta.get("cantidad", 1))
            precio = float(venta.get("precio_unitario", 0))
            total = cantidad * precio
            num = guardar_venta_excel(producto, cantidad, precio, total, vendedor)
            acciones.append(f"✅ Venta #{num} registrada: {producto} x{int(cantidad)} = ${total:,.0f}")
            if precio > 0:
                memoria = cargar_memoria()
                memoria["precios"][producto.lower()] = precio
                guardar_memoria(memoria)
        except Exception as e:
            print(f"Error venta: {e}")
        texto_limpio = texto_limpio.replace(f'[VENTA]{venta_json}[/VENTA]', '')

    # Precios
    for precio_json in re.findall(r'\[PRECIO\](.*?)\[/PRECIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(precio_json.strip())
            memoria = cargar_memoria()
            memoria["precios"][datos["producto"].lower()] = float(datos["precio"])
            guardar_memoria(memoria)
        except Exception as e:
            print(f"Error precio: {e}")
        texto_limpio = texto_limpio.replace(f'[PRECIO]{precio_json}[/PRECIO]', '')

    # Negocio
    for neg_json in re.findall(r'\[NEGOCIO\](.*?)\[/NEGOCIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(neg_json.strip())
            memoria = cargar_memoria()
            memoria["negocio"].update(datos)
            guardar_memoria(memoria)
        except Exception as e:
            print(f"Error negocio: {e}")
        texto_limpio = texto_limpio.replace(f'[NEGOCIO]{neg_json}[/NEGOCIO]', '')

    return texto_limpio.strip(), acciones


# ============================================================
# MANEJADORES
# ============================================================

async def comando_inicio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Hola! Soy tu asistente de negocio.\n\n"
        "Puedo ayudarte con cualquier cosa. Algunos ejemplos:\n\n"
        "🛍️ Registrar ventas — 'Vendi 3 camisas a $50.000'\n"
        "🧠 Recordar precios — 'La camisa vale $50.000'\n"
        "📊 Reportes — 'Cuanto llevamos vendido hoy?'\n"
        "💬 Preguntas — Lo que necesites\n\n"
        "Comandos:\n"
        "/ventas — Ver ultimas ventas\n"
        "/borrar [numero] — Borrar una venta\n"
        "/precios — Ver precios guardados\n"
        "/excel — Descargar archivo Excel"
    )

async def comando_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    inicializar_excel()
    await update.message.reply_text("📎 Aqui esta tu archivo Excel:")
    with open(EXCEL_FILE, "rb") as archivo:
        await update.message.reply_document(document=archivo, filename="ventas.xlsx")

async def comando_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ventas = obtener_ventas_recientes(10)
    if not ventas:
        await update.message.reply_text("No hay ventas registradas este mes.")
        return
    texto = "📋 Ultimas ventas:\n\n"
    for v in ventas:
        num = v[0] if v[0] else "?"
        producto = v[3] if len(v) > 3 else "?"
        total = f"${v[6]:,.0f}" if len(v) > 6 and v[6] else "?"
        vendedor = v[7] if len(v) > 7 else "?"
        texto += f"#{num} — {producto} — {total} — {vendedor}\n"
    texto += "\nUsa /borrar [numero] para eliminar una venta."
    await update.message.reply_text(texto)

async def comando_borrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Indica el numero de venta.\nEjemplo: /borrar 5")
        return
    try:
        numero = int(context.args[0])
        exito, mensaje = borrar_venta_excel(numero)
        await update.message.reply_text(mensaje)
    except ValueError:
        await update.message.reply_text("El numero debe ser entero.\nEjemplo: /borrar 5")

async def comando_precios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"🧠 Precios que recuerdo:\n\n{obtener_precios_como_texto()}")

async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje = update.message.text
    chat_id = update.message.chat_id
    vendedor = update.message.from_user.first_name or "Desconocido"
    if mensaje.startswith("/"):
        return
    await context.bot.send_chat_action(chat_id=chat_id, action="typing")
    try:
        historial = historiales.get(chat_id, [])
        agregar_al_historial(chat_id, "user", f"{vendedor}: {mensaje}")
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {mensaje}", vendedor, historial)
        texto_respuesta, acciones = procesar_acciones(respuesta_raw, vendedor)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)
        if texto_respuesta:
            await update.message.reply_text(texto_respuesta)
        for accion in acciones:
            await update.message.reply_text(accion)
    except Exception as e:
        print(f"Error: {e}")
        await update.message.reply_text("Tuve un problema. Intenta de nuevo.")

async def manejar_audio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    vendedor = update.message.from_user.first_name or "Desconocido"
    chat_id = update.message.chat_id
    await update.message.reply_text("🎤 Escuchando...")
    try:
        archivo_voz = await update.message.voice.get_file()
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            await archivo_voz.download_to_drive(tmp.name)
            ruta_audio = tmp.name
        with open(ruta_audio, "rb") as audio_file:
            transcripcion = openai_client.audio.transcriptions.create(
                model="whisper-1", file=audio_file, language="es"
            )
        os.unlink(ruta_audio)
        texto = transcripcion.text
        await update.message.reply_text(f"📝 Escuche: {texto}")
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")
        historial = historiales.get(chat_id, [])
        agregar_al_historial(chat_id, "user", f"{vendedor}: {texto}")
        respuesta_raw = await procesar_con_claude(f"{vendedor}: {texto}", vendedor, historial)
        texto_respuesta, acciones = procesar_acciones(respuesta_raw, vendedor)
        agregar_al_historial(chat_id, "assistant", texto_respuesta)
        if texto_respuesta:
            await update.message.reply_text(texto_respuesta)
        for accion in acciones:
            await update.message.reply_text(accion)
    except Exception as e:
        print(f"Error audio: {e}")
        await update.message.reply_text("Problema con el audio. Intenta de nuevo.")


# ============================================================
# INICIO
# ============================================================

def main():
    print("🚀 Iniciando bot inteligente...")
    inicializar_excel()
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", comando_inicio))
    app.add_handler(CommandHandler("ayuda", comando_inicio))
    app.add_handler(CommandHandler("excel", comando_excel))
    app.add_handler(CommandHandler("ventas", comando_ventas))
    app.add_handler(CommandHandler("borrar", comando_borrar))
    app.add_handler(CommandHandler("precios", comando_precios))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_mensaje))
    app.add_handler(MessageHandler(filters.VOICE, manejar_audio))
    print("✅ Bot funcionando. Esperando mensajes...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
