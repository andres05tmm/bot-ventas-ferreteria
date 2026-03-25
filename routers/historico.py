"""
Router: Histórico — /historico/*
"""
from __future__ import annotations

import json
import logging
import os
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, Union

import config
from sheets import sheets_leer_ventas_del_dia
from routers.shared import (
    _hoy, _hace_n_dias, _leer_excel_rango, _leer_excel_compras,
    _to_float, _cantidad_a_float, _stock_wayper,
)

logger = logging.getLogger("ferrebot.api")

router = APIRouter()

# ── Historial manual de ventas diarias ────────────────────────────────────────

HISTORICO_FILE  = "historico_ventas.json"
HISTORICO_EXCEL = "historico_ventas.xlsx"
_NOMBRES_MES    = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                   "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# ── Locks y caché en memoria ────────────────────────────────────────────────
import threading as _threading

# Caché para el total de hoy (evita llamar a Sheets en cada GET)
_cache_hoy_lock  = _threading.Lock()
_cache_hoy_valor: float = 0.0
_cache_hoy_fecha: str   = ""      # "YYYY-MM-DD" — se invalida al cambiar de día
_cache_hoy_ts:    float = 0.0     # timestamp UNIX de la última consulta real
_CACHE_TTL = 90                   # segundos — equilibrio entre frescura y peticiones

# Locks de escritura para archivos JSON del histórico
# Evitan corrupción cuando el bot y el API escriben simultáneamente
_historico_lock = _threading.Lock()   # protege historico_ventas.json + .xlsx
_diario_lock    = _threading.Lock()   # protege historico_diario.json

# ── Helpers: total en vivo desde Sheets / Excel ──────────────────────────────

def _total_ventas_hoy_sheets() -> float:
    """
    Calcula el total de ventas del día actual desde Google Sheets (en vivo).
    Resultado cacheado durante _CACHE_TTL segundos para evitar saturar la API
    de Sheets cuando el dashboard refresca frecuentemente.
    Si Sheets falla, intenta Excel como fallback.
    """
    import time as _time
    global _cache_hoy_valor, _cache_hoy_fecha, _cache_hoy_ts

    hoy = _hoy()
    with _cache_hoy_lock:
        ahora = _time.monotonic()
        # Caché válido si es el mismo día y no expiró
        if (
            _cache_hoy_fecha == hoy
            and _cache_hoy_valor > 0
            and (ahora - _cache_hoy_ts) < _CACHE_TTL
        ):
            return _cache_hoy_valor

    # ── Consulta real a Sheets ────────────────────────────────────────────
    total = 0.0
    try:
        ventas = sheets_leer_ventas_del_dia()
        for v in ventas:
            if str(v.get("fecha", ""))[:10] == hoy:
                try:
                    total += float(str(v.get("total", 0)).replace(",", ".") or 0)
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    # ── Fallback: Excel si Sheets no devolvió nada ────────────────────────
    if total <= 0:
        try:
            ventas_xls = _leer_excel_rango(dias=1)
            total = sum(
                float(str(v.get("total", 0)).replace(",", ".") or 0)
                for v in ventas_xls
                if str(v.get("fecha", ""))[:10] == hoy
            )
        except Exception:
            pass

    if total > 0:
        import time as _time2
        with _cache_hoy_lock:
            _cache_hoy_valor = total
            _cache_hoy_fecha = hoy
            _cache_hoy_ts    = _time2.monotonic()
    return total


def _totales_por_dia_excel(dias: int = 30) -> dict:
    """
    Lee ventas del Excel y retorna {fecha: total} para los últimos N días.
    Útil para sincronizar históricos de días pasados.
    """
    try:
        ventas = _leer_excel_rango(dias=dias)
    except Exception:
        return {}
    por_dia: dict[str, float] = defaultdict(float)
    for v in ventas:
        fecha = str(v.get("fecha", ""))[:10]
        if fecha:
            try:
                por_dia[fecha] += float(str(v.get("total", 0)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                pass
    return {k: int(v) for k, v in por_dia.items() if v > 0}


def _sync_historico_hoy() -> dict:
    """
    Sincroniza el total de hoy al histórico persistente.
    Captura ventas (Sheets), desglose por método de pago (Excel),
    gastos del día y abonos a proveedores (memoria.json).
    """
    import json as _json
    hoy   = _hoy()
    total = _total_ventas_hoy_sheets()
    if total <= 0:
        return {"fecha": hoy, "monto": 0, "ok": False, "razon": "sin ventas hoy"}

    monto = int(total)

    # ── Desglose método de pago (desde Excel de ventas) ──────────────────
    efectivo = transferencia = datafono = 0
    n_trans  = 0
    try:
        rows = _leer_excel_rango(hoy, hoy)
        for r in rows:
            mt = str(r.get("metodo_pago", "")).lower()
            v  = float(r.get("total", 0) or 0)
            if "transfer" in mt:
                transferencia += v
            elif "data" in mt or "tarjeta" in mt:
                datafono += v
            else:
                efectivo += v
            n_trans += 1
    except Exception:
        pass

    # ── Gastos y abonos del día — una sola lectura de memoria.json ──────
    gastos_dia = 0.0
    abonos_dia = 0.0
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as _fh:
            _mem = _json.load(_fh)
        for g in _mem.get("gastos", {}).get(hoy, []):
            val = float(g.get("monto", 0))
            gastos_dia += val
            if g.get("categoria") == "abono_proveedor":
                abonos_dia += val
        # Descontar abonos de gastos para no contarlos doble
        gastos_dia = max(0.0, gastos_dia - abonos_dia)
    except Exception:
        pass

    # ── Guardar en historico_diario.json (datos enriquecidos) ─────────────
    _diario_file = "historico_diario.json"
    try:
        with _diario_lock:
            if os.path.exists(_diario_file):
                with open(_diario_file, encoding="utf-8") as _fh:
                    _diario = _json.load(_fh)
            else:
                _diario = {}

            _diario[hoy] = {
                "ventas":               monto,
                "efectivo":             round(efectivo, 2),
                "transferencia":        round(transferencia, 2),
                "datafono":             round(datafono, 2),
                "n_transacciones":      n_trans,
                "gastos":               round(gastos_dia, 2),
                "abonos_proveedores":   round(abonos_dia, 2),
            }
            with open(_diario_file, "w", encoding="utf-8") as _fh:
                _json.dump(_diario, _fh, ensure_ascii=False, indent=2)
        # Drive upload fuera del lock (puede tardar)
        try:
            from drive import subir_a_drive_urgente
            subir_a_drive_urgente(_diario_file)
        except Exception:
            pass
    except Exception:
        pass

    # ── Guardar total en historico principal ──────────────────────────────
    data = _leer_historico()
    data[hoy] = monto
    _guardar_historico(data)

    return {
        "fecha":             hoy,
        "monto":             monto,
        "efectivo":          round(efectivo, 2),
        "transferencia":     round(transferencia, 2),
        "datafono":          round(datafono, 2),
        "n_transacciones":   n_trans,
        "gastos":            round(gastos_dia, 2),
        "abonos":            round(abonos_dia, 2),
        "caja_neta":         monto - round(gastos_dia, 2) - round(abonos_dia, 2),
        "ok":                True,
    }


# ── Helpers internos ──────────────────────────────────────────────────────────

def _excel_a_dict(ruta: str) -> dict:
    """
    Lee historico_ventas.xlsx y retorna {fecha: monto}.
    Columna A (idx 0) = Fecha (YYYY-MM-DD), Columna B (idx 1) = Ventas.
    Orden de columnas hoja 1: Fecha|Ventas|Efectivo|Transferencia|Datáfono|…
    Ignora filas con fecha o monto inválidos.
    """
    data = {}
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha = str(row[0] or "").strip()
            monto = row[1]  # Col B = Ventas (antes era row[4] = Datáfono ← BUG)
            if not fecha or not monto:
                continue
            try:
                # Validar formato YYYY-MM-DD
                parts = fecha.split("-")
                if len(parts) == 3 and len(parts[0]) == 4:
                    m = int(monto)
                    if m > 0:
                        data[fecha] = m
            except Exception:
                pass
    except Exception as e:
        logging.getLogger("ferrebot.api").warning(f"[historico] No se pudo leer Excel: {e}")
    return data

def _dict_a_excel(data: dict, ruta: str) -> None:
    """
    Genera historico_ventas.xlsx con 3 hojas:
      Hoja 1 — Operaciones Diarias  (una fila por día)
      Hoja 2 — Cuentas por Pagar    (una fila por factura)
      Hoja 3 — Resumen Mensual      (una fila por mes, calculada)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    ROJO     = "C0392B"
    AZUL     = "2E4057"
    VERDE    = "1A7A4A"
    BLANCO   = "FFFFFF"
    GRIS_L   = "F2F2F2"
    BORDE_S  = Side(style="thin", color="CCCCCC")
    BORDE    = Border(left=BORDE_S, right=BORDE_S, top=BORDE_S, bottom=BORDE_S)

    def _hdr(ws, fila, cols, color=ROJO):
        """Escribe encabezados con fondo de color."""
        for col, texto in enumerate(cols, 1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.font      = Font(bold=True, color=BLANCO, size=10)
            c.fill      = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = BORDE

    def _celda(ws, fila, col, valor, fmt=None, negrita=False, color_txt=None, bg=None):
        c = ws.cell(row=fila, column=col, value=valor)
        c.border    = BORDE
        c.alignment = Alignment(horizontal="right" if isinstance(valor, (int, float)) else "left",
                                 vertical="center")
        if fmt:        c.number_format = fmt
        if negrita:    c.font = Font(bold=True, color=color_txt or "000000", size=10)
        elif color_txt: c.font = Font(color=color_txt, size=10)
        if bg:         c.fill = PatternFill("solid", fgColor=bg)
        return c

    wb = openpyxl.Workbook()

    # ── HOJA 1: Operaciones Diarias ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Operaciones Diarias"
    ws1.row_dimensions[1].height = 22

    # Título
    ws1.merge_cells("A1:I1")
    t = ws1.cell(row=1, column=1,
                 value="FERRETERÍA PUNTO ROJO — Operaciones Diarias")
    t.font      = Font(bold=True, color=BLANCO, size=12)
    t.fill      = PatternFill("solid", fgColor=ROJO)
    t.alignment = Alignment(horizontal="center", vertical="center")

    cols1 = ["Fecha","Ventas","Efectivo","Transferencia","Datáfono",
             "N° Trans","Gastos","Abonos Prov.","Caja Neta"]
    _hdr(ws1, 2, cols1)

    # Cargar datos enriquecidos si existen (gastos, abonos)
    _datos_dia: dict = {}
    try:
        import json as _j, config as _cfg, os as _os
        if _os.path.exists("historico_diario.json"):
            with open("historico_diario.json", encoding="utf-8") as _fh:
                _datos_dia = _j.load(_fh)
    except Exception:
        pass

    anchos1 = [13, 14, 14, 14, 12, 9, 13, 14, 13]
    for i, w in enumerate(anchos1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    totales1 = [0.0] * 9
    for fila_idx, (fecha, monto) in enumerate(sorted(data.items()), 3):
        bg = GRIS_L if fila_idx % 2 == 0 else None
        dd = _datos_dia.get(fecha, {})
        ventas      = float(monto)
        efectivo    = float(dd.get("efectivo", 0))
        transf      = float(dd.get("transferencia", 0))
        datafono    = float(dd.get("datafono", 0))
        n_trans     = int(dd.get("n_transacciones", 0))
        gastos      = float(dd.get("gastos", 0))
        abonos      = float(dd.get("abonos_proveedores", 0))
        caja_neta   = ventas - gastos - abonos

        vals = [fecha, ventas, efectivo, transf, datafono,
                n_trans, gastos, abonos, caja_neta]
        for ci, val in enumerate(vals, 1):
            fmt = "#,##0" if isinstance(val, float) else None
            color = "1A7A4A" if ci == 9 and caja_neta >= 0 else (
                    "C0392B" if ci == 9 and caja_neta < 0 else None)
            _celda(ws1, fila_idx, ci, val, fmt=fmt, color_txt=color, bg=bg)
        for ci, val in enumerate(vals, 1):
            if isinstance(val, float):
                totales1[ci-1] += val

    # Fila de totales
    fila_tot = max(3, 3 + len(data))
    _celda(ws1, fila_tot, 1, "TOTAL", negrita=True, bg="E8E8E8")
    for ci, tot in enumerate(totales1[1:], 2):
        if ci == 6: continue  # N° Trans no suma
        _celda(ws1, fila_tot, ci, tot, fmt="#,##0", negrita=True, bg="E8E8E8")
    ws1.freeze_panes = "A3"

    # ── HOJA 2: Cuentas por Pagar ────────────────────────────────────────────
    ws2 = wb.create_sheet("Cuentas por Pagar")
    ws2.row_dimensions[1].height = 22
    ws2.merge_cells("A1:H1")
    t2 = ws2.cell(row=1, column=1,
                  value="FERRETERÍA PUNTO ROJO — Cuentas por Pagar")
    t2.font      = Font(bold=True, color=BLANCO, size=12)
    t2.fill      = PatternFill("solid", fgColor=AZUL)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    cols2 = ["ID","Fecha","Proveedor","Descripción","Total","Pagado","Pendiente","Estado"]
    _hdr(ws2, 2, cols2, color=AZUL)

    anchos2 = [10, 13, 20, 30, 14, 14, 14, 11]
    for i, w in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    facturas_data = []
    try:
        from memoria import listar_facturas as _lf
        facturas_data = _lf()
    except Exception:
        pass

    for fi, fac in enumerate(sorted(facturas_data, key=lambda x: x.get("fecha",""))):
        fila_f = fi + 3
        bg = GRIS_L if fila_f % 2 == 0 else None
        estado = fac.get("estado", "pendiente")
        col_estado = {"pagada": "1A7A4A", "parcial": "E67E22", "pendiente": "C0392B"}.get(estado, "000000")

        _celda(ws2, fila_f, 1, fac.get("id",""), bg=bg)
        _celda(ws2, fila_f, 2, fac.get("fecha",""), bg=bg)
        _celda(ws2, fila_f, 3, fac.get("proveedor",""), bg=bg)
        _celda(ws2, fila_f, 4, fac.get("descripcion",""), bg=bg)
        _celda(ws2, fila_f, 5, fac.get("total",0), fmt="#,##0", bg=bg)
        _celda(ws2, fila_f, 6, fac.get("pagado",0), fmt="#,##0", bg=bg)
        _celda(ws2, fila_f, 7, fac.get("pendiente",0), fmt="#,##0",
               negrita=True, color_txt=col_estado if estado != "pagada" else None, bg=bg)
        _celda(ws2, fila_f, 8, estado.upper(), color_txt=col_estado, negrita=True, bg=bg)

        # Si tiene foto, agregar hyperlink en la celda ID
        foto_url = fac.get("foto_url","")
        if foto_url:
            ws2.cell(row=fila_f, column=1).hyperlink = foto_url
            ws2.cell(row=fila_f, column=1).style = "Hyperlink"

    ws2.freeze_panes = "A3"

    # ── HOJA 3: Resumen Mensual ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen Mensual")
    ws3.row_dimensions[1].height = 22
    ws3.merge_cells("A1:F1")
    t3 = ws3.cell(row=1, column=1,
                  value="FERRETERÍA PUNTO ROJO — Resumen Mensual")
    t3.font      = Font(bold=True, color=BLANCO, size=12)
    t3.fill      = PatternFill("solid", fgColor=VERDE)
    t3.alignment = Alignment(horizontal="center", vertical="center")

    cols3 = ["Mes","Ventas","Gastos","Abonos Prov.","Caja Neta","Deuda Total"]
    _hdr(ws3, 2, cols3, color=VERDE)

    anchos3 = [15, 15, 13, 14, 14, 14]
    for i, w in enumerate(anchos3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Agrupar días por mes
    por_mes: dict = defaultdict(lambda: {"ventas":0.0,"gastos":0.0,"abonos":0.0})
    for fecha, monto in data.items():
        mes_k = fecha[:7]
        por_mes[mes_k]["ventas"] += float(monto)
        dd = _datos_dia.get(fecha, {})
        por_mes[mes_k]["gastos"] += float(dd.get("gastos", 0))
        por_mes[mes_k]["abonos"] += float(dd.get("abonos_proveedores", 0))

    # Deuda total vigente al final del período
    deuda_total = sum(f.get("pendiente",0) for f in facturas_data if f.get("estado") != "pagada")

    for mi, (mes_k, vals) in enumerate(sorted(por_mes.items()), 3):
        bg = GRIS_L if mi % 2 == 0 else None
        try:
            año_n, mes_n = int(mes_k[:4]), int(mes_k[5:])
            nom = f"{_NOMBRES_MES[mes_n]} {año_n}"
        except Exception:
            nom = mes_k
        caja_neta = vals["ventas"] - vals["gastos"] - vals["abonos"]
        color_cn = "1A7A4A" if caja_neta >= 0 else "C0392B"
        _celda(ws3, mi, 1, nom, bg=bg)
        _celda(ws3, mi, 2, vals["ventas"], fmt="#,##0", bg=bg)
        _celda(ws3, mi, 3, vals["gastos"], fmt="#,##0", bg=bg)
        _celda(ws3, mi, 4, vals["abonos"], fmt="#,##0", bg=bg)
        _celda(ws3, mi, 5, caja_neta, fmt="#,##0", negrita=True, color_txt=color_cn, bg=bg)
        _celda(ws3, mi, 6, deuda_total, fmt="#,##0", bg=bg)

    ws3.freeze_panes = "A3"

    wb.save(ruta)

def _leer_historico_de_excel() -> dict:
    """
    Descarga el Excel de Drive → lee datos → retorna dict.
    Es la fuente de verdad cuando existe.
    """
    try:
        from drive import descargar_de_drive
        ruta_tmp = HISTORICO_EXCEL + ".tmp"
        if descargar_de_drive(HISTORICO_EXCEL, ruta_tmp):
            data = _excel_a_dict(ruta_tmp)
            try:
                import os as _os; _os.remove(ruta_tmp)
            except Exception:
                pass
            return data
    except Exception:
        pass
    return {}

def _leer_historico() -> dict:
    """
    Lee historial: primero intenta Excel de Drive (fuente de verdad),
    luego JSON local como fallback.
    """
    # 1. Excel en Drive (fuente de verdad)
    data = _leer_historico_de_excel()
    if data:
        # Actualizar caché JSON local
        try:
            with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return data

    # 2. JSON local como fallback
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    # 3. JSON en Drive como último recurso
    try:
        from drive import descargar_de_drive
        if descargar_de_drive(HISTORICO_FILE, HISTORICO_FILE):
            with open(HISTORICO_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _guardar_historico(data: dict) -> None:
    """
    Guarda en JSON local + Excel local, y sube ambos a Drive.
    Excel es la fuente de verdad — siempre se regenera completo.
    Protegido con _historico_lock para evitar corrupción por escrituras concurrentes.
    """
    with _historico_lock:
        # 1. JSON local (caché rápida para el API)
        with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        # 2. Excel local (fuente de verdad editable)
        _dict_a_excel(data, HISTORICO_EXCEL)

    # 3. Subir ambos a Drive (fuera del lock — puede tardar sin bloquear lecturas)
    try:
        from drive import subir_a_drive_urgente
        subir_a_drive_urgente(HISTORICO_FILE)
        subir_a_drive_urgente(HISTORICO_EXCEL)
    except Exception:
        pass

# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/historico/ventas")
def historico_ventas_get(año: int = 0, mes: int = 0):
    """
    Retorna montos del mes/año.
    SIEMPRE inyecta el total en vivo de hoy desde Sheets, para que el
    dashboard muestre el dato real aunque no se haya hecho sync manual.
    """
    data = _leer_historico()

    # ── Inyectar total en vivo del día actual ────────────────────────────
    hoy = _hoy()
    total_hoy = _total_ventas_hoy_sheets()
    if total_hoy > 0:
        data[hoy] = int(total_hoy)

    if not año and not mes:
        return data
    prefijo = f"{año}-{mes:02d}" if mes else str(año)
    return {k: v for k, v in data.items() if k.startswith(prefijo)}

class HistoricoBody(BaseModel):
    año:   int
    mes:   int
    datos: dict   # { "2026-03-01": 850000, ... }

@router.post("/historico/ventas")
def historico_ventas_post(body: HistoricoBody):
    """
    Guarda los montos de un mes.
    1. Lee estado actual del Excel (para no pisar otros meses)
    2. Fusiona con los nuevos datos
    3. Guarda JSON + Excel + sube Drive
    """
    try:
        # Leer estado actual (Excel como fuente de verdad)
        data   = _leer_historico()
        prefijo = f"{body.año}-{body.mes:02d}"

        # Eliminar entradas previas del mes
        data = {k: v for k, v in data.items() if not k.startswith(prefijo)}

        # Agregar los nuevos (solo > 0)
        for fecha, monto in body.datos.items():
            if fecha.startswith(prefijo) and monto and int(monto) > 0:
                data[fecha] = int(monto)

        _guardar_historico(data)
        registros_mes = len([v for k, v in data.items() if k.startswith(prefijo)])
        return {"ok": True, "registros": registros_mes, "total": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/historico/sincronizar-excel")
def historico_sincronizar_excel():
    """
    Lee el Excel de Drive (editado manualmente) y actualiza el JSON.
    Llamar desde dashboard cuando quieras traer cambios hechos en Excel.
    """
    try:
        data = _leer_historico_de_excel()
        if not data:
            return {"ok": False, "error": "No se encontró historico_ventas.xlsx en Drive o está vacío"}
        # Actualizar JSON local con lo que tiene el Excel (protegido con lock)
        with _historico_lock:
            with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            from drive import subir_a_drive_urgente
            subir_a_drive_urgente(HISTORICO_FILE)
        except Exception:
            pass
        return {"ok": True, "registros": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/historico/resumen")
def historico_resumen():
    """Totales mensuales para gráficas — incluye hoy en vivo."""
    data = _leer_historico()

    # Inyectar hoy en vivo
    hoy = _hoy()
    total_hoy = _total_ventas_hoy_sheets()
    if total_hoy > 0:
        data[hoy] = int(total_hoy)

    por_mes: dict = defaultdict(int)
    for fecha, monto in data.items():
        if monto and monto > 0:
            por_mes[fecha[:7]] += monto
    return dict(sorted(por_mes.items()))


@router.post("/historico/auto-sync")
def historico_auto_sync():
    """
    Sincroniza el total de hoy desde Sheets → histórico persistente.
    Llamar desde el dashboard o automáticamente al cierre del día.
    """
    try:
        result = _sync_historico_hoy()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/historico/reconstruir-desglose")
def historico_reconstruir_desglose(dias: int = Query(default=60, ge=1, le=365)):
    """
    Reconstruye historico_diario.json leyendo el Excel de ventas fila por fila.
    Usar cuando el desglose (efectivo/transferencia/datáfono) se muestra en 0
    después de un redeploy que perdió historico_diario.json.
    """
    import json as _json

    _diario_file = "historico_diario.json"

    # Leer diario existente (si hay algo, lo preservamos)
    try:
        with open(_diario_file, encoding="utf-8") as fh:
            diario = _json.load(fh)
    except Exception:
        diario = {}

    # Calcular rango de fechas (strings YYYY-MM-DD)
    desde = _hace_n_dias(dias).strftime("%Y-%m-%d")
    hasta = _hoy()

    # Leer todas las ventas del Excel en ese rango
    try:
        rows = _leer_excel_rango(dias=dias)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo Excel: {e}")

    # Agrupar por fecha y método de pago
    por_dia: dict = defaultdict(lambda: {
        "ventas": 0.0, "efectivo": 0.0,
        "transferencia": 0.0, "datafono": 0.0,
        "n_transacciones": 0, "gastos": 0.0, "abonos_proveedores": 0.0
    })

    for r in rows:
        fecha = str(r.get("fecha", ""))[:10]
        if not fecha or fecha < desde or fecha > hasta:
            continue
        total = float(r.get("total", 0) or 0)
        mt    = str(r.get("metodo", "")).lower()
        por_dia[fecha]["ventas"]          += total
        por_dia[fecha]["n_transacciones"] += 1
        if "transfer" in mt:
            por_dia[fecha]["transferencia"] += total
        elif "data" in mt or "tarjeta" in mt:
            por_dia[fecha]["datafono"] += total
        else:
            por_dia[fecha]["efectivo"] += total

    # Leer gastos y abonos de memoria.json para cada fecha
    try:
        with open(config.MEMORIA_FILE, encoding="utf-8") as fh:
            mem = _json.load(fh)
        gastos_mem = mem.get("gastos", {})
        for fecha, lista in gastos_mem.items():
            if fecha < desde or fecha > hasta:
                continue
            for g in lista:
                val = float(g.get("monto", 0))
                if g.get("categoria") == "abono_proveedor":
                    por_dia[fecha]["abonos_proveedores"] += val
                else:
                    por_dia[fecha]["gastos"] += val
    except Exception:
        pass

    # Redondear y fusionar — solo pisamos días que tienen desglose en 0
    reconstruidos = 0
    for fecha, vals in por_dia.items():
        existente = diario.get(fecha, {})
        ya_tiene_desglose = (
            existente.get("efectivo", 0) > 0
            or existente.get("transferencia", 0) > 0
            or existente.get("datafono", 0) > 0
        )
        if not ya_tiene_desglose:
            diario[fecha] = {
                "ventas":             round(vals["ventas"], 2),
                "efectivo":           round(vals["efectivo"], 2),
                "transferencia":      round(vals["transferencia"], 2),
                "datafono":           round(vals["datafono"], 2),
                "n_transacciones":    vals["n_transacciones"],
                "gastos":             round(vals["gastos"], 2),
                "abonos_proveedores": round(vals["abonos_proveedores"], 2),
            }
            reconstruidos += 1

    # Guardar historico_diario.json localmente
    with _diario_lock:
        with open(_diario_file, "w", encoding="utf-8") as fh:
            _json.dump(diario, fh, ensure_ascii=False, indent=2)

    # Subir a Drive
    try:
        from drive import subir_a_drive_urgente
        subir_a_drive_urgente(_diario_file)
    except Exception:
        pass

    # Regenerar Excel del histórico con el desglose ya reconstruido
    try:
        data_hist = _leer_historico()
        _guardar_historico(data_hist)
    except Exception:
        pass

    return {
        "ok":                    True,
        "dias_escaneados":       dias,
        "fechas_reconstruidas":  reconstruidos,
        "total_dias_en_diario":  len(diario),
    }


@router.post("/historico/sync-rango")
def historico_sync_rango(dias: int = Query(default=30, ge=1, le=365)):
    """
    Sincroniza los totales de los últimos N días desde Excel → histórico.
    Útil para llenar histórico con datos de días pasados que no se guardaron.
    No sobreescribe datos existentes (solo agrega los que faltan).
    """
    try:
        data = _leer_historico()
        totales_excel = _totales_por_dia_excel(dias=dias)

        nuevos = 0
        actualizados = 0
        for fecha, monto in totales_excel.items():
            if fecha not in data:
                data[fecha] = monto
                nuevos += 1
            elif data[fecha] != monto:
                # Solo actualizar si el Excel tiene un monto mayor
                # (probablemente el dato manual estaba incompleto)
                if monto > data[fecha]:
                    data[fecha] = monto
                    actualizados += 1

        # Inyectar hoy en vivo desde Sheets (más preciso que Excel para el día actual)
        hoy = _hoy()
        total_hoy = _total_ventas_hoy_sheets()
        if total_hoy > 0:
            if hoy not in data or int(total_hoy) > data.get(hoy, 0):
                data[hoy] = int(total_hoy)
                if hoy not in totales_excel:
                    nuevos += 1

        _guardar_historico(data)
        return {
            "ok": True,
            "dias_escaneados": dias,
            "nuevos": nuevos,
            "actualizados": actualizados,
            "total_registros": len(data),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
