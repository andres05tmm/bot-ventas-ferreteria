"""
Router: Chat IA — /chat/*, /api/health
"""
from __future__ import annotations

import io
import json
import logging
import os
import secrets
import threading as _threading
import time
import uuid
from collections import defaultdict
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException, Query, Request, UploadFile, File, Depends
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, Union

import config
from routers.shared import (
    _hoy, _hace_n_dias, _leer_excel_rango, _leer_excel_compras,
    _to_float, _cantidad_a_float, _stock_wayper,
)
from routers.deps import get_current_user, get_current_user_optional

logger = logging.getLogger("ferrebot.api")

router = APIRouter()

# ── Consultas de ventas en lenguaje natural ───────────────────────────────────

_temp_exports: dict = {}                           # token → {rows, params, expira}
_temp_exports_lock  = _threading.Lock()


def _limpiar_exports_expirados() -> None:
    ahora = time.time()
    with _temp_exports_lock:
        expirados = [t for t, v in _temp_exports.items() if v["expira"] < ahora]
        for t in expirados:
            _temp_exports.pop(t, None)


def _construir_sql_consulta(params: dict) -> tuple:
    """
    Construye SQL parametrizado a partir de los params que Claude emite en
    [CONSULTA_VENTAS]{...}[/CONSULTA_VENTAS].  Nunca ejecuta SQL literal de Claude.
    Columnas soportadas: periodo, vendedor, metodo, cliente, producto.
    """
    import re as _re
    ahora = datetime.now(config.COLOMBIA_TZ)
    wheres, valores = ["1=1"], []

    periodo = (params.get("periodo") or "").lower().strip()
    if periodo == "hoy":
        wheres.append("v.fecha = %s")
        valores.append(ahora.strftime("%Y-%m-%d"))
    elif periodo in ("semana", "esta semana", "últimos 7 días", "ultimos 7 dias"):
        wheres.append("v.fecha >= %s")
        valores.append((ahora - timedelta(days=7)).strftime("%Y-%m-%d"))
    elif periodo in ("mes", "este mes"):
        wheres.append("v.fecha >= %s")
        valores.append(ahora.replace(day=1).strftime("%Y-%m-%d"))
    else:
        m = _re.match(r"(\d+)\s*(dia|días|dias|semana|semanas)", periodo)
        if m:
            n = int(m.group(1)) * (7 if "semana" in m.group(2) else 1)
            wheres.append("v.fecha >= %s")
            valores.append((ahora - timedelta(days=n)).strftime("%Y-%m-%d"))

    if params.get("vendedor"):
        wheres.append("LOWER(COALESCE(v.vendedor,'')) LIKE %s")
        valores.append(f"%{params['vendedor'].lower().strip()}%")
    if params.get("metodo"):
        wheres.append("LOWER(COALESCE(v.metodo_pago,'')) LIKE %s")
        valores.append(f"%{params['metodo'].lower().strip()}%")
    if params.get("cliente"):
        wheres.append("LOWER(COALESCE(v.cliente_nombre,'')) LIKE %s")
        valores.append(f"%{params['cliente'].lower().strip()}%")
    if params.get("producto"):
        wheres.append("LOWER(COALESCE(d.producto_nombre,'')) LIKE %s")
        valores.append(f"%{params['producto'].lower().strip()}%")

    sql = f"""
        SELECT
            v.fecha::text                                               AS fecha,
            COALESCE(v.hora::text, '')                                  AS hora,
            CASE WHEN v.cliente_id IS NULL THEN 'CF'
                 ELSE v.cliente_id::text END                            AS id_cliente,
            COALESCE(v.cliente_nombre, 'Consumidor Final')              AS cliente,
            COALESCE(
                p.codigo,
                (SELECT p2.codigo FROM productos p2
                 WHERE LOWER(TRIM(p2.nombre)) = LOWER(TRIM(d.producto_nombre))
                 LIMIT 1),
                ''
            )                                                           AS codigo_producto,
            d.producto_nombre                                           AS producto,
            COALESCE(d.unidad_medida, 'Unidad')                        AS unidad_medida,
            d.cantidad::text                                            AS cantidad,
            COALESCE(d.precio_unitario, 0)::float                      AS precio_unitario,
            COALESCE(d.total, 0)::float                                 AS total,
            v.consecutivo                                               AS consecutivo,
            COALESCE(v.vendedor, '')                                    AS vendedor,
            COALESCE(v.metodo_pago, '')                                 AS metodo_pago
        FROM ventas v
        JOIN ventas_detalle d ON d.venta_id = v.id
        LEFT JOIN productos p ON p.id = d.producto_id
        WHERE {' AND '.join(wheres)}
        ORDER BY v.fecha DESC, v.consecutivo DESC, d.id
    """
    return sql, valores


def _formatear_resumen_consulta(rows: list, params: dict) -> str:
    if not rows:
        return "No encontré ventas con esos filtros."

    total_general   = sum(float(r.get("total") or 0) for r in rows)
    n_transacciones = len({r.get("consecutivo") for r in rows if r.get("consecutivo")})
    periodo         = (params.get("periodo") or "").lower()
    periodo_label   = {"hoy": "hoy", "semana": "últimos 7 días", "mes": "este mes"}.get(
        periodo, periodo or "en total"
    )

    lineas = [
        f"{n_transacciones} transacciones ({len(rows)} líneas), "
        f"total ${total_general:,.0f} — {periodo_label}"
    ]

    por_vendedor: dict = {}
    for r in rows:
        v = str(r.get("vendedor") or "Sin vendedor").strip()
        por_vendedor[v] = por_vendedor.get(v, 0) + float(r.get("total") or 0)
    if len(por_vendedor) > 1:
        lineas.append("Por vendedor:")
        for v, t in sorted(por_vendedor.items(), key=lambda x: -x[1]):
            lineas.append(f"  {v}: ${t:,.0f}")

    por_metodo: dict = {}
    for r in rows:
        m = str(r.get("metodo_pago") or "Sin método").strip()
        por_metodo[m] = por_metodo.get(m, 0) + float(r.get("total") or 0)
    lineas.append("Por método de pago:")
    for m, t in sorted(por_metodo.items(), key=lambda x: -x[1]):
        lineas.append(f"  {m}: ${t:,.0f}")

    if not params.get("producto"):
        por_prod: dict = {}
        for r in rows:
            pr = str(r.get("producto") or "").strip()
            if pr:
                por_prod[pr] = por_prod.get(pr, 0) + float(r.get("total") or 0)
        if por_prod:
            lineas.append("Top productos:")
            for pr, t in sorted(por_prod.items(), key=lambda x: -x[1])[:5]:
                lineas.append(f"  {pr}: ${t:,.0f}")

    return "\n".join(lineas)


def _generar_excel_consulta(rows: list, params: dict) -> "io.BytesIO":
    """Genera Excel con el mismo formato visual de export_ventas_xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ahora  = datetime.now(config.COLOMBIA_TZ)

    ROJO_OSCURO  = "C00000"
    ROJO_MEDIO   = "E00000"
    NEGRO_HEADER = "1A1A1A"
    BORDE_OSCURO = "1A1A1A"
    BORDE_INTERNO= "444444"

    fill_banner = PatternFill("solid", fgColor=ROJO_OSCURO)
    fill_franja = PatternFill("solid", fgColor=ROJO_MEDIO)
    fill_header = PatternFill("solid", fgColor=NEGRO_HEADER)
    fill_blanco = PatternFill("solid", fgColor="FFFFFF")
    fill_alt    = PatternFill("solid", fgColor="FFF0F0")

    borde_ext   = Side(border_style="thick", color=BORDE_OSCURO)
    borde_int   = Side(border_style="thin",  color=BORDE_INTERNO)
    borde_none  = Side(border_style=None)

    def _borde(col_idx, n_cols, top=None, bottom=None):
        izq = borde_ext if col_idx == 1      else borde_int
        der = borde_ext if col_idx == n_cols else borde_int
        return Border(left=izq, right=der, top=top or borde_none, bottom=bottom or borde_none)

    periodo       = (params.get("periodo") or "").lower()
    periodo_label = {
        "hoy":    f"Hoy — {ahora.strftime('%d-%m-%Y')}",
        "semana": f"Última semana — hasta {ahora.strftime('%d-%m-%Y')}",
        "mes":    f"{ahora.strftime('%B %Y').capitalize()}",
    }.get(periodo, f"Consulta — {ahora.strftime('%d-%m-%Y')}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = periodo_label[:31]

    COLUMNAS = [
        "fecha", "hora", "id_cliente", "cliente", "codigo_producto", "producto",
        "unidad_medida", "cantidad", "precio_unitario", "total",
        "consecutivo", "vendedor", "metodo_pago",
    ]
    HEADERS = [
        "FECHA", "HORA", "ID CLIENTE", "CLIENTE", "CODIGO DEL PRODUCTO", "PRODUCTO",
        "UNIDAD DE MEDIDA", "CANTIDAD", "VALOR UNITARIO", "TOTAL",
        "CONSECUTIVO DE VENTA", "VENDEDOR", "METODO DE PAGO",
    ]
    ANCHOS = [16.33, 12.55, 14.44, 24.11, 19.44, 26.89, 16.66, 13.33, 17.55, 13.0, 20.33, 18.55, 19.44]
    N = len(COLUMNAS)

    for ci, aw in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = aw

    # Fila 1: Banner rojo
    ws.row_dimensions[1].height = 72.0
    for ci in range(1, N + 1):
        c = ws.cell(row=1, column=ci)
        c.fill   = fill_banner
        c.border = Border(
            left=borde_ext if ci == 1 else borde_none,
            right=borde_ext if ci == N else borde_none,
            top=borde_ext, bottom=borde_none,
        )
    ws.merge_cells("A1:D1")
    ws["A1"].value     = "FERRETERÍA PUNTO ROJO"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["K1"].value     = f"Registro de Ventas — {periodo_label}"
    ws["K1"].font      = Font(name="Calibri", italic=True, size=11, color="FFBABA")
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center")
    ws["K1"].border    = Border(top=borde_ext, right=borde_none)

    # Fila 2: Franja
    ws.row_dimensions[2].height = 9.6
    for ci in range(1, N + 1):
        c = ws.cell(row=2, column=ci)
        c.fill   = fill_franja
        c.border = Border(bottom=borde_ext)

    # Fila 3: Headers
    ws.row_dimensions[3].height = 30.0
    for ci, nombre in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=ci, value=nombre)
        c.font      = Font(name="Sylfaen", bold=True, size=10, color="FFFFFF")
        c.fill      = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(left=borde_ext if ci == 1 else borde_int,
                             right=borde_ext if ci == N else borde_int)

    # Filas de datos
    total_general = 0.0
    for ri, row in enumerate(rows, 4):
        ws.row_dimensions[ri].height = 19.95
        fill_fila = fill_alt if (ri % 2 == 0) else fill_blanco
        for ci, clave in enumerate(COLUMNAS, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(clave))
            c.fill   = fill_fila
            c.font   = Font(name="Calibri", size=10)
            c.border = _borde(ci, N)
            if clave in ("precio_unitario", "total"):
                c.number_format = '"$"#,##0;[Red]("$"#,##0)'
                c.alignment     = Alignment(horizontal="right", vertical="center")
            elif clave == "consecutivo":
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font      = Font(name="Calibri", size=10, bold=True, color="C00000")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        total_general += float(row.get("total") or 0)

    # Fila de totales
    tr = len(rows) + 4
    ws.row_dimensions[tr].height = 22.0
    for ci in range(1, N + 1):
        c = ws.cell(row=tr, column=ci)
        c.fill   = PatternFill("solid", fgColor=NEGRO_HEADER)
        c.border = Border(left=borde_ext if ci == 1 else borde_int,
                          right=borde_ext if ci == N else borde_int,
                          top=borde_ext, bottom=borde_ext)
    ws.merge_cells(f"A{tr}:I{tr}")
    ct = ws[f"A{tr}"]
    ct.value     = f"TOTAL GENERAL  ({len(rows)} registros)"
    ct.font      = Font(name="Sylfaen", bold=True, size=10, color="FFFFFF")
    ct.alignment = Alignment(horizontal="right", vertical="center")
    cj = ws.cell(row=tr, column=10, value=total_general)
    cj.font         = Font(name="Sylfaen", bold=True, size=11, color="FFFFFF")
    cj.number_format = '"$"#,##0;[Red]("$"#,##0)'
    cj.alignment    = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def _procesar_consulta_ventas(texto_claude: str) -> tuple:
    """
    Detecta [CONSULTA_VENTAS]{...}[/CONSULTA_VENTAS] en la respuesta de Claude,
    ejecuta la consulta parametrizada y devuelve (texto_final, url_descarga|None).
    """
    import re as _re
    import json as _json

    m = _re.search(r'\[CONSULTA_VENTAS\](.*?)\[/CONSULTA_VENTAS\]', texto_claude, _re.DOTALL)
    if not m:
        return texto_claude, None

    texto_previo = texto_claude[:m.start()].strip()

    try:
        params = _json.loads(m.group(1).strip())
    except Exception:
        return texto_claude, None

    try:
        from db import query_all as _qa, DB_DISPONIBLE as _db_ok
        if not _db_ok:
            return (texto_previo + "\n" if texto_previo else "") + \
                   "No hay conexión a la base de datos.", None

        sql, valores = _construir_sql_consulta(params)
        rows = [dict(r) for r in _qa(sql, valores if valores else None)]
    except Exception as exc:
        logger.error(f"[consulta_ventas] {exc}")
        return (texto_previo + "\n" if texto_previo else "") + \
               f"Error consultando base de datos: {exc}", None

    tipo = str(params.get("tipo") or "texto").lower()

    if tipo == "exportar":
        _limpiar_exports_expirados()
        token = secrets.token_urlsafe(16)
        with _temp_exports_lock:
            _temp_exports[token] = {
                "rows":   rows,
                "params": params,
                "expira": time.time() + 1800,
            }
        n     = len({r.get("consecutivo") for r in rows if r.get("consecutivo")})
        total = sum(float(r.get("total") or 0) for r in rows)
        url   = f"/chat/export/{token}"
        resp  = (f"{texto_previo}\n\n" if texto_previo else "") + \
                f"Aquí está tu archivo ({n} transacciones — ${total:,.0f}):\n{url}"
        return resp, url

    resumen = _formatear_resumen_consulta(rows, params)
    return (f"{texto_previo}\n\n" if texto_previo else "") + resumen, None


# ── Chat IA desde el Dashboard ────────────────────────────────────────────────

# Convierte session_id (string) en un chat_id negativo único.
# Los chat_id de Telegram son positivos, los del dashboard son negativos → sin colisión.
def _session_chat_id(session_id: str) -> int:
    return -(abs(hash(session_id)) % (10 ** 9))


async def _construir_contexto_dashboard(mensaje: str, tab_activo: str = "") -> str:
    """
    Construye el bloque de contexto enriquecido para el asistente del dashboard.
    Incluye: ventas del día, catálogo completo, top productos, histórico, compras,
    caja, gastos, fiados, inventario crítico, márgenes y memoria persistente.
    """
    from memoria import cargar_memoria, cargar_caja, cargar_gastos_hoy
    from datetime import datetime
    import config

    mem   = cargar_memoria()
    ahora = datetime.now(config.COLOMBIA_TZ)
    fecha_hoy = ahora.strftime("%A %d de %B de %Y, %H:%M")
    hoy_str   = ahora.strftime("%Y-%m-%d")

    # ── Memoria persistente ───────────────────────────────────────────────────
    notas_raw = mem.get("notas", {})
    if isinstance(notas_raw, list):
        notas_raw = {"observaciones": notas_raw} if notas_raw else {}
    contexto_negocio = notas_raw.get("contexto_negocio", "")
    decisiones       = notas_raw.get("decisiones", [])
    observaciones    = notas_raw.get("observaciones", [])

    memoria_partes = []
    if contexto_negocio:
        memoria_partes.append("CONTEXTO DEL NEGOCIO:\n" + contexto_negocio)
    if decisiones:
        memoria_partes.append("DECISIONES GUARDADAS:\n" + "\n".join("- " + d for d in decisiones[-10:]))
    if observaciones:
        memoria_partes.append("OBSERVACIONES:\n" + "\n".join("- " + o for o in observaciones[-10:]))
    memoria_texto = "\n\n".join(memoria_partes) + "\n\n" if memoria_partes else ""

    # ── Caja actual ──────────────────────────────────────────────────────────
    try:
        caja = cargar_caja()
        if caja.get("abierta"):
            ef  = caja.get("efectivo", 0)
            tr  = caja.get("transferencias", 0)
            dat = caja.get("datafono", 0)
            ap  = caja.get("monto_apertura", 0)
            total_caja = ef + tr + dat
            caja_texto = (
                "CAJA (abierta desde " + str(caja.get("fecha", "?")) + "):\n"
                "  Apertura: $" + f"{ap:,.0f}" + " | Efectivo: $" + f"{ef:,.0f}" +
                " | Transferencias: $" + f"{tr:,.0f}" + " | Datafono: $" + f"{dat:,.0f}" + "\n"
                "  Total en caja: $" + f"{total_caja:,.0f}"
            )
        else:
            caja_texto = "CAJA: Cerrada (última fecha: " + str(caja.get("fecha", "sin datos")) + ")"
    except Exception:
        caja_texto = "CAJA: Sin datos"

    # ── Gastos del día ───────────────────────────────────────────────────────
    try:
        gastos_hoy = cargar_gastos_hoy()
        if gastos_hoy:
            total_gastos = sum(float(g.get("monto", 0)) for g in gastos_hoy)
            items_gasto  = "\n".join(
                "  " + str(g.get("hora", "?")) + " " + str(g.get("concepto", "?")) +
                " $" + f"{float(g.get('monto', 0)):,.0f}"
                for g in gastos_hoy
            )
            gastos_texto = "GASTOS HOY (total $" + f"{total_gastos:,.0f}" + "):\n" + items_gasto
        else:
            gastos_texto = "GASTOS HOY: ninguno registrado"
    except Exception:
        gastos_texto = "GASTOS: Sin datos"

    # ── Ventas del día (detalle completo) ────────────────────────────────────
    try:
        from db import query_all
        ventas_hoy_list = []
        try:
            rows = query_all(
                """
                SELECT v.consecutivo     AS num,
                       v.fecha::text     AS fecha,
                       v.hora::text      AS hora,
                       v.cliente_nombre  AS cliente,
                       v.metodo_pago     AS metodo_pago,
                       d.producto_nombre AS producto,
                       d.cantidad,
                       d.unidad_medida,
                       d.precio_unitario,
                       d.total
                FROM ventas v
                JOIN ventas_detalle d ON d.venta_id = v.id
                WHERE v.fecha = %s
                ORDER BY v.consecutivo, d.id
                """,
                (hoy_str,),
            )
            ventas_hoy_list = [dict(r) for r in rows]
        except Exception:
            pass
        if not ventas_hoy_list:
            ventas_hoy_list = [v for v in _leer_excel_rango(dias=1)
                               if str(v.get("fecha", ""))[:10] == hoy_str]

        if ventas_hoy_list:
            total_ventas_hoy = sum(_to_float(v.get("total", 0)) for v in ventas_hoy_list)
            por_prod: dict = {}
            for v in ventas_hoy_list:
                prod = str(v.get("producto", "")).strip()
                if not prod:
                    continue
                if prod not in por_prod:
                    por_prod[prod] = {"uds": 0.0, "total": 0.0}
                por_prod[prod]["uds"]   += _cantidad_a_float(v.get("cantidad", 1))
                por_prod[prod]["total"] += _to_float(v.get("total", 0))
            lineas_ventas = "\n".join(
                f"  {p}: {d['uds']:.2g} uds — ${d['total']:,.0f}"
                for p, d in sorted(por_prod.items(), key=lambda x: -x[1]["total"])
            )
            ef_hoy  = sum(_to_float(v.get("total", 0)) for v in ventas_hoy_list if "efect"  in str(v.get("metodo_pago", "")).lower())
            tr_hoy  = sum(_to_float(v.get("total", 0)) for v in ventas_hoy_list if "transf" in str(v.get("metodo_pago", "")).lower())
            dat_hoy = sum(_to_float(v.get("total", 0)) for v in ventas_hoy_list if "data"   in str(v.get("metodo_pago", "")).lower())
            ventas_texto = (
                f"VENTAS HOY ({len(ventas_hoy_list)} transacciones, total ${total_ventas_hoy:,.0f}):\n"
                + lineas_ventas
                + f"\n  Pago — Efectivo: ${ef_hoy:,.0f} | Transferencia: ${tr_hoy:,.0f} | Datáfono: ${dat_hoy:,.0f}"
            )
        else:
            ventas_texto = "VENTAS HOY: ninguna registrada aún"
    except Exception:
        ventas_texto = "VENTAS HOY: Sin datos"

    # ── Resumen semanal y mensual ─────────────────────────────────────────────
    try:
        ventas_sem = _leer_excel_rango(dias=7)
        total_sem  = sum(_to_float(v.get("total", 0)) for v in ventas_sem)
        dias_con_ventas = len({str(v.get("fecha", ""))[:10] for v in ventas_sem if _to_float(v.get("total", 0)) > 0})
        promedio_dia = total_sem / max(dias_con_ventas, 1)

        ventas_mes = _leer_excel_rango(mes_actual=True)
        total_mes  = sum(_to_float(v.get("total", 0)) for v in ventas_mes)

        por_dia: dict = {}
        for v in ventas_sem:
            dia = str(v.get("fecha", ""))[:10]
            if dia:
                por_dia[dia] = por_dia.get(dia, 0) + _to_float(v.get("total", 0))
        historico_sem = "\n".join(
            f"  {dia}: ${monto:,.0f}" for dia, monto in sorted(por_dia.items())
        )
        resumen_texto = (
            f"RESUMEN SEMANAL (últimos 7 días):\n"
            f"  Total: ${total_sem:,.0f} | Promedio/día: ${promedio_dia:,.0f} | Días activos: {dias_con_ventas}/7\n"
            + historico_sem + "\n"
            f"RESUMEN MENSUAL (mes actual): ${total_mes:,.0f}"
        )
    except Exception:
        resumen_texto = "RESUMEN SEMANAL/MENSUAL: Sin datos"

    # ── Top 10 productos de la semana ─────────────────────────────────────────
    try:
        por_producto: dict = {}
        for v in ventas_sem:
            prod = str(v.get("producto", "")).strip()
            if not prod:
                continue
            if prod not in por_producto:
                por_producto[prod] = {"uds": 0.0, "total": 0.0}
            por_producto[prod]["uds"]   += _cantidad_a_float(v.get("cantidad", 1))
            por_producto[prod]["total"] += _to_float(v.get("total", 0))
        top10 = sorted(por_producto.items(), key=lambda x: -x[1]["total"])[:10]
        if top10:
            top_lineas = "\n".join(
                f"  {i+1}. {p}: {d['uds']:.2g} uds — ${d['total']:,.0f}"
                for i, (p, d) in enumerate(top10)
            )
            top_texto = "TOP 10 PRODUCTOS (semana):\n" + top_lineas
        else:
            top_texto = "TOP PRODUCTOS: Sin ventas esta semana"
    except Exception:
        top_texto = "TOP PRODUCTOS: Sin datos"

    # ── Compras recientes (últimos 30 días) ───────────────────────────────────
    try:
        compras_rec = _leer_excel_compras(dias=30)
        if compras_rec:
            total_compras = sum(_to_float(c.get("costo_total", 0)) for c in compras_rec)
            items_compras = "\n".join(
                "  " + str(c.get("fecha", "?"))[:10] + " " + str(c.get("proveedor", "?")) +
                " — " + str(c.get("producto", "?")) + " $" + f"{_to_float(c.get('costo_total', 0)):,.0f}"
                for c in compras_rec[-10:]
            )
            compras_texto = (
                f"COMPRAS RECIENTES (30 días, total ${total_compras:,.0f}):\n" + items_compras
            )
        else:
            compras_texto = "COMPRAS: Ninguna registrada en los últimos 30 días"
    except Exception:
        compras_texto = "COMPRAS: Sin datos"

    # ── Fiados activos ───────────────────────────────────────────────────────
    try:
        fiados = mem.get("fiados", {})
        fiados_activos = {
            nombre: datos for nombre, datos in fiados.items()
            if float(datos.get("saldo", 0)) > 0
        }
        if fiados_activos:
            total_fiado = sum(float(d.get("saldo", 0)) for d in fiados_activos.values())
            items_fiado = "\n".join(
                "  " + n + ": $" + f"{float(d.get('saldo', 0)):,.0f}"
                for n, d in sorted(fiados_activos.items(), key=lambda x: -float(x[1].get("saldo", 0)))[:15]
            )
            fiados_texto = (
                "FIADOS ACTIVOS (" + str(len(fiados_activos)) +
                " clientes, total $" + f"{total_fiado:,.0f}" + "):\n" + items_fiado
            )
        else:
            fiados_texto = "FIADOS: Sin saldos pendientes"
    except Exception:
        fiados_texto = "FIADOS: Sin datos"

    # ── Catálogo completo (precios, stock, fracciones) ───────────────────────
    try:
        catalogo  = mem.get("catalogo", {})
        inv_mem   = mem.get("inventario", {})
        lineas_cat = []
        for k, p in catalogo.items():
            precio  = p.get("precio_unidad", 0)
            nombre  = p.get("nombre", k)
            cat     = p.get("categoria", "")
            unidad  = p.get("unidad_medida", "Unidad")
            st_raw  = inv_mem.get(k)
            st_val  = st_raw.get("cantidad") if isinstance(st_raw, dict) else st_raw
            st_str  = f" [stock: {st_val}]" if st_val is not None else ""
            fracs   = p.get("precios_fraccion", {})
            fracs_str = ""
            if fracs:
                partes = []
                for fk, fv in fracs.items():
                    fp = fv.get("precio", fv) if isinstance(fv, dict) else fv
                    partes.append(f"{fk}=${fp:,}")
                fracs_str = " [fracs: " + " | ".join(partes) + "]"
            lineas_cat.append(f"  {nombre} ({cat}) — ${precio:,}/{unidad}{st_str}{fracs_str}")
        catalogo_texto = (
            f"CATALOGO COMPLETO ({len(lineas_cat)} productos):\n" +
            "\n".join(lineas_cat)
        )
    except Exception:
        catalogo_texto = "CATALOGO: Sin datos"

    # ── Inventario crítico ───────────────────────────────────────────────────
    try:
        inventario = mem.get("inventario", {})
        if inventario:
            criticos = [
                "  " + k + ": " + str(v.get("cantidad", 0)) + " " + str(v.get("unidad", "u")) +
                " (min: " + str(v.get("minimo", 0)) + ")"
                for k, v in inventario.items()
                if float(v.get("cantidad", 0)) <= float(v.get("minimo", 0)) * 1.2
                and float(v.get("minimo", 0)) > 0
            ]
            if criticos:
                inv_texto = (
                    "INVENTARIO CRITICO (" + str(len(criticos)) + " productos bajo mínimo):\n" +
                    "\n".join(criticos[:15])
                )
            else:
                inv_texto = "INVENTARIO: " + str(len(inventario)) + " productos registrados, todos sobre mínimo"
        else:
            inv_texto = "INVENTARIO: Pendiente de configurar (aún no hay stock registrado)"
    except Exception:
        inv_texto = "INVENTARIO: Sin datos"

    # ── Cuentas por pagar (facturas de proveedores) ─────────────────────────
    try:
        from memoria import listar_facturas as _lf_dash
        _facturas_pend = _lf_dash(solo_pendientes=True)
        if _facturas_pend:
            _total_deuda_dash = sum(f["pendiente"] for f in _facturas_pend)
            _lines_fac = []
            for _f in _facturas_pend[:15]:
                _dias_f = max(0, (datetime.now(config.COLOMBIA_TZ).date() -
                              datetime.fromisoformat(_f["fecha"]).date()).days)
                _lines_fac.append(
                    f"  {_f['id']} | {_f['proveedor']} | "
                    f"Total: ${_f['total']:,.0f} | Pagado: ${_f['pagado']:,.0f} | "
                    f"Pendiente: ${_f['pendiente']:,.0f} | {_dias_f} días"
                )
            cuentas_texto = (
                f"CUENTAS POR PAGAR ({len(_facturas_pend)} facturas, "
                f"deuda total ${_total_deuda_dash:,.0f}):\n"
                + "\n".join(_lines_fac)
            )
        else:
            cuentas_texto = "CUENTAS POR PAGAR: Sin deudas pendientes con proveedores"
    except Exception:
        cuentas_texto = "CUENTAS POR PAGAR: Sin datos"

    # ── Histórico últimos 30 días ─────────────────────────────────────────────
    try:
        import db as _db_hist
        _hist_lines = []
        if _db_hist.DB_DISPONIBLE:
            _hoy_d = datetime.now(config.COLOMBIA_TZ).date()
            _desde = (_hoy_d - timedelta(days=29)).strftime("%Y-%m-%d")
            _rows_hist = await _db_hist.query_all_async(
                """SELECT fecha::text, ventas, gastos, abonos_proveedores
                   FROM historico_ventas
                   WHERE fecha >= %s
                   ORDER BY fecha ASC""",
                (_desde,),
            )
            for _r in _rows_hist:
                _v  = int(_r.get("ventas", 0) or 0)
                _g  = int(_r.get("gastos", 0) or 0)
                _ab = int(_r.get("abonos_proveedores", 0) or 0)
                _cn = _v - _g - _ab
                if _v:
                    _hist_lines.append(
                        f"  {_r['fecha']}: ventas=${_v:,.0f}"
                        + (f" | gastos=${_g:,.0f}" if _g else "")
                        + (f" | abonos=${_ab:,.0f}" if _ab else "")
                        + f" | caja_neta=${_cn:,.0f}"
                    )
        if _hist_lines:
            historico_texto = "HISTÓRICO 30 DÍAS:\n" + "\n".join(_hist_lines)
        else:
            historico_texto = "HISTÓRICO: Sin datos previos registrados"
    except Exception:
        historico_texto = "HISTÓRICO: Sin datos"

    # ── Márgenes ─────────────────────────────────────────────────────────────
    try:
        catalogo_m = mem.get("catalogo", {})
        prods_con_costo = [p for p in catalogo_m.values() if p.get("precio_compra") or p.get("costo")]
        if prods_con_costo:
            margenes_lineas = []
            for p in prods_con_costo[:15]:
                costo = float(p.get("precio_compra") or p.get("costo", 0))
                venta = float(p.get("precio_unidad", 0))
                if costo > 0 and venta > 0:
                    margen = ((venta - costo) / venta) * 100
                    margenes_lineas.append(
                        "  " + str(p["nombre"]) + ": costo $" + f"{costo:,.0f}" +
                        " -> venta $" + f"{venta:,.0f}" + " (margen " + f"{margen:.0f}" + "%)"
                    )
            margenes_texto = (
                "MARGENES (muestra):\n" + "\n".join(margenes_lineas)
            ) if margenes_lineas else "MARGENES: Pendiente (agrega precio_compra al catálogo)"
        else:
            margenes_texto = "MARGENES: Pendiente de configurar (agrega el precio de compra de cada producto)"
    except Exception:
        margenes_texto = "MARGENES: Sin datos"

    # ── Tab activo ───────────────────────────────────────────────────────────
    tab_ctx = (
        "\nTAB ACTIVO EN DASHBOARD: El usuario está mirando '" + tab_activo +
        "'. Ten esto en cuenta para dar contexto relevante."
    ) if tab_activo else ""

    # ── Alertas proactivas calculadas en Python (más confiable que Claude) ──
    alertas = []
    try:
        # Deuda antigua
        from memoria import listar_facturas as _lf2
        for _f2 in _lf2(solo_pendientes=True):
            _dias2 = max(0,(datetime.now(config.COLOMBIA_TZ).date()-
                          datetime.fromisoformat(_f2["fecha"]).date()).days)
            if _dias2 > 30:
                alertas.append(f"DEUDA ANTIGUA: {_f2['id']} ({_f2['proveedor']}) lleva {_dias2} días sin pagar — ${_f2['pendiente']:,.0f} pendiente")
        # Fiados altos
        _fds = mem.get("fiados", {})
        for _fn, _fd in _fds.items():
            if float(_fd.get("saldo", 0)) > 200000:
                alertas.append(f"FIADO ALTO: {_fn} debe ${float(_fd['saldo']):,.0f}")
        # Stock crítico
        for _ik, _iv in mem.get("inventario", {}).items():
            if (isinstance(_iv, dict) and
                float(_iv.get("cantidad",0)) < float(_iv.get("minimo",0)) and
                float(_iv.get("minimo",0)) > 0):
                alertas.append(f"STOCK BAJO: {_ik} — solo {_iv['cantidad']} {_iv.get('unidad','u')} (mín {_iv['minimo']})")
    except Exception:
        pass
    alertas_texto = ("ALERTAS ACTIVAS:\n" + "\n".join("  ⚠ " + a for a in alertas)) if alertas else ""

    return (
        "FECHA Y HORA: " + fecha_hoy + "\n\n"

        "## IDENTIDAD\n"
        "Eres el asistente de negocio de Ferretería Punto Rojo. Eres experto en ferretería "
        "colombiana, conoces cada producto del catálogo, y tienes acceso en tiempo real a todas "
        "las cifras del negocio. No eres un asistente genérico — eres el socio inteligente del dueño.\n\n"

        "## PERSONALIDAD\n"
        "Directo y sin rodeos. Cuando los números son buenos, lo dices. Cuando hay algo malo, "
        "lo señalas sin suavizarlo. Das recomendaciones concretas basadas en datos reales, no "
        "consejos genéricos. Si el negocio está perdiendo plata en algo, lo dices. "
        "Si hay una oportunidad, la identificas. Recuerdas el contexto y las decisiones pasadas.\n\n"

        "## CAPACIDADES\n"
        "REGISTRAR: ventas, gastos, fiados, abonos, facturas de proveedores, abonos a proveedores\n"
        "ANALIZAR: tendencias de ventas, rentabilidad por producto, días buenos/malos, ticket promedio\n"
        "COMPARAR: esta semana vs semana pasada, este mes vs mes anterior\n"
        "ADVERTIR: stock bajo, deudas viejas, fiados altos, días sin ventas\n"
        "RECORDAR: guarda decisiones y observaciones del negocio con [MEMORIA]\n"
        "CONSULTAR: ejecutar consultas de ventas en lenguaje natural con [CONSULTA_VENTAS]\n\n"

        "## FORMATO\n"
        "Para análisis: extenso, con números, con interpretación y recomendación al final.\n"
        "Para registros: compacto — solo JSON de acción, sin texto adicional.\n"
        "Para preguntas cortas: respuesta corta y directa.\n"
        "Sin markdown (sin asteriscos, sin #). Texto plano limpio.\n\n"

        "## ACCIONES DISPONIBLES\n"
        "[VENTA]{...}[/VENTA] — registrar venta\n"
        "[GASTO]{...}[/GASTO] — registrar gasto\n"
        "[FIADO]{...}[/FIADO] — venta a crédito\n"
        "[ABONO_FIADO]{...}[/ABONO_FIADO] — abono de cliente\n"
        "[FACTURA_PROVEEDOR]{...}[/FACTURA_PROVEEDOR] — nueva factura de proveedor\n"
        "[ABONO_PROVEEDOR]{...}[/ABONO_PROVEEDOR] — abono a factura de proveedor\n"
        "[INVENTARIO]{...}[/INVENTARIO] — actualizar stock\n"
        "[MEMORIA]{tipo: decision|observacion, contenido: texto}[/MEMORIA]\n"
        '[CONSULTA_VENTAS]{"periodo":"hoy|semana|mes|N dias","vendedor":"nombre|null","metodo":"nequi|efectivo|datafono|null","cliente":"nombre|null","producto":"nombre|null","tipo":"texto|exportar"}[/CONSULTA_VENTAS] — consultar ventas\n'
        "  Usa [CONSULTA_VENTAS] cuando el usuario pida datos específicos de ventas por vendedor, método de pago, cliente o producto.\n"
        '  tipo="texto" devuelve resumen en el chat. tipo="exportar" genera un Excel descargable.\n'
        "  Ejemplos: 'ventas de Carlos esta semana' → vendedor=Carlos, periodo=semana, tipo=texto\n"
        "            'total por nequi este mes' → metodo=nequi, periodo=mes, tipo=texto\n"
        "            'exportar ventas de hoy' → periodo=hoy, tipo=exportar\n\n"
        "[REPORTE_PDF]{...json con los datos del reporte...}[/REPORTE_PDF] — generar reporte financiero en PDF\n"
        "  Usa [REPORTE_PDF] OBLIGATORIAMENTE cuando el usuario pida un reporte financiero, PDF, reporte del mes o reporte de la semana.\n"
        "  Primero obtén los datos del endpoint /chat/reporte-datos?periodo=mes (o semana), luego responde con el tag incluyendo el JSON completo.\n"
        "  Sin ese tag el PDF no se genera. Siempre incluirlo cuando el usuario pida cualquier tipo de reporte financiero.\n"
        '  Estructura del JSON: {"periodo":"mes|semana","fecha_generacion":"DD/MM/YYYY HH:MM","estado_resultados":{"ingresos":0,"cmv":0,"utilidad_bruta":0,"gastos_operativos":0,"utilidad_neta":0},"indicadores":{"ticket_promedio":0,"transacciones":0},"gastos_categorias":[{"categoria":"...","monto":0}],"top_productos":[{"producto":"...","unidades":0,"ingresos":0}],"proyeccion":{"caja_actual":0,"proyeccion_fin_mes":0,"promedio_diario_ventas":0,"promedio_diario_gastos":0},"cuentas_pagar":[{"proveedor":"...","factura":"...","saldo":0}]}\n\n'

        + (alertas_texto + "\n\n" if alertas_texto else "")

        + "## ESTADO DEL NEGOCIO\n"
        + caja_texto + "\n\n"
        + ventas_texto + "\n\n"
        + gastos_texto + "\n\n"
        + resumen_texto + "\n\n"
        + historico_texto + "\n\n"
        + top_texto + "\n\n"
        + cuentas_texto + "\n\n"
        + fiados_texto + "\n\n"
        + inv_texto + "\n\n"
        + compras_texto + "\n\n"
        + margenes_texto + "\n\n"
        + catalogo_texto + "\n\n"
        + memoria_texto
        + tab_ctx
    )


# ── Endpoint para guardar memoria del negocio ─────────────────────────────────
class MemoriaRequest(BaseModel):
    tipo: str        # "decision" | "observacion" | "contexto_negocio"
    contenido: str


# ── Endpoint para guardar memoria del negocio ─────────────────────────────────
class MemoriaRequest(BaseModel):
    tipo: str        # "decision" | "observacion" | "contexto_negocio"
    contenido: str


@router.post("/chat/memoria")
def guardar_memoria_negocio(req: MemoriaRequest):
    """Guarda una nota/decisión/observación persistente del negocio."""
    from memoria import cargar_memoria, guardar_memoria as _guardar

    if not req.contenido.strip():
        raise HTTPException(status_code=400, detail="Contenido vacío")

    mem   = cargar_memoria()
    notas = mem.get("notas", {})
    if isinstance(notas, list):
        notas = {"observaciones": notas} if notas else {}

    if req.tipo == "contexto_negocio":
        notas["contexto_negocio"] = req.contenido.strip()
    elif req.tipo == "decision":
        from datetime import datetime
        import config
        fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        notas.setdefault("decisiones", []).append(f"[{fecha}] {req.contenido.strip()}")
        notas["decisiones"] = notas["decisiones"][-30:]  # máx 30 decisiones
    else:
        from datetime import datetime
        import config
        fecha = datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")
        notas.setdefault("observaciones", []).append(f"[{fecha}] {req.contenido.strip()}")
        notas["observaciones"] = notas["observaciones"][-30:]

    mem["notas"] = notas
    _guardar(mem, urgente=True)
    return {"ok": True, "tipo": req.tipo}


class ChatRequest(BaseModel):
    mensaje: str
    nombre: str = "Dashboard"
    historial: list = []
    # Si viene con confirmar_pago, se salta Claude y registra directamente
    confirmar_pago: Optional[str] = None   # "efectivo" | "transferencia" | "datafono"
    # ID de sesión único por pestaña del navegador (evita race condition multi-usuario)
    session_id: str = "default"
    # Tab activo en el dashboard (contexto extra para el asistente)
    tab_activo: str = ""
    # Forzar modelo: null=auto, "haiku", "sonnet"
    modelo_preferido: Optional[str] = None


@router.post("/chat")
async def chat_ia(
    req: ChatRequest,
    request: Request,
    current_user=Depends(get_current_user_optional),
):
    """
    Endpoint de chat IA para el dashboard.

    Flujo normal:
      1. procesar_con_claude()      → respuesta con tags [VENTA]
      2. procesar_acciones_async()  → guarda ventas en ventas_pendientes[0]
      3. Si hay ventas pendientes   → devuelve pendiente=True + botones de pago
      4. El frontend muestra botones; el usuario hace clic
      5. Segunda llamada con confirmar_pago="efectivo"|"transferencia"|"datafono"
      6. registrar_ventas_con_metodo_async() → Excel + Sheets + confirmación

    Flujo con método explícito en el mensaje (ej: "venta 3 tornillos efectivo"):
      - Si Claude detecta el método, devuelve PEDIR_CONFIRMACION:efectivo
      - Se registra directamente sin pedir botones
    """
    from ai import procesar_con_claude, procesar_acciones_async
    from ventas_state import ventas_pendientes, registrar_ventas_con_metodo_async, _estado_lock

    log = logging.getLogger("ferrebot.api")
    request_id = getattr(request.state, "request_id", uuid.uuid4().hex[:8])
    t0 = time.perf_counter()

    # ── RAMA B: Confirmación de pago desde botón ─────────────────────────────
    if req.confirmar_pago:
        metodo = req.confirmar_pago.strip().lower()
        if metodo not in ("efectivo", "transferencia", "datafono"):
            raise HTTPException(status_code=400, detail=f"Método de pago inválido: {metodo}")

        _chat_id = _session_chat_id(req.session_id)
        with _estado_lock:
            ventas_pend = list(ventas_pendientes.get(_chat_id, []))

        if not ventas_pend:
            return {
                "ok": True,
                "respuesta": "⚠️ No hay ventas pendientes de confirmar.",
                "acciones": {"ventas": 0, "gastos": 0},
                "pendiente": False,
            }

        confirmacion = await registrar_ventas_con_metodo_async(
            ventas_pend, metodo, req.nombre, _chat_id
        )
        log.info(f"[/chat] ✅ {len(ventas_pend)} venta(s) confirmadas | método: {metodo}")

        return {
            "ok": True,
            "respuesta": "✅ Venta registrada\n" + "\n".join(confirmacion),
            "acciones": {"ventas": len(ventas_pend), "gastos": 0},
            "pendiente": False,
        }

    # ── RAMA A: Mensaje normal ────────────────────────────────────────────────
    if not req.mensaje or not req.mensaje.strip():
        raise HTTPException(status_code=400, detail="Mensaje vacío")

    try:
        mensaje_formateado = f"{req.nombre}: {req.mensaje.strip()}"

        _chat_id = _session_chat_id(req.session_id)

        # ── Contexto dinámico del dashboard (datos reales en cada llamada) ──
        contexto_dash = await _construir_contexto_dashboard(req.mensaje, tab_activo=req.tab_activo)

        # Inyectar flag ##DASHBOARD## para que ai.py active modo dashboard
        mensaje_con_flag = f"##DASHBOARD## {mensaje_formateado}"

        # vendedor_id viene del JWT si el usuario está autenticado, None si no
        _vid_chat = current_user.get("usuario_id") if current_user else None

        # 1. Claude con contexto enriquecido del dashboard
        respuesta_raw = await procesar_con_claude(
            mensaje_usuario=mensaje_con_flag,
            nombre_usuario=req.nombre,
            historial_chat=req.historial,
            contexto_extra=contexto_dash,
            vendedor_id=_vid_chat,
        )

        # 2. Parsear acciones
        texto_limpio, acciones, _ = await procesar_acciones_async(
            respuesta_raw, req.nombre, _chat_id
        )

        pedir_pago   = "PEDIR_METODO_PAGO" in acciones
        confirmacion_accion = next(
            (a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None
        )
        gastos_registrados = sum(
            1 for a in acciones if a.startswith("Gasto registrado:")
        )

        # 3a. Método de pago ya viene en la acción (Claude lo detectó) → registrar directo
        if confirmacion_accion:
            metodo = confirmacion_accion.split(":", 1)[1].strip()
            if metodo not in ("efectivo", "transferencia", "datafono"):
                metodo = "efectivo"

            with _estado_lock:
                ventas_pend = list(ventas_pendientes.get(_chat_id, []))

            if ventas_pend:
                conf = await registrar_ventas_con_metodo_async(
                    ventas_pend, metodo, req.nombre, _chat_id
                )
                return {
                    "ok": True,
                    "respuesta": "✅ Venta registrada\n" + "\n".join(conf),
                    "acciones": {"ventas": len(ventas_pend), "gastos": gastos_registrados},
                    "pendiente": False,
                }

        # 3b. Hay ventas esperando método → devolver botones al frontend
        if pedir_pago:
            with _estado_lock:
                ventas_pend = list(ventas_pendientes.get(_chat_id, []))

            if ventas_pend:
                resumen = "\n".join(
                    f"• {v.get('cantidad',1)} {v.get('producto','?')}  ${float(v.get('total',0)):,.0f}"
                    for v in ventas_pend
                )
                texto_previo = texto_limpio.strip() if texto_limpio and texto_limpio.strip() else ""
                texto_botones = (f"{texto_previo}\n\n" if texto_previo else "") + \
                                f"🧾 {resumen}\n\n¿Cómo pagó?"
                return {
                    "ok": True,
                    "respuesta": texto_botones,
                    "acciones": {"ventas": 0, "gastos": gastos_registrados},
                    "pendiente": True,
                    "opciones_pago": [
                        {"label": "💵 Efectivo",      "valor": "efectivo"},
                        {"label": "📲 Transferencia", "valor": "transferencia"},
                        {"label": "💳 Datafono",      "valor": "datafono"},
                    ],
                }

        # 3c. Consulta de ventas en lenguaje natural
        if "[CONSULTA_VENTAS]" in respuesta_raw:
            texto_consulta, url_descarga = await _procesar_consulta_ventas(respuesta_raw)
            result = {
                "ok":       True,
                "respuesta": texto_consulta,
                "acciones": {"ventas": 0, "gastos": 0},
                "pendiente": False,
            }
            if url_descarga:
                result["url_descarga"] = url_descarga
            log.info(f"[{request_id}] /chat consulta_ventas en {int((time.perf_counter()-t0)*1000)}ms")
            return result

        # 4. Sin ventas pendientes → respuesta normal
        if texto_limpio and texto_limpio.strip():
            texto_final = texto_limpio.strip()
        elif gastos_registrados:
            gasto_msgs = [a for a in acciones if a.startswith("Gasto registrado:")]
            texto_final = "✅ " + "\n".join(gasto_msgs)
        else:
            otras = [a for a in acciones if a not in (
                "PEDIR_METODO_PAGO", "PAGO_PENDIENTE_AVISO", "INICIAR_FLUJO_CLIENTE",
            ) and not a.startswith("PEDIR_CONFIRMACION:")
              and not a.startswith("CLIENTE_DESCONOCIDO:")]
            texto_final = "\n".join(otras) if otras else "(Sin respuesta)"

        result = {
            "ok": True,
            "respuesta": texto_final,
            "acciones": {"ventas": 0, "gastos": gastos_registrados},
            "pendiente": False,
        }
        log.info(f"[{request_id}] /chat completado en {int((time.perf_counter()-t0)*1000)}ms")
        return result

    except Exception as e:
        log.error(f"[{request_id}] /chat ERROR ({int((time.perf_counter()-t0)*1000)}ms): {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))



@router.post("/chat/stream")
async def chat_stream(
    req: ChatRequest,
    request: Request,
    current_user=Depends(get_current_user_optional),
):
    """
    Endpoint SSE para el dashboard con streaming token-a-token.
    Emite eventos:
      data: {"type":"chunk","text":"..."}
      data: {"type":"done","respuesta":"...","acciones":{...},"pendiente":bool}
      data: {"type":"error","message":"..."}
    """
    from ai import procesar_con_claude_stream, procesar_acciones_async
    from ventas_state import ventas_pendientes, registrar_ventas_con_metodo_async, _estado_lock

    log = logging.getLogger("ferrebot.api")
    request_id = getattr(request.state, "request_id", uuid.uuid4().hex[:8])
    t0 = time.perf_counter()

    if not req.mensaje or not req.mensaje.strip():
        async def _err():
            yield f"data: {json.dumps({'type':'error','message':'Mensaje vacío'})}\n\n"
        return StreamingResponse(_err(), media_type="text/event-stream")

    _chat_id = _session_chat_id(req.session_id)
    _vid_stream = current_user.get("usuario_id") if current_user else None

    async def generate():
        try:
            mensaje_formateado = f"{req.nombre}: {req.mensaje.strip()}"
            contexto_dash = await _construir_contexto_dashboard(req.mensaje, tab_activo=req.tab_activo)
            mensaje_con_flag = f"##DASHBOARD## {mensaje_formateado}"
            full_text = ""
            modelo_usado = None  # capturar qué modelo se usó

            async for kind, data in procesar_con_claude_stream(
                mensaje_usuario=mensaje_con_flag,
                nombre_usuario=req.nombre,
                historial_chat=req.historial,
                contexto_extra=contexto_dash,
                modelo_preferido=req.modelo_preferido,
                vendedor_id=_vid_stream,
            ):
                if kind == "model":
                    modelo_usado = "sonnet" if "sonnet" in data else "haiku"
                elif kind == "chunk":
                    full_text += data
                    yield f"data: {json.dumps({'type':'chunk','text':data}, ensure_ascii=False)}\n\n"
                elif kind == "done":
                    full_text = data
                    break
                elif kind == "error":
                    yield f"data: {json.dumps({'type':'error','message':data})}\n\n"
                    return

            # Consulta de ventas en lenguaje natural (antes de procesar otras acciones)
            if "[CONSULTA_VENTAS]" in full_text:
                texto_consulta, url_descarga = await _procesar_consulta_ventas(full_text)
                payload = {
                    "type":      "done",
                    "respuesta": texto_consulta,
                    "acciones":  {"ventas": 0, "gastos": 0},
                    "pendiente": False,
                    "opciones_pago": None,
                    "modelo":    modelo_usado,
                }
                if url_descarga:
                    payload["url_descarga"] = url_descarga
                log.info(f"[{request_id}] /chat/stream consulta_ventas en {int((time.perf_counter()-t0)*1000)}ms")
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                return

            texto_limpio, acciones, _ = await procesar_acciones_async(
                full_text, req.nombre, _chat_id
            )

            pedir_pago          = "PEDIR_METODO_PAGO" in acciones
            confirmacion_accion = next((a for a in acciones if a.startswith("PEDIR_CONFIRMACION:")), None)
            gastos_reg          = sum(1 for a in acciones if a.startswith("Gasto registrado:"))
            ventas_reg          = 0
            pendiente           = False
            opciones_pago       = None

            if confirmacion_accion:
                metodo = confirmacion_accion.split(":", 1)[1].strip()
                if metodo not in ("efectivo", "transferencia", "datafono"):
                    metodo = "efectivo"
                with _estado_lock:
                    vp = list(ventas_pendientes.get(_chat_id, []))
                if vp:
                    conf = await registrar_ventas_con_metodo_async(vp, metodo, req.nombre, _chat_id)
                    texto_limpio = "✅ Venta registrada\n" + "\n".join(conf)
                    ventas_reg = len(vp)

            elif pedir_pago:
                with _estado_lock:
                    vp = list(ventas_pendientes.get(_chat_id, []))
                if vp:
                    resumen = "\n".join(
                        f"• {v.get('cantidad',1)} {v.get('producto','?')}  ${float(v.get('total',0)):,.0f}"
                        for v in vp
                    )
                    tp = texto_limpio.strip() if texto_limpio and texto_limpio.strip() else ""
                    texto_limpio = (f"{tp}\n\n" if tp else "") + f"🧾 {resumen}\n\n¿Cómo pagó?"
                    pendiente = True
                    opciones_pago = [
                        {"label": "💵 Efectivo",      "valor": "efectivo"},
                        {"label": "📲 Transferencia", "valor": "transferencia"},
                        {"label": "💳 Datafono",      "valor": "datafono"},
                    ]

            if not texto_limpio or not texto_limpio.strip():
                if gastos_reg:
                    texto_limpio = "✅ " + "\n".join(a for a in acciones if a.startswith("Gasto registrado:"))
                else:
                    otras = [a for a in acciones if a not in (
                        "PEDIR_METODO_PAGO","PAGO_PENDIENTE_AVISO","INICIAR_FLUJO_CLIENTE"
                    ) and not a.startswith("PEDIR_CONFIRMACION:") and not a.startswith("CLIENTE_DESCONOCIDO:")]
                    texto_limpio = "\n".join(otras) if otras else "(Sin respuesta)"

            payload = {
                "type": "done",
                "respuesta": texto_limpio.strip(),
                "acciones": {"ventas": ventas_reg, "gastos": gastos_reg},
                "pendiente": pendiente,
                "opciones_pago": opciones_pago,
                "modelo": modelo_usado,
            }
            log.info(f"[{request_id}] /chat/stream completado en {int((time.perf_counter()-t0)*1000)}ms")
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

        except Exception as exc:
            log.error(f"[{request_id}] /chat/stream ERROR ({int((time.perf_counter()-t0)*1000)}ms): {exc}", exc_info=True)
            yield f"data: {json.dumps({'type':'error','message':str(exc)})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )



# ── Descarga de Excel temporal (consultas de ventas) ─────────────────────────

@router.get("/chat/export/{token}")
def chat_export_temp(token: str, current_user=Depends(get_current_user)):
    """
    Descarga el Excel temporal generado por una consulta de ventas en el chat.
    El token expira a los 30 minutos.
    """
    _limpiar_exports_expirados()
    with _temp_exports_lock:
        entry = _temp_exports.get(token)

    if not entry:
        raise HTTPException(status_code=404, detail="Enlace no encontrado o expirado")
    if time.time() > entry["expira"]:
        with _temp_exports_lock:
            _temp_exports.pop(token, None)
        raise HTTPException(status_code=410, detail="Enlace expirado")

    try:
        buf = _generar_excel_consulta(entry["rows"], entry["params"])
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {exc}")

    ahora    = datetime.now(config.COLOMBIA_TZ)
    periodo  = (entry["params"].get("periodo") or "consulta").lower().replace(" ", "_")
    filename = f"ventas_{periodo}_{ahora.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Transcripción de audio desde el Dashboard ─────────────────────────────────



# ─────────────────────────────────────────────────────────────────────────────
# MEJORA C: BRIEFING MATUTINO AUTOMÁTICO
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/chat/briefing")
async def briefing_matutino(current_user=Depends(get_current_user)):
    """
    Genera un briefing inteligente del estado del negocio.
    Se llama automáticamente al abrir el dashboard.
    Usa Extended Thinking para un análisis más profundo.
    Incluye: resumen de ayer, alertas críticas, recomendaciones para hoy.
    """
    from ai import procesar_con_claude
    import json as _json

    try:
        ahora    = datetime.now(config.COLOMBIA_TZ)
        hoy_str  = ahora.strftime("%Y-%m-%d")
        ayer_str = (ahora - timedelta(days=1)).strftime("%Y-%m-%d")
        hora     = ahora.hour

        # Saludo según hora
        if hora < 12:
            saludo = "buenos días"
        elif hora < 18:
            saludo = "buenas tardes"
        else:
            saludo = "buenas noches"

        # Construir contexto completo
        contexto = await _construir_contexto_dashboard("briefing", tab_activo="Resumen")

        prompt = (
            f"Andrés, {saludo}. "
            f"Hoy es {ahora.strftime('%A %d de %B')}. "
            f"Dame un briefing breve del negocio en este formato exacto:\n\n"
            f"AYER ({ayer_str}):\n"
            f"[resumen de ventas de ayer: total, productos principales, método de pago predominante]\n\n"
            f"ALERTAS:\n"
            f"[lista de 1-4 alertas críticas: deudas vencidas, stock bajo, fiados altos — solo lo urgente]\n\n"
            f"PARA HOY:\n"
            f"[2-3 recomendaciones concretas y accionables basadas en los datos reales]\n\n"
            f"Si no hay datos de ayer o las alertas están vacías, dilo directamente. "
            f"Máximo 200 palabras en total. Sin asteriscos ni markdown."
        )

        respuesta = await procesar_con_claude(
            mensaje_usuario=f"##DASHBOARD## Dashboard: {prompt}",
            nombre_usuario="Dashboard",
            historial_chat=[],
            contexto_extra=contexto,
            modelo_preferido="sonnet",
            vendedor_id=current_user.get("usuario_id") if current_user else None,
        )

        # Limpiar tags de acción si Claude emitió alguno
        import re as _re
        texto_limpio = _re.sub(r'\[(?:VENTA|GASTO|MEMORIA|INVENTARIO)[^\]]*\].*?\[/(?:VENTA|GASTO|MEMORIA|INVENTARIO)\]',
                               '', respuesta, flags=_re.DOTALL).strip()

        return {
            "ok":       True,
            "briefing": texto_limpio,
            "hora":     ahora.strftime("%H:%M"),
            "fecha":    ahora.strftime("%A %d de %B de %Y"),
        }

    except Exception as e:
        logger.error(f"[briefing] Error: {e}")
        return {
            "ok":       False,
            "briefing": "No pude generar el briefing en este momento.",
            "hora":     datetime.now(config.COLOMBIA_TZ).strftime("%H:%M"),
        }


@router.get("/chat/reporte-datos")
async def reporte_datos(periodo: str = Query(default="mes", pattern="^(semana|mes)$"), current_user=Depends(get_current_user)):
    """
    Agrega en un solo JSON todos los datos necesarios para el reporte financiero PDF.
    Llama internamente a las funciones de DB equivalentes a:
    /ventas/resumen, /ventas/top, /gastos (periodo), /caja, /proveedores/resumen.
    """
    import db as _db
    from collections import defaultdict

    ahora = datetime.now(config.COLOMBIA_TZ)

    if periodo == "semana":
        fecha_inicio = (ahora - timedelta(days=7)).strftime("%Y-%m-%d")
        dias_periodo = 7
    else:
        fecha_inicio = ahora.replace(day=1).strftime("%Y-%m-%d")
        dias_periodo = ahora.day

    fecha_hoy = ahora.strftime("%Y-%m-%d")

    try:
        # ── 1. Ingresos del período ────────────────────────────────────────────
        ingresos_row = _db.query_one(
            """
            SELECT
                COALESCE(SUM(d.total), 0)        AS ingresos,
                COUNT(DISTINCT v.id)             AS transacciones
            FROM ventas v
            JOIN ventas_detalle d ON d.venta_id = v.id
            WHERE v.fecha >= %s
            """,
            [fecha_inicio],
        ) if _db.DB_DISPONIBLE else None

        ingresos      = _to_float(ingresos_row["ingresos"])      if ingresos_row else 0.0
        transacciones = int(ingresos_row["transacciones"])        if ingresos_row else 0
        ticket_prom   = round(ingresos / transacciones, 0)        if transacciones else 0.0

        # ── 2. Gastos operativos del período ──────────────────────────────────
        gastos_rows = _db.query_all(
            """
            SELECT concepto, monto, COALESCE(categoria, 'General') AS categoria
            FROM gastos
            WHERE fecha >= %s
            ORDER BY categoria, monto DESC
            """,
            [fecha_inicio],
        ) if _db.DB_DISPONIBLE else []

        gastos_por_cat: dict = defaultdict(float)
        for g in gastos_rows:
            gastos_por_cat[g["categoria"]] += _to_float(g["monto"])

        gastos_operativos = sum(gastos_por_cat.values())
        gastos_categorias = [
            {"categoria": cat, "monto": monto}
            for cat, monto in sorted(gastos_por_cat.items(), key=lambda x: -x[1])
        ]

        # ── 3. Estado de resultados ────────────────────────────────────────────
        # Nota: sin estructura de costos en la DB, CMV se estima en 0
        # (el negocio no tiene precio de costo registrado)
        cmv            = 0.0
        utilidad_bruta = ingresos - cmv
        utilidad_neta  = utilidad_bruta - gastos_operativos

        # ── 4. Top 5 productos del período ────────────────────────────────────
        top_rows = _db.query_all(
            """
            SELECT
                d.producto_nombre            AS producto,
                COALESCE(SUM(d.cantidad), 0) AS unidades,
                COALESCE(SUM(d.total), 0)    AS ingresos
            FROM ventas v
            JOIN ventas_detalle d ON d.venta_id = v.id
            WHERE v.fecha >= %s
            GROUP BY d.producto_nombre
            ORDER BY SUM(d.total) DESC
            LIMIT 5
            """,
            [fecha_inicio],
        ) if _db.DB_DISPONIBLE else []

        top_productos = [
            {
                "producto": r["producto"],
                "unidades": _to_float(r["unidades"]),
                "ingresos": _to_float(r["ingresos"]),
            }
            for r in top_rows
        ]

        # ── 5. Proyección de cierre de mes ────────────────────────────────────
        caja_row = _db.query_one(
            "SELECT efectivo, transferencias, datafono, monto_apertura FROM caja WHERE fecha = %s",
            [fecha_hoy],
        ) if _db.DB_DISPONIBLE else None

        caja_actual = 0.0
        if caja_row:
            caja_actual = (
                _to_float(caja_row.get("monto_apertura", 0))
                + _to_float(caja_row.get("efectivo", 0))
                - gastos_operativos  # aproximación
            )

        dias_restantes     = max(0, ahora.replace(day=1).replace(
            month=ahora.month % 12 + 1 if ahora.month < 12 else 1,
            year=ahora.year + (1 if ahora.month == 12 else 0),
        ).toordinal() - ahora.toordinal() - 1)

        prom_diario_ventas = round(ingresos / max(dias_periodo, 1), 0)
        prom_diario_gastos = round(gastos_operativos / max(dias_periodo, 1), 0)
        proyeccion_fin_mes = round(
            caja_actual + (prom_diario_ventas - prom_diario_gastos) * dias_restantes, 0
        )

        proyeccion = {
            "caja_actual":            caja_actual,
            "proyeccion_fin_mes":     proyeccion_fin_mes,
            "promedio_diario_ventas": prom_diario_ventas,
            "promedio_diario_gastos": prom_diario_gastos,
        }

        # ── 6. Cuentas por pagar (proveedores) ───────────────────────────────
        import asyncio as _asyncio
        from memoria import listar_facturas
        cuentas_pagar = []
        try:
            facturas = await _asyncio.to_thread(listar_facturas)
            for f in facturas:
                if f.get("estado") != "pagada" and _to_float(f.get("pendiente", 0)) > 0:
                    cuentas_pagar.append({
                        "proveedor": f.get("proveedor", "—"),
                        "factura":   f.get("numero_factura") or f.get("id") or "—",
                        "saldo":     _to_float(f.get("pendiente", 0)),
                    })
        except Exception:
            pass

        # ── 7. Respuesta ──────────────────────────────────────────────────────
        return {
            "ok":      True,
            "periodo": periodo,
            "fecha_generacion": ahora.strftime("%d/%m/%Y %H:%M"),
            "estado_resultados": {
                "ingresos":          ingresos,
                "cmv":               cmv,
                "utilidad_bruta":    utilidad_bruta,
                "gastos_operativos": gastos_operativos,
                "utilidad_neta":     utilidad_neta,
            },
            "indicadores": {
                "ticket_promedio": ticket_prom,
                "transacciones":   transacciones,
            },
            "gastos_categorias": gastos_categorias,
            "top_productos":     top_productos,
            "proyeccion":        proyeccion,
            "cuentas_pagar":     cuentas_pagar,
        }

    except Exception as e:
        logger.error(f"[/chat/reporte-datos] {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/chat/transcribir")
async def transcribir_audio(audio: UploadFile = File(...)):
    """
    Recibe un archivo de audio desde el dashboard, lo transcribe con Whisper
    y devuelve el texto. El frontend luego envía ese texto al /chat/stream.
    """
    import tempfile, os

    if not config.OPENAI_API_KEY:
        raise HTTPException(status_code=503, detail="Transcripción no disponible (sin clave OpenAI)")

    # Validar que sea audio
    content_type = audio.content_type or ""
    if not (content_type.startswith("audio/") or audio.filename.endswith((".ogg", ".webm", ".mp3", ".wav", ".m4a"))):
        raise HTTPException(status_code=400, detail="Formato de audio no soportado")

    # Leer y validar tamaño (~90s de audio webm/opus ≈ 1.5-2 MB, margen hasta 3 MB)
    MAX_AUDIO_BYTES = 3 * 1024 * 1024  # 3 MB
    audio_bytes = await audio.read()
    if len(audio_bytes) > MAX_AUDIO_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Audio demasiado largo (máx ~90 segundos). Tamaño: {len(audio_bytes) / 1024 / 1024:.1f} MB"
        )
    suffix = "." + (audio.filename.rsplit(".", 1)[-1] if "." in audio.filename else "webm")

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(audio_bytes)
        ruta_tmp = tmp.name

    try:
        def _transcribir():
            with open(ruta_tmp, "rb") as f:
                return config.openai_client.audio.transcriptions.create(
                    model="whisper-1", file=f, language="es"
                )

        import asyncio
        resultado = None
        for _tr_intento in range(3):
            try:
                resultado = await asyncio.to_thread(_transcribir)
                break
            except Exception as _tr_e:
                if _tr_intento < 2:
                    logging.getLogger("ferrebot.api").warning(
                        f"[/chat/transcribir] Whisper reintento {_tr_intento + 1}/2 en 2s: {_tr_e}"
                    )
                    await asyncio.sleep(2)
                else:
                    logging.getLogger("ferrebot.api").error(
                        f"[/chat/transcribir] Whisper sin respuesta tras 3 intentos: {_tr_e}"
                    )
                    raise HTTPException(
                        status_code=503,
                        detail="El asistente IA no está disponible ahora. Intenta de nuevo en unos momentos.",
                    )

        texto = resultado.text.strip()

        if not texto:
            return {"ok": False, "texto": ""}

        return {"ok": True, "texto": texto}

    except HTTPException:
        raise
    except Exception as e:
        logging.getLogger("ferrebot.api").error(f"[/chat/transcribir] {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(ruta_tmp)
        except Exception:
            pass
