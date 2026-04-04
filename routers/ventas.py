"""
Router: Ventas — /ventas/* y /venta-rapida
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Depends
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, Union

import config
from memoria import cargar_memoria
from routers.shared import (
    _hoy, _hace_n_dias, _leer_ventas_postgres,
    _to_float, _cantidad_a_float, _stock_wayper,
)
from routers.caja import VentaRapidaPayload, VentaRapidaItem
from routers.deps import get_filtro_efectivo, get_current_user
from routers.events import broadcast

logger = logging.getLogger("ferrebot.api")

router = APIRouter()

# Nombres que representan ajuste de caja (Venta Varia) y deben excluirse de rankings
_PRODUCTOS_EXCLUIR_TOP: frozenset[str] = frozenset({
    "venta varia", "ventas varia", "venta general",
    "no se alcanzó a anotar", "no se alcanzo a anotar",
    "ventas no anotadas", "venta no anotada",
    "no se pudo anotar", "excedente de caja", "sobrante de caja",
})

# Cláusula SQL para excluir ventas-varia del ticket promedio
_SQL_EXCLUIR_VENTA_VARIA = """
    AND NOT EXISTS (
        SELECT 1 FROM ventas_detalle d
        WHERE d.venta_id = v.id
          AND LOWER(TRIM(TRAILING '.,;: ' FROM d.producto_nombre)) = ANY(ARRAY[
              'venta varia','ventas varia','venta general',
              'no se alcanz\u00f3 a anotar','no se alcanzo a anotar',
              'ventas no anotadas','venta no anotada',
              'no se pudo anotar','excedente de caja','sobrante de caja'
          ])
    )
"""

# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/ventas/hoy")
def ventas_hoy(filtro: int | None = Depends(get_filtro_efectivo)):
    try:
        hoy = _hoy()

        try:
            pg_ventas = _leer_ventas_postgres(dias=1)
            filtradas = [v for v in pg_ventas if str(v.get("fecha", ""))[:10] == hoy]
            # Aplicar filtro por usuario_id si es vendedor (admin ve todas)
            if filtro is not None:
                filtradas = [v for v in filtradas if v.get("usuario_id") == filtro]
        except Exception as e_pg:
            raise HTTPException(status_code=503, detail=f"Base de datos no disponible: {e_pg}")

        # Enriquecer con unidad_medida desde el catálogo (solo si falta)
        try:
            necesitan = [v for v in filtradas if not v.get("unidad_medida") or v["unidad_medida"] == "Unidad"]
            if necesitan:
                catalogo = cargar_memoria().get("catalogo", {})

                def _unidad_para(nombre_prod: str) -> str:
                    if not nombre_prod:
                        return "Unidad"
                    n = nombre_prod.lower().strip()
                    for key, prod in catalogo.items():
                        if prod.get("nombre", "").lower().strip() == n or key == n.replace(" ", "_"):
                            return prod.get("unidad_medida", "Unidad") or "Unidad"
                    return "Unidad"

                for v in necesitan:
                    v["unidad_medida"] = _unidad_para(v.get("producto", ""))
        except Exception:
            pass

        return {"fecha": hoy, "ventas": filtradas, "total": len(filtradas), "fuente": "postgres"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ventas/semana")
def ventas_semana(filtro: int | None = Depends(get_filtro_efectivo)):
    try:
        ventas = _leer_ventas_postgres(dias=7)
        if ventas is None:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")
        # Aplicar filtro por usuario_id si es vendedor
        if filtro is not None:
            ventas = [v for v in ventas if v.get("usuario_id") == filtro]
        return {"ventas": ventas, "total": len(ventas)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ventas/top")
def ventas_top(
    periodo: str = Query(default="semana", pattern="^(semana|mes)$"),
    filtro: int | None = Depends(get_filtro_efectivo)
):
    try:
        dias = 7 if periodo == "semana" else None
        mes = periodo == "mes"
        ventas = _leer_ventas_postgres(dias=dias, mes_actual=mes)
        if ventas is None:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")
        # Aplicar filtro por usuario_id si es vendedor
        if filtro is not None:
            ventas = [v for v in ventas if v.get("usuario_id") == filtro]

        cat_unidad: dict[str, str] = {}
        for prod in cargar_memoria().get("catalogo", {}).values():
                nombre_lower = (prod.get("nombre_lower") or prod.get("nombre", "")).lower().strip()
                cat_unidad[nombre_lower] = prod.get("unidad_medida", "Unidad") or "Unidad"

        por_producto: dict[str, dict] = {}
        for v in ventas:
            nombre = str(v.get("producto", "")).strip()
            if not nombre or nombre.lower().strip().rstrip('.,:;') in _PRODUCTOS_EXCLUIR_TOP:
                continue
            cantidad  = _cantidad_a_float(v.get("cantidad", 0))
            total     = _to_float(v.get("total", 0))
            unidad_v  = str(v.get("unidad_medida", "") or "").strip()

            if nombre not in por_producto:
                unidad_cat = cat_unidad.get(nombre.lower(), "")
                por_producto[nombre] = {
                    "unidades":      0.0,
                    "ingresos":      0.0,
                    "frecuencia":    0,
                    "unidad_medida": unidad_cat or unidad_v or "Unidad",
                }
            por_producto[nombre]["unidades"]   += cantidad
            por_producto[nombre]["ingresos"]   += total
            por_producto[nombre]["frecuencia"] += 1

        ranking = sorted(
            [{"producto": k, **v} for k, v in por_producto.items()],
            key=lambda x: x["ingresos"],
            reverse=True,
        )[:10]

        for i, item in enumerate(ranking, 1):
            item["posicion"] = i

        return {"periodo": periodo, "top": ranking}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ventas/resumen")
def ventas_resumen(filtro: int | None = Depends(get_filtro_efectivo)):
    """
    Resumen para las tarjetas del dashboard. 100 % PostgreSQL, sin fallbacks.
    Si es vendedor, filtra por usuario_id.
    """
    import db as _db

    if not _db.DB_DISPONIBLE:
        raise HTTPException(status_code=503, detail="Base de datos no disponible")

    try:
        hoy         = _hoy()
        ahora_local = datetime.now(config.COLOMBIA_TZ)
        logger.debug("/ventas/resumen filtro=%s hoy=%s", filtro, hoy)

        # Construir WHERE clause dinámicamente para filtro de usuario
        where_usuario = "AND v.usuario_id = %s" if filtro is not None else ""
        params_hoy = [hoy, filtro] if filtro is not None else [hoy]

        row_hoy = _db.query_one(
            f"""
            SELECT
                COALESCE(SUM(d.total), 0)          AS total_hoy,
                COUNT(DISTINCT v.id)               AS pedidos_hoy,
                COUNT(DISTINCT v.consecutivo)      AS n_transacciones
            FROM ventas v
            JOIN ventas_detalle d ON d.venta_id = v.id
            WHERE v.fecha = %s {where_usuario}
            """,
            params_hoy,
        )
        logger.debug("row_hoy=%s", row_hoy)
        total_hoy   = _to_float(row_hoy["total_hoy"])   if row_hoy else 0.0
        pedidos_hoy = int(row_hoy["pedidos_hoy"])        if row_hoy else 0

        fecha_7d  = _hace_n_dias(7).strftime("%Y-%m-%d")

        # Si es vendedor, calcular desde detail tables; si admin, usar historico agregado
        if filtro is not None:
            # Vendedor: computar desde ventas_detalle
            rows_sem_detail = _db.query_all(
                f"""
                SELECT v.fecha::text AS fecha, SUM(d.total) AS total_dia
                FROM ventas v
                JOIN ventas_detalle d ON d.venta_id = v.id
                WHERE v.fecha >= %s AND v.usuario_id = %s
                GROUP BY v.fecha
                ORDER BY v.fecha
                """,
                [fecha_7d, filtro]
            )
            ventas_por_dia = {str(r["fecha"])[:10]: _to_float(r["total_dia"]) for r in rows_sem_detail}

            row_tick = _db.query_one(
                f"SELECT COUNT(DISTINCT v.id) AS n, COALESCE(SUM(v.total), 0) AS suma FROM ventas v WHERE v.fecha >= %s AND v.usuario_id = %s{_SQL_EXCLUIR_VENTA_VARIA}",
                [fecha_7d, filtro]
            )
        else:
            # Admin: usar historico para gráfica de días
            rows_sem  = _db.query_all(
                "SELECT fecha::text AS fecha, ventas FROM historico_ventas WHERE fecha >= %s ORDER BY fecha",
                [fecha_7d],
            )
            ventas_por_dia = {str(r["fecha"])[:10]: _to_float(r["ventas"]) for r in rows_sem}

            # Ticket promedio siempre desde ventas directamente, excluyendo ventas-varia
            row_tick = _db.query_one(
                f"SELECT COUNT(DISTINCT v.id) AS n, COALESCE(SUM(v.total), 0) AS suma FROM ventas v WHERE v.fecha >= %s{_SQL_EXCLUIR_VENTA_VARIA}",
                [fecha_7d],
            )

        total_sem = sum(ventas_por_dia.values())
        pedidos_sem = max(int(row_tick["n"]) if row_tick and row_tick["n"] else 1, 1)
        ticket_prom = round(_to_float(row_tick["suma"]) / pedidos_sem, 0) if row_tick else 0

        historico = [
            {"fecha": _hace_n_dias(i).strftime("%Y-%m-%d"),
             "total": ventas_por_dia.get(_hace_n_dias(i).strftime("%Y-%m-%d"), 0)}
            for i in range(6, -1, -1)
        ]

        primer_dia_mes = ahora_local.replace(day=1).strftime("%Y-%m-%d")

        if filtro is not None:
            # Vendedor: computar desde ventas_detalle
            rows_mes_detail = _db.query_all(
                f"""
                SELECT v.fecha::text AS fecha, SUM(d.total) AS total_dia
                FROM ventas v
                JOIN ventas_detalle d ON d.venta_id = v.id
                WHERE v.fecha >= %s AND v.usuario_id = %s
                GROUP BY v.fecha
                ORDER BY v.fecha
                """,
                [primer_dia_mes, filtro]
            )
            ventas_mes_por_dia = {str(r["fecha"])[:10]: _to_float(r["total_dia"]) for r in rows_mes_detail}
        else:
            # Admin: usar historico
            rows_mes = _db.query_all(
                "SELECT fecha::text AS fecha, ventas FROM historico_ventas WHERE fecha >= %s ORDER BY fecha",
                [primer_dia_mes],
            )
            ventas_mes_por_dia = {str(r["fecha"])[:10]: _to_float(r["ventas"]) for r in rows_mes}

        total_mes = sum(ventas_mes_por_dia.values())

        historico_mes = []
        current = ahora_local.replace(day=1)
        while current.date() <= ahora_local.date():
            dia_str = current.strftime("%Y-%m-%d")
            historico_mes.append({"fecha": dia_str, "total": ventas_mes_por_dia.get(dia_str, 0)})
            current += timedelta(days=1)

        return {
            "total_hoy":     total_hoy,
            "pedidos_hoy":   pedidos_hoy,
            "total_semana":  total_sem,
            "ticket_prom":   ticket_prom,
            "historico_7d":  historico,
            "total_mes":     total_mes,
            "historico_mes": historico_mes,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/venta-rapida")
def venta_rapida(payload: VentaRapidaPayload, current_user=Depends(get_current_user)):
    try:
        import db as _db
        import datetime as _dt

        if not _db.DB_DISPONIBLE:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")

        _catalogo_cache = {}
        try:
            _catalogo_cache = cargar_memoria().get("catalogo", {})
        except Exception:
            pass

        def _resolver_unidad(item: VentaRapidaItem) -> str:
            if item.unidad_medida and item.unidad_medida not in ("", "Unidad"):
                return item.unidad_medida
            nombre_norm = item.nombre.lower().strip()
            for prod_key, prod_val in _catalogo_cache.items():
                if prod_val.get("nombre", "").lower().strip() == nombre_norm or prod_key == nombre_norm.replace(" ", "_"):
                    return prod_val.get("unidad_medida", "Unidad")
            return item.unidad_medida or "Unidad"

        # ── Resolver producto_id para cada item (búsqueda por nombre) ──────────
        nombres_productos = [item.nombre for item in payload.productos]
        _prod_id_map: dict[str, int] = {}
        try:
            placeholders = ", ".join(["%s"] * len(nombres_productos))
            rows_prod = _db.query_all(
                f"SELECT id, nombre FROM productos WHERE LOWER(TRIM(nombre)) = ANY(ARRAY[{placeholders}]::text[])",
                [n.lower().strip() for n in nombres_productos],
            )
            for r in rows_prod:
                _prod_id_map[r["nombre"].lower().strip()] = r["id"]
        except Exception:
            pass

        items_calc = []
        for item in payload.productos:
            try:
                from utils import convertir_fraccion_a_decimal
                cant_num = convertir_fraccion_a_decimal(item.cantidad)
            except (ValueError, TypeError):
                cant_num = 1.0
            if not cant_num or cant_num <= 0:
                cant_num = 1.0
            items_calc.append({
                "item":            item,
                "cant_num":        cant_num,
                "precio_unitario": round(item.total / cant_num, 2),
                "unidad":          _resolver_unidad(item),
                "producto_id":     _prod_id_map.get(item.nombre.lower().strip()),
            })

        # ⚠️ Usar hora Colombia (no UTC del servidor Railway) para que la fecha
        # coincida con los filtros del dashboard que también usan COLOMBIA_TZ.
        ahora = _dt.datetime.now(config.COLOMBIA_TZ).replace(tzinfo=None)

        with _db._get_conn() as conn:
            with conn.cursor() as cur:
                # Consecutivo atómico desde PG
                cur.execute("SELECT COALESCE(MAX(consecutivo), 0) + 1 AS siguiente FROM ventas")
                consecutivo = cur.fetchone()["siguiente"]

                cur.execute(
                    """
                    INSERT INTO ventas
                        (consecutivo, fecha, vendedor, metodo_pago,
                         total, cliente_nombre, cliente_id, usuario_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        consecutivo,
                        ahora,
                        payload.vendedor,
                        payload.metodo,
                        sum(i["item"].total for i in items_calc),
                        payload.cliente_nombre or None,
                        payload.cliente_id     or None,
                        current_user.get("usuario_id"),
                    ),
                )
                venta_id = cur.fetchone()["id"]

                for ic in items_calc:
                    cur.execute(
                        """
                        INSERT INTO ventas_detalle
                            (venta_id, producto_nombre, cantidad,
                             precio_unitario, total, unidad_medida, producto_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            venta_id,
                            ic["item"].nombre,
                            ic["cant_num"],
                            ic["precio_unitario"],
                            ic["item"].total,
                            ic["unidad"],
                            ic.get("producto_id"),
                        ),
                    )
            conn.commit()

        for ic in items_calc:
            try:
                from memoria import descontar_inventario
                descontar_inventario(ic["item"].nombre, ic["cant_num"])
            except Exception:
                pass

        broadcast("venta_registrada", {
            "consecutivo": consecutivo,
            "total":       sum(ic["item"].total for ic in items_calc),
            "metodo":      payload.metodo,
            "vendedor":    payload.vendedor,
        })

        return {
            "ok":          True,
            "venta_id":    venta_id,
            "consecutivo": consecutivo,
            "productos":   len(items_calc),
            "total":       sum(ic["item"].total for ic in items_calc),
            "metodo":      payload.metodo,
            "fuente":      "postgres",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ventas/top2")
def ventas_top2(
    periodo:  str = Query(default="semana", pattern="^(semana|mes)$"),
    criterio: str = Query(default="ingresos", pattern="^(ingresos|frecuencia|categoria)$"),
    filtro: int | None = Depends(get_filtro_efectivo)
):
    try:
        dias = 7 if periodo == "semana" else None
        mes  = periodo == "mes"
        ventas = _leer_ventas_postgres(dias=dias, mes_actual=mes)
        if ventas is None:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")
        # Aplicar filtro por usuario_id si es vendedor
        if filtro is not None:
            ventas = [v for v in ventas if v.get("usuario_id") == filtro]

        cat_map: dict[str, str] = {}
        for v in cargar_memoria().get("catalogo", {}).values():
                nombre_lower = v.get("nombre_lower", "").strip()
                cat_map[nombre_lower] = v.get("categoria", "Sin categoría")

        acum: dict[str, dict] = defaultdict(lambda: {
            "ingresos": 0.0, "frecuencia": 0, "categoria": "Sin categoría"
        })
        for v in ventas:
            nombre = str(v.get("producto", "")).strip()
            if not nombre or nombre.lower().strip().rstrip('.,:;') in _PRODUCTOS_EXCLUIR_TOP:
                continue
            total = _to_float(v.get("total", 0))
            acum[nombre]["ingresos"]   += total
            acum[nombre]["frecuencia"] += 1
            if acum[nombre]["categoria"] == "Sin categoría":
                acum[nombre]["categoria"] = cat_map.get(nombre.lower(), "Sin categoría")

        if criterio == "ingresos":
            ranking = sorted(acum.items(), key=lambda x: -x[1]["ingresos"])[:10]
            items = [{"producto": k, "valor": v["ingresos"], "frecuencia": v["frecuencia"],
                      "categoria": v["categoria"], "posicion": i+1}
                     for i, (k, v) in enumerate(ranking)]

        elif criterio == "frecuencia":
            ranking = sorted(acum.items(), key=lambda x: -x[1]["frecuencia"])[:10]
            items = [{"producto": k, "valor": v["frecuencia"], "ingresos": v["ingresos"],
                      "categoria": v["categoria"], "posicion": i+1}
                     for i, (k, v) in enumerate(ranking)]

        else:  # categoria
            por_cat: dict[str, list] = defaultdict(list)
            for nombre, datos in acum.items():
                por_cat[datos["categoria"]].append({"producto": nombre, **datos})
            result_cat = {}
            for cat, prods in por_cat.items():
                top = sorted(prods, key=lambda x: -x["ingresos"])[:5]
                result_cat[cat] = [{"producto": p["producto"], "valor": p["ingresos"],
                                    "frecuencia": p["frecuencia"], "posicion": i+1}
                                   for i, p in enumerate(top)]
            return {"periodo": periodo, "criterio": criterio, "por_categoria": result_cat}

        return {"periodo": periodo, "criterio": criterio, "top": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/ventas/{numero}")
def eliminar_venta(numero: int):
    """
    Elimina todas las filas de un consecutivo de venta.
    """
    try:
        import db as _db

        if not _db.DB_DISPONIBLE:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")

        borradas = _db.execute(
            "DELETE FROM ventas WHERE consecutivo = %s", [numero]
        )

        if not borradas:
            raise HTTPException(status_code=404, detail=f"Consecutivo #{numero} no encontrado")

        broadcast("venta_eliminada", {"consecutivo": numero})
        return {"ok": True, "mensaje": f"Venta #{numero} eliminada"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/ventas/{numero}/linea")
def eliminar_linea_venta(numero: int, producto: str = Query(...)):
    """
    Elimina UNA sola línea (producto) de un consecutivo multi-producto.
    """
    try:
        import db as _db

        if not _db.DB_DISPONIBLE:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")

        borradas = _db.execute(
            """
            DELETE FROM ventas_detalle
            WHERE venta_id = (SELECT id FROM ventas WHERE consecutivo = %s LIMIT 1)
              AND LOWER(producto_nombre) = LOWER(%s)
            """,
            [numero, producto],
        )

        if not borradas:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró '{producto}' en consecutivo #{numero}",
            )

        broadcast("venta_eliminada", {"consecutivo": numero, "producto": producto})
        return {"ok": True, "borradas": borradas, "mensaje": f"'{producto}' eliminado del consecutivo #{numero}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class EditarVentaBody(BaseModel):
    producto:          Union[str, None]   = None
    cantidad:          Union[float, None] = None
    precio_unitario:   Union[float, None] = None
    total:             Union[float, None] = None
    metodo_pago:       Union[str, None]   = None
    cliente:           Union[str, None]   = None
    id_cliente:        Union[str, None]   = None
    vendedor:          Union[str, None]   = None
    producto_original: Union[str, None]   = None  # para identificar fila en multi-producto

@router.patch("/ventas/{numero}")
def editar_venta(numero: int, body: EditarVentaBody):
    """
    Edita los campos de un consecutivo. 100% Postgres.
    """
    try:
        import db as _db

        if not _db.DB_DISPONIBLE:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")

        cambios = {k: v for k, v in body.dict().items() if v is not None and k != "producto_original"}
        if not cambios:
            raise HTTPException(status_code=400, detail="No hay campos para actualizar")

        filtro_producto = body.producto_original.strip().lower() if body.producto_original else None

        campo_col_pg = {
            "producto":        "producto_nombre",
            "cantidad":        "cantidad",
            "precio_unitario": "precio_unitario",
            "total":           "total",
        }
        cabecera_col_pg = {
            "metodo_pago": "metodo_pago",
            "cliente":     "cliente_nombre",
            "vendedor":    "vendedor",
        }

        # Actualizar ventas_detalle
        det_parts, det_params = [], []
        for campo, col in campo_col_pg.items():
            if campo in cambios:
                det_parts.append(f"{col} = %s")
                det_params.append(cambios[campo])
        if det_parts:
            where_det = "WHERE venta_id = (SELECT id FROM ventas WHERE consecutivo = %s LIMIT 1)"
            if filtro_producto:
                where_det += " AND LOWER(producto_nombre) = LOWER(%s)"
                det_params.extend([numero, filtro_producto])
            else:
                det_params.append(numero)
            _db.execute(f"UPDATE ventas_detalle SET {', '.join(det_parts)} {where_det}", det_params)

        # Actualizar ventas (cabecera)
        cab_parts, cab_params = [], []
        for campo, col in cabecera_col_pg.items():
            if campo in cambios:
                cab_parts.append(f"{col} = %s")
                cab_params.append(cambios[campo])
        if cab_parts:
            cab_params.append(numero)
            _db.execute(f"UPDATE ventas SET {', '.join(cab_parts)} WHERE consecutivo = %s", cab_params)

        broadcast("venta_editada", {"consecutivo": numero})
        return {"ok": True, "mensaje": f"Venta #{numero} actualizada"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Venta Varia ───────────────────────────────────────────────────────────────

class VentaVariaRequest(BaseModel):
    monto: float
    metodo_pago: str
    descripcion: str = "Venta Varia"
    vendedor: str = "Dashboard"


@router.post("/ventas/varia")
async def registrar_venta_varia(req: VentaVariaRequest):
    """
    Registra una venta sin detalle de productos para cuadrar caja.
    Marca sin_detalle=True en ventas_detalle — no descuenta inventario
    y queda excluida del análisis de productos.
    """
    from ventas_state import registrar_ventas_con_metodo_async

    metodo = req.metodo_pago.strip().lower()
    if metodo not in ("efectivo", "transferencia", "datafono"):
        raise HTTPException(status_code=400, detail=f"Método de pago inválido: {metodo}")

    if req.monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")

    nombre_prod = req.descripcion.strip() or "Venta Varia"
    venta = {
        "producto":        nombre_prod,
        "cantidad":        1,
        "total":           round(req.monto),
        "precio_unitario": round(req.monto),
        "metodo_pago":     metodo,
        "sin_detalle":     True,   # ← no descuenta inventario, excluida de análisis
    }

    try:
        confirmaciones = await registrar_ventas_con_metodo_async(
            [venta], metodo, req.vendedor, -1
        )
        broadcast("venta_registrada", {
            "tipo":    "varia",
            "monto":   req.monto,
            "metodo":  metodo,
            "vendedor": req.vendedor,
        })
        return {
            "ok": True,
            "mensaje": "Venta varia registrada",
            "detalle": confirmaciones,
        }
    except Exception as e:
        logging.getLogger("ferrebot.api").error(f"[/ventas/varia] {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── Export Excel on-demand ─────────────────────────────────────────────────────

@router.get("/export/ventas.xlsx")
def export_ventas_xlsx(
    periodo: str = Query(default="todo", pattern="^(hoy|semana|mes|todo)$"),
    filtro: int | None = Depends(get_filtro_efectivo)
):
    """
    Genera y descarga un archivo Excel con ventas filtradas por período.
    ?periodo=hoy | semana | mes | todo  (default: todo)
    Usa el formato visual de Ferretería Punto Rojo (banner rojo, tabla oscura).
    Si es vendedor, solo ve sus propias ventas.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        import db as _db
        if not _db.DB_DISPONIBLE:
            raise HTTPException(status_code=503, detail="Base de datos no disponible")

        ahora = datetime.now(config.COLOMBIA_TZ)
        params: list = []
        filtro_fecha = ""

        if periodo == "hoy":
            filtro_fecha = "AND v.fecha = %s"
            params.append(ahora.strftime("%Y-%m-%d"))
        elif periodo == "semana":
            fecha_inicio = (ahora - timedelta(days=7)).strftime("%Y-%m-%d")
            filtro_fecha = "AND v.fecha >= %s"
            params.append(fecha_inicio)
        elif periodo == "mes":
            primer_dia = ahora.replace(day=1).strftime("%Y-%m-%d")
            filtro_fecha = "AND v.fecha >= %s"
            params.append(primer_dia)
        # "todo" → sin filtro de fecha

        # Agregar filtro de usuario si es vendedor
        filtro_usuario = "AND v.usuario_id = %s" if filtro is not None else ""
        if filtro is not None:
            params.append(filtro)

        sql = f"""
            SELECT
                v.fecha::text                                               AS fecha,
                COALESCE(v.hora::text, '')                                  AS hora,
                CASE WHEN v.cliente_id IS NULL THEN 'CF'
                     ELSE v.cliente_id::text END                            AS id_cliente,
                COALESCE(v.cliente_nombre, 'Consumidor Final')              AS cliente,
                COALESCE(
                    p.codigo,
                    (SELECT p2.codigo FROM productos p2
                     WHERE LOWER(TRIM(p2.nombre)) = LOWER(TRIM(d.producto_nombre))
                     LIMIT 1),
                    ''
                )                                                           AS codigo_producto,
                d.producto_nombre                                           AS producto,
                COALESCE(d.unidad_medida, 'Unidad')                        AS unidad_medida,
                d.cantidad::text                                            AS cantidad,
                COALESCE(d.precio_unitario, 0)::float                      AS precio_unitario,
                COALESCE(d.total, 0)::float                                 AS total,
                v.consecutivo                                               AS consecutivo,
                COALESCE(v.vendedor, '')                                    AS vendedor,
                COALESCE(v.metodo_pago, '')                                 AS metodo_pago
            FROM ventas v
            JOIN ventas_detalle d ON d.venta_id = v.id
            LEFT JOIN productos p ON p.id = d.producto_id
            WHERE 1=1 {filtro_fecha} {filtro_usuario}
            ORDER BY v.fecha DESC, v.consecutivo DESC, d.id
        """
        rows = _db.query_all(sql, params if params else None)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando export ventas.xlsx: {e}")
        raise HTTPException(status_code=503, detail=f"Error consultando base de datos: {e}")

    # ── Colores y estilos (extraídos del formato.xlsx) ──────────────────────────
    ROJO_OSCURO  = "C00000"   # fondo banner principal
    ROJO_MEDIO   = "E00000"   # franja separadora
    NEGRO_HEADER = "1A1A1A"   # fondo encabezados de columna
    BORDE_OSCURO = "1A1A1A"   # color borde exterior
    BORDE_INTERNO= "444444"   # color borde entre columnas

    fill_banner  = PatternFill("solid", fgColor=ROJO_OSCURO)
    fill_franja  = PatternFill("solid", fgColor=ROJO_MEDIO)
    fill_header  = PatternFill("solid", fgColor=NEGRO_HEADER)
    fill_blanco  = PatternFill("solid", fgColor="FFFFFF")
    fill_alt     = PatternFill("solid", fgColor="FFF0F0")   # rosa muy suave para filas alternas

    borde_ext_thick = Side(border_style="thick",  color=BORDE_OSCURO)
    borde_int_thin  = Side(border_style="thin",   color=BORDE_INTERNO)
    borde_none      = Side(border_style=None)

    def borde_fila(col_idx, n_cols, top=None, bottom=None):
        izq = borde_ext_thick if col_idx == 1      else borde_int_thin
        der = borde_ext_thick if col_idx == n_cols else borde_int_thin
        return Border(left=izq, right=der,
                      top=top or borde_none, bottom=bottom or borde_none)

    # ── Workbook ────────────────────────────────────────────────────────────────
    wb   = openpyxl.Workbook()
    ws   = wb.active

    labels_periodo = {
        "hoy":    f"Hoy — {ahora.strftime('%d-%m-%Y')}",
        "semana": f"Última semana — hasta {ahora.strftime('%d-%m-%Y')}",
        "mes":    f"{ahora.strftime('%B %Y').capitalize()}",
        "todo":   "Acumulado total",
    }
    titulo_periodo = labels_periodo.get(periodo, "Ventas")
    ws.title = titulo_periodo[:31]   # Excel limita a 31 chars el nombre de hoja

    # ── Columnas: orden igual al formato.xlsx ────────────────────────────────────
    # A=fecha B=hora C=id_cliente D=cliente E=codigo F=producto
    # G=unidad H=cantidad I=precio_unit J=total K=consecutivo L=vendedor M=metodo
    COLUMNAS = [
        "fecha", "hora", "id_cliente", "cliente", "codigo_producto", "producto",
        "unidad_medida", "cantidad", "precio_unitario", "total",
        "consecutivo", "vendedor", "metodo_pago",
    ]
    HEADERS = [
        "FECHA", "HORA", "ID CLIENTE", "CLIENTE", "CODIGO DEL PRODUCTO", "PRODUCTO",
        "UNIDAD DE MEDIDA", "CANTIDAD", "VALOR UNITARIO", "TOTAL",
        "CONSECUTIVO DE VENTA", "VENDEDOR", "METODO DE PAGO",
    ]
    ANCHOS = [16.33, 12.55, 14.44, 24.11, 19.44, 26.89, 16.66, 13.33, 17.55, 13.0, 20.33, 18.55, 19.44]
    N = len(COLUMNAS)

    for col_idx, ancho in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # ── FILA 1: Banner rojo ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 72.0
    for col_idx in range(1, N + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = fill_banner
        top  = borde_ext_thick
        left  = borde_ext_thick if col_idx == 1 else borde_none
        right = borde_ext_thick if col_idx == N else borde_none
        c.border = Border(left=left, right=right, top=top, bottom=borde_none)

    # Celdas A1:D1 fusionadas (espacio para logo / nombre)
    ws.merge_cells("A1:D1")
    ws["A1"].value     = "FERRETERÍA PUNTO ROJO"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Título del informe en K1 (alineado a la derecha como en el template)
    ws["K1"].value     = f"Registro de Ventas — {titulo_periodo}"
    ws["K1"].font      = Font(name="Calibri", italic=True, size=11, color="FFBABA")
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center")
    ws["K1"].border    = Border(top=borde_ext_thick, right=borde_none)

    # ── FILA 2: Franja separadora roja media ─────────────────────────────────────
    ws.row_dimensions[2].height = 9.6
    for col_idx in range(1, N + 1):
        c = ws.cell(row=2, column=col_idx)
        c.fill   = fill_franja
        c.border = Border(bottom=borde_ext_thick)

    # ── FILA 3: Encabezados de columna ───────────────────────────────────────────
    ws.row_dimensions[3].height = 30.0
    for col_idx, nombre in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col_idx, value=nombre)
        c.font      = Font(name="Sylfaen", bold=True, size=10, color="FFFFFF")
        c.fill      = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        izq  = borde_ext_thick if col_idx == 1 else borde_int_thin
        der  = borde_ext_thick if col_idx == N else borde_int_thin
        c.border = Border(left=izq, right=der)

    # ── FILAS DE DATOS ────────────────────────────────────────────────────────────
    total_general = 0.0
    for row_idx, row in enumerate(rows, 4):
        ws.row_dimensions[row_idx].height = 19.95
        es_par = (row_idx % 2 == 0)
        fill_fila = fill_alt if es_par else fill_blanco

        for col_idx, clave in enumerate(COLUMNAS, 1):
            valor = row.get(clave)
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.fill   = fill_fila
            c.font   = Font(name="Calibri", size=10)
            c.border = borde_fila(col_idx, N)

            if clave in ("precio_unitario", "total"):
                c.number_format = '"$"#,##0;[Red]("$"#,##0)'
                c.alignment     = Alignment(horizontal="right", vertical="center")
            elif clave == "cantidad":
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif clave in ("consecutivo",):
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font      = Font(name="Calibri", size=10, bold=True, color="C00000")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

        total_general += float(row.get("total") or 0)

    # ── FILA DE TOTALES ───────────────────────────────────────────────────────────
    total_row = len(rows) + 4
    ws.row_dimensions[total_row].height = 22.0
    for col_idx in range(1, N + 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=NEGRO_HEADER)
        izq  = borde_ext_thick if col_idx == 1 else borde_int_thin
        der  = borde_ext_thick if col_idx == N else borde_int_thin
        c.border = Border(left=izq, right=der, top=borde_ext_thick, bottom=borde_ext_thick)

    # Etiqueta "TOTAL GENERAL" abarcando las primeras 9 columnas
    ws.merge_cells(f"A{total_row}:I{total_row}")
    ct = ws[f"A{total_row}"]
    ct.value     = f"TOTAL GENERAL  ({len(rows)} registros)"
    ct.font      = Font(name="Sylfaen", bold=True, size=10, color="FFFFFF")
    ct.alignment = Alignment(horizontal="right", vertical="center")

    # Valor total en columna J
    cj = ws.cell(row=total_row, column=10, value=total_general)
    cj.font         = Font(name="Sylfaen", bold=True, size=11, color="FFFFFF")
    cj.number_format = '"$"#,##0;[Red]("$"#,##0)'
    cj.alignment    = Alignment(horizontal="right", vertical="center")

    # ── Inmovilizar filas del encabezado ─────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Guardar y devolver ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    labels_archivo = {"hoy": "hoy", "semana": "semana", "mes": "mes", "todo": "acumulado"}
    nombre_archivo = f"ventas_{labels_archivo.get(periodo, 'total')}_{ahora.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )
