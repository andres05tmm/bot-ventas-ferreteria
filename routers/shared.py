"""
routers/shared.py
─────────────────
Helpers y utilidades compartidas por todos los routers de la API.
Ningún router debería reimplementar estas funciones — importarlas desde aquí.

Importar con:  from routers.shared import _hoy, _hace_n_dias, ...
"""
from __future__ import annotations

import json
import logging
import os
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from fastapi import HTTPException

import config
from sheets import sheets_leer_ventas_del_dia

logger = logging.getLogger("ferrebot.api")

def _hoy() -> str:
    return datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d")


def _hace_n_dias(n: int) -> datetime:
    return datetime.now(config.COLOMBIA_TZ) - timedelta(days=n)


# ── Helper: leer Excel histórico ──────────────────────────────────────────────
def _leer_excel_rango(dias: int | None = None, mes_actual: bool = False) -> list[dict]:
    if not os.path.exists(config.EXCEL_FILE):
        return []

    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True, data_only=True)
    except Exception:
        return []

    ahora = datetime.now(config.COLOMBIA_TZ)
    hojas_candidatas = [
        f"{config.MESES[ahora.month]} {ahora.year}",
    ]
    if ahora.month > 1:
        hojas_candidatas.append(f"{config.MESES[ahora.month - 1]} {ahora.year}")
    else:
        hojas_candidatas.append(f"{config.MESES[12]} {ahora.year - 1}")

    if dias is not None:
        limite = (_hace_n_dias(dias)).strftime("%Y-%m-%d")
    else:
        limite = None

    resultado = []
    for nombre_hoja in hojas_candidatas:
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]

        # En modo read_only ws.max_column puede ser None y ws.cell() no es confiable.
        # Leer headers iterando la fila exacta.
        cols: dict[str, int] = {}
        try:
            for fila_hdr in ws.iter_rows(
                min_row=config.EXCEL_FILA_HEADERS,
                max_row=config.EXCEL_FILA_HEADERS,
            ):
                for cell in fila_hdr:
                    if cell.value:
                        cols[str(cell.value).lower().strip()] = cell.column
                break
        except Exception:
            continue

        def _col(*claves) -> int | None:
            for k in claves:
                if k in cols:
                    return cols[k]
            return None

        c_fecha    = _col("fecha")
        c_hora     = _col("hora")
        c_producto = _col("producto")
        c_cantidad = _col("cantidad")
        c_precio   = _col("valor unitario", "precio unitario", "precio")
        c_total    = _col("total")
        c_alias    = _col("alias")
        c_vendedor = _col("vendedor")
        c_metodo   = _col("metodo de pago", "metodo pago", "método pago")
        c_num      = _col("#", "consecutivo", "num", "consecutivo de venta")
        c_unidad   = _col("unidad de medida", "unidad_medida", "unidad")

        for fila in ws.iter_rows(min_row=config.EXCEL_FILA_DATOS, values_only=True):
            if not any(fila):
                continue

            fecha_raw = fila[c_fecha - 1] if (c_fecha and c_fecha <= len(fila)) else None
            if fecha_raw is None:
                continue

            if isinstance(fecha_raw, datetime):
                fecha_str = fecha_raw.strftime("%Y-%m-%d")
            else:
                fecha_str = str(fecha_raw)[:10]

            if limite and fecha_str < limite:
                continue
            if mes_actual and not fecha_str.startswith(f"{ahora.year}-{ahora.month:02d}"):
                continue

            def _v(col_idx):
                if col_idx is None or col_idx > len(fila):
                    return ""
                v = fila[col_idx - 1]
                return v if v is not None else ""

            try:
                total = float(str(_v(c_total)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                total = 0.0

            try:
                precio_unit = float(str(_v(c_precio)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                precio_unit = 0.0

            resultado.append({
                "num":             _v(c_num),
                "fecha":           fecha_str,
                "hora":            str(_v(c_hora)),
                "id_cliente":      "CF",
                "cliente":         "Consumidor Final",
                "codigo_producto": "",
                "producto":        str(_v(c_producto)),
                "cantidad":        str(_v(c_cantidad)),
                "unidad_medida":   str(_v(c_unidad)) or "Unidad",
                "precio_unitario": precio_unit,
                "total":           total,
                "alias":           str(_v(c_alias)),
                "vendedor":        str(_v(c_vendedor)),
                "metodo":          str(_v(c_metodo)),
            })

    try:
        wb.close()
    except Exception:
        pass
    return resultado


# ── Helper: leer ventas desde Postgres ────────────────────────────────────────
def _leer_ventas_postgres(dias: int | None = None, mes_actual: bool = False) -> list[dict] | None:
    """
    Lee ventas desde PostgreSQL (ventas + ventas_detalle) y devuelve el mismo
    formato de dicts que _leer_excel_rango().

    Retorna None si Postgres no está disponible o falla, para que el caller
    pueda recurrir a Excel como fallback.
    """
    try:
        import db as _db
        if not _db.DB_DISPONIBLE:
            return None

        ahora = datetime.now(config.COLOMBIA_TZ)

        sql = """
            SELECT
                v.consecutivo AS num,
                v.fecha::text AS fecha,
                COALESCE(v.hora::text, '') AS hora,
                COALESCE(v.cliente_nombre, 'Consumidor Final') AS cliente,
                CASE WHEN v.cliente_id IS NULL THEN 'CF' ELSE v.cliente_id::text END AS id_cliente,
                d.producto_nombre AS producto,
                d.cantidad::text AS cantidad,
                COALESCE(d.unidad_medida, 'Unidad') AS unidad_medida,
                COALESCE(d.precio_unitario, 0)::float AS precio_unitario,
                COALESCE(d.total, 0)::float AS total,
                COALESCE(d.alias_usado, '') AS alias,
                COALESCE(v.vendedor, '') AS vendedor,
                COALESCE(v.metodo_pago, '') AS metodo
            FROM ventas v
            JOIN ventas_detalle d ON d.venta_id = v.id
            WHERE 1=1
        """
        params: list = []

        if dias is not None:
            fecha_limite = (ahora - timedelta(days=dias)).strftime("%Y-%m-%d")
            sql += " AND v.fecha >= %s"
            params.append(fecha_limite)

        if mes_actual:
            primer_dia = ahora.replace(day=1).strftime("%Y-%m-%d")
            hoy_str = ahora.strftime("%Y-%m-%d")
            sql += " AND v.fecha >= %s AND v.fecha <= %s"
            params.append(primer_dia)
            params.append(hoy_str)

        sql += " ORDER BY v.fecha, v.consecutivo, d.id"

        rows = _db.query_all(sql, params if params else None)

        result = []
        for r in rows:
            result.append({
                "num":             r["num"],
                "fecha":           str(r["fecha"])[:10],
                "hora":            str(r.get("hora", "")),
                "id_cliente":      str(r.get("id_cliente", "CF")),
                "cliente":         str(r.get("cliente", "Consumidor Final")),
                "codigo_producto": "",
                "producto":        str(r.get("producto", "")),
                "cantidad":        str(r.get("cantidad", "")),
                "unidad_medida":   str(r.get("unidad_medida", "Unidad")) or "Unidad",
                "precio_unitario": float(r.get("precio_unitario", 0)),
                "total":           float(r.get("total", 0)),
                "alias":           str(r.get("alias", "")),
                "vendedor":        str(r.get("vendedor", "")),
                "metodo":          str(r.get("metodo", "")),
            })
        return result

    except Exception as e:
        logger.warning(f"Postgres ventas read failed: {e}")
        return None


# ── Redirección de inventario: productos que se almacenan bajo otra clave ─────
# Formato: clave_producto → (clave_inventario_real, divisor_para_mostrar_stock)
#   Waypers: inventario en UNIDADES, se muestra en kg  (divisor = 12)
#   Carbonato x Kg: inventario en KG en la bolsa, se muestra tal cual (divisor = 1)
_WAYPER_KG_KEYS = {
    "wayper_blanco":   ("wayper_blanco_unidad",  12.0),
    "wayper_de_color": ("wayper_de_color_unidad", 12.0),
    # Carbonato por kilo → stock vive en la bolsa de 25 kg (en kg)
    "carbonato_x_kg":  ("carbonato_x_25_kg",       1.0),
}

def _stock_wayper(key: str, inventario: dict):
    """
    Para productos cuyo inventario vive bajo otra clave, aplica la conversión
    correspondiente y devuelve el stock en la unidad de venta.
    Para el resto devuelve el stock directo.
    """
    if key in _WAYPER_KG_KEYS:
        clave_inv, divisor = _WAYPER_KG_KEYS[key]
        inv_raw = inventario.get(clave_inv)
        if inv_raw is not None:
            cantidad = inv_raw.get("cantidad") if isinstance(inv_raw, dict) else inv_raw
            if cantidad is not None:
                return round(cantidad / divisor, 2)
        return None
    raw = inventario.get(key)
    if raw is None:
        return None
    return raw.get("cantidad") if isinstance(raw, dict) else raw


def _leer_excel_compras(dias: int | None = None) -> list[dict]:
    """Lee la hoja 'Compras' del Excel de ventas."""
    if not os.path.exists(config.EXCEL_FILE):
        return []
    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE, read_only=True, data_only=True)
    except Exception:
        return []
    if "Compras" not in wb.sheetnames:
        wb.close()
        return []
    ws     = wb["Compras"]
    ahora  = datetime.now(config.COLOMBIA_TZ)
    limite = (ahora - timedelta(days=dias)).strftime("%Y-%m-%d") if dias else None
    resultado = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        fecha = str(row[0])[:10] if row[0] else ""
        if limite and fecha < limite:
            continue
        resultado.append({
            "fecha":          fecha,
            "hora":           str(row[1] or ""),
            "proveedor":      str(row[2] or "—"),
            "producto":       str(row[3] or ""),
            "cantidad":       _to_float(row[4]),
            "costo_unitario": _to_float(row[5]),
            "costo_total":    _to_float(row[6]),
        })
    wb.close()
    return sorted(resultado, key=lambda x: x["fecha"])


def _to_float(val) -> float:
    try:
        return float(str(val).replace(",", ".") or 0)
    except (ValueError, TypeError):
        return 0.0


def _cantidad_a_float(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        pass
    if "/" in s and " " not in s:
        parts = s.split("/")
        try:
            return float(parts[0]) / float(parts[1])
        except (ValueError, ZeroDivisionError):
            return 0.0
    if " y " in s:
        partes = s.split(" y ")
        try:
            entero = float(partes[0])
            frac_parts = partes[1].split("/")
            frac = float(frac_parts[0]) / float(frac_parts[1])
            return entero + frac
        except (ValueError, IndexError, ZeroDivisionError):
            return 0.0
    return 0.0
