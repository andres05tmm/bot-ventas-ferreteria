"""
Integración con Claude AI (modelo: claude-haiku-4-5-20251001):
- Construcción del system prompt con contexto del negocio
- Llamada a la API de Claude con PROMPT CACHING (ahorro ~60% en tokens de input)
- Parseo y ejecución de acciones embebidas en la respuesta ([VENTA]...[/VENTA], etc.)

OPTIMIZACIONES DE COSTO ACTIVAS:
  1. Prompt caching  — la parte estática del prompt (reglas + catálogo) se cachea 5 min.
                       Costo de tokens cacheados = 10% del precio normal.
  2. Historial corto — se envían solo los últimos 1-4 mensajes (adaptativo).
  3. max_tokens cap  — techo adaptativo de respuesta.
  4. Catálogo simplificado — parte estática solo precio base, fracciones vía MATCH dinámico (~26% menos tokens cacheados).

CORRECCIONES v4:
  - Bug precedencia and/or en filtro tornillos drywall corregido
  - _quitar_tildes (redefinida en loop) eliminada, reemplazada por _normalizar
  - Todos los `import re as _re*` dentro de funciones eliminados (re ya importado al top)
"""

import logging
import os
import asyncio
import json
import re
import traceback
from datetime import datetime

import config
import skill_loader  # Skills separados por archivo .md
import alias_manager
import bypass
import fuzzy_match
# Cache RAM de precios recién actualizados (override del cache de Anthropic, TTL 5 min)
import time as _time
_precios_recientes: dict = {}  # {nombre_lower: (precio, timestamp)}
_PRECIO_TTL = 300  # 5 minutos

def _registrar_precio_reciente(nombre_lower: str, precio: float, fraccion: str = None):
    # Limpiar entradas anteriores del mismo producto antes de guardar
    claves_borrar = [k for k in _precios_recientes if k == nombre_lower or k.startswith(nombre_lower + "___")]
    for k in claves_borrar:
        del _precios_recientes[k]
    key = f"{nombre_lower}___{fraccion}" if fraccion else nombre_lower
    _precios_recientes[key] = (precio, _time.time())

def _get_precios_recientes_activos() -> dict:
    ahora = _time.time()
    return {k: v[0] for k, v in _precios_recientes.items() if ahora - v[1] < _PRECIO_TTL}


from memoria import (
    cargar_memoria, guardar_memoria, invalidar_cache_memoria,
    buscar_producto_en_catalogo, buscar_multiples_en_catalogo,
    buscar_multiples_con_alias,
    obtener_precios_como_texto, obtener_info_fraccion_producto,
    cargar_inventario, cargar_caja, cargar_gastos_hoy,
    obtener_resumen_caja, guardar_gasto,
    guardar_fiado_movimiento, abonar_fiado,
    actualizar_precio_en_catalogo,
)
from excel import (
    obtener_todos_los_datos, obtener_resumen_ventas,
    generar_excel_personalizado, guardar_cliente_nuevo,
    inicializar_excel, buscar_clientes_multiples, _normalizar,
)
from utils import convertir_fraccion_a_decimal, decimal_a_fraccion_legible

# ─────────────────────────────────────────────
# ALIAS DE FERRETERÍA (pre-procesamiento)
# ─────────────────────────────────────────────

_ALIAS_FERRETERIA = [
    # (patrón regex, reemplazo)
    # PUNTILLAS: normalizar abreviaturas y quitar "caja de" → "puntilla X"
    (r'\bcaja[s]?\s+de\s+puntilla[s]?\b', r'puntilla'),   # "caja de puntilla" → "puntilla"
    (r'\bpuntilla[s]?\s+(.*?)\bs\.c\.?\b', r'puntilla \g<1> sin cabeza'),  # s.c → sin cabeza
    (r'\bpuntilla[s]?\s+(.*?)\bc\.c\.?\b', r'puntilla \g<1> con cabeza'),  # c.c → con cabeza
    (r'\bpuntilla[s]?\s+(.*?)\bsc\b',      r'puntilla \g<1> sin cabeza'),  # sc → sin cabeza (sin puntos)
    (r'\bpuntilla[s]?\s+(.*?)\bcc\b',      r'puntilla \g<1> con cabeza'),  # cc → con cabeza (sin puntos)
    (r'\bs\.c\.?\b', r'sin cabeza'),   # s.c genérico
    (r'\bc\.c\.?\b', r'con cabeza'),   # c.c genérico
    (r'\bsc\b(?=.*puntilla|\bpuntilla)', r'sin cabeza'),  # sc genérico cerca de puntilla
    # TORNILLOS DRYWALL: normalizar medidas para evitar confusión (6x3 vs 6x3/4)
    # Importante: estos patrones van PRIMERO para que se apliquen antes de otros
    (r'\btornillo[s]?\s*(?:de\s*)?drywall\s*(\d+)\s*[xX]\s*3\b(?!/)', r'tornillo drywall \g<1>x3'),
    (r'\bdrywall\s*(\d+)\s*[xX]\s*3\b(?!/)', r'drywall \g<1>x3'),
    (r'\b(\d+)\s*[xX]\s*3\b(?!/)\s*(?=.*(?:tornillo|drywall))', r'\g<1>x3'),
    # Rodillo sin medida → Rodillo Convencional
    # Evita que "3 rodillos" matchee "Rodillo de 1"", "Rodillo de 2"", etc.
    # Solo aplica cuando NO va seguido de una medida explícita (número o pulgadas)
    (r'\b(\d+)\s+rodillos?\b(?!\s*(?:de\s+)?\d)', lambda m: f"{m.group(1)} rodillo convencional"),
    # Pegaternit: normalizar variantes de escritura
    (r'\bpagaternit\b', r'pegaternit'),
    (r'\bpega\s*ternit\b', r'pegaternit'),
    (r'\bpegaeternit\b', r'pegaternit'),
    # Esmalte 3en1: normalizar variantes sin espacios
    (r'\b3en1\b', r'3 en 1'),
    (r'\b3-en-1\b', r'3 en 1'),
    # Thinner/Varsol: litro=1/4 galón (8000), botella/botellita=1/10 galón (4000).
    # Convertimos directo a precio total antes de llegar a Claude.
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*botellitas?\s+(?:de\s+)?thinner\b',
        lambda m: f"{int(m.group(1) or 1) * 4000} thinner" if int(m.group(1) or 1) > 1 else "thinner 4000"),
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*botellitas?\s+(?:de\s+)?varsol\b',
        lambda m: f"{int(m.group(1) or 1) * 4000} varsol" if int(m.group(1) or 1) > 1 else "varsol 4000"),
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*botellas?\s+(?:de\s+)?thinner\b',
        lambda m: f"{int(m.group(1) or 1) * 4000} thinner" if int(m.group(1) or 1) > 1 else "thinner 4000"),
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*botellas?\s+(?:de\s+)?varsol\b',
        lambda m: f"{int(m.group(1) or 1) * 4000} varsol" if int(m.group(1) or 1) > 1 else "varsol 4000"),
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*litros?\s+(?:de\s+)?thinner\b',
        lambda m: f"{int(m.group(1) or 1) * 8000} thinner" if int(m.group(1) or 1) > 1 else "thinner 8000"),
    (r'\b(?:un[ao]?\s+)?(\d+)?\s*litros?\s+(?:de\s+)?varsol\b',
        lambda m: f"{int(m.group(1) or 1) * 8000} varsol" if int(m.group(1) or 1) > 1 else "varsol 8000"),
    # Thinner/Varsol por galones (cantidades >= 1/2 galón)
    # "1-1/2 galón de thinner", "1 y medio galón de thinner", "2-1/2 galones thinner"
    (r'\b(\d+)\s*-\s*1/2\s*(?:galon(?:es)?)\s*(?:de\s*)?(thinner|varsol)\b', r'\g<1>.5 galones \g<2>'),
    (r'\b(\d+)\s+y\s+(?:medio|1/2)\s*(?:galon(?:es)?)\s*(?:de\s*)?(thinner|varsol)\b', r'\g<1>.5 galones \g<2>'),
    (r'\b(\d+)\s+(?:galon(?:es)?)\s+y\s+(?:medio|1/2)\s*(?:de\s*)?(thinner|varsol)\b', r'\g<1>.5 galones \g<2>'),
    # "medio galón de thinner", "1/2 galón thinner"  
    (r'\b(?:medio|1/2)\s*(?:galon)?\s*(?:de\s*)?(thinner|varsol)\b', r'0.5 galones \g<1>'),
    # "2 galones de thinner", "1 galón thinner"
    (r'\b(\d+)\s*(?:galon(?:es)?)\s*(?:de\s*)?(thinner|varsol)\b', r'\g<1> galones \g<2>'),
]

def aplicar_alias_ferreteria(mensaje: str) -> str:
    """Transforma alias comunes antes de enviar a Claude."""
    resultado = alias_manager.aplicar_aliases_dinamicos(mensaje)
    for patron, reemplazo in _ALIAS_FERRETERIA:
        if callable(reemplazo):
            # Lambda/función: re.sub la llama directamente con el match
            resultado = re.sub(patron, reemplazo, resultado, flags=re.IGNORECASE)
        elif r'\g<1>' in reemplazo or r'\g<2>' in reemplazo:
            def _hacer_reemplazo(m, repl=reemplazo):
                resultado_repl = repl
                try:
                    g1 = m.group(1) if m.lastindex and m.lastindex >= 1 and m.group(1) else "1"
                    resultado_repl = resultado_repl.replace(r'\g<1>', g1)
                except IndexError:
                    resultado_repl = resultado_repl.replace(r'\g<1>', "1")
                try:
                    if m.lastindex and m.lastindex >= 2 and m.group(2):
                        resultado_repl = resultado_repl.replace(r'\g<2>', m.group(2))
                except IndexError:
                    pass
                return resultado_repl.strip()
            resultado = re.sub(patron, _hacer_reemplazo, resultado, flags=re.IGNORECASE)
        else:
            resultado = re.sub(patron, reemplazo, resultado, flags=re.IGNORECASE)
    return resultado
  
# ─────────────────────────────────────────────
# PARTE ESTÁTICA DEL SYSTEM PROMPT (cacheable)
# ─────────────────────────────────────────────

def _construir_parte_estatica(memoria: dict) -> str:
    """
    Construye la parte del system prompt que NO cambia entre mensajes.
    Al ser idéntica en todas las llamadas, Anthropic la cachea automáticamente.
    """
    catalogo = memoria.get("catalogo", {})

    def _linea_producto_simple(prod):
        # Solo nombre:precio_unidad — las fracciones llegan via MATCH en la parte dinámica
        # Ahorra ~1960 tokens cacheados vs incluir fracciones completas
        pxc = prod.get("precio_por_cantidad")
        if pxc:
            return f"{prod['nombre']}:{pxc['precio_bajo_umbral']}/{pxc['precio_sobre_umbral']}x{pxc['umbral']}"
        else:
            return f"{prod['nombre']}:{prod['precio_unidad']}"

    if catalogo:
        # Catálogo simplificado: precio_unidad solamente (sin fracciones)
        # Las fracciones completas se inyectan en la parte dinámica via MATCH
        # cuando el producto es mencionado en el mensaje
        categorias: dict = {}
        for prod in catalogo.values():
            cat = prod.get("categoria", "Otros")
            categorias.setdefault(cat, []).append(_linea_producto_simple(prod))
        lineas_cat = []
        for cat, items in sorted(categorias.items()):
            lineas_cat.append(f"{cat}:")
            lineas_cat.extend(items[:60])
        precios_texto = "\n".join(lineas_cat)
    else:
        precios_texto = obtener_precios_como_texto()

    precios_fraccion_mem = memoria.get("precios_fraccion", {})
    if precios_fraccion_mem:
        lineas_frac = [
            f"{prod_key} {frac}:{precio}"
            for prod_key, fracs in precios_fraccion_mem.items()
            for frac, precio in fracs.items()
        ]
        precios_fraccion_texto = "FRACCIONES EXTRA:\n" + "\n".join(lineas_frac)
    else:
        precios_fraccion_texto = ""

    # En MODO_MATCH_ONLY: catálogo se omite del estático — llega dinámicamente via MATCH
    # o como fallback completo si MATCH no encuentra nada. Cache estable con ~1235 tokens (reglas).
    # En modo normal: catálogo simplificado (solo precio base, sin fracciones) → 26% menos tokens
    _match_only = os.getenv("MODO_MATCH_ONLY", "false").lower() == "true"

    if _match_only:
        # Solo fracciones extra si las hay — el catálogo llega en la parte dinámica
        catalogo_seccion = precios_fraccion_texto
    else:
        catalogo_seccion = (
            "CATALOGO(nombre:precio_galon_o_unidad. Fracciones exactas en MATCH):\n"
            + precios_texto
            + ("\n" + precios_fraccion_texto if precios_fraccion_texto else "")
        ) if precios_texto else precios_fraccion_texto

    negocio_json = json.dumps(memoria.get("negocio", {}), ensure_ascii=False)

    # Skills estáticos: core + precios_base (siempre necesarios, muy cacheables)
    skills_estaticos = skill_loader.obtener_skills_estaticos()

    return f"""{skills_estaticos}

INFORMACION DEL NEGOCIO: {negocio_json}

{catalogo_seccion}"""

# ─────────────────────────────────────────────
# PARTE DINÁMICA DEL SYSTEM PROMPT (por mensaje)
# ─────────────────────────────────────────────

def _construir_parte_dinamica(mensaje_usuario: str, nombre_usuario: str, memoria: dict) -> str:
    """
    Construye la parte del system prompt que SÍ cambia entre mensajes:
    candidatos del catálogo, cliente encontrado, ventas del día, inventario, caja, etc.
    """
    # ── Resumen de ventas ──
    resumen_sheets_total    = 0
    resumen_sheets_cantidad = 0
    if config.SHEETS_ID and config.SHEETS_DISPONIBLE:
        try:
            from sheets import sheets_leer_ventas_del_dia
            ventas_hoy = sheets_leer_ventas_del_dia()
            for v in ventas_hoy:
                try:
                    resumen_sheets_total    += float(v.get("total", 0) or 0)
                    resumen_sheets_cantidad += 1
                except (ValueError, TypeError):
                    pass
        except Exception:
            pass

    resumen               = obtener_resumen_ventas()
    resumen_excel_total   = resumen["total"]      if resumen else 0
    resumen_excel_cantidad = resumen["num_ventas"] if resumen else 0

    total_mes    = resumen_excel_total
    cantidad_mes = resumen_excel_cantidad

    resumen_texto = (
        f"${total_mes:,.0f} en {cantidad_mes} ventas este mes "
        f"(hoy: ${resumen_sheets_total:,.0f} en {resumen_sheets_cantidad} ventas)"
    ) if cantidad_mes > 0 else "Sin ventas este mes"

    # ── Datos históricos (solo si piden análisis) ──
    palabras_analisis = ["cuanto", "vendimos", "reporte", "analiz", "total",
                         "resumen", "estadistica", "top", "mas vendido"]
    if any(p in mensaje_usuario.lower() for p in palabras_analisis):
        try:
            todos       = obtener_todos_los_datos()
            datos_texto = json.dumps(todos[-100:], ensure_ascii=False, default=str) if todos else "Sin datos aun"
        except Exception:
            datos_texto = "Sin datos aun"
    else:
        datos_texto = "(no cargado)"

    stopwords = {"que", "del", "los", "las", "una", "uno", "con", "por", "para", "como",
                 "fue", "son", "precio", "vale", "cuesta", "cuanto", "la", "el", "de", "en",
                 "galon", "galones", "litro", "litros", "kilo", "kilos", "metro", "metros",
                 "pulgada", "pulgadas", "unidad", "unidades",
                 "botella", "botellas",
                 "vendi", "vendo", "vendimos", "dame", "quiero", "necesito", "par",
                 # palabras de cantidad fraccionaria — no son nombre de producto
                 "y", "un", "cuarto", "medio", "media", "octavo", "tres"}

    def _es_keyword_relevante(p: str) -> bool:
        """Determina si una palabra debe incluirse como keyword de búsqueda."""
        if p in stopwords:
            return False
        if len(p) > 2:
            return True
        if p.isdigit():
            return True
        # Incluir códigos de variante de 2 chars: t1, t2, t3, x1, 6x, 8x, etc.
        _TALLAS_PC = {"xl", "xs", "xxl", "s", "m", "l"}
        if (len(p) == 2 and any(c.isdigit() for c in p)) or p in _TALLAS_PC:
            return True
        return False

    palabras_clave = [p for p in mensaje_usuario.lower().split() if _es_keyword_relevante(p)]
    info_fracciones_extra = ""

    # Mapa decimal → fracción para buscar precio especial
    _dec_a_frac = {0.5: "1/2", 0.25: "1/4", 0.75: "3/4", 0.125: "1/8", 0.0625: "1/16"}
    _frac_a_dec = {"1/2": 0.5, "1/4": 0.25, "3/4": 0.75, "1/8": 0.125, "1/16": 0.0625}

    def _extraer_cantidad_mixta(msg: str):
        """
        Extrae (n_enteros, frac_key, cantidad_total) del mensaje.
        Maneja todas las formas: '2 y medio', '2-1/2', '1.5', '2.5 galones', etc.
        Retorna (None, None, None) si no encuentra cantidad mixta.
        """
        m = msg.lower()

        # Forma decimal: 2.5, 1.5, 3.5 (producida por alias)
        match_dec = re.search(r'(\d+)\.5\b', m)
        if match_dec:
            enteros = int(match_dec.group(1))
            return enteros, "1/2", enteros + 0.5

        match_dec25 = re.search(r'(\d+)\.25\b', m)
        if match_dec25:
            enteros = int(match_dec25.group(1))
            return enteros, "1/4", enteros + 0.25

        match_dec75 = re.search(r'(\d+)\.75\b', m)
        if match_dec75:
            enteros = int(match_dec75.group(1))
            return enteros, "3/4", enteros + 0.75

        # Forma escrita: "2 y medio", "2 y media", "2-1/2", "2 1/2"
        map_frac_texto = {
            r'y\s+medio': "1/2", r'y\s+media': "1/2", r'y\s+un\s+medio': "1/2",
            r'(?<!\d)1/2': "1/2",
            r'y\s+cuarto': "1/4", r'y\s+un\s+cuarto': "1/4", r'(?<!\d)1/4': "1/4",
            r'tres\s+cuartos': "3/4", r'(?<!\d)3/4': "3/4",
            r'y\s+octavo': "1/8", r'y\s+un\s+octavo': "1/8", r'(?<!\d)1/8': "1/8",
        }
        map_enteros_texto = {
            # Solo detectar entero si está SEPARADO de fracciones:
            # - palabras textuales: "un", "dos", "tres"...
            # - número seguido de espacio y luego "y" o palabra (no fracción)
            # Ej: "1 galón y" → sí | "1/4" → no | "24 tornillos" → no
            r'\bun\b(?!\s*/)': 1, r'\buno\b': 1,
            r'\b1\b(?!\s*/)': 1,
            r'\bdos\b': 2, r'\b2\b(?!\s*/)': 2,
            r'\btres\b': 3, r'\b3\b(?!\s*/)': 3,
            r'\bcuatro\b': 4, r'\b4\b(?!\s*/)': 4,
            r'\bcinco\b': 5, r'\b5\b(?!\s*/)': 5,
        }

        # Patrón especial N-1/frac: "2-1/2", "3-1/4", etc. — extraer N directamente
        match_guion = re.search(r'\b(\d+)-1/(\d+)', m)
        if match_guion:
            enteros  = int(match_guion.group(1))
            divisor  = int(match_guion.group(2))
            frac_map = {2: "1/2", 4: "1/4", 8: "1/8"}
            frac_g   = frac_map.get(divisor)
            if frac_g:
                return enteros, frac_g, enteros + _frac_a_dec[frac_g]

        # Detectar si hay fracción — pero SOLO las que vienen precedidas de "y"
        # son candidatas a cantidad mixta. Las fracciones solas (1/4 laca) se ignoran.
        _patrones_con_y = [
            r'y\s+medio', r'y\s+media', r'y\s+un\s+medio',
            r'y\s+cuarto', r'y\s+un\s+cuarto',
            r'tres\s+cuartos',
            r'y\s+octavo', r'y\s+un\s+octavo',
        ]
        _patrones_fraccion_numerica = [
            (r'(?<!\d)1/2', "1/2"), (r'(?<!\d)3/4', "3/4"),
            (r'(?<!\d)1/4', "1/4"), (r'(?<!\d)1/8', "1/8"),
        ]

        frac_key = None
        # Primero buscar fracciones con "y" (garantizado mixto)
        for pat, v in map_frac_texto.items():
            if pat in _patrones_con_y and re.search(pat, m):
                frac_key = v
                break

        # Si no encontró con "y", buscar fracción numérica SOLO si hay un entero antes
        if not frac_key:
            for pat, v in _patrones_fraccion_numerica:
                match_frac = re.search(pat, m)
                if match_frac:
                    # Verificar que haya un número entero ANTES de la fracción
                    texto_antes = m[:match_frac.start()].strip()
                    if re.search(r'\b[1-9]\d*\s+(?:galon|galones|litro|litros|y)\s*$', texto_antes):
                        frac_key = v
                        break

        if not frac_key:
            return None, None, None

        n_enteros = next((v for pat, v in map_enteros_texto.items() if re.search(pat, m)), None)
        if not n_enteros:
            return None, None, None

        return n_enteros, frac_key, n_enteros + _frac_a_dec[frac_key]

    # Detectar si hay fracciones o decimales mixtos en el mensaje
    _msg_lower_pre = mensaje_usuario.lower()
    _tiene_mixto = any(p in _msg_lower_pre for p in
                       ["1/4", "1/2", "3/4", "1/8", "1/16", "cuarto", "medio", "mitad",
                        "octavo", ".5", ".25", ".75"])

    if _tiene_mixto:
        palabras_msg = _msg_lower_pre.split()
        calculos_multi = []  # lista de (prod, n_enteros, frac_key, cantidad_calc, total_calc)

        # Segmentar por producto para manejar multi-producto
        def _norm_seg(s):
            return (s.lower()
                    .replace("á","a").replace("é","e").replace("í","i")
                    .replace("ó","o").replace("ú","u").replace("ñ","n"))

        _segs = re.split(r'[,\n]+', mensaje_usuario.lower())

        for seg in _segs:
            seg = _norm_seg(seg).strip()
            if not seg:
                continue
            # Buscar producto con fracciones en este segmento
            palabras_seg = seg.split()
            # Extraer fracción una sola vez por segmento (no depende del producto)
            n_enteros, frac_key, cantidad_calc = _extraer_cantidad_mixta(seg)

            for largo in [4, 3, 2]:
                encontrado_seg = False
                for i in range(len(palabras_seg) - largo + 1):
                    fragmento = " ".join(palabras_seg[i:i + largo])
                    prod      = buscar_producto_en_catalogo(fragmento)
                    if prod and prod.get("precios_fraccion"):
                        fracs = prod.get("precios_fraccion", {})
                        # Solo usar este producto si tiene la fracción que necesitamos.
                        # Si no, seguir buscando — puede haber un fragmento más específico
                        # adelante en el segmento que sí la tenga.
                        if n_enteros and frac_key and frac_key in fracs:
                            # Precio del galón: fracs["1"] si existe, sino precio_unidad
                            if "1" in fracs:
                                p_galon = fracs["1"]["precio"] if isinstance(fracs.get("1"), dict) else fracs.get("1", 0)
                            else:
                                p_galon = prod.get("precio_unidad", 0)
                            p_frac = fracs[frac_key]["precio"] if isinstance(fracs.get(frac_key), dict) else fracs.get(frac_key, 0)
                            if p_galon and p_frac:
                                total_calc = p_galon * n_enteros + p_frac
                                calculos_multi.append((prod["nombre"], n_enteros, frac_key, cantidad_calc, total_calc, p_galon, p_frac))
                                encontrado_seg = True
                                break  # producto correcto encontrado para este segmento
                        # Producto encontrado pero sin la fracción requerida → seguir buscando
                if encontrado_seg:
                    break

        if calculos_multi:
            lineas = []
            for nombre, n_ent, frac, cant, total, p_gal, p_fr in calculos_multi:
                lineas.append(
                    f"{nombre}: cantidad={cant}, total={total} "
                    f"({n_ent}x${p_gal:,} + {frac}=${p_fr:,})"
                )
            info_fracciones_extra = (
                "TOTALES PRECALCULADOS (USA EXACTAMENTE, NO recalcules):\n"
                + "\n".join(lineas)
            )
            print(f"[PRECALCULADO DEBUG]\n{info_fracciones_extra}")
        else:
            print("[PRECALCULADO DEBUG] No se generó precalculado para este mensaje")

    # ── Precalcular tornillos mayorista ──────────────────────────────────────
    # Estos casos no los cubre el bloque mixto de arriba.
    # Construimos totales determinísticos antes de que Claude los vea.
    _lineas_pre_extra = []
    _segs_pre = re.split(r'[,\n]+', mensaje_usuario.lower())
    for _seg in _segs_pre:
        _seg = _seg.strip()
        if not _seg:
            continue
        # Extraer cantidad entera del segmento (primer número)
        _m_cant = re.match(r'^[^\d]*(\d+)', _seg)
        if not _m_cant:
            continue
        _cant = int(_m_cant.group(1))

        # Buscar producto en el segmento
        _palabras = _seg.split()
        _prod_encontrado = None
        for _largo in [4, 3, 2, 1]:
            for _i in range(len(_palabras) - _largo + 1):
                _frag = " ".join(_palabras[_i:_i+_largo])
                if len(_frag) < 3:
                    continue
                _p = buscar_producto_en_catalogo(_frag)
                if _p:
                    _prod_encontrado = _p
                    break
            if _prod_encontrado:
                break

        if not _prod_encontrado:
            continue

        _pxc = _prod_encontrado.get("precio_por_cantidad")
        _fracs = _prod_encontrado.get("precios_fraccion", {})
        _nombre = _prod_encontrado["nombre"]

        # Tornillos/mayorista: calcular precio correcto según umbral
        if _pxc and _cant > 0:
            _umbral = _pxc.get("umbral", 50)
            _p_bajo = _pxc.get("precio_bajo_umbral", 0)
            _p_sobre = _pxc.get("precio_sobre_umbral", 0)
            if _p_bajo and _p_sobre:
                _precio_u = _p_sobre if _cant >= _umbral else _p_bajo
                _total = _cant * _precio_u
                _tier = f"mayorista x{_umbral}+" if _cant >= _umbral else "normal"
                _lineas_pre_extra.append(
                    f"{_nombre}: cantidad={_cant}, precio_unit={_precio_u}({_tier}), total={_total}"
                )

    if _lineas_pre_extra:
        _bloque_pre = (
            "TOTALES PRECALCULADOS (USA EXACTAMENTE, NO recalcules):\n"
            + "\n".join(_lineas_pre_extra)
        )
        if info_fracciones_extra:
            info_fracciones_extra += "\n" + _bloque_pre
        else:
            info_fracciones_extra = _bloque_pre
        print(f"[PRECALCULADO EXTRA]\n{_bloque_pre}")

    # ── Candidatos del catálogo para este mensaje específico ──
    info_candidatos_extra = ""
    # palabras_clave ya definida arriba con _es_keyword_relevante (incluye t1/t2/t3)

    # Detectar fracciones y cantidades mixtas mencionadas en el mensaje
    _fracs_mencionadas = set()
    _msg_lower = mensaje_usuario.lower()
    _mapa_palabras = {
        "cuarto": "1/4", "un cuarto": "1/4",
        "medio": "1/2",  "media": "1/2",  "un medio": "1/2",
        "octavo": "1/8", "un octavo": "1/8",
        "tres cuartos": "3/4",
        "botella": "botella", "botellas": "botella",
        "litro": "litro",    "litros": "litro",
    }
    for palabra, frac in _mapa_palabras.items():
        if palabra in _msg_lower:
            _fracs_mencionadas.add(frac)
    for token in _msg_lower.split():
        if token in ("1/4","1/2","3/4","1/8","1/16","3/8"):
            _fracs_mencionadas.add(token)

    # Tokenizar mensaje para detectar fracciones adyacentes a productos
    _tokens = mensaje_usuario.lower().replace(",","").split()
    _fracs_set = {"1/4","1/2","3/4","1/8","1/16","3/8","botella","litro"}

    def _linea_candidato(p: dict) -> str:
        # Formato comprimido: sin "  - ", sin "$", sin comas, fraccion relevante marcada con *
        fracs = p.get("precios_fraccion", {})
        pxc   = p.get("precio_por_cantidad")
        if fracs:
            nl = p.get("nombre_lower", "")
            palabras_prod = [w for w in nl.split() if len(w) > 3]
            frac_este_prod = None
            _tok = _msg_lower.replace(",","").split()
            for idx_t, tok in enumerate(_tok):
                if tok in _fracs_set:
                    ventana = " ".join(_tok[idx_t:idx_t+5])
                    if any(pp in ventana for pp in palabras_prod):
                        frac_este_prod = tok
                        break
            lineas_frac = []
            # Siempre incluir precio_unidad como "1=X" primero — es el precio del
            # galón/unidad completa. Sin esto, Claude no puede calcular fracciones
            # mixtas como "1 galón y un cuarto" porque no tiene el precio base.
            precio_unidad = p.get("precio_unidad", 0)
            if precio_unidad:
                marca_1 = "*" if frac_este_prod == "1" else ""
                lineas_frac.append(f"1={precio_unidad}{marca_1}")
            for k, v in fracs.items():
                if k == "1":
                    continue  # evitar duplicar si ya está en precios_fraccion
                precio = v['precio'] if isinstance(v, dict) else v
                marca = "*" if k == frac_este_prod else ""
                lineas_frac.append(f"{k}={precio}{marca}")
            return f"{p['nombre']}:" + "|".join(lineas_frac)
        elif pxc:
            return f"{p['nombre']}:{pxc['precio_bajo_umbral']}/{pxc['precio_sobre_umbral']}x{pxc['umbral']}"
        else:
            return f"{p['nombre']}:{p['precio_unidad']}"

    if palabras_clave:
        # FIX MULTI-PRODUCTO: segmentar el mensaje por producto para que cada uno
        # tenga garantizado su candidato, sin que unos "aplasten" a otros.
        # Ej: "1/4 vinilo blanco, 1/2 laca miel, 3/4 thinner" → 3 segmentos independientes
        # Separar solo por coma. NO separar por 'y' — puede ser parte de cantidad mixta
        # Ej: '1 galón y un cuarto vinilo' NO debe partirse
        _segmentos_raw = re.split(r'[,\n]+', mensaje_usuario.lower())
        _segmentos = []
        for seg in _segmentos_raw:
            seg = seg.strip()
            if len(seg) > 3:
                _segmentos.append(seg)

        combinados = {}
        _candidatos_garantizados = {}  # nl → prod: el mejor hit por segmento, siempre incluido

        # Familias donde hay múltiples tallas/variantes — necesitamos límite más alto
        _familias_con_tallas = {"brocha", "rodillo", "lija", "disco", "tornillo", "chazo",
                                 "tuerca", "arandela", "bisagra", "candado", "manguera",
                                 "lampara", "foco", "cable", "codo", "tee", "reduccion"}

        # Palabras de acción al inicio del mensaje que no son producto
        _palabras_accion = {"vendi", "vende", "vendí", "vender", "cobré", "cobre", "cobrar",
                             "dame", "deme", "dar", "quiero", "necesito", "compre", "compré"}

        # Stemming mínimo: quitar 's' final para que "lijas"→"lija", "discos"→"disco"
        def _stem(w):
            if w.endswith("les") and len(w) > 5:
                return w[:-2]   # aerosoles → aerosol
            if w.endswith("es") and len(w) > 4:
                return w[:-2]
            if w.endswith("s") and len(w) > 4:
                return w[:-1]   # tornillos → tornillo
            return w

        # 1. Buscar candidato por cada segmento de producto (garantiza uno por producto)
        for seg in _segmentos:
            # Limpiar símbolos de puntuación y basura al inicio del segmento
            seg = re.sub(r'^[^\w]+', '', seg).strip()
            # Normalizar tildes usando _normalizar (ya importada desde utils)
            seg = _normalizar(seg)
            # Quitar palabras de acción y cantidades iniciales del segmento
            palabras_raw = seg.split()
            # Saltar tokens no-alfanuméricos, palabras de acción y cantidades al inicio
            while palabras_raw and (
                not re.search(r'[\w\d]', palabras_raw[0]) or
                palabras_raw[0] in _palabras_accion or
                re.match(r'^[\d/\.]+$', palabras_raw[0])
            ):
                palabras_raw = palabras_raw[1:]
            # Saltar palabras de volumen/unidad inmediatas tras la cantidad
            _unidades_volumen = {"galon", "galones", "cuarto", "cuartos", "litro", "litros",
                                  "kilo", "kilos", "gramo", "gramos", "metro", "metros",
                                  "unidad", "unidades", "caja", "cajas", "bolsa", "bolsas",
                                  "rollo", "rollos", "par", "pares", "y", "un", "una",
                                  "botella", "botellas",
                                  "medio", "media", "cuarto"}
            while palabras_raw and palabras_raw[0] in _unidades_volumen:
                palabras_raw = palabras_raw[1:]

            # Incluir: palabras del nombre del producto + números que son tallas (van DESPUÉS del nombre base)
            # Los números como "3", "80", "100" solo se incluyen si no son la primera palabra
            # (para evitar que "3 cuartos de vinilo T1" incluya "3" que matchearía T3)
            palabras_seg = []
            nombre_producto_encontrado = False
            for p in palabras_raw:
                if p in stopwords:
                    continue
                # Tokens cortos: alfanuméricos (t1,t2) y tallas (xl,xs,s,m,l)
                _TALLAS_SEG = {"xl", "xs", "xxl", "s", "m", "l"}
                if (len(p) == 2 and any(c.isdigit() for c in p)) or p in _TALLAS_SEG:
                    palabras_seg.append(p)
                    nombre_producto_encontrado = True
                elif len(p) > 2 and not p.replace('.','').replace(',','').isdigit():
                    palabras_seg.append(_stem(p))  # con stemming
                    nombre_producto_encontrado = True
                elif re.match(r'^\d+x\d+', p):  # formatos como 3x3, 8x1
                    palabras_seg.append(p)
                    nombre_producto_encontrado = True
                elif nombre_producto_encontrado and p.isdigit() and 1 <= int(p) <= 999:
                    # Número de talla SOLO después de haber encontrado el nombre del producto
                    palabras_seg.append(p)

            if not palabras_seg:
                continue

            # Detectar si el segmento es de familia con tallas → usar límite más alto
            es_familia = any(f in seg.lower() or _stem(f) in seg.lower() for f in _familias_con_tallas)
            _limite_seg = 8 if es_familia else 3

            # Palabras que no identifican un producto por sí solas (no usar en largo=1)
            _SOLO_ADJETIVOS = {"pequeno", "pequeña", "pequeño", "grande", "economico",
                                "economica", "simple", "corriente", "basico", "normal",
                                "plastico", "plastica", "metalico", "metalica", "acero",
                                "madera", "hierro", "metal", "comun", "especial"}
            for largo in [4, 3, 2, 1]:
                encontrado_seg = False
                for i in range(len(palabras_seg) - largo + 1):
                    fragmento = " ".join(palabras_seg[i:i + largo])
                    if len(fragmento) < 3:
                        continue
                    # En largo=1, no buscar con palabras que son solo adjetivos/materiales
                    if largo == 1 and fragmento in _SOLO_ADJETIVOS:
                        continue
                    resultados = buscar_multiples_con_alias(fragmento, limite=_limite_seg)
                    primer = True
                    for prod in resultados:
                        nl = prod["nombre_lower"]
                        combinados[nl] = prod
                        if primer:
                            # El primer resultado es el mejor match — garantizarlo en la lista final
                            _candidatos_garantizados[nl] = prod
                            primer = False
                            print(f"[SEG DEBUG] seg='{seg[:30]}' frag='{fragmento}' garantizado='{prod['nombre']}'")
                        encontrado_seg = True
                    if encontrado_seg:
                        break
                if encontrado_seg:
                    break

        # 2. Búsqueda global adicional (fragmentos del mensaje completo)
        for largo in [4, 3, 2]:
            for i in range(len(palabras_clave) - largo + 1):
                frag_exact = " ".join(palabras_clave[i:i + largo])
                if len(frag_exact) < 4:
                    continue
                for prod in buscar_multiples_en_catalogo(frag_exact, limite=1):
                    if frag_exact in prod["nombre_lower"]:
                        combinados[prod["nombre_lower"]] = prod

        # 2.5 FILTRO TORNILLOS: evitar confusión entre medidas similares (6x3 vs 6x3/4)
        # Si el mensaje menciona una medida exacta como "6x3", eliminar productos con medidas
        # que la contengan pero sean diferentes (como "6x3/4", "6x3-1/2")
        _medidas_exactas = re.findall(r'\b(\d+)\s*[xX]\s*(\d+)\b(?![/\-])', _msg_lower)
        if _medidas_exactas and ("tornillo" in _msg_lower or "drywall" in _msg_lower):
            for calibre, largo_med in _medidas_exactas:
                medida_exacta = f"{calibre}x{largo_med}"
                # Filtrar productos que tengan la medida como substring pero NO sean exactos
                # Ej: si busca "6x3", eliminar "6x3/4" y "6x3-1/2" pero mantener "6x3"
                productos_a_eliminar = []
                for nl, prod in list(combinados.items()):
                    if "tornillo" in nl or "drywall" in nl:
                        # Buscar la medida en el nombre del producto
                        medida_prod = re.search(r'(\d+)[xX](\d+(?:[/-]\d+(?:/\d+)?)?)', nl)
                        if medida_prod:
                            medida_completa = medida_prod.group(0).lower()
                            # Si la medida del producto es más larga que la buscada, es diferente
                            if medida_exacta in medida_completa and medida_completa != medida_exacta:
                                productos_a_eliminar.append(nl)
                for nl in productos_a_eliminar:
                    if nl in combinados and nl not in _candidatos_garantizados:
                        del combinados[nl]

        # 3. Ordenar: más palabras del mensaje completo en el nombre = mayor prioridad
        #    Los candidatos garantizados (mejor hit por segmento) siempre se incluyen primero.
        #    Luego se agregan hasta 25 adicionales del pool general, ordenados por relevancia.
        _garantizados_lista = list(_candidatos_garantizados.values())
        _garantizados_nls   = set(_candidatos_garantizados.keys())
        _resto = sorted(
            [p for p in combinados.values() if p["nombre_lower"] not in _garantizados_nls],
            key=lambda p: sum(1 for w in palabras_clave if w in p["nombre_lower"]),
            reverse=True
        )
        candidatos = _garantizados_lista + _resto
        candidatos = candidatos[:max(len(_garantizados_lista), 25)]

        # Filtro de relevancia: se aplica SOLO al _resto (pool global), nunca a los
        # candidatos garantizados. Un candidato garantizado ya fue validado por el
        # segmentador (tiene su propio segmento que lo respaldó) — filtrarlo aquí con
        # las palabras del mensaje completo sería incorrecto en ventas multi-producto,
        # donde el mensaje tiene muchas palabras que no pertenecen a ese candidato.
        #
        # Para el _resto: comparar contra las palabras del segmento más cercano
        # es complejo, así que se usa el mensaje completo pero con umbral ajustado:
        #   - umbral fijo de 1 hit alfab: cualquier palabra del nombre debe aparecer en
        #     algún lugar del mensaje. Esto filtra productos completamente ajenos (ej:
        #     un tornillo que entró por un número suelto) sin afectar productos legítimos.

        def _stem_simple(w):
            if w.endswith("les") and len(w) > 5: return w[:-2]
            if w.endswith("es") and len(w) > 4:  return w[:-2]
            if w.endswith("s") and len(w) > 4:   return w[:-1]
            return w

        _palabras_alfab_msg = [
            w for w in palabras_clave
            if len(w) > 2 and not w.replace("/","").replace("-","").isdigit()
        ]

        def _es_relevante_resto(prod_nombre_lower):
            """Filtro para candidatos del pool global (no garantizados).
            Exige al menos 1 palabra alfabética del nombre en el mensaje completo."""
            if not _palabras_alfab_msg:
                return True
            nl = _normalizar(prod_nombre_lower)
            return any(
                _normalizar(w) in nl or _stem_simple(_normalizar(w)) in nl
                for w in _palabras_alfab_msg
            )

        # Garantizados: pasan siempre. Resto: pasan solo si son relevantes.
        candidatos = (
            _garantizados_lista +
            [p for p in _resto if _es_relevante_resto(p.get("nombre_lower", ""))]
        )

        if candidatos:
            lineas = [_linea_candidato(p) for p in candidatos]
            info_candidatos_extra = "MATCH:\n" + "\n".join(lineas)
            print(f"[CANDIDATOS DEBUG]\n{info_candidatos_extra}")
        else:
            _frag_fuzzy = " ".join(palabras_clave[:4]) if palabras_clave else ""
            _sugs = fuzzy_match.buscar_fuzzy(_frag_fuzzy) if _frag_fuzzy else []
            if _sugs:
                _lf = [f"  {_p['nombre']}:{_p.get('precio_unidad',0)} ({_s:.0f}% similar)"
                       for _p, _s in _sugs]
                info_candidatos_extra = (
                    "MATCH_DIFUSO (similares, NO exactos — pregunta cuál es):\n"
                    + "\n".join(_lf)
                )
            else:
                info_candidatos_extra = "MATCH: (sin resultados — producto no encontrado en catalogo)"
                print("[CANDIDATOS DEBUG] MATCH vacío — producto no en catálogo")

    # ── Clientes recientes ──
    clientes_recientes_texto = ""
    palabras_recientes = ["ultimo", "ultimos", "reciente", "recientes", "nuevo", "nuevos",
                          "anadido", "anadidos", "agregado", "agregados", "registrado", "registrados"]
    _msg_norm = _normalizar(mensaje_usuario)
    if any(p in _msg_norm for p in palabras_recientes) and "cliente" in _msg_norm:
        try:
            from excel import obtener_clientes_recientes
            recientes = obtener_clientes_recientes(5)
            if recientes:
                lineas = []
                for c in recientes:
                    nombre = c.get("Nombre tercero", "")
                    id_c   = c.get("Identificacion", "") or c.get("Identificacion", "")
                    tipo   = c.get("Tipo de identificacion", "") or c.get("Tipo de identificacion", "")
                    fecha  = c.get("Fecha registro", "Sin fecha")
                    lineas.append(f"  - {nombre} ({tipo}: {id_c}) — registrado: {fecha}")
                clientes_recientes_texto = (
                    "ULTIMOS 5 CLIENTES REGISTRADOS EN EL SISTEMA:\n" + "\n".join(lineas)
                )
        except Exception as e:
            print(f"Error clientes recientes: {e}")

    # ── Búsqueda de cliente si el mensaje lo indica ──
    clientes_texto      = ""
    _indicadores_cliente = [
        "cliente", "para ", "de parte", "a nombre", "factura", "facturar",
        "a credito", "fiado", "cuenta de",
    ]
    _menciona_cliente = any(ind in mensaje_usuario.lower() for ind in _indicadores_cliente)
    if _menciona_cliente:
        try:
            from excel import buscar_cliente_con_resultado
            # Extraer nombre despues de "para", "a nombre de", "de parte de", etc.
            _match_nombre = re.search(
                r'(?:para|a nombre de|de parte de|cuenta de)\s+([A-Za-záéíóúÁÉÍÓÚñÑ]+(?:\s+[A-Za-záéíóúÁÉÍÓÚñÑ]+){0,3})',
                mensaje_usuario, re.IGNORECASE
            )
            if _match_nombre:
                termino_cliente = _match_nombre.group(1).strip()
            else:
                palabras_cliente = [p for p in mensaje_usuario.lower().split()
                                    if len(p) > 3 and p not in stopwords]
                termino_cliente = " ".join(palabras_cliente[:4]) if palabras_cliente else ""
            if termino_cliente:
                cliente_unico, candidatos_cli = buscar_cliente_con_resultado(termino_cliente)

                if len(candidatos_cli) == 1:
                    c      = candidatos_cli[0]
                    nombre = c.get("Nombre tercero", "")
                    id_c   = c.get("Identificacion", "")
                    tipo   = c.get("Tipo de identificacion", "")
                    # Solo asignar si hay 2+ palabras en comun con el nombre buscado
                    palabras_buscadas    = set(_normalizar(termino_cliente).split())
                    palabras_encontradas = set(_normalizar(nombre).split())
                    coincidencias = palabras_buscadas & palabras_encontradas
                    if len(coincidencias) >= 2 or (len(palabras_buscadas) == 1 and coincidencias):
                        clientes_texto = (
                            f"CLIENTE ENCONTRADO EN EL SISTEMA (usar este directamente):\n"
                            f"  - {nombre} ({tipo}: {id_c})"
                        )
                    else:
                        # Coincidencia parcial — marcar para preguntar ANTES de confirmar
                        clientes_texto = (
                            f"CLIENTE NO IDENTIFICADO: usa exactamente \"cliente\": \"{termino_cliente}\" en el JSON. "
                            f"NO uses \"{nombre}\". El sistema preguntara si es cliente nuevo o existente."
                        )
                elif len(candidatos_cli) > 1:
                    lineas_cli = []
                    for c in candidatos_cli:
                        nombre = c.get("Nombre tercero", "")
                        id_c   = c.get("Identificacion", "")
                        tipo   = c.get("Tipo de identificacion", "")
                        lineas_cli.append(f"  - {nombre} ({tipo}: {id_c})")
                    clientes_texto = (
                        "MULTIPLES CLIENTES ENCONTRADOS — pregunta al usuario cual es:\n"
                        + "\n".join(lineas_cli)
                        + "\nEjemplo: 'Te refieres a NOMBRE1 (CC: 123) o NOMBRE2 (CC: 456)?'"
                    )
        except Exception:
            clientes_texto = ""

    # ── Inventario, caja y gastos ──
    palabras_inv     = ["inventario", "stock", "queda", "quedan", "hay", "cuanto hay", "existencia"]
    inventario_texto = (
        f"INVENTARIO ACTUAL:\n{json.dumps(cargar_inventario(), ensure_ascii=False)}"
        if any(p in mensaje_usuario.lower() for p in palabras_inv) else ""
    )

    palabras_caja_kw = ["caja", "gasto", "gastos", "apertura", "cierre", "efectivo", "cuanto hay en caja"]
    if any(p in mensaje_usuario.lower() for p in palabras_caja_kw):
        caja_texto   = f"ESTADO CAJA:\n{obtener_resumen_caja()}"
        gastos_texto = f"GASTOS DE HOY:\n{json.dumps(cargar_gastos_hoy(), ensure_ascii=False, default=str)}"
    else:
        caja_texto   = ""
        gastos_texto = ""

    aviso_drive = (
        "AVISO: Google Drive no disponible. Los datos se guardan localmente."
        if not config._get_drive_disponible() else ""
    )

    msg_l = mensaje_usuario.lower()

    # ── Acronal: precalcular total en Python ──
    acronal_calculado = ""
    _acronal_precio_kg = 13000
    _acronal_precio_medio = 7000
    # precios_fraccion esta en la RAIZ de memoria, no dentro del objeto catalogo.
    # El catalogo puede tener precio_unidad desactualizado — precios_fraccion manda.
    _frac_mem = memoria.get("precios_fraccion", {}).get("acronal", {})
    if _frac_mem:
        _v1 = _frac_mem.get("1", 0)
        _acronal_precio_kg = int(_v1) if _v1 else _acronal_precio_kg
        _vm = _frac_mem.get("1/2", 0)
        _acronal_precio_medio = int(_vm) if _vm else _acronal_precio_medio
    else:
        # fallback: precio_unidad del catalogo si no hay precios_fraccion definidos
        for _ak, _av in memoria.get("catalogo", {}).items():
            if "acronal" in _av.get("nombre_lower", ""):
                _pu = _av.get("precio_unidad", 0)
                if _pu:
                    _acronal_precio_kg = int(_pu)
                break
    if "acronal" in msg_l:
        # Normalizar "kilos y medio" -> "X.5", "medio kilo" -> "0.5"
        msg_ac = msg_l
        msg_ac = re.sub(r'(\d+)\s+(?:kilo[s]?\s+)?y\s+medio', lambda m: str(int(m.group(1))) + '.5', msg_ac)
        msg_ac = msg_ac.replace('medio kilo', '0.5').replace('kilo y medio', '1.5')
        # Buscar cantidad: "2-1/2", "2.5", "4", etc.
        # Detectar "1/2 kg" o "medio" antes del regex numerico
        if re.search(r'(?:^|\s)(?:1/2|medio)\s*(?:kilo[s]?|kg)?\s*(?:de\s+)?acronal|acronal\s*(?:1/2|medio)', msg_ac):
            acronal_calculado = f"ACRONAL PRECALCULADO: 0.5kg = ${_acronal_precio_medio:,} (precio especial). USA cantidad=0.5, total={_acronal_precio_medio} EXACTAMENTE."
            continue_ac = False
        else:
            continue_ac = True
        m_ac = re.search(r'([\d]+(?:[.,]\d+)?(?:-1/2|-1/4)?)\s*(?:kilo[s]?|kg)?\s*(?:de\s+)?acronal|acronal\s*(?:kilo[s]?|kg)?\s*([\d]+(?:[.,]\d+)?(?:-1/2|-1/4)?)', msg_ac) if continue_ac else None
        if m_ac:
            raw = (m_ac.group(1) or m_ac.group(2) or '').strip()
            raw = raw.replace(',', '.').replace('-1/2', '.5').replace('-1/4', '.25')
            try:
                kg = float(raw)
                enteros = int(kg)
                medio   = kg - enteros
                if abs(medio - 0.5) < 0.01:
                    total_ac = enteros * _acronal_precio_kg + _acronal_precio_medio
                elif abs(medio - 0.25) < 0.01:
                    total_ac = enteros * _acronal_precio_kg + round(_acronal_precio_medio / 2)
                else:
                    total_ac = round(kg * _acronal_precio_kg)
                acronal_calculado = (
                    f"ACRONAL PRECALCULADO: {kg}kg = ${total_ac:,} "
                    f"(1kg={_acronal_precio_kg},1/2kg={_acronal_precio_medio}). "
                    f"CRITICO: USA cantidad={kg}, total={total_ac} SIN MODIFICAR. PROHIBIDO recalcular."
                )
            except Exception:
                pass

    # ── Thinner y Varsol: precalcular fraccion en Python ──
    # CASO 1: "N litros/botellas de thinner/varsol" — el alias convirtió a "N litros/botellas thinner/varsol"
    # CASO 2: "thinner/varsol 8000" (precio por fracción de galón) — tabla normal.
    # Ambos productos usan la misma tabla de precios.
    thinner_calculado = ""
    _tabla_tv = {3000:"1/12",4000:"1/10",5000:"1/8",6000:"1/6",8000:"1/4",
                 10000:"1/3",13000:"1/2",20000:"3/4",26000:"1 galon"}
    _dec_tv   = {3000:1/12,4000:0.1,5000:0.125,6000:1/6,8000:0.25,
                 10000:1/3,13000:0.5,20000:0.75,26000:1.0}

    for _producto_tv in ("thinner", "varsol"):
        if _producto_tv not in msg_l:
            continue
        _m_litros   = re.search(rf'(\d+)\s+litros?\s+{_producto_tv}', msg_l)
        _m_botellas = re.search(rf'(\d+)\s+botellas?\s+{_producto_tv}', msg_l)
        if _m_litros:
            _n = int(_m_litros.group(1))
            _total_t = _n * 8000
            thinner_calculado += (
                f"{_producto_tv.upper()} PRECALCULADO: {_n} litro{'s' if _n > 1 else ''} = "
                f"${_total_t:,} total (cantidad={_n} litros, precio_litro=8000). "
                f"USA cantidad={_n}, total={_total_t} SIN MODIFICAR.\n"
            )
        elif _m_botellas:
            _n = int(_m_botellas.group(1))
            _total_t = _n * 4000
            thinner_calculado += (
                f"{_producto_tv.upper()} PRECALCULADO: {_n} botella{'s' if _n > 1 else ''} = "
                f"${_total_t:,} total (cantidad={_n} botellas, precio_botella=4000). "
                f"USA cantidad={_n}, total={_total_t} SIN MODIFICAR.\n"
            )
        else:
            # CASO 2: precio por fracción de galón
            _m_precio = re.search(
                rf'(\d[\d\.]*)\s*(?:de\s+)?{_producto_tv}|{_producto_tv}\s+(\d[\d\.]*)', msg_l
            )
            if _m_precio:
                precio_t = int(float(_m_precio.group(1) or _m_precio.group(2)))
                if precio_t in _tabla_tv:
                    frac_t = _tabla_tv[precio_t]
                    dec_t  = _dec_tv[precio_t]
                    thinner_calculado += (
                        f"{_producto_tv.upper()} PRECALCULADO: ${precio_t:,} = {frac_t} galon "
                        f"(cantidad={dec_t:.4f}, total={precio_t}). USA EXACTAMENTE estos valores.\n"
                    )
    thinner_calculado = thinner_calculado.strip()

    # ── Tornillos drywall: precalcular precio correcto ──
    tornillo_calculado = ""
    if "drywall" in msg_l or "tornillo" in msg_l:
        tabla_drywall = {
            "6x1/2":(25,25),"6x3/4":(58,30),"6x1":(38,35),"6x1-1/4":(42,40),
            "6x1-1/2":(58,55),"6x2":(67,60),"6x2-1/2":(75,70),"6x3":(83,80),
            "8x3/4":(33,30),"8x1":(38,35),"8x1-1/2":(58,55),"8x2":(67,60),"8x3":(83,80),
            "10x1":(83,70),"10x1-1/2":(125,100),"10x2":(150,120),"10x2-1/2":(167,160),
            "10x3":(167,160),"10x3-1/2":(208,200),"10x4":(208,200),
        }
        voz_medida = [
            ("3 y medio","3-1/2"),("3 y media","3-1/2"),("3½","3-1/2"),
            ("2 y medio","2-1/2"),("2 y media","2-1/2"),
            ("1 y medio","1-1/2"),("1 y media","1-1/2"),("1 y cuarto","1-1/4"),
        ]
        # Normalizar voz a fraccion
        msg_norm = msg_l
        for voz, frac in voz_medida:
            msg_norm = msg_norm.replace(voz, frac)
        m = re.search(r'(\d+)\s+tornillo[s]?\s+drywall\s+(\d+)\s+[xXpor]+\s+([\d\-/½]+)', msg_norm)
        if m:
            cant   = int(m.group(1))
            cal    = m.group(2)
            medida = m.group(3).strip()
            key    = f"{cal}x{medida}"
            if key in tabla_drywall:
                p1, p2 = tabla_drywall[key]
                precio_u = p1 if cant < 50 else p2
                total_t  = cant * precio_u
                tornillo_calculado = (
                    f"TORNILLO PRECALCULADO: {cant} TORNILLO DRYWALL {cal.upper()}X{medida.upper()} "
                    f"({'<' if cant < 50 else '>='} 50 uds → ${precio_u}/u) = total {total_t}. "
                    f"USA EXACTAMENTE estos valores."
                )

    # "DATOS HISTORICOS" solo se incluye cuando hay datos reales — no enviar "(no cargado)" innecesariamente
    datos_historicos_item = f"DATOS HISTORICOS:\n{datos_texto}" if datos_texto != "(no cargado)" else ""

    # Precios recién actualizados en RAM — override del cache de Anthropic (TTL 5 min)
    precios_modificados_texto = ""
    _pm_ram = _get_precios_recientes_activos()
    if _pm_ram and palabras_clave:
        overrides = []
        for clave_pm, val in _pm_ram.items():
            nombre_pm, _, frac = clave_pm.partition("___")
            if any(p in nombre_pm for p in palabras_clave):
                if frac:
                    overrides.append(f"{nombre_pm} {frac}={val}")
                else:
                    overrides.append(f"{nombre_pm}={val}")
        if overrides:
            precios_modificados_texto = "PRECIOS ACTUALIZADOS (usar estos, ignorar el catalogo):\n" + "\n".join(overrides)

    # Regla siempre activa en multi-producto: Claude debe listar lo que no pudo registrar
    _es_multiproducto = "\n" in mensaje_usuario.strip() or mensaje_usuario.count(",") >= 2
    # La regla aplica SIEMPRE (producto único o múltiple) para que el formato ⚠️
    # sea consistente y el sistema pueda guardarlo automáticamente en /pendientes.
    _regla_no_encontrado = (
        "REGLA MULTI-PRODUCTO: Para CADA producto en el mensaje, verifica si existe en el MATCH. "
        "Registra con [VENTA] SOLO los que SÍ encontraste con buena coincidencia. "
        "Si un producto NO tiene match claro en el catálogo (o el match es dudoso/diferente), "
        "NO lo registres y agrégalo a una línea final: "
        "⚠️ No encontré en catálogo: [nombre(s)]. "
        "NUNCA omitas en silencio productos no encontrados ni uses matches dudosos."
        if _es_multiproducto else
        "Si el MATCH está vacío o no coincide con el producto pedido, responde EXACTAMENTE: "
        "⚠️ No encontré en catálogo: [nombre del producto]. "
        "NUNCA uses otra frase como 'No tengo X' o 'X no está disponible'."
    )

    partes = [
        p for p in [
            precios_modificados_texto,
            info_fracciones_extra,
            acronal_calculado,
            thinner_calculado,
            tornillo_calculado,
            info_candidatos_extra,
            clientes_recientes_texto,
            clientes_texto,
            f"VENTAS MES:{resumen_texto}",
            datos_historicos_item,
            inventario_texto,
            caja_texto,
            gastos_texto,
            aviso_drive,
            f"Vendedor:{nombre_usuario}",
            _regla_no_encontrado,
            # Skills dinámicos: solo se inyectan cuando el mensaje los necesita
            skill_loader.obtener_skills_dinamicos(mensaje_usuario),
        ] if p
    ]
    return "\n\n".join(partes)


# ─────────────────────────────────────────────
# FUNCIÓN AUXILIAR: calcular historial adaptativo
# ─────────────────────────────────────────────

def _calcular_historial(mensaje: str) -> int:
    """
    Determina cuántos mensajes de historial enviar según el contexto.
    OPTIMIZACIÓN: ventas simples solo necesitan 1 mensaje, ahorrando ~100 tokens.
    """
    msg_l = mensaje.lower().strip()
    
    # Respuesta corta (1-2 palabras) = probablemente respuesta a pregunta anterior
    # Necesita contexto completo para no perder la venta original
    # Ej: "blanco", "rojo", "si", "no", "2 pulgadas"
    palabras = msg_l.split()
    if len(palabras) <= 2:
        return 4
    
    # Necesita contexto completo (cliente, correcciones, fiados)
    if any(k in msg_l for k in ("cliente", "fiado", "para ", "a nombre", 
                                 "corrig", "modific", "error", "equivoque",
                                 "cambia", "quita", "agrega")):
        return 4
    
    # Análisis, reportes o consultas complejas
    _kw_contexto = {"cuanto", "vendimos", "reporte", "analiz", "resumen", "estadistica",
                    "inventario", "grafica", "top", "mas vendido", "caja", "gasto"}
    if any(k in msg_l for k in _kw_contexto):
        return 4
    
    # Multi-producto (comas o saltos de línea)
    if "," in mensaje or mensaje.count("\n") > 0:
        return 2
    
    # Venta simple: solo el mensaje actual basta
    return 1


# ─────────────────────────────────────────────
# LLAMADA A CLAUDE CON PROMPT CACHING
# ─────────────────────────────────────────────

async def _llamar_claude_con_reintentos(cliente, max_tokens, system, messages, max_reintentos=5):
    """
    Wrapper para llamar a Claude con reintentos adicionales para error 529 (overloaded).
    El SDK ya hace 3 reintentos internos, pero agregamos una capa extra con backoff.
    """
    import random
    from anthropic import APIError
    
    ultimo_error = None
    for intento in range(max_reintentos):
        try:
            loop = asyncio.get_event_loop()
            respuesta = await asyncio.wait_for(
                loop.run_in_executor(
                    None,
                    lambda: cliente.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=max_tokens,
                        system=system,
                        messages=messages,
                    )
                ),
                timeout=45.0,  # timeout más generoso
            )
            return respuesta
        except asyncio.TimeoutError:
            ultimo_error = RuntimeError("La IA tardó demasiado en responder (>45s).")
            # No reintentar en timeout, probablemente es un problema de red
            if intento >= 2:
                raise ultimo_error
        except Exception as e:
            ultimo_error = e
            error_str = str(e).lower()
            # Solo reintentar en errores 529 (overloaded) o 503 (service unavailable)
            if "529" in str(e) or "overload" in error_str or "503" in str(e) or "unavailable" in error_str:
                if intento < max_reintentos - 1:
                    # Backoff exponencial con jitter: 2^intento + random(0-1) segundos
                    espera = (2 ** intento) + random.uniform(0, 1)
                    logging.getLogger("ferrebot.ai").warning(
                        f"[CLAUDE] Error 529/503, reintento {intento+1}/{max_reintentos} en {espera:.1f}s..."
                    )
                    await asyncio.sleep(espera)
                    continue
            # Otros errores: no reintentar
            raise
    
    # Si llegamos aquí, agotamos los reintentos
    raise ultimo_error or RuntimeError("Error desconocido al llamar a Claude")


async def procesar_con_claude(mensaje_usuario: str, nombre_usuario: str, historial_chat: list) -> str:
    # BYPASS PYTHON — ANTES de alias_ferreteria (que transforma fracciones y rompería el match)
    # Solo se aplican aliases DINÁMICOS (simples word-substitutions: tiner→thinner, etc.)
    # El mensaje llega como "{vendedor}: {texto}" — stripear prefijo antes del bypass
    import re as _re
    _msg_bypass = _re.sub(r'^[^:]+:\s*', '', mensaje_usuario).strip()
    _msg_bypass = alias_manager.aplicar_aliases_dinamicos(_msg_bypass)
    memoria = cargar_memoria()
    _bypass = bypass.intentar_bypass_python(_msg_bypass, memoria.get("catalogo", {}))
    if _bypass:
        import json as _jbp
        _txt, _venta = _bypass
        # Multi-producto: expandir a múltiples tags [VENTA]
        if _venta.get("multi"):
            _tags = ""
            for _item in _venta.get("items", []):
                _v = {
                    "producto":        _item["producto"],
                    "cantidad":        _item["cantidad"],
                    "total":           _item["total"],
                    "precio_unitario": _item["precio_unitario"],
                    "metodo_pago":     "",
                }
                _tags += f"[VENTA]{_jbp.dumps(_v, ensure_ascii=False)}[/VENTA]"
            return f"{_txt}\n{_tags}"
        # Single producto
        return f"{_txt}\n[VENTA]{_jbp.dumps(_venta, ensure_ascii=False)}[/VENTA]"

    logging.getLogger("ferrebot.ai").info(f"[→ CLAUDE] '{_msg_bypass[:60]}'")

    # Alias solo para Claude — después de que bypass descartó el mensaje
    mensaje_usuario = aplicar_alias_ferreteria(mensaje_usuario)

    parte_estatica = _construir_parte_estatica(memoria)
    parte_dinamica = _construir_parte_dinamica(mensaje_usuario, nombre_usuario, memoria)

    # BLOQUEO PYTHON: si el MATCH está vacío y el mensaje parece una venta
    # (no es consulta, no es reporte), responder directamente sin llamar a Claude.
    # Esto evita que el bot registre productos inexistentes con total:0.
    _SEÑAL_MATCH_VACIO = "MATCH: (sin resultados — producto no encontrado en catalogo)"
    _kw_no_venta = {"cuanto","vendimos","reporte","analiz","resumen","estadistica",
                    "top","mas vendido","gasto","caja","inventario","cliente",
                    "precio","vale","cuesta","cuanto vale","hay","stock","quedan"}
    _es_consulta = any(p in mensaje_usuario.lower() for p in _kw_no_venta)

    if _SEÑAL_MATCH_VACIO in parte_dinamica and not _es_consulta:
        # Extraer nombre del producto del mensaje para respuesta clara
        _msg_limpio = mensaje_usuario.strip().lower()
        # Quitar cantidades y unidades del inicio para aislar el nombre
        _msg_limpio = re.sub(r'^[\d\s/\.]+', '', _msg_limpio).strip()
        _msg_limpio = re.sub(r'^(kilo|kilos|galon|galones|metro|metros|unidad|unidades|litro|litros)\s*', '', _msg_limpio).strip()
        return f"No tengo {_msg_limpio} en el catálogo."

    _modo = "MATCH+SIMPLE-CAT 💡"  # fracciones en MATCH, precio_unidad en estático

    # Historial adaptativo: usa _calcular_historial para determinar cuántos mensajes
    _n_hist = _calcular_historial(mensaje_usuario)

    messages = []
    for msg in historial_chat[-_n_hist:]:
        if isinstance(msg, dict) and "role" in msg and "content" in msg:
            messages.append({"role": str(msg["role"]), "content": str(msg["content"])})
    messages.append({"role": "user", "content": str(mensaje_usuario)})

    # max_tokens adaptativo por tipo de mensaje:
    # - Venta simple (1 producto, sin comas ni saltos): solo JSON → 400 tok
    # - Venta multi-producto: JSON × N productos + posible texto → 250 × lineas
    # - Consulta/reporte/modificacion: respuesta larga → 2000 mínimo
    _kw_reporte = {"cuanto","vendimos","reporte","analiz","resumen","estadistica",
                   "grafica","top","mas vendido","gasto","caja","inventario"}
    _kw_edicion = {"modificar","corregir","cambia","quita","agrega","error",
                   "equivoque","fiado","debe","abono","borrar","eliminar"}
    num_lineas = mensaje_usuario.count("\n") + mensaje_usuario.count(",") + 1
    _msg_low   = mensaje_usuario.lower()
    if any(p in _msg_low for p in _kw_reporte):
        max_tokens = 2000          # reportes necesitan espacio
    elif any(p in _msg_low for p in _kw_edicion):
        max_tokens = 1200          # ediciones: algo de texto + JSON
    elif num_lineas == 1 and "," not in mensaje_usuario:
        max_tokens = 450           # venta simple: solo JSON, ~150 tok reales
    else:
        max_tokens = min(3000, max(800, num_lineas * 220))  # multi-producto

    system = [
        {
            "type": "text",
            "text": parte_estatica,
            "cache_control": {"type": "ephemeral"},
        },
        {
            "type": "text",
            "text": parte_dinamica,
        },
    ]

    respuesta = await _llamar_claude_con_reintentos(
        config.claude_client, max_tokens, system, messages
    )

    # ── Log de uso de tokens y cache ──
    uso = respuesta.usage
    cache_read    = getattr(uso, "cache_read_input_tokens",    0) or 0
    cache_created = getattr(uso, "cache_creation_input_tokens", 0) or 0
    input_normal  = getattr(uso, "input_tokens",               0) or 0
    output_tokens = getattr(uso, "output_tokens",              0) or 0

    if cache_read > 0 or cache_created > 0:
        costo_input   = (input_normal  / 1_000_000) * 1.00
        costo_cached  = (cache_read    / 1_000_000) * 0.10
        costo_created = (cache_created / 1_000_000) * 1.25
        costo_output  = (output_tokens / 1_000_000) * 5.00
        costo_total   = costo_input + costo_cached + costo_created + costo_output
        logging.getLogger("ferrebot.cache").info(
            f"[CACHE] ✅ hit={cache_read} tok | created={cache_created} tok | "
            f"input={input_normal} tok | output={output_tokens} tok | "
            f"costo≈${costo_total:.5f}"
        )
    else:
        logging.getLogger("ferrebot.cache").warning(
            f"[CACHE] ⚠️ SIN CACHE — input={input_normal} tok | output={output_tokens} tok"
        )

    return respuesta.content[0].text

# ─────────────────────────────────────────────
# PARSEO Y EJECUCIÓN DE ACCIONES
# ─────────────────────────────────────────────

def procesar_acciones(texto_respuesta: str, vendedor: str, chat_id: int) -> tuple[str, list, list]:
    from ventas_state import ventas_pendientes, registrar_ventas_con_metodo, _estado_lock, mensajes_standby

    acciones:       list[str] = []
    archivos_excel: list[str] = []
    texto_limpio = texto_respuesta

    # ── Ventas ──
    ventas_con_metodo = []
    ventas_sin_metodo = []

    with _estado_lock:
        esperando_pago = bool(ventas_pendientes.get(chat_id))

    # ── Helper: conversión para productos vendidos por mililitro (MLT) ──────
    def _convertir_venta_mlt(venta: dict) -> dict:
        """
        Para productos con unidad_medida='MLT', maneja tres casos:

        CASO 1 — cantidad viene en tarros (entero pequeño ≤ 20):
          Si total ≈ cantidad × precio_1000ml → son tarros → × 1000
          Ej: {"cantidad":1,"total":26000} precio_ml=26 → cantidad=1000

        CASO 2 — cantidad viene en pesos (bot repitió mismo número):
          Si cantidad == total → ml = total / precio_por_ml
          Ej: {"cantidad":2000,"total":2000} → cantidad=76.9

        CASO 3 — cantidad ya en ml → no tocar.
          Ej: {"cantidad":500,"total":13000} precio_ml=26 → 500×26=13000 ✅
        """
        try:
            prod = buscar_producto_en_catalogo(venta.get("producto", ""))
            if not prod:
                return venta
            if prod.get("unidad_medida") != "MLT":
                return venta

            precio_por_ml = prod.get("precio_unidad", 0)
            if not precio_por_ml:
                return venta

            cantidad = float(venta.get("cantidad", 1))
            total    = float(venta.get("total", 0))

            if total <= 0:
                return venta

            # ── CASO 1: cantidad parece tarros (entero ≤ 20, total ≈ tarros × 1000ml × precio) ──
            precio_tarro = precio_por_ml * 1000
            if (cantidad <= 20
                    and cantidad == int(cantidad)
                    and abs(total - cantidad * precio_tarro) / max(total, 1) < 0.05):
                ml = int(cantidad * 1000)
                venta = dict(venta)
                venta["cantidad"] = ml
                logging.getLogger("ferrebot.ai").info(
                    "[MLT] Tarros→ml: %s | %d tarro(s) → %d ml | $%.0f",
                    prod.get("nombre"), int(cantidad), ml, total
                )
                return venta

            # ── CASO 2: cantidad == total o parece monto en pesos ──
            cantidad_parece_pesos = (
                cantidad == total
                or (cantidad > precio_por_ml * 10 and cantidad % 500 == 0)
            )
            if cantidad_parece_pesos:
                ml = round(total / precio_por_ml, 1)
                venta = dict(venta)
                venta["cantidad"] = ml
                logging.getLogger("ferrebot.ai").info(
                    "[MLT] Pesos→ml: %s | $%.0f / $%.2f por ml = %.1f ml",
                    prod.get("nombre"), total, precio_por_ml, ml
                )
                return venta

            # ── CASO 3: cantidad ya en ml → no tocar ──

        except Exception as e:
            logging.getLogger("ferrebot.ai").warning("[MLT] Error conversión: %s", e)
        return venta

    for venta_json in re.findall(r'\[VENTA\](.*?)\[/VENTA\]', texto_respuesta, re.DOTALL):
        try:
            if esperando_pago:
                print(f"[VENTA] ignorado — esperando selección de pago para chat {chat_id}")
            else:
                venta = json.loads(venta_json.strip())
                logging.getLogger("ferrebot.ai").debug(f"[VENTA] JSON recibido: {venta}")
                # Aplicar conversión ml si aplica
                venta = _convertir_venta_mlt(venta)
                if venta.get("metodo_pago"):
                    ventas_con_metodo.append(venta)
                else:
                    ventas_sin_metodo.append(venta)
        except Exception as e:
            logging.getLogger("ferrebot.ai").error(f"Error parseando venta: {e} | JSON raw: {repr(venta_json.strip())}")
        texto_limpio = texto_limpio.replace(f'[VENTA]{venta_json}[/VENTA]', '')

    if esperando_pago and ventas_con_metodo:
        ventas_con_metodo.clear()

    def _tiene_cliente_desconocido(ventas: list) -> str | None:
        from excel import buscar_cliente_con_resultado
        for v in ventas:
            nombre_cliente = v.get("cliente", "").strip()
            if not nombre_cliente or nombre_cliente.lower() in ("consumidor final", "cf", ""):
                continue
            try:
                _, candidatos = buscar_cliente_con_resultado(nombre_cliente)
                if not candidatos:
                    return nombre_cliente
                # Verificar que algún candidato coincida con al menos 2 palabras
                palabras_buscadas = set(_normalizar(nombre_cliente).split())
                match_exacto = False
                for c in candidatos:
                    palabras_encontradas = set(_normalizar(c.get("Nombre tercero", "")).split())
                    coincidencias = palabras_buscadas & palabras_encontradas
                    if len(coincidencias) >= 2 or (len(palabras_buscadas) == 1 and coincidencias):
                        match_exacto = True
                        break
                if not match_exacto:
                    return nombre_cliente
            except Exception:
                pass
        return None

    todas_las_ventas_nuevas = ventas_con_metodo + ventas_sin_metodo
    cliente_desconocido     = _tiene_cliente_desconocido(todas_las_ventas_nuevas) if todas_las_ventas_nuevas else None

    if cliente_desconocido and not esperando_pago:
        with _estado_lock:
            ventas_pendientes[chat_id] = todas_las_ventas_nuevas
        acciones.append(f"CLIENTE_DESCONOCIDO:{cliente_desconocido}")
        ventas_con_metodo.clear()
        ventas_sin_metodo.clear()

    if ventas_con_metodo:
        metodo_conocido = ventas_con_metodo[0].get("metodo_pago", "efectivo").lower()
        with _estado_lock:
            ventas_pendientes[chat_id] = ventas_con_metodo
        acciones.append(f"PEDIR_CONFIRMACION:{metodo_conocido}")

    ventas_ignoradas = esperando_pago and bool(
        re.findall(r'\[VENTA\](.*?)\[/VENTA\]', texto_respuesta, re.DOTALL)
    )
    if ventas_ignoradas or (ventas_sin_metodo and esperando_pago):
        acciones.append("PAGO_PENDIENTE_AVISO")
    elif ventas_sin_metodo:
        with _estado_lock:
            ventas_pendientes[chat_id] = ventas_sin_metodo
        acciones.append("PEDIR_METODO_PAGO")

    # ── Cliente nuevo (datos completos) ──
    for cli_json in re.findall(r'\[CLIENTE_NUEVO\](.*?)\[/CLIENTE_NUEVO\]', texto_respuesta, re.DOTALL):
        try:
            datos  = json.loads(cli_json.strip())
            nombre = datos.get("nombre", "").strip()
            id_num = str(datos.get("identificacion", "")).strip()
            if nombre and id_num:
                ok = guardar_cliente_nuevo(
                    nombre, datos.get("tipo_id", "Cedula de ciudadania"), id_num,
                    datos.get("tipo_persona", "Natural"),
                    datos.get("correo", ""), datos.get("telefono", ""),
                )
                acciones.append(
                    f"Cliente creado: {nombre.upper()} — {datos.get('tipo_id','')}: {id_num}"
                    if ok else f"No pude guardar el cliente {nombre}."
                )
        except Exception as e:
            logging.getLogger("ferrebot.ai").error(f"Error cliente nuevo: {e}")
        texto_limpio = texto_limpio.replace(f'[CLIENTE_NUEVO]{cli_json}[/CLIENTE_NUEVO]', '')

    # ── Iniciar flujo paso a paso de cliente ──
    for ini_json in re.findall(r'\[INICIAR_CLIENTE\](.*?)\[/INICIAR_CLIENTE\]', texto_respuesta, re.DOTALL):
        try:
            datos  = json.loads(ini_json.strip())
            nombre = datos.get("nombre", "").strip()
            from ventas_state import clientes_en_proceso, ventas_esperando_cliente, _estado_lock as _lock
            with _lock:
                clientes_en_proceso[chat_id] = {
                    "nombre":         nombre,
                    "tipo_id":        None,
                    "identificacion": None,
                    "tipo_persona":   None,
                    "correo":         None,
                    "paso":           "nombre" if not nombre else "tipo_id",
                    "vendedor":       vendedor,
                }
                if chat_id in ventas_pendientes and ventas_pendientes[chat_id]:
                    ventas_esperando_cliente[chat_id] = {
                        "ventas":   ventas_pendientes.pop(chat_id),
                        "metodo":   None,
                        "vendedor": vendedor,
                    }
                elif ventas_sin_metodo:
                    ventas_esperando_cliente[chat_id] = {
                        "ventas":   list(ventas_sin_metodo),
                        "metodo":   None,
                        "vendedor": vendedor,
                    }
                    ventas_sin_metodo.clear()
            acciones.append("INICIAR_FLUJO_CLIENTE")
        except Exception as e:
            print(f"Error iniciando flujo cliente: {e}")
        texto_limpio = texto_limpio.replace(f'[INICIAR_CLIENTE]{ini_json}[/INICIAR_CLIENTE]', '')

    # ── Borrar cliente ──
    for bc_json in re.findall(r'\[BORRAR_CLIENTE\](.*?)\[/BORRAR_CLIENTE\]', texto_respuesta, re.DOTALL):
        try:
            datos  = json.loads(bc_json.strip())
            nombre = datos.get("nombre", "").strip()
            if nombre:
                from excel import borrar_cliente
                exito, msg = borrar_cliente(nombre)
                acciones.append(msg)
        except Exception as e:
            print(f"Error borrando cliente: {e}")
        texto_limpio = texto_limpio.replace(f'[BORRAR_CLIENTE]{bc_json}[/BORRAR_CLIENTE]', '')

    # ── Precio fraccion ──
    for pf_json in re.findall(r'\[PRECIO_FRACCION\](.*?)\[/PRECIO_FRACCION\]', texto_respuesta, re.DOTALL):
        try:
            datos    = json.loads(pf_json.strip())
            producto = datos.get("producto", "").strip()
            fraccion = datos.get("fraccion", "").strip()
            precio   = float(datos.get("precio", 0))
            if producto and fraccion and precio:
                # Intentar actualizar en catálogo (fuente única de verdad)
                en_cat = actualizar_precio_en_catalogo(producto, precio, fraccion)
                if en_cat:
                    # Override RAM 5 min + encolar Excel vía precio_sync (sin hilo manual)
                    _pf_prod = buscar_producto_en_catalogo(producto)
                    _pf_key  = _pf_prod.get("nombre_lower", producto.lower()) if _pf_prod else producto.lower()
                    _registrar_precio_reciente(_pf_key, precio, fraccion)
                    invalidar_cache_memoria()
                    from precio_sync import actualizar_precio as _ap_frac
                    _ap_frac(producto, precio, fraccion)  # encola Excel internamente
                else:
                    # Producto no en catálogo: guardar en precios_fraccion como fallback
                    mem = cargar_memoria()
                    mem.setdefault("precios_fraccion", {}).setdefault(producto.lower(), {})[fraccion] = round(precio)
                    guardar_memoria(mem, urgente=True)
                acciones.append(f"Precio de fracción guardado: {producto} {fraccion} = ${precio:,.0f}")
        except Exception as e:
            print(f"Error precio fraccion: {e}")
        texto_limpio = texto_limpio.replace(f'[PRECIO_FRACCION]{pf_json}[/PRECIO_FRACCION]', '')

    # ── Precio ──
    for precio_json in re.findall(r'\[PRECIO\](.*?)\[/PRECIO\]', texto_respuesta, re.DOTALL):
        try:
            datos    = json.loads(precio_json.strip())
            producto = datos["producto"]
            precio   = float(datos["precio"])
            fraccion = datos.get("fraccion")  # opcional: "1/4", "1/2", etc.

            # Actualizar en catálogo (fuente única de verdad) + encolar Excel
            from precio_sync import actualizar_precio as _ap_precio
            en_catalogo, _ = _ap_precio(producto, precio, fraccion)

            # Override RAM 5 min
            prod_encontrado = buscar_producto_en_catalogo(producto)
            nombre_lower_pc = prod_encontrado.get("nombre_lower", producto.lower()) if prod_encontrado else producto.lower()
            _registrar_precio_reciente(nombre_lower_pc, precio, fraccion)
            invalidar_cache_memoria()

            if fraccion:
                acciones.append(f"🧠 Precio actualizado: {producto} {fraccion} = ${precio:,.0f}")
            else:
                acciones.append(f"🧠 Precio actualizado: {producto} = ${precio:,.0f}")
        except Exception as e:
            print(f"Error precio: {e}")
        texto_limpio = texto_limpio.replace(f'[PRECIO]{precio_json}[/PRECIO]', '')

    # ── Precio mayorista (tornillería) ──
    for pm_json in re.findall(r'\[PRECIO_MAYORISTA\](.*?)\[/PRECIO_MAYORISTA\]', texto_respuesta, re.DOTALL):
        try:
            datos       = json.loads(pm_json.strip())
            producto    = datos["producto"]
            p_unidad    = float(datos.get("precio_unidad", 0) or 0)
            p_mayorista = float(datos.get("precio_mayorista", 0) or 0)
            umbral      = int(datos.get("umbral", 50))

            prod = buscar_producto_en_catalogo(producto)
            if not prod:
                acciones.append(f"⚠️ Producto no encontrado: {producto}")
            else:
                from memoria import cargar_memoria, guardar_memoria, invalidar_cache_memoria as _inv
                mem = cargar_memoria()
                cat = mem.get("catalogo", {})
                clave = next((k for k, v in cat.items() if v.get("nombre_lower") == prod.get("nombre_lower")), None)
                if clave:
                    if p_unidad > 0:
                        cat[clave]["precio_unidad"] = round(p_unidad)
                    pxc = cat[clave].get("precio_por_cantidad", {})
                    if p_unidad > 0:
                        pxc["precio_bajo_umbral"] = round(p_unidad)
                    if p_mayorista > 0:
                        pxc["precio_sobre_umbral"] = round(p_mayorista)
                    if umbral:
                        pxc["umbral"] = umbral
                    cat[clave]["precio_por_cantidad"] = pxc
                    mem["catalogo"] = cat
                    guardar_memoria(mem, urgente=True)
                    _inv()
                    # Encolar Excel para precio unidad
                    if p_unidad > 0:
                        from precio_sync import actualizar_precio as _ap_m
                        _ap_m(producto, p_unidad, None)
                    nombre_display = prod.get("nombre", producto)
                    msg = f"🧠 {nombre_display}: unidad=${p_unidad:,.0f}" if p_unidad else f"🧠 {nombre_display}"
                    if p_mayorista > 0:
                        msg += f" | mayorista ×{umbral}=${p_mayorista:,.0f}"
                    acciones.append(msg)
        except Exception as e:
            print(f"Error precio_mayorista: {e}")
        texto_limpio = texto_limpio.replace(f'[PRECIO_MAYORISTA]{pm_json}[/PRECIO_MAYORISTA]', '')

    # ── Código producto ──
    for cp_json in re.findall(r'\[CODIGO_PRODUCTO\](.*?)\[/CODIGO_PRODUCTO\]', texto_respuesta, re.DOTALL):
        try:
            datos  = json.loads(cp_json.strip())
            nombre = datos.get("producto", "").strip()
            codigo = datos.get("codigo", "").strip()
            if nombre and codigo:
                mem      = cargar_memoria()
                catalogo = mem.get("catalogo", {})
                prod     = buscar_producto_en_catalogo(nombre)
                if prod:
                    for k, v in catalogo.items():
                        if v.get("nombre_lower") == prod.get("nombre_lower"):
                            catalogo[k]["codigo"] = codigo
                            break
                    mem["catalogo"] = catalogo
                    guardar_memoria(mem)
                    acciones.append(f"Código guardado: {nombre} = {codigo}")
        except Exception as e:
            print(f"Error código producto: {e}")
        texto_limpio = texto_limpio.replace(f'[CODIGO_PRODUCTO]{cp_json}[/CODIGO_PRODUCTO]', '')

    # ── Negocio ──
    for neg_json in re.findall(r'\[NEGOCIO\](.*?)\[/NEGOCIO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(neg_json.strip())
            mem   = cargar_memoria()
            mem["negocio"].update(datos)
            guardar_memoria(mem)
        except Exception as e:
            print(f"Error negocio: {e}")
        texto_limpio = texto_limpio.replace(f'[NEGOCIO]{neg_json}[/NEGOCIO]', '')

    # ── Caja ──
    for caja_json in re.findall(r'\[CAJA\](.*?)\[/CAJA\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(caja_json.strip())
            caja  = cargar_caja()
            if datos.get("accion") == "apertura":
                caja.update({
                    "abierta": True,
                    "fecha":   datetime.now(config.COLOMBIA_TZ).strftime("%Y-%m-%d"),
                    "monto_apertura": float(datos.get("monto", 0)),
                    "efectivo": 0, "transferencias": 0, "datafono": 0,
                })
                from memoria import guardar_caja
                guardar_caja(caja)
                acciones.append(f"Caja abierta con ${float(datos.get('monto', 0)):,.0f}")
            elif datos.get("accion") == "cierre":
                acciones.append(f"Caja cerrada.\n{obtener_resumen_caja()}")
                caja["abierta"] = False
                from memoria import guardar_caja
                guardar_caja(caja)
        except Exception as e:
            print(f"Error caja: {e}")
        texto_limpio = texto_limpio.replace(f'[CAJA]{caja_json}[/CAJA]', '')

    # ── Gastos ──
    for gasto_json in re.findall(r'\[GASTO\](.*?)\[/GASTO\]', texto_respuesta, re.DOTALL):
        try:
            datos = json.loads(gasto_json.strip())
            gasto = {
                "concepto":  datos.get("concepto", ""),
                "monto":     float(datos.get("monto", 0)),
                "categoria": datos.get("categoria", "varios"),
                "origen":    datos.get("origen", "externo"),
                "hora":      datetime.now(config.COLOMBIA_TZ).strftime("%H:%M"),
            }
            guardar_gasto(gasto)
            acciones.append(f"Gasto registrado: {gasto['concepto']} — ${gasto['monto']:,.0f} ({gasto['origen']})")
        except Exception as e:
            print(f"Error gasto: {e}")
        texto_limpio = texto_limpio.replace(f'[GASTO]{gasto_json}[/GASTO]', '')

    # ── Fiado ──
    for fiado_json in re.findall(r'\[FIADO\](.*?)\[/FIADO\]', texto_respuesta, re.DOTALL):
        try:
            datos    = json.loads(fiado_json.strip())
            cliente  = datos.get("cliente", "").strip()
            concepto = datos.get("concepto", "")
            cargo    = float(datos.get("cargo", 0))
            abono    = float(datos.get("abono", 0))
            if cliente and cargo > 0:
                saldo = guardar_fiado_movimiento(cliente, concepto, cargo, abono)
                from excel import registrar_fiado_en_excel
                registrar_fiado_en_excel(cliente, concepto, cargo, abono, saldo)
                acciones.append(f"Fiado registrado: {cliente} debe ${saldo:,.0f}")
        except Exception as e:
            print(f"Error fiado: {e}")
        texto_limpio = texto_limpio.replace(f'[FIADO]{fiado_json}[/FIADO]', '')

    # ── Abono fiado ──
    for abono_json in re.findall(r'\[ABONO_FIADO\](.*?)\[/ABONO_FIADO\]', texto_respuesta, re.DOTALL):
        try:
            datos   = json.loads(abono_json.strip())
            cliente = datos.get("cliente", "").strip()
            monto   = float(datos.get("monto", 0))
            if cliente and monto > 0:
                ok, msg = abonar_fiado(cliente, monto)
                if ok:
                    from excel import registrar_fiado_en_excel
                    from memoria import cargar_fiados
                    fiados      = cargar_fiados()
                    cliente_key = next((k for k in fiados if k.lower() in cliente.lower() or cliente.lower() in k.lower()), cliente)
                    saldo       = fiados.get(cliente_key, {}).get("saldo", 0)
                    registrar_fiado_en_excel(cliente_key, "Abono", 0, monto, saldo)
                acciones.append(msg)
        except Exception as e:
            print(f"Error abono fiado: {e}")
        texto_limpio = texto_limpio.replace(f'[ABONO_FIADO]{abono_json}[/ABONO_FIADO]', '')

    # ── Inventario ──
    for inv_json in re.findall(r'\[INVENTARIO\](.*?)\[/INVENTARIO\]', texto_respuesta, re.DOTALL):
        try:
            datos      = json.loads(inv_json.strip())
            inventario = cargar_inventario()
            producto   = datos.get("producto", "").lower()
            accion     = datos.get("accion", "actualizar")
            if accion == "actualizar":
                cantidad = convertir_fraccion_a_decimal(datos.get("cantidad", 0))
                minimo   = convertir_fraccion_a_decimal(datos.get("minimo", 0.5))
                unidad   = datos.get("unidad", "unidades")
                inventario[producto] = {
                    "cantidad": cantidad, "minimo": minimo, "unidad": unidad,
                    "nombre_original": datos.get("producto", producto),
                }
                from memoria import guardar_inventario as _guardar_inv
                _guardar_inv(inventario)
                acciones.append(f"Inventario: {datos['producto']} — {decimal_a_fraccion_legible(cantidad)} {unidad}")
            elif accion == "descontar" and producto in inventario:
                descuento = convertir_fraccion_a_decimal(datos.get("cantidad", 0))
                inventario[producto]["cantidad"] = max(0, inventario[producto]["cantidad"] - descuento)
                from memoria import guardar_inventario as _guardar_inv
                _guardar_inv(inventario)
            from memoria import verificar_alertas_inventario
            acciones.extend(verificar_alertas_inventario())
        except Exception as e:
            print(f"Error inventario: {e}")
        texto_limpio = texto_limpio.replace(f'[INVENTARIO]{inv_json}[/INVENTARIO]', '')

    # ── Excel personalizado ──
    for excel_json in re.findall(r'\[EXCEL\](.*?)\[/EXCEL\]', texto_respuesta, re.DOTALL):
        try:
            datos  = json.loads(excel_json.strip())
            nombre = f"reporte_{datetime.now(config.COLOMBIA_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
            generar_excel_personalizado(
                datos.get("titulo", "Reporte"),
                datos.get("encabezados", []),
                datos.get("filas", []),
                nombre,
            )
            archivos_excel.append(nombre)
        except Exception as e:
            print(f"Error generando Excel: {e}")
        texto_limpio = texto_limpio.replace(f'[EXCEL]{excel_json}[/EXCEL]', '')

    return texto_limpio.strip(), acciones, archivos_excel

# ─────────────────────────────────────────────
# EDICIÓN DE EXCEL CON CLAUDE
# ─────────────────────────────────────────────

async def editar_excel_con_claude(instruccion: str, ruta_excel: str, nombre_excel: str,
                                   vendedor: str, chat_id: int) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(ruta_excel)
    info_hojas = []
    for hoja_nombre in wb.sheetnames:
        ws = wb[hoja_nombre]
        encabezados   = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        filas_ejemplo = []
        for fila in ws.iter_rows(
            min_row=config.EXCEL_FILA_DATOS,
            max_row=min(config.EXCEL_FILA_DATOS + 2, ws.max_row),
            values_only=True,
        ):
            filas_ejemplo.append(list(fila))
        info_hojas.append({
            "hoja": hoja_nombre, "encabezados": encabezados,
            "ejemplo_filas": filas_ejemplo, "total_filas": ws.max_row - 1,
        })

    prompt = f"""Eres un experto en Python y openpyxl. El usuario tiene un archivo Excel llamado '{nombre_excel}' con esta estructura:

{json.dumps(info_hojas, ensure_ascii=False, default=str)}

El usuario quiere: {instruccion}

Genera SOLO el código Python necesario para modificar el archivo usando openpyxl.
- El archivo ya está cargado, usa: wb = openpyxl.load_workbook('{ruta_excel}')
- Al final guarda con: wb.save('{ruta_excel}')
- Usa colores en formato hex sin # (ej: 'FF0000' para rojo)
- Solo tienes disponibles: openpyxl y json. NO uses os, sys, subprocess ni ninguna otra librería.
- Solo el código, sin explicaciones ni comentarios ni bloques ```
- Si la instrucción no tiene sentido para un Excel, devuelve solo la palabra: IMPOSIBLE"""

    loop     = asyncio.get_event_loop()
    respuesta = await loop.run_in_executor(
        None,
        lambda: config.claude_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}],
        )
    )
    codigo = respuesta.content[0].text.strip()
    if "```python" in codigo:
        codigo = codigo.split("```python")[1].split("```")[0].strip()
    elif "```" in codigo:
        codigo = codigo.split("```")[1].split("```")[0].strip()
    return codigo

# ─────────────────────────────────────────────
# VERSIÓN ASYNC DE PROCESAR_ACCIONES
# ─────────────────────────────────────────────

async def procesar_acciones_async(texto_respuesta: str, vendedor: str, chat_id: int) -> tuple[str, list, list]:
    """
    Wrapper async de procesar_acciones para compatibilidad con handlers async.
    Ejecuta procesar_acciones en un executor para no bloquear el event loop.
    """
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        None,
        lambda: procesar_acciones(texto_respuesta, vendedor, chat_id)
    )
